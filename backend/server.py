from voice_ai_service import get_voice_ai_service
from whatsapp_service import send_whatsapp_template, detect_language_from_phone
from funnel_registry import is_frontend_owned, is_valid_event
from funnel_service import ensure_funnel_indexes, record_funnel_event, track_webhook_event
from aha_helpers import pick_aha_slot, send_aha_whatsapp_with_retry
from meta_capi_service import get_meta_capi_service, split_full_name
from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Request, Response, File, UploadFile, Form, Query, BackgroundTasks
from fastapi.responses import JSONResponse, HTMLResponse, RedirectResponse
from fastapi.encoders import jsonable_encoder
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from pathlib import Path
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import asyncio
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Literal, Dict
import uuid
from datetime import datetime, timezone, timedelta
import requests
from zoneinfo import ZoneInfo
import re
import xml.etree.ElementTree as ET
import hashlib
import hmac
import base64
import json
import secrets
import io
import math
from PIL import Image as PilImage
import openpyxl
import csv

from contextlib import asynccontextmanager
from passlib.context import CryptContext
from jose import JWTError, jwt
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.interval import IntervalTrigger
import socketio
import sib_api_v3_sdk
from sib_api_v3_sdk.rest import ApiException
from pywebpush import webpush, WebPushException

# (Cache ve Rate Limit importları, sizin projenizden alındı)
from cache import init_redis, invalidate_cache, cache_result
from rate_limit import initialize_limiter, rate_limit, LIMITS
from slowapi.errors import RateLimitExceeded
from slowapi import _rate_limit_exceeded_handler

import firebase_admin
from firebase_admin import credentials, messaging

# === SENTRY ERROR TRACKING ===
import sentry_sdk
sentry_sdk.init(
    dsn=os.environ.get("SENTRY_DSN", ""),
    send_default_pii=True,
    traces_sample_rate=0.2,  # %20 performance tracing (production için yeterli)
    profiles_sample_rate=0.1,
    environment=os.environ.get("SENTRY_ENV", "production"),
)

# Firebase Initialization moved to lifespan() for proper logging

# AI Service Import
import ai_service
from ai_service import chat_with_ai

# === LOGGING AYARLARI ===
logging.basicConfig(
    level=logging.INFO,  # INFO level for production
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('/tmp/backend.log')
    ],
    force=True  # Mevcut yapılandırmayı zorla güncelle
)
logger = logging.getLogger("server")  # Logger adını sabit yap
logger.setLevel(logging.INFO)  # Logger seviyesini açıkça ayarla
# Enable socketio server logging
logging.getLogger('socketio.server').setLevel(logging.INFO)
logging.getLogger('engineio.server').setLevel(logging.INFO)

# === GÜVENLİK AYARLARI ===
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'default_karmaşık_bir_secret_key_ekleyin_mutlaka')
if SECRET_KEY == 'default_karmaşık_bir_secret_key_ekleyin_mutlaka':
    logging.warning("WARNING: JWT_SECRET_KEY is using default value! Please set a secure secret key in production.")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24
ACCESS_TOKEN_EXPIRE_MINUTES_REMEMBER = 60 * 24 * 30
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/token")

# --- ROOT DİZİN VE .ENV YÜKLEME ---
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# --- SABİT SMS AYARLARI ---
ILETIMERKEZI_API_KEY = os.environ.get('ILETIMERKEZI_API_KEY')
ILETIMERKEZI_HASH = os.environ.get('ILETIMERKEZI_HASH')
ILETIMERKEZI_SENDER = os.environ.get('ILETIMERKEZI_SENDER', 'FatihSenyuz') 
SMS_ENABLED = os.environ.get('SMS_ENABLED', 'false').lower() in ('1', 'true', 'yes')
AUDIT_LOGS_ENABLED = os.environ.get('AUDIT_LOGS_ENABLED', 'false').lower() in ('1', 'true', 'yes')

# --- STRIPE ÖDEME AYARLARI ---
import stripe

STRIPE_SECRET_KEY = os.environ.get("STRIPE_SECRET_KEY")
STRIPE_PUBLISHABLE_KEY = os.environ.get("STRIPE_PUBLISHABLE_KEY") 
STRIPE_WEBHOOK_SECRET = os.environ.get("STRIPE_WEBHOOK_SECRET")

# Stripe'ı yapılandır
if STRIPE_SECRET_KEY:
    stripe.api_key = STRIPE_SECRET_KEY
    logger.info("✅ Stripe API key configured")
else:
    logger.warning("⚠️ STRIPE_SECRET_KEY not configured. Payment features will not work.")

from bulk_session_idempotency import BulkSessionCreate, SessionSlot, canonical_bulk_session_hash as _canonical_bulk_session_hash
from bulk_package import BulkPackageCreate
from session_wave_planner import plan_wave_remaining_sessions
from appointment_refund_flags import (
    REFUNDABLE_MERCHANT_STATES,
    attach_refund_eligible,
    session_group_ids_from_merchant_transactions,
)
from cancel_selected_support import staff_cannot_access_appointments
from capabilities import (
    CAP_APPOINTMENT_CANCEL,
    CAP_APPOINTMENT_REFUND,
    CAP_SESSION_PLAN,
    assert_capability_for_user,
)
from operation_audit import log_operation_audit

# Success ve Cancel URL'leri
PAYMENT_SUCCESS_URL = "https://plannapp.co/"
PAYMENT_SUCCESS_URL_NATIVE = "plannapp://payment-success"
# Native app için deep link, web için normal URL
PAYMENT_CANCEL_URL = "https://plannapp.co/dashboard"
PAYMENT_CANCEL_URL_NATIVE = "plannapp://dashboard"

# --- PUSH NOTIFICATION AYARLARI ---
VAPID_PUBLIC_KEY = os.environ.get('VAPID_PUBLIC_KEY', 'BB4nvNoHrWlWS6KTM1ybHUZD260l8b7Nnr2bMHvwnbflCJ4OJVd68Dqmw1hpaOFFUNmRFySvP3Ewzm596xjqF7g')
# Private key - base64 encoded raw key (pywebpush format)
VAPID_PRIVATE_KEY = os.environ.get('VAPID_PRIVATE_KEY', '5V68PoLSQ16LLymN-Wur5MwrJk1Q9b3HvS6ijdGDtAM')
if VAPID_PRIVATE_KEY:
    logger.info("✅ VAPID private key configured")
else:
    logger.warning("⚠️ VAPID_PRIVATE_KEY not configured. Push notifications will not work.")
VAPID_CLAIMS = {
    "sub": "mailto:info@plannapp.co"
}

# --- BREVO EMAIL AYARLARI ---
BREVO_API_KEY = os.environ.get('BREVO_API_KEY')
if BREVO_API_KEY:
    try:
        brevo_configuration = sib_api_v3_sdk.Configuration()
        brevo_configuration.api_key['api-key'] = BREVO_API_KEY
        brevo_api_instance = sib_api_v3_sdk.TransactionalEmailsApi(sib_api_v3_sdk.ApiClient(brevo_configuration))
        logging.info("✅ Brevo API instance başarıyla oluşturuldu.")
    except Exception as e:
        logging.error(f"❌ Brevo API instance oluşturulamadı: {str(e)}")
        brevo_api_instance = None
else:
    brevo_api_instance = None
    logging.warning("⚠️ BREVO_API_KEY bulunamadı! E-posta gönderimi devre dışı.")

async def send_email(to_email: str, subject: str, html_content: str, to_name: str = None, sender_name: str = "PLANN", sender_email: str = "info@plannapp.co"):
    """Brevo API ile e-posta gönder - Global helper fonksiyon (async)"""
    global brevo_api_instance
    try:
        logging.info(f"📧 [SEND_EMAIL] E-posta gönderme başlatılıyor: {to_email} - Subject: {subject}")
        logging.info(f"📧 [SEND_EMAIL] Sender: {sender_name} <{sender_email}>")
        logging.info(f"📧 [SEND_EMAIL] To: {to_name or to_email}")
        
        # Runtime'da API key'i tekrar kontrol et
        current_api_key = os.environ.get('BREVO_API_KEY')
        logging.info(f"🔑 BREVO_API_KEY kontrol: {'Var' if current_api_key else 'YOK'} - Uzunluk: {len(current_api_key) if current_api_key else 0}")
        
        if not brevo_api_instance:
            logging.warning("❌ Brevo API instance bulunamadı! E-posta gönderilemedi.")
            # Runtime'da instance oluşturmayı dene
            if current_api_key:
                try:
                    logging.info("🔄 Runtime'da Brevo API instance oluşturuluyor...")
                    brevo_configuration = sib_api_v3_sdk.Configuration()
                    brevo_configuration.api_key['api-key'] = current_api_key
                    brevo_api_instance = sib_api_v3_sdk.TransactionalEmailsApi(sib_api_v3_sdk.ApiClient(brevo_configuration))
                    logging.info("✅ Runtime'da Brevo API instance oluşturuldu!")
                except Exception as e:
                    logging.error(f"❌ Runtime'da Brevo API instance oluşturulamadı: {e}")
                    return False
            else:
                return False
        
        sender = {"name": sender_name, "email": sender_email}
        to = [{"email": to_email, "name": to_name or to_email}]
        
        send_smtp_email = sib_api_v3_sdk.SendSmtpEmail(
            to=to,
            sender=sender,
            subject=subject,
            html_content=html_content
        )
        
        # Async context'te sync API çağrısını thread pool'da çalıştır
        import asyncio
        logging.info(f"📤 Brevo API'ye e-posta gönderiliyor...")
        api_response = await asyncio.to_thread(brevo_api_instance.send_transac_email, send_smtp_email)
        logging.info(f"✅ E-posta başarıyla gönderildi: {to_email} - Subject: {subject} - Message ID: {api_response.message_id}")
        return True
    except ApiException as e:
        logging.error(f"❌ E-posta gönderilirken Brevo API hatası: {e.status} - {e.reason} - {e.body}")
        import traceback
        logging.error(traceback.format_exc())
        return False
    except Exception as e:
        logging.error(f"❌ E-posta gönderilirken beklenmedik hata: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return False


# ============================================================================
# === ORTAK MARKA E-POSTA ŞABLONU ===
# ============================================================================
# Tüm e-postalar (kayıt/hoş geldin, abonelik, duyuru vb.) aynı PLANN marka
# şablonunu kullanır: logolu açık gri başlık, beyaz içerik gövdesi, alt bilgi.
# `content_html` gövde hücresinin içine yerleştirilir.

# --- Premium marka tipografisi (Uptime/System Monitor şablonu ile aynı dil) ---
_BRAND_FONT_SERIF = "Georgia, 'Times New Roman', Times, serif"
_BRAND_FONT_SANS = "-apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif"


def _brand_shell(content_html: str, lang: str = "tr") -> str:
    """İçeriği PLANN premium e-posta kabuğuna sarar.

    Krem zemin, beyaz kart, tipografik 'P L A N N' başlık, serif gövde ve
    minimalist alt bilgi. Tüm stiller inline (Gmail/Brevo uyumu için)."""
    year = datetime.now(timezone.utc).year
    return f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="margin:0;padding:0;background-color:#fbfbfa;color:#1a1a1a;-webkit-font-smoothing:antialiased;font-family:{_BRAND_FONT_SERIF};">
  <center style="width:100%;background-color:#fbfbfa;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="width:100%;background-color:#fbfbfa;">
      <tr>
        <td align="center" style="padding:60px 16px;">
          <table role="presentation" width="600" cellpadding="0" cellspacing="0" border="0" style="width:600px;max-width:600px;margin:0 auto;background-color:#ffffff;border:1px solid #e5e5e3;box-shadow:0 4px 20px rgba(0,0,0,0.02);">
            <tr>
              <td style="padding:60px 50px 40px 50px;text-align:center;border-bottom:1px solid #f2f2f0;">
                <h1 style="margin:0;font-family:{_BRAND_FONT_SERIF};font-size:30px;font-weight:300;letter-spacing:10px;color:#1a1a1a;text-transform:uppercase;">P L A N N</h1>
              </td>
            </tr>
            <tr>
              <td style="padding:50px 50px 40px 50px;">
                {content_html}
              </td>
            </tr>
            <tr>
              <td style="padding:40px 50px;text-align:center;font-family:{_BRAND_FONT_SANS};font-size:10px;letter-spacing:2px;color:#a3a3a0;text-transform:uppercase;border-top:1px solid #f2f2f0;">
                PLANNAPP LTD. ALL RIGHTS RESERVED. &copy; {year}
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </center>
</body>
</html>"""


def _brand_body(
    *,
    badge: str = "",
    statement: str = "",
    paragraphs: Optional[list] = None,
    meta_rows: Optional[list] = None,
    button: Optional[dict] = None,
    note: str = "",
) -> str:
    """Premium gövde bloğu üretir: kategori etiketi, italik ana cümle, paragraflar,
    etiket/değer meta tablosu, çerçeveli buton ve isteğe bağlı dipnot."""
    parts = []
    if badge:
        parts.append(
            f'<div style="font-family:{_BRAND_FONT_SANS};font-size:10px;letter-spacing:3px;'
            f'text-transform:uppercase;color:#8c8c88;margin:0 0 30px 0;text-align:center;font-weight:500;">{badge}</div>'
        )
    if statement:
        parts.append(
            f'<div style="font-family:{_BRAND_FONT_SERIF};font-size:23px;line-height:1.6;color:#1a1a1a;'
            f'text-align:center;margin:0 0 45px 0;font-weight:400;font-style:italic;">{statement}</div>'
        )
    for p in (paragraphs or []):
        parts.append(
            f'<p style="font-family:{_BRAND_FONT_SERIF};font-size:17px;line-height:1.75;color:#3a3a3a;margin:0 0 20px 0;">{p}</p>'
        )
    if meta_rows:
        # Etiket/değer satırları ince çerçeveli bir kutu içinde (premium meta paneli)
        rows = ""
        n = len(meta_rows)
        for i, (label, value) in enumerate(meta_rows):
            border = "" if i == n - 1 else "border-bottom:1px solid #f2f2f0;"
            rows += (
                f'<tr><td style="padding:18px 24px;{border}font-family:{_BRAND_FONT_SANS};font-size:11px;'
                f'text-transform:uppercase;letter-spacing:2px;color:#8c8c88;font-weight:500;width:30%;vertical-align:top;">{label}</td>'
                f'<td style="padding:18px 24px 18px 0;{border}font-family:{_BRAND_FONT_SERIF};font-size:16px;color:#1a1a1a;vertical-align:top;">{value}</td></tr>'
            )
        parts.append(
            f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" '
            f'style="width:100%;margin:12px 0 45px 0;border:1px solid #e5e5e3;border-collapse:collapse;">{rows}</table>'
        )
    if note:
        parts.append(
            f'<p style="font-family:{_BRAND_FONT_SANS};font-size:13px;line-height:1.6;color:#8c8c88;'
            f'margin:28px 0 0 0;padding-top:20px;border-top:1px solid #f2f2f0;">{note}</p>'
        )
    if button and button.get("url") and button.get("label"):
        parts.append(
            f'<div style="text-align:center;margin-top:40px;">'
            f'<a href="{button["url"]}" target="_blank" style="border:1px solid #1a1a1a;color:#1a1a1a;'
            f'text-decoration:none;padding:16px 42px;font-family:{_BRAND_FONT_SANS};font-size:11px;'
            f'font-weight:600;text-transform:uppercase;letter-spacing:3px;display:inline-block;">{button["label"]}</a></div>'
        )
    return "\n".join(parts)


def _brand_email(*, lang: str = "tr", **body_kwargs) -> str:
    """Kısayol: premium gövdeyi premium kabuğa sarıp tam HTML döndürür."""
    return _brand_shell(_brand_body(**body_kwargs), lang=lang)


def _wrap_brand_email(content_html: str, lang: str = "tr") -> str:
    """Geriye dönük uyum: serbest içerik HTML'ini premium kabuğa sarar."""
    body = (
        f'<div style="font-family:{_BRAND_FONT_SERIF};font-size:17px;line-height:1.75;color:#3a3a3a;">'
        f'{content_html}</div>'
    )
    return _brand_shell(body, lang=lang)


# Subscription email helpers
def _get_domain_from_request(request: Request) -> str:
    try:
        from urllib.parse import urlparse

        raw_host = (request.headers.get('host') or '').strip()
        raw_origin = (request.headers.get('origin') or '').strip()
        raw_referer = (request.headers.get('referer') or '').strip()

        def _extract_host(value: str) -> str:
            if not value:
                return ''
            value = value.strip()
            if value.startswith('http://') or value.startswith('https://'):
                parsed = urlparse(value)
                value = parsed.netloc
            # strip port
            if ':' in value:
                value = value.split(':', 1)[0]
            return value.lower()

        # Prefer Origin/Referer (browser) to preserve www/non-www exactly, then Host
        candidates = [
            _extract_host(raw_origin),
            _extract_host(raw_referer),
            _extract_host(raw_host),
        ]

        allowed_suffixes = (
            'plannapp.co.uk',
            'plannapp.co',
            'plannapp.com.tr',
            'plannapp.com',
        )

        for c in candidates:
            if not c:
                continue
            if any(c == sfx or c.endswith('.' + sfx) for sfx in allowed_suffixes):
                return c

        return 'plannapp.co'
    except Exception:
        return 'plannapp.co'


def _get_lang_from_domain(domain: str) -> str:
    domain = (domain or '').lower()
    if 'plannapp.co.uk' in domain:
        return 'en'
    return 'tr'


def _build_subscription_email(lang: str, full_name: str, plan_name: str, billing_cycle: str, dashboard_url: str) -> tuple[str, str]:
    safe_name = full_name or ''
    safe_plan = plan_name or ''
    cycle = (billing_cycle or 'monthly').lower()
    if lang == 'en':
        subject = "Your PLANN subscription is now active"
        cycle_label = "Yearly" if cycle == 'yearly' else "Monthly"
        html_content = _brand_email(
            lang='en',
            badge="Subscription",
            statement="Your subscription is now active.",
            paragraphs=[
                f"Hello {safe_name or 'there'},",
                "Your payment was received and your PLANN subscription has been successfully activated.",
                "You can now continue using PLANN with your new plan.",
            ],
            meta_rows=[
                ("Plan", safe_plan or "—"),
                ("Billing", cycle_label),
            ],
            button={"label": "Go to Dashboard", "url": dashboard_url},
        )
        return subject, html_content

    subject = "PLANN aboneliğiniz aktif edildi"
    cycle_label = "Yıllık" if cycle == 'yearly' else "Aylık"
    html_content = _brand_email(
        lang='tr',
        badge="Abonelik",
        statement="Aboneliğiniz aktif edildi.",
        paragraphs=[
            f"Merhaba {safe_name or ''},",
            "Ödemeniz başarıyla alındı ve aboneliğiniz aktif edildi.",
            "Artık yeni paketinizle PLANN'ı kullanmaya devam edebilirsiniz.",
        ],
        meta_rows=[
            ("Paket", safe_plan or "—"),
            ("Dönem", cycle_label),
        ],
        button={"label": "Panele git", "url": dashboard_url},
    )
    return subject, html_content


async def _send_subscription_email_once(
    db,
    request: Request,
    organization_id: str,
    session_id: str,
    plan_name: str,
    billing_cycle: str,
    payment_log: Optional[dict] = None,
    admin_user: Optional[dict] = None,
):
    if not session_id:
        return
    if payment_log and payment_log.get('subscription_email_sent') is True:
        return

    if not admin_user:
        admin_user = await db.users.find_one({"organization_id": organization_id, "role": "admin"})
    if not admin_user:
        logger.warning(f"Subscription email: admin user bulunamadı - org={organization_id}")
        return

    to_email = admin_user.get('username')
    to_name = admin_user.get('full_name')
    if not to_email:
        logger.warning(f"Subscription email: admin email bulunamadı - org={organization_id}")
        return

    domain = _get_domain_from_request(request)
    lang = _get_lang_from_domain(domain)
    dashboard_url = f"https://{domain}"
    subject, html_content = _build_subscription_email(lang, to_name, plan_name, billing_cycle, dashboard_url)

    sent_ok = await send_email(to_email=to_email, subject=subject, html_content=html_content, to_name=to_name)
    if sent_ok:
        now = datetime.now(timezone.utc).isoformat()
        await db.payment_logs.update_one(
            {"session_id": session_id},
            {"$set": {"subscription_email_sent": True, "subscription_email_sent_at": now, "subscription_email_lang": lang}},
            upsert=True
        )


# ══════════════════════════════════════════════════════════════════════════
# AKTİVASYON NUDGE — kayıt +1 gün, hiç GERÇEK randevu oluşturmayan admin'e
# uygulama içine çekmeye yönelik tek seferlik premium e-posta.
# ══════════════════════════════════════════════════════════════════════════
ACTIVATION_NUDGE_MIN_AGE_HOURS = int(os.getenv("ACTIVATION_NUDGE_MIN_AGE_HOURS", "24"))
ACTIVATION_NUDGE_MAX_AGE_DAYS = int(os.getenv("ACTIVATION_NUDGE_MAX_AGE_DAYS", "7"))
# Akıllı buton: cihaza göre uygulamayı/mağazayı, masaüstünde web panelini açar
# (hoş geldin e-postasıyla aynı /api/app/open yönlendirmesi).
ACTIVATION_NUDGE_APP_URL_TR = os.getenv(
    "ACTIVATION_NUDGE_APP_URL_TR",
    "https://plannapp.co/api/app/open?web=https%3A%2F%2Fplannapp.co",
)
ACTIVATION_NUDGE_APP_URL_EN = os.getenv(
    "ACTIVATION_NUDGE_APP_URL_EN",
    "https://plannapp.co/api/app/open?web=https%3A%2F%2Fplannapp.co.uk",
)


def _parse_iso_dt(value):
    """ISO string / datetime → tz-aware UTC datetime (naive ise UTC varsayılır)."""
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if isinstance(value, str):
        try:
            d = datetime.fromisoformat(value.replace("Z", "+00:00"))
            return d if d.tzinfo else d.replace(tzinfo=timezone.utc)
        except Exception:
            return None
    return None


def _build_activation_nudge_email(lang: str, full_name: str, app_url: str) -> tuple[str, str]:
    """Kayıt olup henüz randevu oluşturmamış kullanıcıya nazik hatırlatma e-postası."""
    safe = (full_name or "").strip()
    if lang == "en":
        subject = "Put an end to empty appointment slots"
        html = _brand_email(
            lang="en",
            badge="First Step",
            statement="Put an end to empty appointment slots.",
            paragraphs=[
                f"Hello {safe or 'there'},",
                "Waiting for a client while wondering &ldquo;Will they actually show up?&rdquo; — or losing hours to no-shows because someone simply forgot — is frustrating. And texting everyone one by one to confirm is a job in itself.",
                "<strong>PLANN was built to eliminate exactly this problem.</strong>",
                "You just add the appointment; PLANN takes it from there. The automatic WhatsApp confirmations and reminders sent to your clients don't only prevent forgotten bookings — <strong>they take your business's image to the next level.</strong> A client who receives a professional notification the moment they book will see you as a far more established and trustworthy brand.",
                "You've created your account, but you haven't experienced this ease yet. Take just one minute: open PLANN, tap the <strong>(+)</strong> button, and add your first client.",
                "Let us handle the reminders and your professional image — you just focus on growing your business.",
            ],
            note="You received this because you signed up for PLANN and haven't created an appointment yet.",
            button={"label": "Create my first appointment", "url": app_url},
        )
        return subject, html
    subject = "Boş kalan randevu saatlerine son verin"
    html = _brand_email(
        lang="tr",
        badge="İlk Adım",
        statement="Boş kalan randevu saatlerine son verin.",
        paragraphs=[
            f"Merhaba {safe or ''},",
            "Müşterinizi beklerken &ldquo;Acaba gelecek mi?&rdquo; diye düşünmek veya randevuyu unuttukları için boşa giden saatlerin maliyeti can sıkıcıdır. Herkese tek tek mesaj atıp teyit almak ise ayrı bir mesai gerektirir.",
            "<strong>PLANN tam olarak bu sorunu ortadan kaldırmak için tasarlandı.</strong>",
            "Siz sadece randevuyu eklersiniz; gerisini PLANN devralır. Müşterilerinize giden otomatik WhatsApp onay ve hatırlatma mesajları sadece unutkanlıkların önüne geçmekle kalmaz; <strong>işletmenizin imajını da bir üst seviyeye taşır.</strong> Randevusunu aldığı an profesyonel bir bilgilendirme mesajı alan müşterilerinizin gözünde çok daha kurumsal ve güvenilir bir marka olarak konumlanırsınız.",
            "Hesabınızı oluşturdunuz ama bu rahatlığı henüz deneyimlemediniz. Sadece bir dakikanızı ayırın: PLANN'ı açın, alttaki <strong>(+)</strong> butonuna dokunun ve ilk müşterinizi ekleyin.",
            "Hatırlatmaları ve profesyonel görünümünüzü biz halledelim, siz sadece işinizi büyütmeye odaklanın.",
        ],
        note="Bu e-postayı, PLANN'a kaydolduğunuz hâlde henüz randevu oluşturmadığınız için aldınız.",
        button={"label": "İlk randevumu oluştur", "url": app_url},
    )
    return subject, html


async def send_activation_nudges(db):
    """Kayıttan ~1 gün sonra hiç gerçek (demo olmayan) randevu oluşturmamış admin'lere
    tek seferlik aktivasyon e-postası gönderir. `activation_nudge_sent` ile idempotent."""
    try:
        now = datetime.now(timezone.utc)
        # DISTRIBUTED LOCK: 17 worker → günde tek çalıştırma (aksi 17x gönderim riski)
        redis_client = getattr(getattr(_app_instance, 'state', None), 'redis_client', None)
        if redis_client:
            try:
                lock_key = f"activation_nudge_lock:{now.strftime('%Y%m%d')}"
                if not await redis_client.set(lock_key, "1", nx=True, ex=23 * 3600):
                    logging.debug("Activation nudge skipped - another worker holds the daily lock")
                    return 0
            except Exception as lock_err:
                logging.warning(f"Activation nudge lock unavailable, proceeding: {lock_err}")
        min_created = now - timedelta(days=ACTIVATION_NUDGE_MAX_AGE_DAYS)
        max_created = now - timedelta(hours=ACTIVATION_NUDGE_MIN_AGE_HOURS)
        sent = 0
        cursor = db.users.find(
            {"role": "admin", "activation_nudge_sent": {"$ne": True}},
            {"username": 1, "full_name": 1, "organization_id": 1, "created_at": 1, "user_id": 1},
        )
        async for u in cursor:
            created = _parse_iso_dt(u.get("created_at"))
            if not created or not (min_created <= created <= max_created):
                continue
            org = u.get("organization_id")
            email = (u.get("username") or "").strip()
            uid = u.get("user_id")
            if not org or "@" not in email:
                continue
            # Gerçek (demo olmayan) randevu var mı?
            real = await db.appointments.find_one(
                {"organization_id": org, "is_demo": {"$ne": True}, "source": {"$ne": "aha_activation"}},
                {"_id": 1},
            )
            if real:
                await db.users.update_one(
                    {"user_id": uid},
                    {"$set": {"activation_nudge_sent": True, "activation_nudge_skipped": "has_appointment"}},
                )
                continue
            # Dil: settings.support_phone +44 ise EN, aksi TR
            lang = "tr"
            try:
                s = await db.settings.find_one({"organization_id": org}, {"support_phone": 1})
                if ((s or {}).get("support_phone") or "").startswith("+44"):
                    lang = "en"
            except Exception:
                pass
            app_url = ACTIVATION_NUDGE_APP_URL_EN if lang == "en" else ACTIVATION_NUDGE_APP_URL_TR
            subject, html = _build_activation_nudge_email(lang, u.get("full_name"), app_url)
            ok = await send_email(to_email=email, subject=subject, html_content=html, to_name=u.get("full_name"))
            await db.users.update_one(
                {"user_id": uid},
                {"$set": {
                    "activation_nudge_sent": bool(ok),
                    "activation_nudge_sent_at": now.isoformat() if ok else None,
                    "activation_nudge_lang": lang,
                }},
            )
            if ok:
                sent += 1
        if sent:
            logging.info(f"Activation nudge: {sent} e-posta gönderildi")
        return sent
    except Exception as e:
        logging.error(f"send_activation_nudges failed: {e}", exc_info=True)
        return 0


# === SMS REMINDER SCHEDULER ===
scheduler = AsyncIOScheduler()
_app_instance = None  # Global app instance for scheduler

# Hatırlatma poll aralığı (saniye). Küçük değer = daha hassas zamanlama.
REMINDER_POLL_SECONDS = int(os.getenv("REMINDER_POLL_SECONDS", "60"))


async def check_and_send_reminders():
    """Yaklaşan randevuları kontrol edip WhatsApp hatırlatması gönderir.

    Zamanlama: her REMINDER_POLL_SECONDS sn'de bir çalışır; bir randevu için
    (apt_time - reminder_hours) anına ULAŞILDIĞINDA ilk turda gönderir → poll
    aralığı kadar hassas. Kaçan turlar 'due/overdue' mantığıyla telafi edilir
    (randevu henüz geçmediyse geç de olsa gönderilir)."""
    try:
        logging.info("=== WhatsApp Reminder Check Started ===")
        # Global app instance'ından db'yi al
        global _app_instance
        if _app_instance is None:
            logging.warning("App instance not available, skipping reminder check")
            return
        
        db = getattr(_app_instance, 'db', None)
        if db is None:
            logging.warning("MongoDB not available, skipping reminder check")
            return
        
        turkey_tz = ZoneInfo("Europe/Istanbul")
        now = datetime.now(turkey_tz)

        # DISTRIBUTED LOCK: multi-worker duplicate send prevention
        # Kova poll aralığına göre (saniye hassasiyetli). Dakika bazlı eski kova
        # 30 sn poll'da aynı dakikadaki 2. turu yanlışlıkla atlatıyordu.
        redis_client = getattr(getattr(_app_instance, 'state', None), 'redis_client', None)
        if redis_client:
            try:
                _bucket = int(now.timestamp() // REMINDER_POLL_SECONDS)
                lock_key = f"reminder_lock:{_bucket}"
                if not await redis_client.set(lock_key, "1", nx=True, ex=REMINDER_POLL_SECONDS + 10):
                    logging.debug("Reminder check skipped - another worker holds the lock")
                    return
                logging.debug(f"Distributed lock acquired: {lock_key}")
            except Exception as lock_err:
                logging.warning(f"Redis lock unavailable, proceeding without lock: {lock_err}")

        logging.info(f"Current time (Turkey): {now.strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Tüm organization'ların ayarlarını al
        all_settings = await db.settings.find({}, {"_id": 0}).to_list(1000)
        logging.info(f"Found {len(all_settings)} organizations to check")
        
        for setting in all_settings:
            org_id = setting.get('organization_id')
            reminder_hours = setting.get('sms_reminder_hours', 1.0)
            company_name = setting.get('company_name', 'İşletmeniz')
            support_phone = setting.get('support_phone', 'Destek')
            
            logging.info(f"Checking org {org_id}: reminder_hours={reminder_hours}, company={company_name}")
            
            # Yakın ufuk: bugünden itibaren 3 gün (reminder_hours büyük olsa bile kapsar).
            # 30 sn'de bir tarama yaptığımız için sorguyu dar tutuyoruz.
            _today = now.strftime("%Y-%m-%d")
            _horizon = (now + timedelta(days=3)).strftime("%Y-%m-%d")

            # Bekleyen, hatırlatması gitmemiş, yakın tarihli randevular
            appointments = await db.appointments.find({
                "organization_id": org_id,
                "status": "Bekliyor",
                "reminder_sent": {"$ne": True},
                "appointment_date": {"$gte": _today, "$lte": _horizon},
            }, {"_id": 0}).to_list(1000)
            
            logging.info(f"  Found {len(appointments)} pending appointments without reminder")
            
            for apt in appointments:
                try:
                    # Randevu zamanını parse et
                    apt_datetime_str = f"{apt['appointment_date']} {apt['appointment_time']}"
                    apt_datetime = datetime.strptime(apt_datetime_str, "%Y-%m-%d %H:%M").replace(tzinfo=turkey_tz)
                    
                    logging.debug(f"  Appointment {apt.get('id')}: {apt_datetime_str} (parsed: {apt_datetime.strftime('%Y-%m-%d %H:%M:%S')})")
                    
                    # Hatırlatma zamanı: apt_time - reminder_hours. Bu ana ULAŞILDIYSA
                    # ve randevu henüz geçmediyse gönder → poll aralığı kadar hassas,
                    # kaçırılan turlarda bile 'geç ama gönderilmiş' garantisi.
                    remind_at = apt_datetime - timedelta(hours=reminder_hours)
                    if remind_at <= now < apt_datetime:
                        # Atomic claim: only one worker can proceed (race condition fix)
                        claimed = await db.appointments.find_one_and_update(
                            {"id": apt['id'], "reminder_sent": {"$ne": True}},
                            {"$set": {"reminder_sent": True}},
                        )
                        if claimed is None:
                            logging.info(f"  Appointment {apt.get('id')} already claimed by another worker, skipping")
                            continue
                        logging.info(f"  Appointment {apt.get('id')} claimed. Sending WhatsApp...")
                        import asyncio
                        business = setting or {}
                        settings = business.get('settings') if isinstance(business.get('settings'), dict) else business
                        location = (settings or {}).get('location') or {}
                        coords = (location or {}).get('coordinates', {})
                        lat = coords.get('lat')
                        lng = coords.get('lng')
                        try:
                            wa_result = await asyncio.to_thread(
                                send_whatsapp_template,
                                apt['phone'], "REMINDER", apt['customer_name'],
                                company_name, apt['appointment_date'], apt['appointment_time'],
                                apt['service_name'], support_phone,
                                business_lat=lat, business_lng=lng,
                                business_address=location.get('address'),
                            )
                            logging.info(f"  WhatsApp reminder sent to {apt['customer_name']} ({apt['phone']}) apt={apt['id']}")
                            try:
                                import time as _time
                                _phone = str(apt['phone']).lstrip('+')
                                await db.whatsapp_message_logs.update_one(
                                    {"message_id": wa_result},
                                    {"$set": {"message_id": wa_result, "recipient": _phone,
                                              "status": "sent", "timestamp": int(_time.time()),
                                              "recorded_at": datetime.utcnow().isoformat(), "source": "reminder"}},
                                    upsert=True
                                )
                            except Exception:
                                pass
                        except Exception as send_err:
                            # Rollback so next scheduler run can retry
                            await db.appointments.update_one({"id": apt['id']}, {"$set": {"reminder_sent": False}})
                            raise send_err
                    else:
                        logging.debug(f"  - Appointment {apt.get('id')} is not in reminder window (time: {apt_datetime.strftime('%Y-%m-%d %H:%M:%S')})")
                
                except Exception as e:
                    logging.error(f"Error sending reminder for appointment {apt.get('id')}: {e}", exc_info=True)
                    try:
                        import time as _time
                        _phone = str(apt.get('phone', '')).lstrip('+')
                        await db.whatsapp_message_logs.update_one(
                            {"message_id": f"fail_{_phone}_{int(_time.time())}"},
                            {"$set": {
                                "recipient": _phone,
                                "status": "failed",
                                "timestamp": int(_time.time()),
                                "recorded_at": datetime.utcnow().isoformat(),
                                "errors": [{"code": "SEND_ERROR", "title": str(e)[:300]}],
                                "source": "reminder"
                            }},
                            upsert=True
                        )
                    except Exception:
                        pass
        
        logging.info("=== WhatsApp Reminder Check Completed ===")
    
    except Exception as e:
        logging.error(f"Error in check_and_send_reminders: {e}", exc_info=True)

# === Recurring Payment Checker (Her gün çalışır) ===
async def check_and_process_recurring_payments():
    """Vadesi gelen recurring payment'leri işle"""
    try:
        logging.info("=== Recurring Payment Check Started ===")
        
        if _app_instance is None or _app_instance.db is None:
            logging.warning("Database not available for recurring payment check")
            return
        
        db = _app_instance.db
        today = datetime.now(timezone.utc).date()
        today_str = today.isoformat()
        
        # Bugün ödeme günü olan organizasyonları bul
        cursor = db.organization_plans.find({
            "card_saved": True,
            "next_billing_date": {"$lte": datetime.now(timezone.utc).isoformat()},
            "plan_id": {"$ne": "tier_trial"}  # Trial paketleri hariç
        })
        
        organizations = await cursor.to_list(length=1000)
        logging.info(f"Found {len(organizations)} organizations with due payments")
        
        for org_plan in organizations:
            organization_id = org_plan.get('organization_id')
            plan_id = org_plan.get('plan_id')
            
            try:
                logging.info(f"Processing recurring payment for organization: {organization_id}, plan: {plan_id}")
                
                # Ödeme çek (internal API call simulation)
                # Gerçek implementasyonda recurring payment endpoint'ini çağır
                utoken = org_plan.get('payment_utoken')
                ctoken = org_plan.get('payment_ctoken')
                
                if not utoken or not ctoken:
                    logging.error(f"Missing payment tokens for organization: {organization_id}")
                    continue
                
                # Plan bilgilerini al
                plan_info = await get_plan_info(plan_id)
                if not plan_info:
                    logging.error(f"Plan info not found: {plan_id}")
                    continue
                
                price_monthly = plan_info.get('price_monthly', 0)
                payment_amount_kurus = int(price_monthly * 100)
                
                # Organization admin'ini bul
                user = await db.users.find_one({"organization_id": organization_id, "role": "admin"})
                if not user:
                    logging.error(f"Admin user not found for organization: {organization_id}")
                    continue
                
                user_email = user.get('username', 'noreply@royalpremiumcare.com')
                user_name = user.get('full_name', 'Kullanıcı')
                
                # Settings'den bilgi al
                settings = await db.settings.find_one({"organization_id": organization_id})
                user_address = settings.get('address', 'Adres Bilgisi Yok') if settings else 'Adres Bilgisi Yok'
                user_phone = settings.get('support_phone', '05000000000') if settings else '05000000000'
                
                # Merchant OID oluştur
                org_id_clean = organization_id.replace('-', '')
                timestamp_str = str(int(datetime.now(timezone.utc).timestamp()))
                merchant_oid = f"AUTO{org_id_clean}{timestamp_str}"
                
                # Sepet bilgisi
                plan_name = plan_info.get('name', 'Plan')
                user_basket = base64.b64encode(json.dumps([
                    [plan_name, str(price_monthly), 1]
                ]).encode('utf-8')).decode('utf-8')
                
                # PayTR parametreleri
                user_ip = "127.0.0.1"  # Sistem iç çağrısı
                payment_type = 'card'
                currency = 'TL'
                test_mode = '0'
                non_3d = '1'
                
                # Hash oluştur
                hash_str = f"{PAYTR_MERCHANT_ID}{user_ip}{merchant_oid}{user_email}{payment_amount_kurus}{payment_type}0{currency}{test_mode}{non_3d}"
                paytr_token = base64.b64encode(hmac.new(
                    PAYTR_MERCHANT_KEY.encode('utf-8'), 
                    hash_str.encode('utf-8') + PAYTR_MERCHANT_SALT.encode('utf-8'), 
                    hashlib.sha256
                ).digest()).decode('utf-8')
                
                # PayTR'a istek gönder
                post_data = {
                    'merchant_id': PAYTR_MERCHANT_ID,
                    'user_ip': user_ip,
                    'merchant_oid': merchant_oid,
                    'email': user_email,
                    'payment_type': payment_type,
                    'payment_amount': payment_amount_kurus,
                    'currency': currency,
                    'test_mode': test_mode,
                    'non_3d': non_3d,
                    'merchant_ok_url': PAYTR_SUCCESS_URL,
                    'merchant_fail_url': PAYTR_FAIL_URL,
                    'user_name': user_name,
                    'user_address': user_address[:400],
                    'user_phone': user_phone[:20],
                    'user_basket': user_basket,
                    'debug_on': '1',
                    'paytr_token': paytr_token,
                    'utoken': utoken,
                    'ctoken': ctoken,
                    'installment_count': '0'
                }
                
                # Payment log oluştur
                payment_log = {
                    "merchant_oid": merchant_oid,
                    "organization_id": organization_id,
                    "plan_id": plan_id,
                    "amount": price_monthly,
                    "status": "pending",
                    "payment_type": "auto_recurring",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.payment_logs.insert_one(payment_log)
                
                # PayTR'a istek gönder
                response = requests.post("https://www.paytr.com/odeme", data=post_data, timeout=15)
                
                if response.status_code == 200:
                    res_data = response.json() if response.headers.get('content-type', '').startswith('application/json') else {}
                    
                    if res_data.get('status') == 'success':
                        logging.info(f"✓ Auto recurring payment successful for organization: {organization_id}")
                        
                        # Bir sonraki ödeme tarihini güncelle (30 gün sonra)
                        next_billing = datetime.now(timezone.utc) + timedelta(days=30)
                        await db.organization_plans.update_one(
                            {"organization_id": organization_id},
                            {"$set": {
                                "next_billing_date": next_billing.isoformat(),
                                "last_payment_date": datetime.now(timezone.utc).isoformat(),
                                "updated_at": datetime.now(timezone.utc).isoformat()
                            }}
                        )
                    else:
                        error_msg = res_data.get('reason', 'Bilinmeyen hata')
                        logging.error(f"✗ Auto recurring payment failed for {organization_id}: {error_msg}")
                        
                        # Başarısız ödeme kaydını güncelle
                        await db.payment_logs.update_one(
                            {"merchant_oid": merchant_oid},
                            {"$set": {"status": "failed", "failed_reason": error_msg}}
                        )
                        
                        # Retry için 3 gün sonraya ertele
                        retry_date = datetime.now(timezone.utc) + timedelta(days=3)
                        await db.organization_plans.update_one(
                            {"organization_id": organization_id},
                            {"$set": {
                                "next_billing_date": retry_date.isoformat(),
                                "payment_retry_count": org_plan.get('payment_retry_count', 0) + 1,
                                "last_payment_attempt": datetime.now(timezone.utc).isoformat()
                            }}
                        )
                        
                        # TODO: Admin'e e-posta gönder
                else:
                    logging.error(f"✗ PayTR HTTP error for {organization_id}: {response.status_code}")
                
            except Exception as e:
                logging.error(f"Error processing recurring payment for {organization_id}: {e}", exc_info=True)
        
        logging.info("=== Recurring Payment Check Completed ===")
    
    except Exception as e:
        logging.error(f"Error in check_and_process_recurring_payments: {e}", exc_info=True)

# === MongoDB ve Redis Yaşam Döngüsü (Lifespan) --- SYNTAX HATASI DÜZELTİLDİ ===
@asynccontextmanager
async def lifespan(app: FastAPI):
    global _app_instance
    _app_instance = app  # Global app instance'ı sakla (scheduler için)
    app.mongodb_client = None; app.db = None; app.redis_client = None
    try:
        logging.info("Step 1: Connecting to MongoDB..."); mongo_url = os.environ.get('MONGO_URL'); db_name = os.environ.get('DB_NAME', 'royal_koltuk_dev')
        if not mongo_url:
            logging.error("CRITICAL: MONGO_URL environment variable is required!"); logging.warning("MongoDB connection will be lazy-initialized on first request.")
        else:
            try:
                app.mongodb_client = AsyncIOMotorClient(mongo_url, serverSelectionTimeoutMS=5000); await app.mongodb_client.admin.command('ping')
                app.db = app.mongodb_client[db_name]; logging.info(f"Step 1 SUCCESS: Successfully connected to MongoDB ({db_name})!")
            except Exception as mongo_error:
                logging.warning(f"MongoDB connection failed during startup: {mongo_error}"); logging.info("MongoDB will be lazy-initialized on first request.")
    except Exception as e:
        logging.warning(f"WARNING during MongoDB connection: {type(e).__name__}: {str(e)}"); app.mongodb_client = None; app.db = None
    try:
        logging.info("Step 2: Initializing Redis..."); app.state.redis_client = await init_redis() # <--- BURASI DEĞİŞTİ
        if app.state.redis_client:
            try: await app.state.redis_client.ping(); logging.info("Step 2 SUCCESS: Redis initialized and ping successful.")
            except Exception as redis_ping_error: logging.warning(f"Redis ping failed: {redis_ping_error}"); app.state.redis_client = None
        else: logging.warning("WARNING: Redis client could not be initialized (init_redis returned None).")
    except Exception as e:
        logging.warning(f"WARNING during Redis connection: {type(e).__name__}: {str(e)}"); app.state.redis_client = None
    try:
        logging.info("Step 3: Initializing Rate Limiter...")
        if app.redis_client is None: logging.warning("WARNING: Using dummy Rate Limiter due to failed Redis connection.")
        app.state.limiter = initialize_limiter(app.redis_client); logging.info("Step 3 SUCCESS: Rate Limiter initialized.")
    except Exception as e:
        logging.warning(f"WARNING during Rate Limiter initialization: {type(e).__name__}: {str(e)}"); app.state.limiter = None
    
    try:
        logging.info("Step 4: Starting Schedulers...")
        # WhatsApp Reminder Job - Her 5 dakikada bir (WhatsApp hatırlatmaları için)
        # NOT: Bu job WhatsApp gönderiyor, SMS değil. SMS_ENABLED kontrolü yapılmıyor.
        scheduler.add_job(
            check_and_send_reminders, 
            IntervalTrigger(seconds=REMINDER_POLL_SECONDS), 
            id='whatsapp_reminder_job',
            replace_existing=True,
            max_instances=1  # Aynı anda sadece bir instance çalışsın
        )
        logging.info(f"  - WhatsApp Reminder: Enabled (Every {REMINDER_POLL_SECONDS}s, hassas zamanlama)")
        
        # Recurring Payment Job - Her gün saat 02:00'de (UTC)
        from apscheduler.triggers.cron import CronTrigger
        scheduler.add_job(
            check_and_process_recurring_payments,
            CronTrigger(hour=2, minute=0),  # Her gün 02:00
            id='recurring_payment_job',
            replace_existing=True,
            max_instances=1
        )

        # Aktivasyon nudge — her gün 08:00 UTC (~11:00 TR): kayıttan +1 gün geçmiş,
        # hiç gerçek randevu oluşturmamış admin'lere tek seferlik e-posta.
        if app.db is not None:
            _nudge_db = app.db

            async def _run_activation_nudges():
                try:
                    await send_activation_nudges(_nudge_db)
                except Exception as e:
                    logging.error(f"activation_nudges job failed: {e}", exc_info=True)

            scheduler.add_job(
                _run_activation_nudges,
                CronTrigger(hour=8, minute=0),
                id='activation_nudge_job',
                replace_existing=True,
                max_instances=1,
            )
            logging.info("  - Activation Nudge: Enabled (daily 08:00 UTC)")
        
        # Step 4b: Financial Engine Cron Jobs
        if app.db is not None:
            try:
                from financial.cron_jobs import register_financial_cron_jobs
                register_financial_cron_jobs(scheduler, app.db)
                logging.info("  - Financial Engine: 8 cron jobs registered")
            except Exception as fin_err:
                logging.warning(f"Financial cron jobs registration failed: {fin_err}")

        # Step 4c: PLANN v2 — Migrations + new cron jobs + DLQ handlers
        if app.db is not None:
            # 4c.1 Run any pending migrations
            try:
                from financial.migrations import runner as mig_runner
                mig_result = await mig_runner.apply_all(app.db)
                logging.info(
                    "  - PLANN v2 migrations: applied=%s errors=%s",
                    mig_result.get("applied"), mig_result.get("errors"),
                )
            except Exception as mig_err:
                logging.warning(f"PLANN v2 migrations failed: {mig_err}")

            # 4c.2 Register DLQ retry handlers (failure_type → callable)
            try:
                from financial import dead_letter_queue as dlq_mod
                from financial import wednesday_batch as wb_mod
                from financial.refund_reconciliation_service import (
                    resolve_refund_reconciliation_by_id,
                )

                async def _dlq_reconciliation_retry(db, doc):
                    src_id = doc.get("source_entity_id")
                    if not src_id:
                        raise RuntimeError("no source_entity_id")
                    result = await resolve_refund_reconciliation_by_id(
                        db, src_id, locked_by="dlq:retry", request=None,
                    )
                    if not result.get("ok"):
                        raise RuntimeError(f"reconciliation retry failed: {result}")

                async def _dlq_wise_fund_retry(db, doc):
                    # For wise_fund_failed: we don't auto-retry funding (manual admin only)
                    # Mark as requiring manual review by raising
                    raise RuntimeError("wise_fund requires manual admin intervention")

                dlq_mod.register_handler(
                    dlq_mod.FailureType.RECONCILIATION_MISMATCH.value,
                    _dlq_reconciliation_retry,
                )
                dlq_mod.register_handler(
                    dlq_mod.FailureType.WISE_FUND_FAILED.value,
                    _dlq_wise_fund_retry,
                )
                logging.info("  - PLANN v2 DLQ: handlers registered")
            except Exception as dlq_err:
                logging.warning(f"PLANN v2 DLQ handler registration failed: {dlq_err}")

            # 4c.3 New cron jobs (reserve expiry, batch prep, refund SLA, DLQ retry)
            try:
                from financial.expiry_service import run_expiry_crons
                from financial.wednesday_batch import tuesday_batch_prepare_cron
                from financial.refund_requests import run_sla_monitor_cron
                from financial.dead_letter_queue import run_retry_cron as dlq_run_retry_cron

                _db = app.db

                async def _run_expiry():
                    try:
                        await run_expiry_crons(_db)
                    except Exception as e:
                        logging.error(f"run_expiry_crons failed: {e}", exc_info=True)

                async def _run_tuesday_batch():
                    try:
                        await tuesday_batch_prepare_cron(_db)
                    except Exception as e:
                        logging.error(f"tuesday_batch_prepare_cron failed: {e}", exc_info=True)

                async def _run_refund_sla():
                    try:
                        await run_sla_monitor_cron(_db)
                    except Exception as e:
                        logging.error(f"run_sla_monitor_cron failed: {e}", exc_info=True)

                async def _run_dlq_retry():
                    try:
                        await dlq_run_retry_cron(_db)
                    except Exception as e:
                        logging.error(f"dlq_retry_cron failed: {e}", exc_info=True)

                scheduler.add_job(
                    _run_expiry, IntervalTrigger(hours=1),
                    id='v2_expiry_crons', replace_existing=True, max_instances=1,
                )
                scheduler.add_job(
                    _run_tuesday_batch, CronTrigger(day_of_week='tue', hour=23, minute=0),
                    id='v2_tuesday_batch_prepare', replace_existing=True, max_instances=1,
                )
                scheduler.add_job(
                    _run_refund_sla, IntervalTrigger(minutes=15),
                    id='v2_refund_sla_monitor', replace_existing=True, max_instances=1,
                )
                scheduler.add_job(
                    _run_dlq_retry, IntervalTrigger(minutes=5),
                    id='v2_dlq_retry', replace_existing=True, max_instances=1,
                )
                logging.info("  - PLANN v2 cron jobs: 4 registered (expiry, batch, sla, dlq)")
            except Exception as cron_err:
                logging.warning(f"PLANN v2 cron registration failed: {cron_err}")

            # Step 4d: PLANN Asistan — zeka motoru cron işleri (08:30 / 22:00 / Pazar 22:00)
            try:
                from assistant import register_assistant_jobs
                register_assistant_jobs(
                    scheduler, app.db,
                    redis_client=getattr(app.state, "redis_client", None),
                    push_fn=send_push_notification,
                )
            except Exception as asst_err:
                logging.warning(f"PLANN Asistan cron registration failed: {asst_err}")

            # Step 4e: Pazarlama Otopilotu — WhatsApp geri-kazanım (enqueue + dispatch)
            try:
                from marketing_autopilot import register_autopilot_jobs
                register_autopilot_jobs(
                    scheduler, app.db,
                    redis_client=getattr(app.state, "redis_client", None),
                )
            except Exception as autopilot_err:
                logging.warning(f"Pazarlama Otopilotu cron registration failed: {autopilot_err}")

        scheduler.start()
        logging.info("Step 4 SUCCESS: Schedulers started")
        logging.info("  - Recurring Payments: Daily at 02:00 UTC")
        
        # İlk WhatsApp hatırlatma kontrolünü hemen yap
        import asyncio
        asyncio.create_task(check_and_send_reminders())
    except Exception as e:
        logging.error(f"ERROR during Scheduler initialization: {type(e).__name__}: {str(e)}", exc_info=True)
    
    try:
        logging.info("Step 5: Creating Database Indexes...")
        if app.db is not None:
            # Appointments indexes - Performance optimization
            await app.db.appointments.create_index([("organization_id", 1), ("appointment_date", -1)])
            await app.db.appointments.create_index([("organization_id", 1), ("staff_member_id", 1)])
            await app.db.appointments.create_index([("organization_id", 1), ("phone", 1)])
            await app.db.appointments.create_index([("organization_id", 1), ("status", 1)])
            # Dashboard aggregate'leri (org + status + appointment_date) filtreliyor → compound index.
            await app.db.appointments.create_index([("organization_id", 1), ("status", 1), ("appointment_date", 1)])
            # GET /api/appointments cursor sıralaması: sort([date:1, time:1, id:1]).
            # Bu index olmadan planlayıcı organizasyonun TÜM randevularını çekip belleğe
            # alıyor ve orada sıralıyordu (SORT stage) — limit=50 istense bile. Index
            # sıralı geldiği için artık limit kadar okuyup duruyor.
            try:
                await app.db.appointments.create_index(
                    [("organization_id", 1), ("appointment_date", 1), ("appointment_time", 1), ("id", 1)],
                    name="appointments_org_date_time_id",
                )
            except Exception as idx_err:
                logging.debug(f"appointments (org, date, time, id) index skipped: {idx_err}")
            
            # Customers indexes — cursor pagination + Türkçe collation ile arama.
            #
            # NOT: (organization_id, phone) üzerinde GLOBAL unique index BİLİNÇLİ olarak
            # zorlanmıyor. Domain gereği bu alanlar tekrar edebilir:
            #   • Telefonsuz walk-in müşteriler `phone:"0"` ile saklanıyor → org başına
            #     birden çok farklı kişi aynı "0" değerine sahip olabilir (veri değil, tasarım).
            #   • Aynı telefon farklı isimle ayrı müşteri olabilir (ör. anne telefonuyla
            #     kızına randevu) → paylaşılan telefon geçerli bir senaryo.
            # Bu yüzden aşağıdaki NON-UNIQUE index efektif olandır. Unique'i yine de bir kez
            # deniyoruz (kurulursa bonus veri bütünlüğü); kurulamazsa bu beklenen bir durum,
            # bu nedenle DEBUG seviyesinde loglanır (WARNING değil — prod loglarını kirletmesin).
            try:
                await app.db.customers.create_index(
                    [("organization_id", 1), ("phone", 1)],
                    unique=True,
                    name="customers_org_phone_unique",
                )
            except Exception as idx_err:
                logging.debug(
                    f"customers (org_id, phone) unique index atlandı (beklenen: paylaşılan/boş telefonlar): {idx_err}"
                )
                try:
                    await app.db.customers.create_index(
                        [("organization_id", 1), ("phone", 1)],
                        name="customers_org_phone_nonunique",
                    )
                except Exception:
                    pass
            try:
                await app.db.customers.create_index(
                    [("organization_id", 1), ("name", 1), ("_id", 1)],
                    collation={"locale": "tr", "strength": 2},
                    name="customers_org_name_id_tr",
                )
            except Exception as idx_err:
                logging.debug(f"customers (org_id, name, _id) tr collation index skipped: {idx_err}")
            try:
                await app.db.customers.create_index(
                    [("organization_id", 1), ("last_appointment_at", -1)],
                    name="customers_org_last_appt_desc",
                )
            except Exception as idx_err:
                logging.debug(f"customers (org_id, last_appointment_at) index skipped: {idx_err}")

            # Users indexes
            await app.db.users.create_index([("organization_id", 1), ("role", 1)])
            try:
                await app.db.users.create_index([("slug", 1)], unique=True, sparse=True)
            except Exception as idx_err:
                # Index might already exist or have duplicate null values, skip
                logging.debug(f"Users slug index creation skipped: {idx_err}")
            
            # Settings indexes
            await app.db.settings.create_index([("organization_id", 1)], unique=True)
            try:
                await app.db.settings.create_index([("slug", 1)], unique=True, sparse=True)
            except Exception as idx_err:
                # Index might already exist or have duplicate null values, skip
                logging.debug(f"Settings slug index creation skipped: {idx_err}")
            
            # Contact requests indexes
            await app.db.contact_requests.create_index([("created_at", -1)])
            await app.db.contact_requests.create_index([("status", 1)])

            # Pazarlama Otopilotu — görev kuyruğu indeksleri
            try:
                from marketing_autopilot import ensure_indexes as _ensure_autopilot_indexes
                await _ensure_autopilot_indexes(app.db)
            except Exception as _ap_idx_err:
                logging.debug(f"autopilot index creation skipped: {_ap_idx_err}")

            # Funnel events ledger (activation + dropoff + subscription funnel)
            try:
                await ensure_funnel_indexes(app.db)
            except Exception as idx_funnel:
                logging.debug(f"funnel_events index creation skipped: {idx_funnel}")

            await app.db.operation_audit_logs.create_index(
                [("organization_id", 1), ("created_at", -1)]
            )

            try:
                await app.db.refund_reconciliation_pending.create_index(
                    [("organization_id", 1), ("status", 1), ("created_at", -1)]
                )
                await app.db.refund_reconciliation_pending.create_index(
                    [("stripe_refund_id", 1)], sparse=True
                )
                await app.db.refund_reconciliation_pending.create_index(
                    [("status", 1), ("next_retry_at", 1)], sparse=True
                )
            except Exception as idx_rr:
                logging.debug(f"refund_reconciliation_pending index: {idx_rr}")

            try:
                await app.db.merchant_transactions.create_index(
                    [("stripe_refund_id", 1)],
                    unique=True,
                    sparse=True,
                    name="merchant_tx_stripe_refund_id_unique_sparse",
                )
            except Exception as idx_mt:
                logging.debug(f"merchant_transactions stripe_refund_id index: {idx_mt}")

            # Marketing leads — havuz/aramalarım listesi için bileşik index'ler.
            # Önceki sürümde 5000 doc tarama vardı, bu index'ler ile <50ms'ye iner.
            try:
                # Havuz listesi: status="pool" + (batch_id) + sort by created_at asc.
                await app.db.leads.create_index(
                    [("status", 1), ("batch_id", 1), ("created_at", 1)],
                    name="leads_pool_batch_created",
                )
                # Aramalarım: claimed_by + status filtre + sort by updated_at desc.
                await app.db.leads.create_index(
                    [("claimed_by", 1), ("status", 1), ("updated_at", -1)],
                    name="leads_claimed_status_updated",
                )
                # Sektör filtresi (her iki tab'da kullanılıyor).
                await app.db.leads.create_index(
                    [("sector", 1)], sparse=True, name="leads_sector_sparse"
                )
                # Lookup'lar: lead_id ile claim/status/release.
                await app.db.leads.create_index(
                    [("id", 1)], unique=True, sparse=True, name="leads_id_unique"
                )
            except Exception as idx_leads:
                logging.debug(f"leads indexes skipped: {idx_leads}")

            # PLANN Asistan — mesaj havuzu index'leri + idempotent seed + bildirim/snapshot index'leri
            try:
                from assistant.seed_messages import ensure_indexes as _asst_ensure_idx, seed_messages as _asst_seed
                await _asst_ensure_idx(app.db)
                await app.db.assistant_notifications.create_index(
                    [("organization_id", 1), ("scenario", 1), ("sent_at", -1)]
                )
                await app.db.assistant_notifications.create_index(
                    [("organization_id", 1), ("sent_at", -1)]
                )
                await app.db.assistant_daily_snapshots.create_index(
                    [("organization_id", 1), ("date", -1)]
                )
                _asst_seed_res = await _asst_seed(app.db)
                logging.info(f"  - PLANN Asistan: mesaj havuzu seed edildi ({_asst_seed_res.get('total')} metin)")
            except Exception as idx_asst:
                logging.warning(f"PLANN Asistan index/seed skipped: {idx_asst}")

            logging.info("Step 5 SUCCESS: Database indexes created")
        else:
            logging.warning("Step 5 SKIPPED: Database not available")
    except Exception as e:
        logging.warning(f"WARNING during Index creation: {type(e).__name__}: {str(e)}")

    # Step 6: Firebase Admin SDK
    try:
        logging.info("Step 6: Initializing Firebase Admin SDK...")
        if not firebase_admin._apps:
            firebase_cred_path = os.path.join(os.path.dirname(__file__), "firebase-admin-key.json")
            logging.info(f"  Firebase key path: {firebase_cred_path}, exists: {os.path.exists(firebase_cred_path)}")
            if os.path.exists(firebase_cred_path):
                cred = credentials.Certificate(firebase_cred_path)
                firebase_admin.initialize_app(cred)
                logging.info("Step 6 SUCCESS: ✅ Firebase Admin SDK initialized")
            else:
                logging.warning("Step 6 SKIPPED: ⚠️ firebase-admin-key.json not found")
        else:
            logging.info("Step 6 SUCCESS: Firebase Admin SDK already initialized")
    except Exception as e:
        logging.error(f"Step 6 FAILED: ❌ Firebase initialization error: {e}", exc_info=True)

    yield

    # --- Cleanup blokları ---
    # NOT: Scheduler'ı cleanup'ta kapatma - uygulama çalışırken scheduler aktif kalmalı
    # Sadece uygulama kapanırken kapatılmalı
    logging.info("Application shutdown initiated...")
    # Global app instance'ı temizle (global değişken, direkt atama yapılabilir)
    # _app_instance zaten global scope'ta tanımlı, burada sadece None yapıyoruz
    if scheduler.running:
        logging.info("Stopping WhatsApp Reminder Scheduler...")
        try: 
            scheduler.shutdown(wait=False)
            logging.info("WhatsApp Reminder Scheduler stopped")
        except Exception as e:
            logging.error(f"Error stopping scheduler: {e}")
    if app.mongodb_client:
        logging.info("Closing MongoDB connection...")
        try: app.mongodb_client.close()
        except: pass
    # hasattr veya getattr kullanarak güvenli bir şekilde kapatıyoruz
    if getattr(app.state, 'redis_client', None):
        logging.info("Closing Redis connection...")
        try: await app.state.redis_client.close()
        except: pass

# Create the main app
app = FastAPI(title="Randevu SaaS API", description="... (Açıklamanız buradaydı) ...", version="1.4.2 (Final Fixes)", lifespan=lifespan)
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# === SOCKET.IO SETUP ===
# Get CORS origins for Socket.IO (same as FastAPI CORS)
cors_origins_for_socketio = os.environ.get('CORS_ORIGINS', '*')
if cors_origins_for_socketio == '*':
    socketio_cors_origins = '*'
else:
    socketio_cors_origins = [origin.strip() for origin in cors_origins_for_socketio.split(',') if origin.strip()]

mgr = socketio.AsyncRedisManager('redis://plann_redis:6379/0')

sio = socketio.AsyncServer(
    async_mode='asgi',
    cors_allowed_origins="*",
    client_manager=mgr,  # 🚀 İŞTE BU! 4 işçiyi birbirine bağlayan hat.
    logger=True,
    engineio_logger=True
)
socket_app = socketio.ASGIApp(sio, socketio_path='/api/socket.io', other_asgi_app=app)

# Set socketio instance for AI service
ai_service.set_socketio(sio)

# --- Router prefix'i kaldırıldı ---
api_router = APIRouter()


# === HEALTH CHECK ENDPOINT ===
@api_router.get("/health")
async def health_check(request: Request):
    """Comprehensive health check for all system components."""
    import time as _time
    checks = {}
    overall_ok = True

    # 1. MongoDB
    try:
        t0 = _time.monotonic()
        db = getattr(request.app, "db", None)
        if db is None:
            raise Exception("db not initialised")
        await db.command("ping")
        checks["mongodb"] = {"status": "ok", "latency_ms": round((_time.monotonic() - t0) * 1000, 1)}
    except Exception as e:
        checks["mongodb"] = {"status": "error", "detail": str(e)}
        overall_ok = False

    # 2. Redis
    try:
        t0 = _time.monotonic()
        redis_client = getattr(request.app.state, "redis_client", None) or getattr(request.app, "redis_client", None)
        if redis_client is None:
            raise Exception("redis not initialised")
        await redis_client.ping()
        checks["redis"] = {"status": "ok", "latency_ms": round((_time.monotonic() - t0) * 1000, 1)}
    except Exception as e:
        checks["redis"] = {"status": "error", "detail": str(e)}
        overall_ok = False

    # 3. Scheduler
    try:
        checks["scheduler"] = {
            "status": "ok" if scheduler.running else "error",
            "jobs": len(scheduler.get_jobs()),
        }
        if not scheduler.running:
            overall_ok = False
    except Exception as e:
        checks["scheduler"] = {"status": "error", "detail": str(e)}
        overall_ok = False

    # 4. Stripe
    stripe_key = os.environ.get("STRIPE_SECRET_KEY", "")
    checks["stripe"] = {
        "status": "ok" if stripe_key.startswith("sk_live_") else "warning",
        "mode": "live" if stripe_key.startswith("sk_live_") else "test" if stripe_key.startswith("sk_test_") else "missing",
    }

    # 5. Wise
    wise_token = os.environ.get("WISE_API_TOKEN", "")
    wise_env = os.environ.get("WISE_ENVIRONMENT", "sandbox")
    checks["wise"] = {
        "status": "ok" if wise_token and wise_env == "production" else "warning",
        "environment": wise_env,
        "token_set": bool(wise_token),
    }

    # 6. Brevo (Email)
    checks["brevo_email"] = {"status": "ok" if os.environ.get("BREVO_API_KEY") else "error"}

    # 7. WhatsApp (Meta)
    checks["whatsapp"] = {"status": "ok" if os.environ.get("META_ACCESS_TOKEN") else "error"}

    # 8. Gemini AI
    checks["gemini_ai"] = {"status": "ok" if os.environ.get("GOOGLE_GEMINI_KEY") else "warning"}

    # 9. Sentry
    checks["sentry"] = {
        "status": "ok" if os.environ.get("SENTRY_DSN") else "warning",
        "env": os.environ.get("SENTRY_ENV", "unknown"),
    }

    # 10. Database collections
    if checks["mongodb"]["status"] == "ok":
        try:
            col_names = await db.list_collection_names()
            checks["db_collections"] = {"status": "ok", "count": len(col_names)}
        except Exception:
            checks["db_collections"] = {"status": "error"}

    from starlette.status import HTTP_200_OK, HTTP_503_SERVICE_UNAVAILABLE
    status_code = HTTP_200_OK if overall_ok else HTTP_503_SERVICE_UNAVAILABLE

    return JSONResponse(
        status_code=status_code,
        content={
            "status": "healthy" if overall_ok else "degraded",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "version": "1.0.0",
            "checks": checks,
        },
    )


# === SOCKET.IO EVENT HANDLERS ===
@sio.event
async def connect(sid, environ, *args):
    """Client connected - with authentication"""
    logger.info(f"🔵 [CONNECT] WebSocket client attempting connection: {sid}")
    logger.info(f"🔍 [CONNECT] Args received: {args}")
    
    # Token'ı bul (auth parametresi, query string veya header'dan)
    token = None
    
    # 1. Socket.IO auth parametresinden token al (args'da gelebilir)
    if args and len(args) > 0:
        logger.info(f"📦 [CONNECT] Args[0] type: {type(args[0])}, content: {args[0]}")
        auth_data = args[0]
        if isinstance(auth_data, dict):
            token = auth_data.get('token')
            logger.info(f"🔑 [CONNECT] Token from auth dict: {token[:20] if token else 'None'}...")
        elif isinstance(auth_data, str):
            token = auth_data
            logger.info(f"🔑 [CONNECT] Token from auth string: {token[:20] if token else 'None'}...")
    
    # 2. Query string'den token al (client query: {token: ...} kullanırsa)
    if not token:
        query_string = environ.get('QUERY_STRING', '')
        logger.info(f"❓ [CONNECT] Query string: {query_string}")
        if query_string:
            from urllib.parse import parse_qs
            params = parse_qs(query_string)
            token_list = params.get('token', [])
            if token_list:
                token = token_list[0]
                logger.info(f"🔑 [CONNECT] Token from query: {token[:20]}...")
    
    # 3. Header'dan token al (HTTP_AUTHORIZATION)
    if not token:
        auth_header = environ.get('HTTP_AUTHORIZATION', '')
        logger.info(f"📋 [CONNECT] Auth header: {auth_header[:30] if auth_header else 'None'}...")
        if auth_header.startswith('Bearer '):
            token = auth_header[7:]
            logger.info(f"🔑 [CONNECT] Token from header: {token[:20]}...")
    
    # Public sayfa için token olmayabilir - bu durumda sadece public room'lara join edebilir
    is_public_connection = False
    if not token:
        logger.info(f"🌐 [CONNECT] No token provided by {sid} - treating as public connection")
        is_public_connection = True
        # Public connection için minimal session oluştur
        await sio.save_session(sid, {
            'is_public': True
        })
        await sio.emit('connection_established', {'status': 'connected', 'public': True}, room=sid)
        return True  # Public bağlantıya izin ver
    
    # Token varsa normal authentication yap
    
    # Token'ı doğrula
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        organization_id = payload.get("org_id")
        
        if not username or not organization_id:
            logger.warning(f"⚠️ [CONNECT] Eksik payload, public'e düşürülüyor: {sid}")
            await sio.save_session(sid, {'is_public': True}) # Buraya ekle
            return True # False'u True yap!
        
        await sio.save_session(sid, {
            'username': username,
            'organization_id': organization_id,
            'role': payload.get('role')
        })
        logger.info(f"✓ [CONNECT] Authenticated: {username}")

        try:
            room_name = f"org_{organization_id}"
            await sio.enter_room(sid, room_name)
            logger.info(f"✓ [CONNECT] Auto-joined {sid} to room: {room_name}")
            await sio.emit('joined_organization', {'organization_id': organization_id}, room=sid)
        except Exception as join_error:
            logger.error(f"✗ [CONNECT] Failed to auto-join org room: {join_error}", exc_info=True)
        return True
        
    except (JWTError, Exception) as e:
        logger.warning(f"⚠️ [CONNECT] Auth hatası ({type(e).__name__}), public devam ediliyor: {sid}")
        await sio.save_session(sid, {'is_public': True})
        return True

@sio.event
async def disconnect(sid):
    """Client disconnected"""
    logger.info(f"WebSocket client disconnected: {sid}")
    
    # Voice session cleanup
    if sid in _voice_sessions:
        try:
            voice_service = get_voice_ai_service()
            if voice_service:
                session_info = _voice_sessions[sid]
                voice_session = session_info['session']
                receive_task = session_info.get('receive_task')
                
                # Receive loop'u durdur
                if receive_task and not receive_task.done():
                    receive_task.cancel()
                    try:
                        await receive_task
                    except asyncio.CancelledError:
                        pass
                
                await voice_service.close_session(voice_session)
                del _voice_sessions[sid]
                logger.info(f"🧹 [VOICE] Voice session cleaned up for disconnected client {sid}")
        except Exception as e:
            logger.error(f"Error cleaning up voice session: {e}")

@sio.event
async def join_organization(sid, data):
    """Join organization room for real-time updates - with authorization"""
    logger.info(f"🟢 [JOIN_ORG] join_organization event received from {sid} with data: {data}")
    
    try:
        # Session'dan kullanıcı bilgilerini al
        session = await sio.get_session(sid)
        if not session:
            logger.warning(f"✗ [JOIN_ORG] No session found for {sid} - connection not authenticated")
            await sio.emit('error', {'message': 'Not authenticated'}, room=sid)
            return
        
        user_org_id = session.get('organization_id')
        if not user_org_id:
            logger.warning(f"✗ [JOIN_ORG] No organization_id in session for {sid}")
            await sio.emit('error', {'message': 'Invalid session'}, room=sid)
            return
        
        # İstenen organization_id
        requested_org_id = data.get('organization_id')
        if not requested_org_id:
            logger.warning(f"⚠ [JOIN_ORG] join_organization called without organization_id from {sid}")
            await sio.emit('error', {'message': 'organization_id required'}, room=sid)
            return
        
        # KRİTİK: Kullanıcının organization_id'si ile istenen organization_id eşleşmeli
        if user_org_id != requested_org_id:
            logger.warning(f"✗ [JOIN_ORG] Authorization failed: User {session.get('username')} (org: {user_org_id}) tried to join org {requested_org_id}")
            await sio.emit('error', {'message': 'Unauthorized: Cannot join this organization'}, room=sid)
            return
        
        # Doğrulama başarılı - odaya katıl
        room_name = f"org_{requested_org_id}"
        await sio.enter_room(sid, room_name)
        logger.info(f"✓ [JOIN_ORG] Client {sid} (user: {session.get('username')}) joined organization room: {room_name}")
        
        await sio.emit('joined_organization', {'organization_id': requested_org_id}, room=sid)
        logger.info(f"✓ [JOIN_ORG] Sent joined_organization confirmation to {sid}")
        
    except Exception as e:
        logger.error(f"✗ [JOIN_ORG] Error in join_organization: {e}", exc_info=True)
        await sio.emit('error', {'message': 'Internal server error'}, room=sid)

@sio.event
async def leave_organization(sid, data):
    """Leave organization room"""
    organization_id = data.get('organization_id')
    if organization_id:
        await sio.leave_room(sid, f"org_{organization_id}")
        logger.info(f"Client {sid} left organization room: org_{organization_id}")

@sio.event
async def join_public_organization(sid, data):
    """Public sayfa için organization room'a katıl - slug ile doğrulama"""
    logger.info(f"🌐 [JOIN_PUBLIC] join_public_organization event received from {sid} with data: {data}")
    
    try:
        slug = data.get('slug')
        organization_id = data.get('organization_id')
        
        if not slug and not organization_id:
            logger.warning(f"✗ [JOIN_PUBLIC] join_public_organization called without slug or organization_id from {sid}")
            await sio.emit('error', {'message': 'slug or organization_id required'}, room=sid)
            return
        
        # Database connection - app instance'ından al
        from motor.motor_asyncio import AsyncIOMotorDatabase
        global _app_instance
        if _app_instance is None:
            logger.error(f"✗ [JOIN_PUBLIC] App instance not available")
            await sio.emit('error', {'message': 'Internal server error'}, room=sid)
            return
        
        db = getattr(_app_instance, 'db', None)
        if db is None:
            logger.error(f"✗ [JOIN_PUBLIC] Database not available")
            await sio.emit('error', {'message': 'Internal server error'}, room=sid)
            return
        
        # Slug varsa organization_id'yi bul
        if slug and not organization_id:
            admin_user = await db.users.find_one({"slug": slug}, {"_id": 0, "organization_id": 1})
            if not admin_user:
                logger.warning(f"✗ [JOIN_PUBLIC] Slug not found: {slug}")
                await sio.emit('error', {'message': 'Invalid slug'}, room=sid)
                return
            organization_id = admin_user.get('organization_id')
        
        if not organization_id:
            logger.warning(f"✗ [JOIN_PUBLIC] organization_id not found for slug: {slug}")
            await sio.emit('error', {'message': 'Invalid organization'}, room=sid)
            return
        
        # Slug ile organization_id doğrulaması yap (güvenlik için)
        if slug:
            admin_user = await db.users.find_one({"slug": slug, "organization_id": organization_id}, {"_id": 0, "organization_id": 1})
            if not admin_user:
                logger.warning(f"✗ [JOIN_PUBLIC] Slug and organization_id mismatch: {slug} vs {organization_id}")
                await sio.emit('error', {'message': 'Invalid slug or organization_id'}, room=sid)
                return
        
        # Public room'a katıl (org_{organization_id} - aynı room, authenticated ve public client'lar birlikte)
        room_name = f"org_{organization_id}"
        await sio.enter_room(sid, room_name)
        logger.info(f"✓ [JOIN_PUBLIC] Client {sid} joined public organization room: {room_name}")
        
        await sio.emit('joined_public_organization', {'organization_id': organization_id}, room=sid)
        logger.info(f"✓ [JOIN_PUBLIC] Sent joined_public_organization confirmation to {sid}")
        
    except Exception as e:
        logger.error(f"✗ [JOIN_PUBLIC] Error in join_public_organization: {e}", exc_info=True)
        await sio.emit('error', {'message': 'Internal server error'}, room=sid)

# === VOICE AI WEBSOCKET HANDLERS ===

# Active voice sessions dictionary
_voice_sessions = {}

async def _voice_receive_loop(sid, voice_session, voice_service):
    """
    Background task: AI'dan gelen sesleri sürekli dinle ve client'a gönder
    """
    try:
        logger.info(f"🔊 [VOICE] Receive loop started for {sid}")
        
        loop_count = 0
        while sid in _voice_sessions:
            loop_count += 1
            logger.info(f"🔄 [VOICE] Receive loop iteration {loop_count} for {sid}")
            
            # AI'dan ses cevabını al
            logger.info(f"⏳ [VOICE] Waiting for AI response for {sid}...")
            response_audio = await voice_service.receive_audio_response(voice_session)
            logger.info(f"✅ [VOICE] Received response from AI for {sid}, has_audio: {bool(response_audio)}")
            
            if response_audio:
                # Client'a ses cevabını gönder
                logger.info(f"📤 [VOICE] Sending audio response to {sid}, size: {len(response_audio)}")
                await sio.emit('voice_response', {
                    'audio': response_audio
                }, room=sid)
                
                logger.info(f"✅ [VOICE] Audio response sent to {sid}")
            else:
                logger.warning(f"⚠️ [VOICE] No audio in response for {sid}")
            
            # Küçük bir bekleme (CPU'yu aşırı yüklememek için)
            await asyncio.sleep(0.01)
    
    except asyncio.CancelledError:
        logger.info(f"🛑 [VOICE] Receive loop cancelled for {sid}")
        raise
    except Exception as e:
        logger.error(f"❌ [VOICE] Receive loop error for {sid}: {e}", exc_info=True)
        await sio.emit('voice_error', {
            'message': f'Voice receive error: {str(e)}'
        }, room=sid)

@sio.on('voice_start')
async def handle_voice_start(sid, data):
    """
    Sesli görüşme oturumunu başlat
    
    Client'tan gelen data:
    {
        "organization_id": "...",
        "user_role": "admin",
        "username": "..."
    }
    """
    try:
        logger.info(f"🎤 [VOICE] Voice session start request from {sid}")
        
        # Session kontrolü
        session_data = await sio.get_session(sid)
        if not session_data:
            await sio.emit('voice_error', {
                'message': 'Not authenticated'
            }, room=sid)
            return
        
        organization_id = data.get('organization_id')
        user_role = data.get('user_role', 'staff')
        username = data.get('username', 'User')
        
        # Voice AI service'i al
        voice_service = get_voice_ai_service()
        if not voice_service:
            await sio.emit('voice_error', {
                'message': 'Voice AI service not available'
            }, room=sid)
            return
        
        # System instruction oluştur (metin chatbot'takine benzer)
        system_instruction = f"""
Sen PLANN Asistan'sın. Randevu yönetim sistemi için sesli asistansın.

Organizasyon: {organization_id}
Kullanıcı Rolü: {user_role}
Kullanıcı: {username}

Görevlerin:
- Randevu bilgilerini sesli olarak açıkla
- Kullanıcı sorularına doğal ve akıcı yanıt ver
- Gerektiğinde metin chatbot'a yönlendir (karmaşık işlemler için)

Yanıtların kısa, net ve doğal olsun. Türkçe konuş.
"""
        
        # Yeni sesli oturum oluştur
        voice_session = await voice_service.create_session(
            system_instruction=system_instruction
        )
        
        # Background receive loop başlat
        receive_task = asyncio.create_task(
            _voice_receive_loop(sid, voice_session, voice_service)
        )
        
        # Session'ı kaydet
        _voice_sessions[sid] = {
            'session': voice_session,
            'organization_id': organization_id,
            'user_role': user_role,
            'username': username,
            'receive_task': receive_task  # Task'ı kaydet (cleanup için)
        }
        
        # Client'a başarı bildirimi gönder
        await sio.emit('voice_ready', {
            'status': 'ready',
            'message': 'Voice session initialized'
        }, room=sid)
        
        logger.info(f"✅ [VOICE] Voice session created for {sid}")
    
    except Exception as e:
        logger.error(f"❌ [VOICE] Error starting voice session: {e}", exc_info=True)
        await sio.emit('voice_error', {
            'message': f'Failed to start voice session: {str(e)}'
        }, room=sid)


@sio.on('voice_audio')
async def handle_voice_audio(sid, data):
    """
    Kullanıcıdan gelen sesi işle ve AI cevabını döndür
    
    Client'tan gelen data:
    {
        "audio": "BASE64_ENCODED_AUDIO_DATA"
    }
    """
    logger.info(f"🎤 [VOICE] handle_voice_audio called for {sid}")
    try:
        # Session kontrolü
        logger.info(f"🔍 [VOICE] Checking session for {sid}, active sessions: {list(_voice_sessions.keys())}")
        if sid not in _voice_sessions:
            logger.warning(f"⚠️ [VOICE] No session found for {sid}")
            await sio.emit('voice_error', {
                'message': 'No active voice session'
            }, room=sid)
            return
        
        logger.info(f"✅ [VOICE] Session found for {sid}")
        
        audio_base64 = data.get('audio')
        if not audio_base64:
            logger.warning(f"⚠️ [VOICE] No audio data in request from {sid}")
            await sio.emit('voice_error', {
                'message': 'No audio data provided'
            }, room=sid)
            return
        
        logger.info(f"🎤 [VOICE] Audio received from {sid}: {len(audio_base64)} chars")
        
        # Voice service ve session'ı al
        voice_service = get_voice_ai_service()
        session_info = _voice_sessions[sid]
        voice_session = session_info['session']
        
        # AI'ya sesi gönder (sadece gönder, receive loop zaten çalışıyor)
        logger.info(f"📨 [VOICE] Calling send_audio for {sid}...")
        await voice_service.send_audio(voice_session, audio_base64)
        
        logger.info(f"✅ [VOICE] Audio sent to AI from {sid}")
        
        # Receive loop otomatik olarak cevabı gönderecek
        # Burada await etmeye gerek yok
    
    except Exception as e:
        logger.error(f"❌ [VOICE] Error processing audio for {sid}: {e}", exc_info=True)
        await sio.emit('voice_error', {
            'message': f'Failed to process audio: {str(e)}'
        }, room=sid)


@sio.on('voice_stop')
async def handle_voice_stop(sid):
    """
    Sesli görüşme oturumunu kapat
    """
    try:
        if sid not in _voice_sessions:
            return
        
        logger.info(f"🛑 [VOICE] Voice session stop request from {sid}")
        
        # Voice service ve session'ı al
        voice_service = get_voice_ai_service()
        session_info = _voice_sessions[sid]
        voice_session = session_info['session']
        receive_task = session_info.get('receive_task')
        
        # Receive loop'u durdur
        if receive_task and not receive_task.done():
            receive_task.cancel()
            try:
                await receive_task
            except asyncio.CancelledError:
                pass
        
        # Session'ı kapat
        await voice_service.close_session(voice_session)
        
        # Session'ı sil
        del _voice_sessions[sid]
        
        # Client'a bildirim gönder
        await sio.emit('voice_stopped', {
            'status': 'stopped',
            'message': 'Voice session closed'
        }, room=sid)
        
        logger.info(f"✅ [VOICE] Voice session closed for {sid}")
    
    except Exception as e:
        logger.error(f"❌ [VOICE] Error stopping voice session: {e}", exc_info=True)

# Helper function to make data JSON serializable (convert ObjectId, datetime, etc.)
def make_json_serializable(obj):
    """Recursively convert MongoDB ObjectId and datetime objects to JSON-serializable types"""
    import json
    from bson import ObjectId
    from datetime import datetime, date
    
    if isinstance(obj, ObjectId):
        return str(obj)
    elif isinstance(obj, (datetime, date)):
        return obj.isoformat() if hasattr(obj, 'isoformat') else str(obj)
    elif isinstance(obj, dict):
        return {key: make_json_serializable(value) for key, value in obj.items()}
    elif isinstance(obj, list):
        return [make_json_serializable(item) for item in obj]
    elif isinstance(obj, tuple):
        return tuple(make_json_serializable(item) for item in obj)
    elif isinstance(obj, set):
        return {make_json_serializable(item) for item in obj}
    else:
        # Try to convert to JSON-serializable type
        try:
            json.dumps(obj)
            return obj
        except (TypeError, ValueError):
            # If not serializable, convert to string
            return str(obj)

# Helper function to emit events to organization rooms
async def emit_to_organization(organization_id: str, event: str, data: dict):
    """Emit event to all clients in an organization room"""
    try:
        room_name = f"org_{organization_id}"
        logger.info(f"📤 [EMIT] About to emit {event} to room {room_name}")
        
        # Convert data to JSON-serializable format
        serializable_data = make_json_serializable(data)
        
        # Get all sockets in the room to verify
        try:
            # Note: Socket.IO doesn't have a direct way to list room members, but we can try to emit
            await sio.emit(event, serializable_data, room=room_name)
            logger.info(f"✓ [EMIT] Successfully emitted {event} to room {room_name} with data keys: {list(serializable_data.keys())}")
            
            # Debug: Try to get room info (if available in python-socketio)
            try:
                # Check if room exists by trying to get room info
                # Note: python-socketio AsyncServer doesn't expose room member count directly
                # But we can log that we attempted the emit
                logger.info(f"🔍 [EMIT] Event {event} sent to room {room_name} - waiting for client receipt")
            except Exception as debug_error:
                logger.warning(f"⚠ [EMIT] Debug check failed: {debug_error}")
        except Exception as emit_error:
            logger.error(f"✗ [EMIT] Error during emit to {room_name}: {emit_error}", exc_info=True)
            raise
    except Exception as e:
        logger.error(f"✗ [EMIT] Error emitting {event} to org_{organization_id}: {e}", exc_info=True)

# === GÜVENLİK YARDIMCI FONKSİYONLARI (Aynı kaldı) ===
_mongo_client = None; _mongo_db = None
async def ensure_db_connection(request: Request):
    global _mongo_client, _mongo_db; db = getattr(request.app, 'db', None)
    if db is None and _mongo_db is None:
        mongo_url = os.environ.get('MONGO_URL'); db_name = os.environ.get('DB_NAME', 'royal_koltuk_dev')
        if not mongo_url: raise HTTPException(status_code=503, detail="Database connection not configured.")
        try:
            new_client = AsyncIOMotorClient(mongo_url, serverSelectionTimeoutMS=5000); await new_client.admin.command('ping')
            new_db = new_client[db_name]; _mongo_client = new_client; _mongo_db = new_db
            request.app.mongodb_client = new_client; request.app.db = new_db
            logging.info(f"MongoDB connection established (lazy initialization) to {db_name}")
        except Exception as e:
            logging.error(f"Failed to establish MongoDB connection: {e}"); raise HTTPException(status_code=503, detail="Database connection failed.")
    elif db is None and _mongo_db is not None:
        request.app.mongodb_client = _mongo_client; request.app.db = _mongo_db; db = _mongo_db
    try:
        client_to_check = getattr(request.app, 'mongodb_client', _mongo_client)
        if client_to_check: await client_to_check.admin.command('ping')
    except Exception as e:
        logging.warning(f"MongoDB connection check failed: {e}, reconnecting..."); _mongo_client = None; _mongo_db = None; await ensure_db_connection(request) 
async def get_db(request: Request):
    await ensure_db_connection(request); db = getattr(request.app, 'db', None)
    if db is None: raise HTTPException(status_code=503, detail="Database connection failed.")
    return db
async def get_db_from_request(request: Request):
    await ensure_db_connection(request); db = getattr(request.app, 'db', None)
    if db is None: raise HTTPException(status_code=503, detail="Database connection failed.")
    return db
def verify_password(plain_password, hashed_password): return pwd_context.verify(plain_password, hashed_password)
def get_password_hash(password): return pwd_context.hash(password)
def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta: expire = datetime.now(timezone.utc) + expires_delta
    else: expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire}); encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt
async def _ensure_user_identity_fields(db, user: dict) -> dict:
    """
    Lazy backfill: mevcut kullanıcılarda eksik olan stable `user_id` ve `activation_state`
    alanlarını ilk okuma anında doldurur. Migration script gerekmez.

    `user_id` (UUID): PostHog distinct_id'nin sabit kaynağı; username değişse bile kalır.
    `activation_state`: Aha onboarding state machine.
    """
    updates = {}
    if not user.get("user_id"):
        updates["user_id"] = str(uuid.uuid4())
    if not user.get("activation_state"):
        role = user.get("role") or "admin"
        onboarding_done = bool(user.get("onboarding_completed"))
        if role != "admin":
            updates["activation_state"] = "not_applicable"
        elif onboarding_done:
            # Mevcut admin zaten klasik turu görmüş — sürpriz overlay olmasın
            updates["activation_state"] = "aha_skipped"
            updates["aha_skipped_reason"] = "legacy_backfill"
        else:
            updates["activation_state"] = "pending_aha_appointment"
    if updates:
        try:
            await db.users.update_one(
                {"username": user.get("username")},
                {"$set": updates},
            )
            user.update(updates)
        except Exception as e:
            logging.warning(f"Identity backfill failed for {user.get('username')}: {e}")
    return user


async def get_user_from_db(request: Request, username: str, db=None):
    if db is None:
        await ensure_db_connection(request); db = getattr(request.app, 'db', None)
        if db is None: raise HTTPException(status_code=503, detail="Database connection failed.")
    
    # Önce tam eşleşme dene
    user = await db.users.find_one({"username": username}, {"_id": 0})
    if user:
        try:
            user = await _ensure_user_identity_fields(db, user)
            return UserInDB(**user)
        except Exception as e: logging.warning(f"Kullanıcı veritabanında, ancak UserInDB modeline uymuyor: {e}"); return None
    
    # Tam eşleşme yoksa, case-insensitive arama yap (email için)
    if "@" in username:  # Email adresi gibi görünüyorsa
        import re
        user = await db.users.find_one({"username": {"$regex": f"^{re.escape(username)}$", "$options": "i"}}, {"_id": 0})
        if user:
            try:
                user = await _ensure_user_identity_fields(db, user)
                return UserInDB(**user)
            except Exception as e: logging.warning(f"Kullanıcı veritabanında (case-insensitive), ancak UserInDB modeline uymuyor: {e}"); return None
    
    return None
async def get_current_user(request: Request, token: str = Depends(oauth2_scheme), db = Depends(get_db)):
    credentials_exception = HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Could not validate credentials", headers={"WWW-Authenticate": "Bearer"})
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM]); username: str = payload.get("sub")
        if username is None: raise credentials_exception
    except JWTError: raise credentials_exception
    user = await get_user_from_db(request, username, db=db) 
    if user is None: raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not found")
    return user

async def get_superadmin_user(request: Request, token: str = Depends(oauth2_scheme), db = Depends(get_db)):
    """Sadece superadmin rolüne sahip kullanıcılar için dependency"""
    user = await get_current_user(request, token, db)
    if user.role != "superadmin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Bu işlem için superadmin yetkisi gereklidir")
    return user

async def get_marketing_user(request: Request, token: str = Depends(oauth2_scheme), db = Depends(get_db)):
    """Sadece marketing rolüne sahip kullanıcılar için dependency"""
    user = await get_current_user(request, token, db)
    if user.role != "marketing":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Bu işlem için marketing yetkisi gereklidir")
    return user

# --- KOTA YÖNETİM FONKSİYONLARI ---
async def get_organization_plan(db, organization_id: str) -> Optional[dict]:
    """Organization'ın plan bilgisini getir. Yoksa trial oluştur."""
    plan_doc = await db.organization_plans.find_one({"organization_id": organization_id})
    if not plan_doc:
        # Yeni kayıt - Trial paketi oluştur
        trial_start = datetime.now(timezone.utc)
        trial_end = trial_start + timedelta(days=7)
        quota_reset = trial_start + timedelta(days=30)
        new_plan = OrganizationPlan(
            organization_id=organization_id,
            plan_id="tier_trial",
            quota_usage=0,
            quota_reset_date=quota_reset,
            quota_last_reset_date=trial_start,  # Lazy reset için başlangıç tarihi
            trial_start_date=trial_start,
            trial_end_date=trial_end,
            is_first_month=True
        )
        plan_doc = new_plan.model_dump()
        plan_doc['trial_start_date'] = plan_doc['trial_start_date'].isoformat()
        plan_doc['trial_end_date'] = plan_doc['trial_end_date'].isoformat()
        plan_doc['quota_reset_date'] = plan_doc['quota_reset_date'].isoformat()
        plan_doc['created_at'] = plan_doc['created_at'].isoformat()
        plan_doc['updated_at'] = plan_doc['updated_at'].isoformat()
        await db.organization_plans.insert_one(plan_doc)
        return plan_doc
    
    # Datetime string'lerini parse et
    if isinstance(plan_doc.get('trial_start_date'), str):
        plan_doc['trial_start_date'] = datetime.fromisoformat(plan_doc['trial_start_date'].replace('Z', '+00:00'))
    if isinstance(plan_doc.get('trial_end_date'), str):
        plan_doc['trial_end_date'] = datetime.fromisoformat(plan_doc['trial_end_date'].replace('Z', '+00:00'))
    if isinstance(plan_doc.get('quota_reset_date'), str):
        plan_doc['quota_reset_date'] = datetime.fromisoformat(plan_doc['quota_reset_date'].replace('Z', '+00:00'))
    
    return plan_doc

async def check_and_reset_quota(db, organization_id: str) -> None:
    """
    TASK 2: Lazy Quota Reset - Yearly kullanıcılar için aylık reset kontrolü.
    
    Stripe yearly subscription'lar için aylık webhook göndermediği için,
    her ay başında quota'yu sıfırlamak için internal trigger gerekiyor.
    
    Bu fonksiyon quota_last_reset_date ile mevcut ay/yıl'ı karşılaştırır.
    Eğer ay değişmişse, quota_usage'i sıfırlar.
    """
    plan_doc = await get_organization_plan(db, organization_id)
    if not plan_doc:
        return
    
    quota_last_reset = plan_doc.get('quota_last_reset_date')
    if not quota_last_reset:
        # quota_last_reset_date yoksa, şimdi olarak ayarla
        now = datetime.now(timezone.utc)
        await db.organization_plans.update_one(
            {"organization_id": organization_id},
            {"$set": {"quota_last_reset_date": now.isoformat()}}
        )
        return
    
    # String ise datetime'a çevir
    if isinstance(quota_last_reset, str):
        quota_last_reset = datetime.fromisoformat(quota_last_reset.replace('Z', '+00:00'))
    
    # Mevcut tarih
    now = datetime.now(timezone.utc)
    
    # Ay ve yıl karşılaştırması
    last_reset_month = quota_last_reset.month
    last_reset_year = quota_last_reset.year
    current_month = now.month
    current_year = now.year
    
    # Ay değişmiş mi kontrol et
    if (current_year > last_reset_year) or (current_year == last_reset_year and current_month > last_reset_month):
        # Ay değişmiş, quota'yu sıfırla
        logger.info(f"🔄 Monthly quota reset triggered for Org ID: {organization_id} (Last reset: {last_reset_month}/{last_reset_year}, Current: {current_month}/{current_year})")
        
        await db.organization_plans.update_one(
            {"organization_id": organization_id},
            {
                "$set": {
                    "quota_usage": 0,
                    "quota_last_reset_date": now.isoformat(),
                    "updated_at": now.isoformat()
                }
            }
        )

async def check_quota_and_increment(db, organization_id: str) -> tuple[bool, str]:
    """Kota kontrolü yap ve kullanılırsa artır. (success, error_message)"""
    # TASK 2: Lazy quota reset kontrolü (yearly kullanıcılar için)
    await check_and_reset_quota(db, organization_id)
    
    plan_doc = await get_organization_plan(db, organization_id)
    if not plan_doc:
        return False, "Plan bilgisi bulunamadı"
    
    plan_id = plan_doc.get('plan_id', 'tier_trial')
    plan_info = next((p for p in PLANS if p['id'] == plan_id), None)
    if not plan_info:
        return False, "Plan bilgisi geçersiz"
    
    # Trial kontrolü
    if plan_id == 'tier_trial':
        trial_end = plan_doc.get('trial_end_date')
        if isinstance(trial_end, str):
            trial_end = datetime.fromisoformat(trial_end.replace('Z', '+00:00'))
        if trial_end and datetime.now(timezone.utc) > trial_end:
            return False, "Deneme süreniz doldu. Devam etmek için lütfen bir paket seçin."
    
    # GÜVENLİK: Ücretli aboneliklerde süre + grace period (SUBSCRIPTION_GRACE_DAYS)
    # kontrolü. Bitiş tarihi geçip toleransı da aşan abonelikler, kotaları kalsa
    # bile randevu oluşturamaz. SuperAdmin paneliyle (_compute_subscription_status)
    # birebir tutarlı davranır.
    if plan_id != 'tier_trial':
        _now = datetime.now(timezone.utc)
        _has_deadline = any(
            _parse_plan_dt(plan_doc.get(_k)) is not None
            for _k in ("next_billing_date", "next_payment_date", "quota_reset_date")
        )
        if not _has_deadline:
            # Fail-open: tarih yoksa engelleme yok ama sessiz de geçme.
            logger.warning("UYARI: %s için abonelik bitiş tarihi bulunamadı!", organization_id)
        else:
            _, _sub_key, _ = _compute_subscription_status(plan_doc, _now)
            if _sub_key == "expired":
                return False, "Abonelik süreniz doldu, lütfen ödeme yönteminizi güncelleyin."
    
    # Kota reset kontrolü (eski mantık - quota_reset_date bazlı)
    quota_reset = plan_doc.get('quota_reset_date')
    if isinstance(quota_reset, str):
        quota_reset = datetime.fromisoformat(quota_reset.replace('Z', '+00:00'))
    
    current_usage = plan_doc.get('quota_usage', 0)
    
    # quota_limit'i plan_doc'dan al (webhook'tan kaydedilmiş olmalı)
    # Fallback: plan_info'dan al
    quota_limit = plan_doc.get('quota_limit')
    if not quota_limit:
        quota_limit = plan_info.get('quota_monthly_appointments', 50)
        logger.warning(f"⚠️ quota_limit plan_doc'da yok, plan_info'dan alındı: {quota_limit}")
    # Plan tanımı sınırsız ise (-1) DB kaydından bağımsız olarak override et
    if plan_info.get('quota_monthly_appointments') == -1:
        quota_limit = -1
    
    # Eğer reset tarihi geçmişse, kullanımı sıfırla
    if quota_reset and datetime.now(timezone.utc) > quota_reset:
        current_usage = 0
        # Yeni reset tarihi ayarla (bir ay sonra)
        new_reset_date = datetime.now(timezone.utc) + timedelta(days=30)
        await db.organization_plans.update_one(
            {"organization_id": organization_id},
            {
                "$set": {
                    "quota_usage": 0,
                    "ai_usage_count": 0,  # AI mesaj kotasını da sıfırla
                    "quota_reset_date": new_reset_date.isoformat(),
                    "is_first_month": False,  # İlk ay indirimi sadece bir kez
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
    
    # Kota kontrolü (-1 = sınırsız)
    if quota_limit != -1 and current_usage >= quota_limit:
        return False, f"Aylık randevu limitinize ulaştınız ({quota_limit} randevu). Paketinizi yükseltmeniz gerekmektedir."
    
    # Kullanımı artır
    await db.organization_plans.update_one(
        {"organization_id": organization_id},
        {
            "$set": {
                "quota_usage": current_usage + 1,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    return True, ""

async def get_plan_info(plan_id: str) -> Optional[dict]:
    """Plan bilgisini getir"""
    return next((p for p in PLANS if p['id'] == plan_id), None)

# --- TÜRKÇE TARİH FORMATI (push bildirimleri için) ---
_TR_MONTHS = {
    1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan", 5: "Mayıs", 6: "Haziran",
    7: "Temmuz", 8: "Ağustos", 9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık",
}
_TR_WEEKDAYS = {
    0: "Pazartesi", 1: "Salı", 2: "Çarşamba", 3: "Perşembe",
    4: "Cuma", 5: "Cumartesi", 6: "Pazar",
}

def format_date_tr(date_str: str) -> str:
    """'YYYY-MM-DD' -> '14 Mayıs Çarşamba' (Türkçe gün/ay adlarıyla).

    locale bağımsız (container'da tr_TR yüklü olmayabilir) — manuel sözlük.
    Parse edilemezse girdiyi olduğu gibi döndürür.
    """
    try:
        d = datetime.strptime(str(date_str), "%Y-%m-%d")
        return f"{d.day} {_TR_MONTHS[d.month]} {_TR_WEEKDAYS[d.weekday()]}"
    except (ValueError, TypeError, KeyError):
        return str(date_str)

# --- SMS FONKSİYONU ---
def build_sms_message(company_name: str, customer_name: str, date: str, time: str, service: str, support_phone: str, hours_until: Optional[float] = None, sms_type: str = "confirmation") -> str:
    """SMS mesajı oluşturur. Template desteği kaldırıldı, sadece default format kullanılıyor."""
    # Tarih formatını YYYY-MM-DD'den DD.MM.YYYY'ye çevir
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%d.%m.%Y")
    except (ValueError, TypeError):
        # Eğer parse edilemezse olduğu gibi kullan
        formatted_date = date
    
    if sms_type == "cancellation":
        # İptal SMS'i için default
        return f"{company_name}: Randevunuz iptal edildi.\nHizmet: {service}\nTarih: {formatted_date}\nSaat: {time}\nBilgi: {support_phone}"
    elif hours_until is not None:
        # Hatırlatma SMS'i için default
        return f"Sayın {customer_name},\n{company_name} randevunuzu hatırlatmak isteriz.\nHizmet: {service}\nTarih: {formatted_date}\nSaat: {time}\nBilgi/İptal: {support_phone}"
    else:
        # Onay SMS'i için default
        return f"{company_name}: Randevunuz onaylandı.\nHizmet: {service}\nTarih: {formatted_date}\nSaat: {time}\nBilgi/İptal: {support_phone}"

def send_sms(to_phone: str, message: str):
    try:
        if not SMS_ENABLED: logging.info("SMS sending is disabled via SMS_ENABLED env. Skipping."); return True
        clean_phone = re.sub(r'\D', '', to_phone); 
        if clean_phone.startswith('90'): clean_phone = clean_phone[2:]
        if clean_phone.startswith('0'): clean_phone = clean_phone[1:]
        if not clean_phone.startswith('5') or len(clean_phone) != 10: logging.error(f"Invalid Turkish phone number format: {to_phone} -> {clean_phone}"); return False
        
        # Newline karakterlerini koru, sadece fazla boşlukları temizle
        # Önce newline'ları geçici bir karakterle değiştir
        temp_message = message.replace('\n', '|||NEWLINE|||')
        # Fazla boşlukları temizle
        temp_message = re.sub(r'[ \t]+', ' ', temp_message)
        # Newline'ları geri getir
        sanitized = temp_message.replace('|||NEWLINE|||', '\n').strip()
        MAX_LEN = 480
        if len(sanitized) > MAX_LEN: sanitized = sanitized[:MAX_LEN]
            
        api_url = "https://api.iletimerkezi.com/v1/send-sms/get/"
        params = {
            'key': ILETIMERKEZI_API_KEY, 'hash': ILETIMERKEZI_HASH, 'text': sanitized,
            'receipents': clean_phone, 'sender': ILETIMERKEZI_SENDER, 
            'iys': '1', 'iysList': 'BIREYSEL'
        }
        response = requests.get(api_url, params=params, timeout=10)
        try:
            root = ET.fromstring(response.text); status_code = root.find('.//status/code').text
            status_message = root.find('.//status/message').text
            if status_code == '200': logging.info(f"SMS sent successfully to {clean_phone} (Title: {ILETIMERKEZI_SENDER})."); return True
            else: logging.error(f"SMS failed to {clean_phone} (Title: {ILETIMERKEZI_SENDER}). Code: {status_code}, Message: {status_message}"); return False
        except ET.ParseError as e:
            logging.error(f"Failed to parse İletimerkezi response (status={response.status_code}): {response.text} | Error: {str(e)}"); return False
    except Exception as e:
        logging.error(f"Failed to send SMS to {to_phone}: {str(e)}"); return False

# === YARDIMCI FONKSİYONLAR ===
def slugify(text: str) -> str:
    """Türkçe karakterleri dönüştürerek URL-friendly slug oluşturur"""
    turkish_map = {
        'ı': 'i', 'İ': 'i', 'ğ': 'g', 'Ğ': 'g', 'ü': 'u', 'Ü': 'u',
        'ş': 's', 'Ş': 's', 'ö': 'o', 'Ö': 'o', 'ç': 'c', 'Ç': 'c'
    }
    text = text.lower()
    for turkish_char, latin_char in turkish_map.items():
        text = text.replace(turkish_char, latin_char)
    text = re.sub(r'[^a-z0-9]+', '', text)
    return text

# === AUDIT LOG HELPER ===
def clean_dict_for_audit(data: Optional[dict]) -> Optional[dict]:
    """MongoDB ObjectID'lerini temizle"""
    if not data:
        return data
    cleaned = {}
    for key, value in data.items():
        if key == '_id':
            continue  # MongoDB _id'yi atla
        if isinstance(value, dict):
            cleaned[key] = clean_dict_for_audit(value)
        elif isinstance(value, list):
            cleaned[key] = [clean_dict_for_audit(item) if isinstance(item, dict) else item for item in value]
        else:
            cleaned[key] = value
    return cleaned

async def create_audit_log(
    db,
    organization_id: str,
    user_id: str,
    user_full_name: str,
    action: str,
    resource_type: str,
    resource_id: str,
    old_value: Optional[dict] = None,
    new_value: Optional[dict] = None,
    ip_address: Optional[str] = None
):
    """Denetim günlüğü kaydı oluştur"""
    if not AUDIT_LOGS_ENABLED:
        return
    try:
        # Clean values
        cleaned_old = clean_dict_for_audit(old_value)
        cleaned_new = clean_dict_for_audit(new_value)
        
        audit_log = AuditLog(
            organization_id=organization_id,
            user_id=user_id,
            user_full_name=user_full_name,
            action=action,
            resource_type=resource_type,
            resource_id=resource_id,
            old_value=cleaned_old,
            new_value=cleaned_new,
            ip_address=ip_address
        )
        doc = audit_log.model_dump()
        doc['timestamp'] = doc['timestamp'].isoformat()
        await db.audit_logs.insert_one(doc)
        logger.info(f"Audit log created: {action} {resource_type} by {user_id}")
    except Exception as e:
        logger.error(f"Failed to create audit log: {e}")

# === ÇOKLU HİZMET: Cache Invalidation Yardımcıları ===
async def _invalidate_redis_pattern(request: Request, pattern: str):
    """Belirtilen pattern'e uyan tüm Redis anahtarlarını siler."""
    redis_client = getattr(request.app.state, 'redis_client', None) if hasattr(request, 'app') else None
    if not redis_client:
        return
    try:
        keys = await redis_client.keys(pattern)
        if keys:
            await redis_client.delete(*keys)
            logging.info(f"🧹 CACHE TEMİZLENDİ: {len(keys)} anahtar (pattern: {pattern})")
    except Exception as e:
        logging.error(f"Redis pattern silme hatası ({pattern}): {e}")

async def invalidate_fee_cache(request: Request, org_id: str):
    """Fee-preview cache'ini (org bazlı) temizler. Fiyat/komisyon/kur değişiminde çağrılmalı."""
    await _invalidate_redis_pattern(request, f"plann:fee:{org_id}:*")

async def invalidate_availability_cache(request: Request, org_id: str):
    """Availability cache'ini (org bazlı) temizler. Randevu create/update/delete'te çağrılmalı."""
    await _invalidate_redis_pattern(request, f"plann:avail:{org_id}:*")

async def invalidate_service_caches(request: Request, org_id: str):
    """Hizmet/kategori değişiminde fee + availability cache'ini birlikte temizler."""
    await invalidate_fee_cache(request, org_id)
    await invalidate_availability_cache(request, org_id)


# === VERİ MODELLERİ (Aynı kaldı) ===
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: Optional[str] = None; user_id: Optional[str] = None; username: str; full_name: Optional[str] = None; language: Optional[str] = None; organization_id: str = Field(default_factory=lambda: str(uuid.uuid4())); role: str = "admin"; slug: Optional[str] = None; permitted_service_ids: List[str] = []; payment_type: Optional[str] = "salary"; payment_amount: Optional[float] = 0.0; status: Optional[str] = "active"; invitation_token: Optional[str] = None; days_off: List[str] = Field(default_factory=list); onboarding_completed: bool = False; activation_state: Optional[str] = None; aha_skipped_reason: Optional[str] = None; aha_skipped_at: Optional[str] = None; aha_wa_message_id: Optional[str] = None; aha_wa_last_error: Optional[str] = None; aha_appointment_id: Optional[str] = None; first_appointment_at: Optional[str] = None; breaks: List[dict] = Field(default_factory=list); can_view_all_appointments: bool = False
class UserInDB(User): hashed_password: Optional[str] = None
class UserCreate(BaseModel): username: str; password: str; full_name: Optional[str] = None; organization_name: Optional[str] = None; support_phone: Optional[str] = None; sector: Optional[str] = None
class Token(BaseModel): access_token: str; token_type: str
class ForgotPasswordRequest(BaseModel): username: str
class ResetPasswordRequest(BaseModel): token: str; new_password: str
class SetupPasswordRequest(BaseModel): token: str; new_password: str
class SsoCodeResponse(BaseModel): code: str; expires_in: int
class SsoExchangeRequest(BaseModel): code: str
class PlanUpdateRequest(BaseModel): 
    plan_id: str
    billing_cycle: str = "monthly"
    platform: str = "web"  # 'web', 'android', 'ios'
    currency: Optional[str] = None  # 'gbp' veya 'try' - otomatik algılanır
class ContactRequest(BaseModel): name: str = Field(..., min_length=1); phone: str = Field(..., min_length=10); email: Optional[str] = None; message: Optional[str] = None
class ContactStatusUpdate(BaseModel): status: Literal["pending", "contacted", "resolved"]
# === Hizmet Kategorileri ===
class Category(BaseModel):
    model_config = ConfigDict(extra="ignore"); organization_id: str; id: str = Field(default_factory=lambda: str(uuid.uuid4())); name: str; order: Optional[int] = None; created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
class CategoryCreate(BaseModel): name: str; order: Optional[int] = None
class CategoryUpdate(BaseModel): name: Optional[str] = None; order: Optional[int] = None
class Service(BaseModel):
    model_config = ConfigDict(extra="ignore"); organization_id: str; id: str = Field(default_factory=lambda: str(uuid.uuid4())); name: str; price: float; duration: int = 30; order: Optional[int] = None; category_id: Optional[str] = None; version: int = 1; is_active: bool = True; session_count: Optional[int] = None; payment_rule: Optional[str] = None; deposit_percentage: Optional[int] = None; deposit_amount: Optional[float] = None; created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
class ServiceCreate(BaseModel): name: str; price: float; duration: int = 30; order: Optional[int] = None; category_id: Optional[str] = None; session_count: Optional[int] = None; payment_rule: Optional[str] = None; deposit_percentage: Optional[int] = None; deposit_amount: Optional[float] = None
class ServiceUpdate(BaseModel): name: Optional[str] = None; price: Optional[float] = None; duration: Optional[int] = None; order: Optional[int] = None; category_id: Optional[str] = None; session_count: Optional[int] = None; payment_rule: Optional[str] = None; deposit_percentage: Optional[int] = None; deposit_amount: Optional[float] = None
# === Çoklu Hizmet: değişmez (immutable) snapshot kalemi ===
class AppointmentService(BaseModel):
    model_config = ConfigDict(extra="ignore")
    service_id: str
    sequence: int = 0          # uygulanma sırası (order/sort_order DEĞİL)
    name_snapshot: str
    slug_snapshot: Optional[str] = None  # analytics için stabil anahtar (örn. "hair-cut")
    duration_snapshot: int = 30
    price_snapshot: float = 0.0
    service_version: int = 1   # hizmet sürümü (Haircut v3/v4 takibi)

class Appointment(BaseModel):
    model_config = ConfigDict(extra="ignore"); organization_id: str; id: str = Field(default_factory=lambda: str(uuid.uuid4())); customer_name: str; phone: str; service_id: str; service_name: str; service_price: float; appointment_date: str; appointment_time: str; notes: str = ""; status: str = "Bekliyor"; staff_member_id: Optional[str] = None; created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc)); completed_at: Optional[str] = None; service_duration: Optional[int] = None; source: Optional[str] = None; session_group_id: Optional[str] = None; session_number: Optional[int] = None; session_total: Optional[int] = None; payment_status: Optional[str] = None; refund_eligible: Optional[bool] = None; services: Optional[List[AppointmentService]] = None; selection_type: str = "INDIVIDUAL"
class AppointmentCreate(BaseModel):
    customer_name: str; phone: str; service_id: Optional[str] = None; service_ids: Optional[List[str]] = None; appointment_date: str; appointment_time: str; notes: str = ""; staff_member_id: Optional[str] = None; turnstile_token: Optional[str] = None; session_group_id: Optional[str] = None; session_number: Optional[int] = None; session_total: Optional[int] = None; payment_status: Optional[str] = None; skip_confirmation_whatsapp: bool = False
class AppointmentUpdate(BaseModel):
    customer_name: Optional[str] = None; phone: Optional[str] = None; address: Optional[str] = None; service_id: Optional[str] = None; service_ids: Optional[List[str]] = None; appointment_date: Optional[str] = None; appointment_time: Optional[str] = None; notes: Optional[str] = None; status: Optional[str] = None; staff_member_id: Optional[str] = None; session_group_id: Optional[str] = None; session_number: Optional[int] = None; session_total: Optional[int] = None; payment_status: Optional[str] = None
class Transaction(BaseModel):
    model_config = ConfigDict(extra="ignore"); organization_id: str; id: str = Field(default_factory=lambda: str(uuid.uuid4())); appointment_id: str; customer_name: str; service_name: str; amount: float; date: str; staff_member_id: Optional[str] = None; created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
class TransactionUpdate(BaseModel): amount: float
class BusinessHoursDay(BaseModel):
    is_open: bool = True
    open_time: str = "09:00"
    close_time: str = "18:00"

class LocationCoordinates(BaseModel):
    lat: float
    lng: float


class BusinessLocation(BaseModel):
    address: str
    coordinates: LocationCoordinates


class Settings(BaseModel):
    model_config = ConfigDict(extra="ignore"); organization_id: str; id: str = Field(default_factory=lambda: str(uuid.uuid4())); work_start_hour: int = 7; work_end_hour: int = 3; appointment_interval: int = 30
    company_name: str = "İşletmeniz"; support_phone: str = "05000000000"; slug: Optional[str] = None; customer_can_choose_staff: bool = False
    logo_url: Optional[str] = None; sms_reminder_hours: float = 1.0; sector: Optional[str] = None; admin_provides_service: bool = False
    images: List[str] = Field(default_factory=list)
    show_service_duration_on_public: bool = True; show_service_price_on_public: bool = True
    multi_service_enabled: bool = False
    break_limit_minutes: int = 60; break_limit_count: int = 2
    # Pazarlama Otopilotu (WhatsApp geri-kazanım): varsayılan KAPALI.
    # marketing_autopilot_days: {service_id: gün}. Haritada olmayan hizmet için
    # varsayılan 21 gün uygulanır. enabled_at, otopilot AÇILDIĞI an set edilir ve
    # yalnızca o andan SONRA tamamlanan randevular kuyruğa alınır (geçmiş patlaması yok).
    marketing_autopilot_enabled: bool = False
    marketing_autopilot_days: dict = Field(default_factory=dict)
    marketing_autopilot_enabled_at: Optional[str] = None
    location: Optional[BusinessLocation] = None
    business_hours: Optional[dict] = Field(default_factory=lambda: {
        "monday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
        "tuesday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
        "wednesday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
        "thursday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
        "friday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
        "saturday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
        "sunday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"}
    })

# Push Notification Subscription Model
class PushSubscription(BaseModel):
    endpoint: str
    keys: dict  # p256dh ve auth keys
    platform: Optional[str] = None # 'android', 'ios' or None (web)

class PushSubscriptionCreate(BaseModel):
    subscription: dict  # {endpoint, keys: {p256dh, auth}, platform}

# Verilen abonelik listesine push teslim eden ortak yardımcı (native FCM + web VAPID)
async def _deliver_push_to_subscriptions(db, subscriptions, title: str, body: str, data: dict = None):
    """Abonelik listesine push gönderir. Döner: gönderim istatistikleri."""
    stats = {"native_sent": 0, "web_sent": 0, "failed": 0, "removed": 0}

    notification_payload = json.dumps({
        "title": title,
        "body": body,
        "icon": "https://plannapp.co/icons/icon-192x192.png",
        "badge": "https://plannapp.co/icons/badge-mono-96x96.png",
        # ÜST SEVİYE url: eski service worker sürümleri notification'ı üst seviye
        # data.url'den kurar (nested data'yı okumaz). Böylece SW güncellenmemiş
        # tarayıcılarda bile derin bağlantı (ör. /?randevu=<id>) çalışır.
        "url": (data or {}).get("url", "/"),
        "data": data or {},
        "timestamp": datetime.now(timezone.utc).isoformat()
    })

    for sub in subscriptions:
        # Native Platform (Android/iOS) - Firebase Cloud Messaging (FCM)
        platform = sub.get("platform")
        if platform in ['android', 'ios']:
            try:
                # FCM data payload: TÜM key/value string olmalı (FCM zorunluluğu)
                fcm_data = {str(k): str(v) for k, v in (data or {}).items() if v is not None}

                message = messaging.Message(
                    token=sub["endpoint"], # Native için endpoint token'dır
                    notification=messaging.Notification(
                        title=title,
                        body=body,
                    ),
                    data=fcm_data, # String key-value olmalı
                    android=messaging.AndroidConfig(
                        priority='high',
                        notification=messaging.AndroidNotification(
                            icon='ic_stat_icon', # Native resource adı
                            color='#2563eb',
                            click_action='FLUTTER_NOTIFICATION_CLICK' # Capacitor için standart
                        ),
                    ),
                    apns=messaging.APNSConfig(
                        payload=messaging.APNSPayload(
                            aps=messaging.Aps(
                                badge=1,
                                sound='default'
                            )
                        )
                    )
                )
                response = messaging.send(message)
                stats["native_sent"] += 1
                logger.info(f"✅ Native Push sent ({platform}): {response}")
            except Exception as e:
                stats["failed"] += 1
                logger.error(f"❌ Native Push Error for {sub.get('user_id', 'unknown')}: {e}")
                if 'registration-token-not-registered' in str(e) or 'invalid-argument' in str(e):
                    await db.push_subscriptions.delete_one({"_id": sub["_id"]})
                    stats["removed"] += 1
                    logger.info(f"Removed invalid native subscription")
            continue # Native gönderildi, web push kısmını atla

        # Web Push (VAPID)
        if not VAPID_PRIVATE_KEY:
            continue
        try:
            subscription_info = {
                "endpoint": sub["endpoint"],
                "keys": sub["keys"]
            }
            webpush(
                subscription_info=subscription_info,
                data=notification_payload,
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims=VAPID_CLAIMS.copy(),
                ttl=86400,
                headers={"Urgency": "high"}
            )
            stats["web_sent"] += 1
            logger.info(f"✅ Web Push sent to user: {sub.get('user_id', 'unknown')}")
        except WebPushException as e:
            stats["failed"] += 1
            logger.error(f"❌ WebPushException for {sub.get('user_id', 'unknown')}: {e}")
            if e.response and e.response.status_code in [404, 410]:
                await db.push_subscriptions.delete_one({"_id": sub["_id"]})
                stats["removed"] += 1
                logger.info(f"Removed invalid push subscription: {sub.get('user_id', 'unknown')}")
        except Exception as e:
            stats["failed"] += 1
            logger.error(f"❌ Unexpected error for {sub.get('user_id', 'unknown')}: {type(e).__name__}: {e}")

    return stats


# Push notification gönderme fonksiyonu
async def send_push_notification(db, organization_id: str, title: str, body: str, data: dict = None, user_id: str = None):
    """Belirli bir organizasyondaki tüm kullanıcılara veya belirli bir kullanıcıya push notification gönder"""
    try:
        # Abonelikleri bul
        query = {"organization_id": organization_id}
        if user_id:
            query["user_id"] = user_id

        subscriptions = await db.push_subscriptions.find(query).to_list(None)

        if not subscriptions:
            logger.info(f"No push subscriptions found for org: {organization_id}")
            return

        logger.info(f"🔔 Sending push to {len(subscriptions)} subscriptions")
        await _deliver_push_to_subscriptions(db, subscriptions, title, body, data)
    except Exception as e:
        logger.error(f"Error sending push notifications: {e}")

# === ABONELİK PAKETLERİ ===
PLANS = [
    {
        "id": "tier_trial",
        "name": "Trial",
        "price_monthly": 0,
        "quota_monthly_appointments": 50,
        "ai_message_limit": 100,
        "trial_days": 7,
        "features": [
            "50 Randevu veya 7 Gün (Hangisi önce)",
            "Randevu Hatırlatma Dahil",
            "Sınırsız Personel",
            "Sınırsız Müşteri",
            "Online Randevu",
            "İstatistikler",
            "Yapay Zeka Akıllı Asistan (Test)"
        ],
        "target_audience_tr": "Yeni kullanıcılar için deneme paketi."
    },
    {
        "id": "tier_1_standard",
        "name": "Standart",
        "price_monthly": 710,
        "price_yearly": 7100,
        "quota_monthly_appointments": 100,
        "ai_message_limit": 500,
        "features": [
            "100 Randevu/Ay",
            "Randevu Hatırlatma Dahil",
            "Sınırsız Personel",
            "Sınırsız Müşteri",
            "Online Randevu",
            "Online Ödeme Alın (Kapora veya Tam Ödeme)",
            "Seans Paketi Yönetimi",
            "İstatistikler",
            "Yapay Zeka Akıllı Asistan (500 Mesaj/Ay)",
            "Akıllı Pazarlama Otopilotu"
        ],
        "target_audience_tr": "Yeni başlayanlar, tek kişilik veya butik işletmeler için ideal başlangıç paketi."
    },
    {
        "id": "tier_2_profesyonel",
        "name": "Profesyonel",
        "price_monthly": 1020,
        "price_yearly": 10200,
        "quota_monthly_appointments": 300,
        "ai_message_limit": 3000,
        "features": [
            "300 Randevu/Ay",
            "Randevu Hatırlatma Dahil",
            "Sınırsız Personel",
            "Sınırsız Müşteri",
            "Online Randevu",
            "Online Ödeme Alın (Kapora veya Tam Ödeme)",
            "Seans Paketi Yönetimi",
            "İstatistikler",
            "Yapay Zeka Akıllı Asistan (3.000 Mesaj/Ay)",
            "Akıllı Pazarlama Otopilotu"
        ],
        "target_audience_tr": "Büyümekte olan ve müşteri kitlesini oturtmaya başlamış salonlar için."
    },
    {
        "id": "tier_3_premium",
        "name": "Premium",
        "price_monthly": 1530,
        "price_yearly": 15300,
        "quota_monthly_appointments": 600,
        "ai_message_limit": 10000,
        "features": [
            "600 Randevu/Ay",
            "Randevu Hatırlatma Dahil",
            "Sınırsız Personel",
            "Sınırsız Müşteri",
            "Online Randevu",
            "Online Ödeme Alın (Kapora veya Tam Ödeme)",
            "Seans Paketi Yönetimi",
            "İstatistikler",
            "Yapay Zeka Akıllı Asistan (10.000 Mesaj/Ay)",
            "Akıllı Pazarlama Otopilotu"
        ],
        "target_audience_tr": "Düzenli ve sabit bir müşteri hacmine sahip, yerleşik işletmeler için."
    },
    {
        "id": "tier_4_business",
        "name": "Business",
        "price_monthly": 1940,
        "price_yearly": 19400,
        "quota_monthly_appointments": 900,
        "ai_message_limit": -1,
        "features": [
            "900 Randevu/Ay",
            "Randevu Hatırlatma Dahil",
            "Sınırsız Personel",
            "Sınırsız Müşteri",
            "Online Randevu",
            "Online Ödeme Alın (Kapora veya Tam Ödeme)",
            "Seans Paketi Yönetimi",
            "İstatistikler",
            "Yapay Zeka Akıllı Asistan (Limitsiz)",
            "Akıllı Pazarlama Otopilotu"
        ],
        "target_audience_tr": "Yoğun tempolu, orta ölçekli salonlar ve merkezler için en popüler seçim."
    },
    {
        "id": "tier_5_enterprise",
        "name": "Enterprise",
        "price_monthly": 2350,
        "price_yearly": 23500,
        "quota_monthly_appointments": 1200,
        "ai_message_limit": -1,
        "features": [
            "1.200 Randevu/Ay",
            "Randevu Hatırlatma Dahil",
            "Sınırsız Personel",
            "Sınırsız Müşteri",
            "Online Randevu",
            "Online Ödeme Alın (Kapora veya Tam Ödeme)",
            "Seans Paketi Yönetimi",
            "İstatistikler",
            "Yapay Zeka Akıllı Asistan (Limitsiz)",
            "Akıllı Pazarlama Otopilotu"
        ],
        "target_audience_tr": "Yüksek hacimli, birden fazla uzman/personel çalıştıran salonlar ve klinikler için."
    },
    {
        "id": "tier_6_kurumsal",
        "name": "Kurumsal",
        "price_monthly": 3580,
        "price_yearly": 35800,
        "quota_monthly_appointments": -1,
        "ai_message_limit": -1,
        "features": [
            "Sınırsız Randevu/Ay",
            "Randevu Hatırlatma Dahil",
            "Sınırsız Personel",
            "Sınırsız Müşteri",
            "Online Randevu",
            "Online Ödeme Alın (Kapora veya Tam Ödeme)",
            "Seans Paketi Yönetimi",
            "İstatistikler",
            "Yapay Zeka Akıllı Asistan (Limitsiz)",
            "Akıllı Pazarlama Otopilotu"
        ],
        "target_audience_tr": "Sektörün en yoğun klinikleri, poliklinikler ve büyük ölçekli işletmeler için tam çözüm."
    }
]

class OrganizationPlan(BaseModel):
    """Organization'ın abonelik planı ve kota bilgisi"""
    model_config = ConfigDict(extra="ignore")
    organization_id: str
    plan_id: str = "tier_trial"  # Default trial
    quota_usage: int = 0  # Bu ay kullanılan randevu sayısı
    ai_usage_count: int = 0  # Bu ay atılan AI mesaj sayısı
    quota_reset_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc) + timedelta(days=30))  # Kota sıfırlama tarihi
    quota_last_reset_date: Optional[datetime] = Field(default_factory=lambda: datetime.now(timezone.utc))  # Son quota reset tarihi (lazy reset için)
    trial_start_date: Optional[datetime] = None  # Trial başlangıç tarihi
    trial_end_date: Optional[datetime] = None  # Trial bitiş tarihi
    is_first_month: bool = True  # İlk ay indirimi için
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AuditLog(BaseModel):
    """Denetim günlüğü modeli - Kritik işlemleri kaydeder"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    user_id: str  # username
    user_full_name: str
    action: str  # CREATE, UPDATE, DELETE
    resource_type: str  # APPOINTMENT, SETTINGS, CUSTOMER, SERVICE, STAFF
    resource_id: str
    old_value: Optional[dict] = None
    new_value: Optional[dict] = None
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    ip_address: Optional[str] = None

# === GÜVENLİK API ENDPOINT'LERİ ===

async def _validate_register_input(request: Request, user_in: UserCreate, db) -> tuple[str, str]:
    """Kayıt validasyonları — email/telefon/slug uniqueness + dil çıkarımı.

    Hem `/register` (eski, doğrudan hesap açma) hem `/register-initiate`
    (WhatsApp OTP akışı) tarafından paylaşılır. Validation hatası HTTPException
    fırlatır; başarılıysa `(slug, lang)` döner.

    Telefon uniqueness'i kritik güvenlik kontrolüdür: aynı support_phone ile
    iki işletme açılırsa, kötü niyetli bir kullanıcı başka bir işletmenin
    numarasıyla org açıp Aha akışı / public booking üzerinden o numaraya
    WhatsApp mesajı tetikleyebilir.
    """
    existing_user = await get_user_from_db(request, user_in.username, db=db)
    if existing_user:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="This username is already registered.")

    support_phone = (user_in.support_phone or "").strip()
    if support_phone:
        normalized = re.sub(r'\D', '', support_phone)
        if normalized and len(normalized) >= 10:
            # Public booking ile aynı flex-regex deseni: depolanan numara
            # boşluk/tire/parantez/+ içerebilir, son 10 hane karşılaştırılır.
            suffix = normalized[-10:]
            flex_regex = r"\D*" + r"\D*".join(re.escape(d) for d in suffix) + r"$"
            existing_phone_org = await db.settings.find_one(
                {"support_phone": {"$regex": flex_regex}},
                {"_id": 0, "organization_id": 1},
            )
            if existing_phone_org:
                logging.warning(
                    f"🛡️ Register blocked: phone {support_phone} already used by org {existing_phone_org.get('organization_id')}"
                )
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Bu telefon numarası başka bir işletme tarafından kullanılıyor. Lütfen farklı bir numara girin.",
                )

    base_slug = slugify(user_in.organization_name or user_in.username)
    if not base_slug:
        raise HTTPException(status_code=400, detail="İşletme adı geçerli karakter içermelidir.")
    existing_slug = await db.users.find_one({"slug": base_slug})
    if existing_slug:
        raise HTTPException(
            status_code=400,
            detail="Bu işletme adı zaten kullanılıyor veya mevcut bir işletmeyle çok benzer. Lütfen farklı bir işletme adı seçin."
        )

    accept_language = request.headers.get('Accept-Language', '')
    request_lang = get_request_language(request) if accept_language else None
    lang = request_lang or get_email_language(user_in.username)
    if lang not in ['tr', 'en']:
        lang = 'tr'

    return base_slug, lang


@api_router.post("/register", response_model=User)
@rate_limit(LIMITS['register']) 
async def register_user(request: Request, user_in: UserCreate, db = Depends(get_db)):
    base_slug, lang = await _validate_register_input(request, user_in, db)
    unique_slug = base_slug

    # Yeni organization ID oluştur
    new_org_id = str(uuid.uuid4())
    hashed_password = get_password_hash(user_in.password)
    # Stable user_id (UUID) — PostHog distinct_id'nin sabit kaynağı.
    # Username değişse bile funnel parçalanmaz.
    new_user_id = str(uuid.uuid4())

    # support_phone yoksa Aha akışı atlanır; klasik tour başlar.
    support_phone_value = (user_in.support_phone or "").strip()
    initial_activation_state = "pending_aha_appointment" if support_phone_value else "skip_aha_no_phone"

    user_db_data = user_in.model_dump(exclude={"organization_name", "support_phone"})
    user_db = UserInDB(**user_db_data, user_id=new_user_id, hashed_password=hashed_password, organization_id=new_org_id, role="admin", slug=unique_slug, permitted_service_ids=[], onboarding_completed=False, activation_state=initial_activation_state, language=lang)
    user_doc = user_db.model_dump()
    user_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.users.insert_one(user_doc)
    
    # Bu kullanıcı için varsayılan Settings oluştur (kayıt bilgileriyle)
    default_settings = Settings(
        organization_id=new_org_id,
        company_name=user_in.organization_name or "İşletmeniz",
        support_phone=user_in.support_phone or "05000000000",
        slug=unique_slug,
        customer_can_choose_staff=False,
        sector=getattr(user_in, 'sector', None)
    )
    await db.settings.insert_one(default_settings.model_dump())
    
    # Yeni kayıt için Trial paketi oluştur
    trial_start = datetime.now(timezone.utc)
    trial_end = trial_start + timedelta(days=7)
    quota_reset = trial_start + timedelta(days=30)
    trial_plan = OrganizationPlan(
        organization_id=new_org_id,
        plan_id="tier_trial",
        quota_usage=0,
        quota_reset_date=quota_reset,
        quota_last_reset_date=trial_start,  # Lazy reset için başlangıç tarihi
        trial_start_date=trial_start,
        trial_end_date=trial_end,
        is_first_month=True
    )
    plan_doc = trial_plan.model_dump()
    plan_doc['trial_start_date'] = plan_doc['trial_start_date'].isoformat()
    plan_doc['trial_end_date'] = plan_doc['trial_end_date'].isoformat()
    plan_doc['quota_reset_date'] = plan_doc['quota_reset_date'].isoformat()
    plan_doc['created_at'] = plan_doc['created_at'].isoformat()
    plan_doc['updated_at'] = plan_doc['updated_at'].isoformat()
    await db.organization_plans.insert_one(plan_doc)
    
    # Sektör bazlı default services ekle ve admin'e ata
    sector = getattr(user_in, 'sector', None)
    service_ids = []
    
    if sector and sector != "Diğer/Boş":
        # Dil belirleme - önce request'den, yoksa kayıt dilinden (fallback)
        lang_for_services = lang
        if not lang_for_services:
            lang_for_services = get_request_language(request)
        if not lang_for_services:
            lang_for_services = 'tr'
        
        # Default services'i dil bazlı al
        services_to_add = get_sector_default_services(sector, lang_for_services)
        
        # Eğer dil bazlı bulunamazsa, Türkçe'den al
        if not services_to_add:
            services_to_add = get_sector_default_services(sector, 'tr')
        
        # Price'ları ekle (dil bazlı değil, sabit)
        price_map = {
            "Kuaför": {"Saç Kesimi": 150, "Saç Boyama": 300, "Sakal Traşı": 80, "Haircut": 150, "Hair Colouring": 300, "Beard Trim": 80},
            "Güzellik Salonu": {"Manikür": 100, "Pedikür": 120, "Cilt Bakımı": 250, "Kaş Dizaynı": 80, "Manicure": 100, "Pedicure": 120, "Facial Treatment": 250, "Eyebrow Design": 80},
            "Masaj / SPA": {"Klasik Masaj": 300, "Aromaterapi Masajı": 350, "İsveç Masajı": 400, "Classic Massage": 300, "Aromatherapy Massage": 350, "Swedish Massage": 400},
            "Diyetisyen": {"İlk Danışma": 300, "Kontrol Muayenesi": 200, "Diyet Planı": 250, "Initial Consultation": 300, "Follow-up Consultation": 200, "Diet Plan": 250},
            "Psikolog / Danışmanlık": {"Bireysel Terapi": 500, "Çift Terapisi": 700, "Aile Danışmanlığı": 600, "Individual Therapy": 500, "Couples Therapy": 700, "Family Counselling": 600},
            "Diş Klinikleri": {"Muayene": 200, "Dolgu": 400, "Diş Temizliği": 300, "Beyazlatma": 1500, "Examination": 200, "Filling": 400, "Teeth Cleaning": 300, "Teeth Whitening": 1500},
        }
        
        # Price'ları ekle
        for service in services_to_add:
            service_name = service.get("name", "")
            if sector in price_map and service_name in price_map[sector]:
                service["price"] = price_map[sector][service_name]
        
        # Services'leri oluştur
        for service_data in services_to_add:
            service_id = str(uuid.uuid4())
            service = Service(
                id=service_id,
                organization_id=new_org_id,
                **service_data
            )
            await db.services.insert_one(service.model_dump())
            service_ids.append(service_id)
    else:
        # Sektör "Diğer/Boş" (veya boş) → sektör bazlı default hizmet yok. Yine de
        # takvim baştan boş kalmasın ve aha demo çalışsın diye tek bir genel hizmet ekle.
        gen_lang = lang or get_request_language(request) or 'tr'
        gen_name = "Appointment" if str(gen_lang).lower().startswith('en') else "Randevu"
        try:
            gen_id = str(uuid.uuid4())
            gen_service = Service(id=gen_id, organization_id=new_org_id, name=gen_name, price=0, duration=30)
            await db.services.insert_one(gen_service.model_dump())
            service_ids.append(gen_id)
        except Exception as _gen_err:
            logging.warning(f"Genel default hizmet eklenemedi ({user_in.username}): {_gen_err}")
    
    # Admin'e tüm hizmetleri ata
    if service_ids:
        await db.users.update_one(
            {"username": user_in.username},
            {"$set": {"permitted_service_ids": service_ids}}
        )

    # Aha akışı için owner-customer seed:
    # support_phone varsa, /onboarding/aha-test-appointment endpoint'i bu profili
    # bularak ilk randevuyu owner'ın kendi numarasına gönderir.
    if support_phone_value:
        try:
            owner_customer_doc = {
                "id": str(uuid.uuid4()),
                "organization_id": new_org_id,
                "name": user_in.full_name or user_in.username,
                "phone": support_phone_value,
                "source": "owner_self_seed",
                "is_owner_profile": True,
                "owner_user_id": new_user_id,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.customers.insert_one(owner_customer_doc)
        except Exception as e:
            logging.warning(f"Owner-customer seed başarısız ({user_in.username}): {e}")

    # Brevo ile hoş geldin e-postası gönder
    try:
        # Dil belirleme (kayıt sırasında belirlenen dil)
        lang_for_email = lang or get_email_language(user_in.username)
        if lang_for_email not in ['tr', 'en']:
            lang_for_email = 'tr'
        
        logo_url = "https://plannapp.co/api/static/logo.png"
        user_name = user_in.full_name or user_in.username
        # Akıllı buton: cihaza göre uygulamayı açar (iOS/Android), yoksa mağazaya,
        # masaüstünde web paneline yönlendirir.
        if lang_for_email == 'en':
            login_url = "https://plannapp.co/api/app/open?web=https%3A%2F%2Fplannapp.co.uk"
            subject = "Welcome to PLANN! Your Free Trial Has Started."
            html_content = _brand_email(
                lang='en',
                badge="Welcome",
                statement="We're thrilled you chose to grow your business with PLANN.",
                paragraphs=[
                    f"Hello {user_name},",
                    "To let you explore the system we designed to save you time — with no limits — we've started your <strong>7-day free trial</strong>.",
                    "Just sign in to your dashboard to create your first appointment and see how easily you can manage them.",
                ],
                button={"label": "Sign In", "url": login_url},
            )
        else:
            login_url = "https://plannapp.co/api/app/open?web=https%3A%2F%2Fplannapp.co"
            subject = "PLANN'a Hoş Geldiniz! Ücretsiz Deneme Sürümünüz Başladı."
            html_content = _brand_email(
                lang='tr',
                badge="Hoş Geldiniz",
                statement="İşletmenizi PLANN ile büyütmeyi seçtiğiniz için çok mutluyuz.",
                paragraphs=[
                    f"Merhaba {user_name},",
                    "Size zaman kazandırmak için tasarladığımız sistemimizi hiçbir kısıtlama olmadan inceleyebilmeniz için <strong>7 günlük ücretsiz deneme</strong> sürenizi başlattık.",
                    "Kontrol panelinize giriş yaparak ilk randevunuzu oluşturabilir, randevularınızı ne kadar kolay yönetebileceğinizi hemen test edebilirsiniz.",
                ],
                button={"label": "Giriş Yap", "url": login_url},
            )
        
        await send_email(
            to_email=user_in.username,
            subject=subject,
            html_content=html_content,
            to_name=user_name
        )
    except Exception as e:
        logging.error(f"E-posta gönderme sırasında beklenmedik hata: {e}")
        # E-posta gönderilemese bile kayıt başarılı olmalı
    
    return User(**user_db.model_dump())


# ============================================================================
# === REGISTER OTP (WhatsApp doğrulama) ===
# ============================================================================
# Yeni hesap kaydında, public booking ile aynı pattern: önce numaraya 6 haneli
# WhatsApp kodu gönderilir, kullanıcı kodu girince hesap oluşturulur. Mevcut
# `randevu_dogrulama` template'i tekrar kullanılır (ek Meta approval gerekmez).

class RegisterVerifyRequest(BaseModel):
    code: str = Field(..., min_length=6, max_length=6)
    phone: str = Field(..., min_length=5, max_length=32)


def _mask_phone(phone: str) -> str:
    digits = re.sub(r'\D', '', phone or "")
    if len(digits) < 4:
        return "***"
    return f"***{digits[-4:]}"


async def _send_register_whatsapp_code(phone: str, code: str) -> bool:
    """Public booking ile aynı `randevu_dogrulama` template'i ile kod gönder."""
    try:
        from whatsapp_service import format_phone_number, detect_language_from_phone
        phone_clean = format_phone_number(phone)
        wa_lang = detect_language_from_phone(phone)
        wa_template_name = "randevu_dogrulama" if wa_lang == "TR" else "randevu_dogrulama_en"
        wa_lang_code = "tr" if wa_lang == "TR" else "en"

        wa_api_url = f"https://graph.facebook.com/{os.getenv('META_API_VERSION', 'v21.0')}/{os.getenv('META_PHONE_NUMBER_ID')}/messages"
        wa_headers = {
            "Authorization": f"Bearer {os.getenv('META_ACCESS_TOKEN')}",
            "Content-Type": "application/json",
        }
        wa_payload = {
            "messaging_product": "whatsapp",
            "to": phone_clean,
            "type": "template",
            "template": {
                "name": wa_template_name,
                "language": {"code": wa_lang_code},
                "components": [
                    {"type": "body", "parameters": [{"type": "text", "text": code}]},
                    {"type": "button", "sub_type": "url", "index": "0",
                     "parameters": [{"type": "text", "text": code}]},
                ],
            },
        }
        wa_resp = await asyncio.to_thread(
            lambda: requests.post(wa_api_url, json=wa_payload, headers=wa_headers, timeout=10)
        )
        if wa_resp.status_code == 200:
            logging.info(f"📱 Register OTP gönderildi: {phone} → {code}")
            return True
        logging.warning(f"⚠️ Register OTP gönderilemedi: {wa_resp.status_code} {wa_resp.text[:200]}")
        return False
    except Exception as e:
        logging.error(f"❌ Register OTP gönderim hatası: {e}")
        return False


@api_router.post("/register-initiate")
@rate_limit(LIMITS['register'])
async def register_initiate(request: Request, user_in: UserCreate, db = Depends(get_db)):
    """Hesap KAYIT için WhatsApp OTP gönder. Hesap henüz oluşturulmaz."""
    redis_client = getattr(request.app.state, 'redis_client', None)
    if not redis_client:
        raise HTTPException(status_code=503, detail="Doğrulama servisi şu an kullanılamıyor. Lütfen birkaç dakika sonra deneyin.")

    # Aynı uniqueness/format validations
    await _validate_register_input(request, user_in, db)

    support_phone = (user_in.support_phone or "").strip()
    if not support_phone:
        raise HTTPException(status_code=400, detail="Doğrulama için telefon numarası gerekli.")
    digits = re.sub(r'\D', '', support_phone)
    if len(digits) < 10:
        raise HTTPException(status_code=400, detail="Geçerli bir telefon numarası girin.")

    # Flood koruması: aynı telefondan saatte en fazla 3 kod isteği
    wa_key = f"plann:rate:reg_wa:{support_phone}"
    wa_count = await redis_client.incr(wa_key)
    if wa_count == 1:
        await redis_client.expire(wa_key, 3600)
    if wa_count > 3:
        raise HTTPException(status_code=429, detail="Çok fazla doğrulama kodu istediniz. Lütfen 1 saat sonra tekrar deneyin.")

    # 6 haneli kod üret + redis'e payload yaz (TTL 15dk)
    code = str(secrets.randbelow(900000) + 100000)
    code_key = f"plann:reg:pending:{code}"
    phone_map_key = f"plann:reg:pending:phone:{support_phone}"
    ttl = 900

    # Aynı telefon için eski pending varsa eziyoruz (upsert)
    old_code = await redis_client.get(phone_map_key)
    if old_code:
        old_str = old_code.decode() if isinstance(old_code, bytes) else old_code
        await redis_client.delete(f"plann:reg:pending:{old_str}")

    payload = user_in.model_dump()
    await redis_client.setex(code_key, ttl, json.dumps(payload, ensure_ascii=False))
    await redis_client.setex(phone_map_key, ttl, code)

    # WhatsApp gönder
    sent = await _send_register_whatsapp_code(support_phone, code)
    if not sent:
        # Redis temizle ki kullanıcı tekrar deneyebilsin (flood counter da düşsün)
        await redis_client.delete(code_key)
        await redis_client.delete(phone_map_key)
        await redis_client.decr(wa_key)
        raise HTTPException(status_code=502, detail="Doğrulama mesajı gönderilemedi. Lütfen tekrar deneyin veya birkaç dakika sonra tekrar girin.")

    return {
        "verification_required": True,
        "expires_in": ttl,
        "phone_masked": _mask_phone(support_phone),
    }


@api_router.post("/register-verify", response_model=Token)
@rate_limit("20/hour")
async def register_verify(request: Request, body: RegisterVerifyRequest, background_tasks: BackgroundTasks, db = Depends(get_db)):
    """Kayıt OTP'sini doğrula, hesabı oluştur, JWT token dön (otomatik login)."""
    redis_client = getattr(request.app.state, 'redis_client', None)
    if not redis_client:
        raise HTTPException(status_code=503, detail="Doğrulama servisi şu an kullanılamıyor.")

    code = (body.code or "").strip()
    phone = (body.phone or "").strip()
    if not re.fullmatch(r"\d{6}", code):
        raise HTTPException(status_code=400, detail="Kod 6 haneli olmalıdır.")
    if not phone:
        raise HTTPException(status_code=400, detail="Telefon numarası gerekli.")

    code_key = f"plann:reg:pending:{code}"
    payload_raw = await redis_client.get(code_key)
    if not payload_raw:
        raise HTTPException(status_code=400, detail="Doğrulama kodu geçersiz veya süresi dolmuş.")

    payload_str = payload_raw.decode() if isinstance(payload_raw, bytes) else payload_raw
    try:
        payload = json.loads(payload_str)
    except json.JSONDecodeError:
        await redis_client.delete(code_key)
        raise HTTPException(status_code=500, detail="Doğrulama verisi okunamadı. Lütfen tekrar başlatın.")

    saved_phone = (payload.get("support_phone") or "").strip()
    if saved_phone != phone:
        raise HTTPException(status_code=400, detail="Telefon numarası eşleşmiyor.")

    # Redis'i ŞİMDİ temizle — race'e karşı (aynı kodla iki çağrı gelirse ikincisi 400 alır)
    await redis_client.delete(code_key)
    await redis_client.delete(f"plann:reg:pending:phone:{saved_phone}")

    # UserCreate'e dönüştür ve hesabı oluştur (mevcut register_user akışı)
    try:
        user_in = UserCreate(**payload)
    except Exception as e:
        logging.error(f"Register-verify payload parse error: {e}")
        raise HTTPException(status_code=500, detail="Kayıt verisi okunamadı. Lütfen tekrar başlatın.")

    # register_user, /register endpoint fonksiyonu — içeriden çağırırken
    # validations tekrar çalışır (race koruması için iyi).
    user_response = await register_user(request, user_in, db)

    # --- Meta CAPI: CompleteRegistration (server-side) ---
    try:
        svc = get_meta_capi_service()
        if svc.enabled:
            names = split_full_name(user_in.full_name)
            background_tasks.add_task(
                svc.send_event,
                event_name="CompleteRegistration",
                event_source_url="https://plannapp.co/register",
                action_source="website",
                client_ip=request.client.host if request.client else None,
                client_user_agent=request.headers.get("user-agent"),
                user_data={
                    "em": user_in.username,
                    "ph": user_in.support_phone or saved_phone,
                    "fn": names["fn"],
                    "ln": names["ln"],
                },
                fbp=request.cookies.get("_fbp"),
                fbc=request.cookies.get("_fbc"),
                external_id=user_response.organization_id,
            )
    except Exception as exc:
        logging.warning(f"[register-verify] Meta CAPI CompleteRegistration hatası: {exc}")

    # JWT üret — login_for_access_token ile aynı payload yapısı
    access_token = create_access_token(data={
        "sub": user_in.username,
        "org_id": user_response.organization_id,
        "role": user_response.role,
        "onboarding_completed": getattr(user_response, "onboarding_completed", False),
        "full_name": getattr(user_response, "full_name", "") or "",
    })
    return Token(access_token=access_token, token_type="bearer")


# ============================================================================
# === AHA ACTIVATION + FUNNEL EVENTS ===
# ============================================================================

class AhaSkipRequest(BaseModel):
    reason: Optional[str] = Field(default=None, max_length=64)

class FunnelEventIn(BaseModel):
    event_name: str = Field(..., max_length=64)
    insert_id: str = Field(..., min_length=6, max_length=64)
    properties: Optional[dict] = None


@api_router.post("/onboarding/aha-test-appointment")
@rate_limit(LIMITS.get('aha_activation', "5/hour"))
async def aha_test_appointment(
    request: Request,
    current_user: UserInDB = Depends(get_current_user),
    db = Depends(get_db),
):
    """
    Aha moment endpoint'i:
    1. Owner'ın kendi numarasına gerçek WhatsApp CONFIRMATION mesajı + appointment kaydı oluşturur.
    2. WA başarısız olursa 1 kez sessiz retry dener; hala fail ise state=`aha_wa_failed`.
    3. İdempotent: state=`aha_completed` ise aynı randevuyu/msg_id'yi döndürür.
    4. State=`aha_wa_failed` ise: randevuyu silmeden sadece WA'yı yeniden dener.
    5. Quota'ya yansımaz.

    Yalnızca admin role'u çağırabilir; staff 403 alır.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can trigger the Aha test appointment.")

    state = current_user.activation_state or "pending_aha_appointment"

    # Terminal state: idempotent success
    if state == "aha_completed":
        existing_apt_id = current_user.aha_appointment_id
        existing_apt = None
        if existing_apt_id:
            existing_apt = await db.appointments.find_one(
                {"id": existing_apt_id, "organization_id": current_user.organization_id},
                {"_id": 0},
            )
        return {
            "ok": True,
            "idempotent": True,
            "activation_state": "aha_completed",
            "appointment": existing_apt,
            "wa_message_id": current_user.aha_wa_message_id,
        }

    if state == "aha_skipped":
        raise HTTPException(status_code=409, detail="Aha activation was skipped. Cannot re-trigger.")

    if state == "not_applicable":
        raise HTTPException(status_code=403, detail="This account is not eligible for Aha activation.")

    # Settings ve support phone kontrolü
    settings_doc = await db.settings.find_one({"organization_id": current_user.organization_id})
    if not settings_doc:
        raise HTTPException(status_code=400, detail="Settings not configured for this organization.")

    support_phone = (settings_doc.get("support_phone") or "").strip()
    if not support_phone or support_phone == "05000000000":
        # Owner kayıtta telefon vermedi → skip state'e al, frontend klasik tura yönlendirir
        await db.users.update_one(
            {"username": current_user.username},
            {"$set": {
                "activation_state": "skip_aha_no_phone",
                "aha_skipped_reason": "no_support_phone",
                "aha_skipped_at": datetime.now(timezone.utc).isoformat(),
            }}
        )
        raise HTTPException(status_code=400, detail="support_phone_missing")

    company_name = settings_doc.get("company_name", "İşletmeniz")
    location = settings_doc.get("location") or {}
    coords = location.get("coordinates") or {}
    biz_lat = coords.get("lat") if isinstance(coords, dict) else None
    biz_lng = coords.get("lng") if isinstance(coords, dict) else None
    biz_address = location.get("address") if isinstance(location, dict) else None

    # İlk hizmeti seç (owner'ın kendine atadığı; register_user default service'leri ekledi)
    service = await db.services.find_one(
        {"organization_id": current_user.organization_id},
        {"_id": 0},
        sort=[("order", 1), ("created_at", 1)],
    )
    if not service:
        # Sektörü "Diğer/Boş" seçen org'larda hiç default hizmet oluşmuyordu →
        # aha demo "No service available" ile patlıyordu (aha başarısızlıklarının %90'ı).
        # Hata vermek yerine genel bir hizmet oluştur: hem demo çalışır hem org kullanılabilir olur.
        svc_lang = (getattr(current_user, "language", None) or "tr").lower()
        svc_name = "Appointment" if svc_lang == "en" else "Randevu"
        try:
            fallback_service = Service(
                id=str(uuid.uuid4()),
                organization_id=current_user.organization_id,
                name=svc_name,
                price=0,
                duration=30,
            )
            await db.services.insert_one(fallback_service.model_dump())
            service = fallback_service.model_dump()
            service.pop("_id", None)
            # Admin'e bu hizmeti ata (permitted_service_ids)
            try:
                await db.users.update_one(
                    {"username": current_user.username},
                    {"$addToSet": {"permitted_service_ids": fallback_service.id}},
                )
            except Exception:
                pass
            logging.info(f"aha: fallback hizmet oluşturuldu org={current_user.organization_id}")
        except Exception as _svc_err:
            logging.error(f"aha fallback service creation failed: {_svc_err}", exc_info=True)
            raise HTTPException(status_code=400, detail="No service available for the demo appointment.")

    # State=aha_wa_failed iken mevcut randevu varsa onu reuse et, sadece WA tekrar dene
    existing_apt = None
    if state == "aha_wa_failed" and current_user.aha_appointment_id:
        existing_apt = await db.appointments.find_one(
            {"id": current_user.aha_appointment_id, "organization_id": current_user.organization_id},
            {"_id": 0},
        )

    if existing_apt:
        apt_doc = existing_apt
        apt_id = apt_doc["id"]
    else:
        apt_date, apt_time = pick_aha_slot(settings_doc)
        apt_id = str(uuid.uuid4())
        apt_doc = {
            "id": apt_id,
            "organization_id": current_user.organization_id,
            "customer_name": current_user.full_name or current_user.username,
            "phone": support_phone,
            "service_id": service.get("id"),
            "service_name": service.get("name", "Demo"),
            "service_price": float(service.get("price") or 0),
            "appointment_date": apt_date,
            "appointment_time": apt_time,
            "status": "Bekliyor",
            "staff_member_id": None,
            "notes": "",
            "source": "aha_activation",
            "is_demo": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "reminder_sent": False,
        }
        await db.appointments.insert_one(dict(apt_doc))
        # Not: check_quota_and_increment burada ÇAĞRILMAZ — demo randevusu quota'ya yansımaz.

        # Denormalize sayaç — aha demo randevusu da müşteri sayaçlarına yansısın.
        try:
            await _apply_customer_delta(
                db,
                organization_id=current_user.organization_id,
                phone=apt_doc.get("phone") or "",
                name=apt_doc.get("customer_name"),
                delta_total=1,
                delta_completed=0,
                appointment_date=apt_doc.get("appointment_date"),
                appointment_time=apt_doc.get("appointment_time"),
            )
        except Exception as _delta_err:
            logging.warning(f"customer delta (aha_test_appointment) skipped: {_delta_err}")

    # WhatsApp gönderimi (retry'li)
    wa_ok, wa_msg_id, wa_err = await send_aha_whatsapp_with_retry(
        send_callable=send_whatsapp_template,
        to_number=support_phone,
        customer_name=apt_doc["customer_name"],
        company_name=company_name,
        appointment_date=apt_doc["appointment_date"],
        appointment_time=apt_doc["appointment_time"],
        service_name=apt_doc.get("service_name", "Demo"),
        support_phone=support_phone,
        business_lat=biz_lat,
        business_lng=biz_lng,
        business_address=biz_address,
    )

    now_iso = datetime.now(timezone.utc).isoformat()

    # WebSocket emit: Aha randevusu da Dashboard'a anlık düşsün.
    # `existing_apt` reuse durumunda sadece WA tekrar deniyoruz; yeni doc yok,
    # zaten Dashboard daha önce göstermiş olabilir. Yine de re-emit zararsız
    # (frontend duplicate id'yi dedupe eder).
    try:
        apt_clean = {k: v for k, v in apt_doc.items() if k != "_id"}
        await emit_to_organization(
            current_user.organization_id,
            "appointment_created",
            {"appointment": apt_clean},
        )
    except Exception as emit_err:
        logging.warning(f"Aha appointment_created emit failed: {emit_err}")

    if wa_ok:
        await db.users.update_one(
            {"username": current_user.username},
            {"$set": {
                "activation_state": "aha_completed",
                "aha_wa_message_id": wa_msg_id or "",
                "aha_wa_last_error": None,
                "aha_appointment_id": apt_id,
                "first_appointment_at": now_iso,
            }}
        )
        return {
            "ok": True,
            "idempotent": False,
            "activation_state": "aha_completed",
            "appointment": apt_doc,
            "wa_message_id": wa_msg_id,
        }

    # WA başarısız — randevu kaydı duruyor, state=aha_wa_failed
    await db.users.update_one(
        {"username": current_user.username},
        {"$set": {
            "activation_state": "aha_wa_failed",
            "aha_wa_last_error": (wa_err or "unknown")[:240],
            "aha_appointment_id": apt_id,
        }}
    )
    return {
        "ok": False,
        "idempotent": False,
        "activation_state": "aha_wa_failed",
        "wa_status": "failed",
        "wa_error": wa_err,
        "appointment": apt_doc,
    }


@api_router.post("/onboarding/aha-skip")
async def aha_skip(
    body: AhaSkipRequest,
    current_user: UserInDB = Depends(get_current_user),
    db = Depends(get_db),
):
    """
    Owner Aha'yı atlıyor → terminal state `aha_skipped`. İdempotent.
    Tekrar tetiklenmez; frontend klasik 7-step tour'a geçer.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can skip Aha activation.")

    state = current_user.activation_state or "pending_aha_appointment"

    # Zaten terminal → no-op idempotent
    if state in ("aha_completed", "aha_skipped"):
        return {"ok": True, "idempotent": True, "activation_state": state}

    if state == "not_applicable":
        raise HTTPException(status_code=403, detail="This account is not eligible for Aha activation.")

    reason = (body.reason or "user_skipped")[:64]
    await db.users.update_one(
        {"username": current_user.username},
        {"$set": {
            "activation_state": "aha_skipped",
            "aha_skipped_reason": reason,
            "aha_skipped_at": datetime.now(timezone.utc).isoformat(),
        }}
    )
    return {"ok": True, "idempotent": False, "activation_state": "aha_skipped"}


@api_router.post("/funnel/event", status_code=204)
async def funnel_event_sink(
    payload: FunnelEventIn,
    current_user: UserInDB = Depends(get_current_user),
    db = Depends(get_db),
):
    """
    Frontend-owned funnel event'lerini MongoDB ledger'a yazar.
    PostHog'a yazım tarayıcıda yapılır; burası ground-truth dedupe sink'idir.
    `insert_id` unique index'i E11000 → idempotent no-op.
    """
    if not is_valid_event(payload.event_name):
        raise HTTPException(status_code=400, detail="Unknown funnel event name.")

    if not is_frontend_owned(payload.event_name):
        raise HTTPException(status_code=400, detail="Event is not frontend-owned; cannot be recorded here.")

    distinct_id = current_user.user_id or current_user.username
    await record_funnel_event(
        db,
        event_name=payload.event_name,
        insert_id=payload.insert_id,
        distinct_id=distinct_id,
        organization_id=current_user.organization_id,
        user_id=current_user.user_id,
        source="frontend",
        properties=payload.properties or {},
    )
    return Response(status_code=204)


# ============================================================================

@api_router.post("/token", response_model=Token)
@rate_limit(LIMITS['login']) 
async def login_for_access_token(request: Request, form_data: OAuth2PasswordRequestForm = Depends(), db = Depends(get_db)):
    try:
        user = await get_user_from_db(request, form_data.username, db=db)
        if not user or not verify_password(form_data.password, user.hashed_password):
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Incorrect username or password", headers={"WWW-Authenticate": "Bearer"})
        
        # Pending (bekleyen) kullanıcılar giriş yapamaz
        if user.status == "pending":
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Hesabınız henüz aktif değil. Lütfen e-postanızdaki davet linkine tıklayarak şifrenizi belirleyin.")
        
        remember_me = "remember_me" in (form_data.scopes or [])
        expire_minutes = ACCESS_TOKEN_EXPIRE_MINUTES_REMEMBER if remember_me else ACCESS_TOKEN_EXPIRE_MINUTES
        access_token_expires = timedelta(minutes=expire_minutes)
        settings = await db.settings.find_one({"organization_id": user.organization_id})
        company_name = settings.get("company_name", "") if settings else ""
        
        token_data = {
            "sub": user.username, 
            "user_id": user.user_id,
            "org_id": user.organization_id, 
            "company_name": company_name,
            "role": user.role, 
            "onboarding_completed": user.onboarding_completed,
            "activation_state": user.activation_state,
            "full_name": user.full_name or None,
            "can_view_all_appointments": user.can_view_all_appointments
        }
        access_token = create_access_token(data=token_data, expires_delta=access_token_expires)
        return {"access_token": access_token, "token_type": "bearer"}
    except HTTPException: raise
    except Exception as e:
        logging.error(f"Login error: {type(e).__name__}: {str(e)}"); import traceback; logging.error(traceback.format_exc())
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="An error occurred during login. Please try again later.")

@api_router.post("/sso/create", response_model=SsoCodeResponse)
@rate_limit(LIMITS.get('sso_create', LIMITS['login']))
async def create_sso_code(request: Request, current_user: UserInDB = Depends(get_current_user)):
    try:
        redis_client = getattr(getattr(request.app, 'state', None), 'redis_client', None) or getattr(request.app, 'redis_client', None)

        ttl_seconds = int(os.environ.get('SSO_CODE_TTL_SECONDS', '60'))
        code = secrets.token_urlsafe(32)
        payload = {
            "sub": current_user.username,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        if redis_client:
            redis_key = f"sso:{code}"
            await redis_client.set(redis_key, json.dumps(payload), ex=ttl_seconds)
        else:
            db_inst = getattr(request.app, 'db', None)
            if db_inst is not None:
                from datetime import timedelta
                await db_inst.sso_codes.insert_one({"code": code, "payload": json.dumps(payload), "expire_at": datetime.now(timezone.utc) + timedelta(seconds=ttl_seconds)})
            else:
                raise HTTPException(status_code=503, detail="SSO unavailable")
        return {"code": code, "expires_in": ttl_seconds}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"SSO create error: {type(e).__name__}: {str(e)}");
        raise HTTPException(status_code=500, detail="Failed to create SSO code")

@api_router.post("/sso/exchange", response_model=Token)
@rate_limit(LIMITS.get('sso_exchange', LIMITS['login']))
async def exchange_sso_code(request: Request, body: SsoExchangeRequest, db = Depends(get_db)):
    try:
        redis_client = getattr(getattr(request.app, 'state', None), 'redis_client', None) or getattr(request.app, 'redis_client', None)

        code = (body.code or '').strip()
        if not code or len(code) < 10:
            raise HTTPException(status_code=400, detail="Invalid code")

        raw = None
        if redis_client:
            redis_key = f"sso:{code}"
            try:
                raw = await redis_client.getdel(redis_key)
            except Exception:
                raw = await redis_client.get(redis_key)
                if raw is not None:
                    await redis_client.delete(redis_key)
        else:
            sso_doc = await db.sso_codes.find_one_and_delete({"code": code})
            if sso_doc:
                expire_at = sso_doc.get("expire_at")
                if expire_at and expire_at.replace(tzinfo=timezone.utc) > datetime.now(timezone.utc):
                    raw = sso_doc.get("payload")

        if not raw:
            raise HTTPException(status_code=400, detail="Code expired or already used")

        payload = json.loads(raw)
        username = payload.get('sub')
        if not username:
            raise HTTPException(status_code=400, detail="Invalid code")

        user = await get_user_from_db(request, username, db=db)
        if not user:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Could not validate credentials", headers={"WWW-Authenticate": "Bearer"})

        if user.status == "pending":
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Hesabınız henüz aktif değil.")

        settings = await db.settings.find_one({"organization_id": user.organization_id})
        company_name = settings.get("company_name", "") if settings else ""
        
        access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        token_data = {
            "sub": user.username,
            "user_id": user.user_id,
            "org_id": user.organization_id,
            "company_name": company_name,
            "role": user.role,
            "onboarding_completed": user.onboarding_completed,
            "activation_state": user.activation_state,
            "full_name": user.full_name or None,
            "can_view_all_appointments": user.can_view_all_appointments
        }
        access_token = create_access_token(data=token_data, expires_delta=access_token_expires)
        return {"access_token": access_token, "token_type": "bearer"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"SSO exchange error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail="SSO exchange failed")

# === E-POSTA GÖNDERME FONKSİYONLARI ===
def get_email_language(email: str) -> str:
    """Email domain'ine göre dil belirler"""
    if email and ('.co.uk' in email.lower() or '@' in email and email.split('@')[1].startswith('uk')):
        return 'en'
    return 'tr'

def get_request_language(request: Request) -> str:
    """Request header'dan dil belirler"""
    accept_language = request.headers.get('Accept-Language', '')
    if not accept_language:
        return 'tr'

    raw = accept_language.lower().strip()
    first = raw.split(',')[0].strip()
    if first.startswith('en'):
        return 'en'
    if first.startswith('tr'):
        return 'tr'

    if 'en' in raw and 'tr' not in raw:
        return 'en'
    if 'tr' in raw:
        return 'tr'
    return 'tr'

def detect_currency(request: Request, user_phone: Optional[str] = None) -> str:
    """
    Para birimini algılar: Telefon prefix veya IP bazlı.
    
    Args:
        request: FastAPI Request objesi
        user_phone: Kullanıcı telefon numarası (opsiyonel)
    
    Returns:
        'gbp' veya 'try'
    """
    # 1. Telefon numarası prefix kontrolü
    if user_phone:
        clean_phone = re.sub(r'[^\d+]', '', user_phone)
        if clean_phone.startswith('+44') or clean_phone.startswith('44'):
            return 'gbp'
        if clean_phone.startswith('+90') or clean_phone.startswith('90'):
            return 'try'
        if clean_phone.startswith('0') and len(clean_phone) == 11:
            return 'try'
    
    # 2. IP bazlı algılama (basit - UK IP'leri için)
    # Not: Production'da daha gelişmiş bir IP geolocation servisi kullanılabilir
    client_ip = request.client.host if request.client else None
    if client_ip:
        # Basit kontrol: localhost veya private IP'ler için TRY
        if client_ip in ['127.0.0.1', 'localhost'] or client_ip.startswith('192.168.') or client_ip.startswith('10.'):
            return 'try'
        # Production'da burada bir IP geolocation API çağrısı yapılabilir
    
    # 3. Accept-Language header kontrolü
    accept_language = request.headers.get('Accept-Language', '')
    if accept_language and 'en' in accept_language.lower() and 'tr' not in accept_language.lower():
        # İngilizce dil tercihi varsa GBP'ye yönlendir (UK pazarı için)
        return 'gbp'
    
    # 4. Varsayılan: TRY
    return 'try'

def get_stripe_lookup_key(plan_id: str, billing_cycle: str, currency: str) -> str:
    """
    Stripe lookup key oluşturur (currency ile birlikte).
    
    Format: {plan_name}_{billing_cycle}_{currency}
    Örnek: standard_monthly_gbp, professional_monthly_try
    
    Note: Stripe'da her currency için ayrı lookup key kullanılıyor.
    Bu sayede doğrudan ilgili fiyatı çekebiliriz, filtrelemeye gerek yok.
    
    Args:
        plan_id: Plan ID (tier_1_standard, tier_2_profesyonel, vb.)
        billing_cycle: 'monthly' veya 'yearly'
        currency: 'gbp' veya 'try'
    
    Returns:
        Lookup key string (currency ile birlikte)
    """
    # Plan ID'den plan adını çıkar
    # Not: Stripe'da "standart" kullanılıyor (Türkçe), "standard" değil
    plan_name_map = {
        'tier_1_standard': 'standart',  # Stripe'da "standart" olarak kayıtlı
        'tier_2_profesyonel': 'professional',
        'tier_3_premium': 'premium',
        'tier_4_business': 'business',
        'tier_5_enterprise': 'enterprise',
        'tier_6_kurumsal': 'corporate'
    }
    
    plan_name = plan_name_map.get(plan_id, plan_id.replace('tier_', '').replace('_', ''))
    currency_lower = currency.lower()
    
    return f"{plan_name}_{billing_cycle}_{currency_lower}"

def get_stripe_price_by_lookup_key(lookup_key: str) -> Optional[stripe.Price]:
    """
    Stripe'dan lookup key ile fiyat çeker.
    
    Lookup key zaten currency içeriyor (örn: professional_monthly_gbp),
    bu yüzden doğrudan ilgili fiyatı çekebiliriz, filtrelemeye gerek yok.
    
    Args:
        lookup_key: Stripe lookup key (örn: standard_monthly_gbp, professional_monthly_try)
    
    Returns:
        Stripe Price objesi veya None
    """
    try:
        if not STRIPE_SECRET_KEY:
            logger.error("STRIPE_SECRET_KEY tanımlı değil!")
            return None
        
        logger.info(f"🔍 Stripe'da lookup key aranıyor: '{lookup_key}'")
        
        # Stripe'da lookup key ile fiyat ara (currency zaten key'de)
        prices = stripe.Price.list(
            lookup_keys=[lookup_key],
            active=True,
            limit=1  # Currency-specific key olduğu için sadece 1 sonuç bekleniyor
        )
        
        if not prices.data or len(prices.data) == 0:
            logger.warning(f"❌ Stripe'da lookup key bulunamadı: '{lookup_key}'")
            logger.warning(f"   💡 Kontrol edin: Stripe Dashboard'da bu lookup key ile bir fiyat var mı?")
            logger.warning(f"   💡 Lookup key formatı: '{lookup_key}' (örn: 'standard_monthly_gbp', 'professional_monthly_try')")
            return None
        
        # İlk (ve muhtemelen tek) fiyatı döndür
        price = prices.data[0]
        logger.info(f"✅ Stripe fiyat bulundu: lookup_key={lookup_key}, currency={price.currency}, price_id={price.id}, amount={price.unit_amount}")
        return price
        
    except stripe.error.StripeError as e:
        logger.error(f"❌ Stripe lookup key hatası: {e}")
        logger.error(f"   Lookup key: '{lookup_key}'")
        return None
    except Exception as e:
        logger.error(f"❌ Fiyat çekme hatası: {e}")
        logger.error(f"   Lookup key: '{lookup_key}'")
        return None

async def get_exchange_rate(from_currency: str, to_currency: str) -> float:
    """
    Gerçek kur bilgisini API'den çeker.
    
    Args:
        from_currency: Kaynak para birimi (örn: 'TRY')
        to_currency: Hedef para birimi (örn: 'GBP')
    
    Returns:
        Kur oranı (örn: 1 TRY = X GBP)
    """
    # Aynı para birimiyse 1.0 döndür
    if from_currency.upper() == to_currency.upper():
        return 1.0
    
    try:
        # ExchangeRate-API (ücretsiz tier) kullan
        # Alternatif: exchangerate-api.com veya fixer.io
        api_url = f"https://api.exchangerate-api.com/v4/latest/{from_currency.upper()}"
        
        response = requests.get(api_url, timeout=5)
        response.raise_for_status()
        
        data = response.json()
        rates = data.get('rates', {})
        
        if to_currency.upper() in rates:
            rate = rates[to_currency.upper()]
            logger.info(f"✅ Kur bilgisi alındı: 1 {from_currency.upper()} = {rate} {to_currency.upper()}")
            return float(rate)
        else:
            logger.warning(f"⚠️ Kur bulunamadı: {to_currency.upper()}, gerçek kur fallback kullanılıyor")
            # Fallback: GBP/TRY için gerçek kur (güncel: 2025-12-19)
            if from_currency.upper() == 'TRY' and to_currency.upper() == 'GBP':
                return 0.0175  # 1 TRY = 0.0175 GBP (gerçek kur)
            elif from_currency.upper() == 'GBP' and to_currency.upper() == 'TRY':
                return 57.14  # 1 GBP = 57.14 TRY (gerçek kur)
            return 1.0
            
    except requests.exceptions.RequestException as e:
        logger.error(f"❌ Kur API hatası: {e}, gerçek kur fallback kullanılıyor")
        # Fallback: GBP/TRY için gerçek kur (güncel: 2025-12-19)
        if from_currency.upper() == 'TRY' and to_currency.upper() == 'GBP':
            return 0.0175  # 1 TRY = 0.0175 GBP (gerçek kur)
        elif from_currency.upper() == 'GBP' and to_currency.upper() == 'TRY':
            return 57.14  # 1 GBP = 57.14 TRY (gerçek kur)
        return 1.0
    except Exception as e:
        logger.error(f"❌ Kur hesaplama hatası: {e}, gerçek kur fallback kullanılıyor")
        # Fallback: GBP/TRY için gerçek kur (güncel: 2025-12-19)
        if from_currency.upper() == 'TRY' and to_currency.upper() == 'GBP':
            return 0.0175  # 1 TRY = 0.0175 GBP (gerçek kur)
        elif from_currency.upper() == 'GBP' and to_currency.upper() == 'TRY':
            return 57.14  # 1 GBP = 57.14 TRY (gerçek kur)
        return 1.0

def get_sector_default_services(sector: str, lang: str = 'tr') -> list:
    """Sektör bazlı default hizmetleri dil bazlı döndürür"""
    if lang == 'en':
        sector_defaults = {
            "Hair Salon": [
                {"name": "Haircut", "price": 0, "duration": 30},
                {"name": "Hair Colouring", "price": 0, "duration": 60},
                {"name": "Beard Trim", "price": 0, "duration": 20}
            ],
            "Beauty Salon": [
                {"name": "Manicure", "price": 0, "duration": 30},
                {"name": "Pedicure", "price": 0, "duration": 40},
                {"name": "Facial Treatment", "price": 0, "duration": 60},
                {"name": "Eyebrow Design", "price": 0, "duration": 20}
            ],
            "Massage / SPA": [
                {"name": "Classic Massage", "price": 0, "duration": 60},
                {"name": "Aromatherapy Massage", "price": 0, "duration": 90},
                {"name": "Swedish Massage", "price": 0, "duration": 60}
            ],
            "Dietitian": [
                {"name": "Initial Consultation", "price": 0, "duration": 45},
                {"name": "Follow-up Consultation", "price": 0, "duration": 30},
                {"name": "Diet Plan", "price": 0, "duration": 60}
            ],
            "Psychologist / Counselling": [
                {"name": "Individual Therapy", "price": 0, "duration": 60},
                {"name": "Couples Therapy", "price": 0, "duration": 90},
                {"name": "Family Counselling", "price": 0, "duration": 90}
            ],
            "Dental Clinics": [
                {"name": "Examination", "price": 0, "duration": 30},
                {"name": "Filling", "price": 0, "duration": 45},
                {"name": "Teeth Cleaning", "price": 0, "duration": 40},
                {"name": "Teeth Whitening", "price": 0, "duration": 60}
            ],
        }
        # Türkçe sektör isimlerini İngilizce'ye map et
        sector_map = {
            "Kuaför": "Hair Salon",
            "Güzellik Salonu": "Beauty Salon",
            "Masaj / SPA": "Massage / SPA",
            "Diyetisyen": "Dietitian",
            "Psikolog / Danışmanlık": "Psychologist / Counselling",
            "Diş Klinikleri": "Dental Clinics"
        }
        mapped_sector = sector_map.get(sector, sector)
        return sector_defaults.get(mapped_sector, [])
    else:
        sector_defaults = {
            "Kuaför": [
                {"name": "Saç Kesimi", "price": 0, "duration": 30},
                {"name": "Saç Boyama", "price": 0, "duration": 60},
                {"name": "Sakal Traşı", "price": 0, "duration": 20}
            ],
            "Güzellik Salonu": [
                {"name": "Manikür", "price": 0, "duration": 30},
                {"name": "Pedikür", "price": 0, "duration": 40},
                {"name": "Cilt Bakımı", "price": 0, "duration": 60},
                {"name": "Kaş Dizaynı", "price": 0, "duration": 20}
            ],
            "Masaj / SPA": [
                {"name": "Klasik Masaj", "price": 0, "duration": 60},
                {"name": "Aromaterapi Masajı", "price": 0, "duration": 90},
                {"name": "İsveç Masajı", "price": 0, "duration": 60}
            ],
            "Diyetisyen": [
                {"name": "İlk Danışma", "price": 0, "duration": 45},
                {"name": "Kontrol Muayenesi", "price": 0, "duration": 30},
                {"name": "Diyet Planı", "price": 0, "duration": 60}
            ],
            "Psikolog / Danışmanlık": [
                {"name": "Bireysel Terapi", "price": 0, "duration": 60},
                {"name": "Çift Terapisi", "price": 0, "duration": 90},
                {"name": "Aile Danışmanlığı", "price": 0, "duration": 90}
            ],
            "Diş Klinikleri": [
                {"name": "Muayene", "price": 0, "duration": 30},
                {"name": "Dolgu", "price": 0, "duration": 45},
                {"name": "Diş Temizliği", "price": 0, "duration": 40},
                {"name": "Beyazlatma", "price": 0, "duration": 60}
            ],
        }
        return sector_defaults.get(sector, [])

async def send_personnel_invitation_email(recipient_email: str, recipient_name: str, admin_name: str, organization_name: str, invitation_token: str, lang: Optional[str] = None):
    """Personel davet e-postası gönderir."""
    try:
        # Dil belirleme
        lang = (lang or '').strip().lower()
        if lang not in ['tr', 'en']:
            lang = get_email_language(recipient_email)
        
        # Invitation link oluştur (setup-password route'u kullan)
        invitation_link = f"https://plannapp.co/setup-password?token={invitation_token}" if lang == 'tr' else f"https://plannapp.co.uk/setup-password?token={invitation_token}"
        
        if lang == 'en':
            subject = "PLANN Invitation: Create Your Account"
            html_content = _brand_email(
                lang='en',
                badge="Invitation",
                statement="You have been invited to PLANN.",
                paragraphs=[
                    f"Hello {recipient_name},",
                    f"<strong>{admin_name}</strong> has added you as staff to <strong>{organization_name}</strong>'s PLANN appointment system.",
                    "Please click the button below to activate your account and set your password.",
                ],
                button={"label": "Set Password and Log In", "url": invitation_link},
            )
        else:
            subject = "PLANN Davetiyesi: Hesabınızı Oluşturun"
            html_content = _brand_email(
                lang='tr',
                badge="Davet",
                statement="PLANN'a davet edildiniz.",
                paragraphs=[
                    f"Merhaba {recipient_name},",
                    f"<strong>{admin_name}</strong> sizi <strong>{organization_name}</strong> işletmesinin PLANN randevu sistemine personel olarak ekledi.",
                    "Hesabınızı aktif etmek ve şifrenizi belirlemek için lütfen aşağıdaki butona tıklayın.",
                ],
                button={"label": "Şifremi Belirle ve Giriş Yap", "url": invitation_link},
            )
        
        logging.info(f"📧 Personel davet e-postası gönderiliyor: {recipient_email} (Token: {invitation_token[:8]}...)")
        
        result = await send_email(
            to_email=recipient_email,
            subject=subject,
            html_content=html_content,
            to_name=recipient_name
        )
        
        if result:
            logging.info(f"✅ Personel davet e-postası başarıyla gönderildi: {recipient_email}")
        else:
            logging.error(f"❌ Personel davet e-postası gönderilemedi: {recipient_email}")
        
        return result
    except Exception as e:
        logging.error(f"❌ E-posta gönderme sırasında beklenmedik hata: {e}")
        import traceback
        logging.error(traceback.format_exc())
        return False

async def send_password_reset_email(user_email: str, user_name: str, reset_link: str, lang: Optional[str] = None):
    """Kullanıcıya şifre sıfırlama linkini içeren kurumsal e-postayı gönderir."""
    try:
        # user_name kontrolü
        if not user_name or user_name.strip() == "":
            user_name = user_email.split("@")[0]  # Email'den isim çıkar
            logging.warning(f"⚠️ [SEND_PASSWORD_RESET_EMAIL] user_name boş, email'den çıkarıldı: {user_name}")
        
        logging.info(f"📧 [SEND_PASSWORD_RESET_EMAIL] user_email: {user_email}, user_name: {user_name}, reset_link: {reset_link[:50]}...")
        
        # Dil belirleme
        lang = (lang or '').strip().lower()
        if lang not in ['tr', 'en']:
            lang = get_email_language(user_email)
        
        if lang == 'en':
            subject = "PLANN Password Reset Request"
            html_content = _brand_email(
                lang='en',
                badge="Security",
                statement="Forgot your password?",
                paragraphs=[
                    f"Hello {user_name},",
                    "We received a password reset request for your PLANN account. Please click the button below to regain access to your account.",
                    "This link will expire after <strong>30 minutes</strong> for security reasons.",
                ],
                button={"label": "Reset Password", "url": reset_link},
                note="If you did not make this request, please ignore this email. Your account will remain secure.",
            )
        else:
            subject = "PLANN Şifre Sıfırlama Talebi"
            html_content = _brand_email(
                lang='tr',
                badge="Güvenlik",
                statement="Şifrenizi mi unuttunuz?",
                paragraphs=[
                    f"Merhaba {user_name},",
                    "PLANN hesabınız için bir şifre sıfırlama talebi aldık. Hesabınıza yeniden erişim sağlamak için lütfen aşağıdaki butona tıklayın.",
                    "Bu link, güvenlik nedeniyle <strong>30 dakika</strong> sonra geçerliliğini yitirecektir.",
                ],
                button={"label": "Şifremi Sıfırla", "url": reset_link},
                note="Eğer bu talebi siz yapmadıysanız, bu e-postayı dikkate almayınız. Hesabınız güvende kalmaya devam edecektir.",
            )
        
        # HTML içeriğinin uzunluğunu kontrol et
        html_length = len(html_content)
        logging.info(f"📧 [SEND_PASSWORD_RESET_EMAIL] HTML içerik uzunluğu: {html_length} karakter")
        
        if html_length < 100:
            logging.error(f"❌ [SEND_PASSWORD_RESET_EMAIL] HTML içerik çok kısa! ({html_length} karakter)")
            logging.error(f"❌ [SEND_PASSWORD_RESET_EMAIL] HTML içerik: {html_content[:200]}")
            return False
        
        result = await send_email(
            to_email=user_email,
            subject=subject,
            html_content=html_content,
            to_name=user_name,
            sender_name="PLANN Destek"
        )
        
        logging.info(f"📧 [SEND_PASSWORD_RESET_EMAIL] Email gönderim sonucu: {result}")
        return result
    except Exception as e:
        logging.error(f"❌ [SEND_PASSWORD_RESET_EMAIL] E-posta gönderme sırasında beklenmedik hata: {e}", exc_info=True)
        return False

async def send_contact_notification_email(contact_name: str, contact_phone: str, contact_email: Optional[str], contact_message: Optional[str]):
    """Yeni iletişim talebi için bildirim e-postası gönderir (admin'e)"""
    try:
        # Admin e-posta adresi - environment variable'dan al veya default kullan
        admin_email = os.environ.get('ADMIN_EMAIL', 'fatihsenyuz12@gmail.com')
        
        # Dil belirleme - contact_email veya contact_phone'a göre
        lang = get_email_language(contact_email or contact_phone) if contact_email else 'tr'
        
        phone_html = f'<a href="tel:{contact_phone}" style="color:#1a1a1a;text-decoration:underline;">{contact_phone}</a>'

        if lang == 'en':
            subject = "PLANN - New Contact Request"
            email_display = contact_email if contact_email else "<em>Not provided</em>"
            meta_rows = [
                ("Full Name", contact_name),
                ("Phone", phone_html),
                ("Email", email_display),
            ]
            if contact_message:
                meta_rows.append(("Message", contact_message))
            html_content = _brand_email(
                lang='en',
                badge="New Contact Request",
                statement="A new contact request has been received.",
                meta_rows=meta_rows,
                note="Please contact the customer as soon as possible.",
            )
        else:
            subject = "PLANN - Yeni İletişim Talebi"
            email_display = contact_email if contact_email else "<em>Belirtilmemiş</em>"
            meta_rows = [
                ("Ad Soyad", contact_name),
                ("Telefon", phone_html),
                ("E-posta", email_display),
            ]
            if contact_message:
                meta_rows.append(("Mesaj", contact_message))
            html_content = _brand_email(
                lang='tr',
                badge="Yeni İletişim Talebi",
                statement="Yeni bir iletişim talebi alındı.",
                meta_rows=meta_rows,
                note="Lütfen en kısa sürede müşteri ile iletişime geçin.",
            )
        
        result = await send_email(
            to_email=admin_email,
            subject=subject,
            html_content=html_content,
            to_name="PLANN Yönetim",
            sender_name="PLANN Sistem"
        )
        
        logging.info(f"📧 [SEND_CONTACT_NOTIFICATION] Email gönderim sonucu: {result}")
        return result
    except Exception as e:
        logging.error(f"❌ [SEND_CONTACT_NOTIFICATION] E-posta gönderme sırasında beklenmedik hata: {e}", exc_info=True)
        return False

async def send_contact_confirmation_email(contact_name: str, contact_email: str):
    """Kullanıcıya iletişim talebi onay e-postası gönderir"""
    try:
        dashboard_url = "https://plannapp.co"
        subject = "PLANN - Talebiniz Alındı"
        html_content = _brand_email(
            lang='tr',
            badge="Talebiniz Alındı",
            statement="Talebiniz bize ulaştı.",
            paragraphs=[
                f"Merhaba {contact_name},",
                "PLANN iletişim formunu doldurduğunuz için teşekkür ederiz.",
                "İletişim bilgileriniz kaydedildi ve en kısa sürede sizinle iletişime geçeceğiz.",
            ],
            button={"label": "PLANN'ı Keşfet", "url": dashboard_url},
        )
        
        result = await send_email(
            to_email=contact_email,
            subject=subject,
            html_content=html_content,
            to_name=contact_name,
            sender_name="PLANN"
        )
        
        logging.info(f"📧 [SEND_CONTACT_CONFIRMATION] Kullanıcıya onay e-postası gönderildi: {contact_email} - Sonuç: {result}")
        return result
    except Exception as e:
        logging.error(f"❌ [SEND_CONTACT_CONFIRMATION] Kullanıcıya e-posta gönderme hatası: {e}", exc_info=True)
        return False

@api_router.post("/contact")
async def submit_contact(request: Request, contact_data: ContactRequest, db = Depends(get_db)):
    """Landing page'den gelen iletişim formu"""
    try:
        logging.info(f"📞 [CONTACT] Yeni iletişim talebi: {contact_data.name} - {contact_data.phone}")
        
        # IP adresini al
        client_ip = None
        if request.client:
            client_ip = request.client.host
        # Nginx proxy arkasındaysa X-Forwarded-For'dan al
        forwarded_for = request.headers.get("X-Forwarded-For")
        if forwarded_for:
            client_ip = forwarded_for.split(",")[0].strip()
        
        contact_doc = {
            "id": str(uuid.uuid4()),
            "name": contact_data.name.strip(),
            "phone": contact_data.phone.strip(),
            "email": contact_data.email.strip() if contact_data.email else None,
            "message": contact_data.message.strip() if contact_data.message else None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "status": "pending",  # pending, contacted, resolved
            "ip_address": client_ip
        }
        
        await db.contact_requests.insert_one(contact_doc)
        logging.info(f"✅ [CONTACT] İletişim talebi kaydedildi: {contact_doc['id']}")
        
        # Admin'e e-posta bildirimi gönder
        try:
            await send_contact_notification_email(
                contact_name=contact_data.name,
                contact_phone=contact_data.phone,
                contact_email=contact_data.email,
                contact_message=contact_data.message
            )
        except Exception as email_error:
            logging.error(f"⚠️ [CONTACT] Admin'e e-posta bildirimi gönderilemedi: {email_error}")
            # E-posta gönderilemese bile kayıt başarılı olmalı
        
        # Kullanıcıya onay e-postası gönder (eğer e-posta girildiyse)
        if contact_data.email and contact_data.email.strip():
            try:
                await send_contact_confirmation_email(
                    contact_name=contact_data.name,
                    contact_email=contact_data.email.strip()
                )
            except Exception as user_email_error:
                logging.error(f"⚠️ [CONTACT] Kullanıcıya onay e-postası gönderilemedi: {user_email_error}")
                # Kullanıcı e-postası gönderilemese bile kayıt başarılı olmalı
        
        return {"success": True, "message": "Talebiniz alındı, en kısa sürede size ulaşacağız."}
    except Exception as e:
        logging.error(f"❌ [CONTACT] İletişim talebi kaydedilemedi: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Bir hata oluştu, lütfen tekrar deneyin.")

@api_router.post("/forgot-password")
@rate_limit("3/hour")
async def forgot_password(request: Request, forgot_request: ForgotPasswordRequest, db = Depends(get_db)):
    """Kullanıcıya şifre sıfırlama e-postası gönderir."""
    try:
        logging.info(f"🔐 [FORGOT_PASSWORD] Request alındı: {forgot_request.username}")
        # Kullanıcıyı bul
        user = await get_user_from_db(request, forgot_request.username, db=db)
        if not user:
            # Güvenlik nedeniyle kullanıcı yoksa da başarılı mesajı döndür
            logging.warning(f"⚠️ [FORGOT_PASSWORD] Kullanıcı bulunamadı: {forgot_request.username}")
            return {"message": "Eğer bu e-posta adresi kayıtlıysa, şifre sıfırlama linki gönderildi."}
        
        logging.info(f"✅ [FORGOT_PASSWORD] Kullanıcı bulundu: {user.username}")

        accept_language = request.headers.get('Accept-Language', '')
        request_lang = get_request_language(request) if accept_language else None
        user_lang = getattr(user, 'language', None)
        lang = user_lang or request_lang or get_email_language(user.username)
        if lang not in ['tr', 'en']:
            lang = 'tr'
        
        # Benzersiz token oluştur
        reset_token = str(uuid.uuid4()) + str(uuid.uuid4()).replace('-', '')
        
        # Token'ı veritabanına kaydet (30 dakika geçerli)
        expires_at = datetime.now(timezone.utc) + timedelta(minutes=30)
        await db.password_reset_tokens.insert_one({
            "username": user.username,
            "token": reset_token,
            "expires_at": expires_at.isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "used": False
        })
        
        # Reset link oluştur
        reset_link = f"https://plannapp.co.uk/reset-password?token={reset_token}" if lang == 'en' else f"https://plannapp.co/reset-password?token={reset_token}"
        
        # E-posta gönder
        user_name = user.full_name or user.username
        logging.info(f"🔐 [FORGOT_PASSWORD] Token oluşturuldu, email gönderiliyor: {user.username}")
        logging.info(f"🔐 [FORGOT_PASSWORD] Reset link: {reset_link}")
        try:
            email_sent = await send_password_reset_email(user.username, user_name, reset_link, lang=lang)
            logging.info(f"🔐 [FORGOT_PASSWORD] send_password_reset_email çağrıldı, sonuç: {email_sent}")
        except Exception as email_error:
            logging.error(f"❌ [FORGOT_PASSWORD] Email gönderim hatası: {email_error}", exc_info=True)
            email_sent = False
        
        if email_sent:
            logging.info(f"✅ [FORGOT_PASSWORD] Şifre sıfırlama token'ı oluşturuldu ve e-posta gönderildi: {user.username}")
        else:
            logging.error(f"❌ [FORGOT_PASSWORD] Şifre sıfırlama token'ı oluşturuldu ancak e-posta gönderilemedi: {user.username}")
        
        # Güvenlik nedeniyle her zaman başarılı mesajı döndür
        return {"message": "Eğer bu e-posta adresi kayıtlıysa, şifre sıfırlama linki gönderildi."}
    except Exception as e:
        logging.error(f"Şifre sıfırlama talebi hatası: {e}")
        import traceback
        logging.error(traceback.format_exc())
        # Güvenlik nedeniyle hata durumunda da başarılı mesajı döndür
        return {"message": "Eğer bu e-posta adresi kayıtlıysa, şifre sıfırlama linki gönderildi."}

@api_router.post("/auth/setup-password")
@rate_limit(LIMITS['register'])
async def setup_password(request: Request, setup_request: SetupPasswordRequest, db = Depends(get_db)):
    """Personel davet token'ı ile şifre belirleme."""
    try:
        # Token ile kullanıcıyı bul
        user = await db.users.find_one({"invitation_token": setup_request.token})
        if not user:
            raise HTTPException(status_code=400, detail="Geçersiz veya süresi dolmuş davet linki")
        
        # Kullanıcı zaten aktif mi kontrol et
        if user.get("status") == "active":
            raise HTTPException(status_code=400, detail="Bu hesap zaten aktif edilmiş")
        
        # Şifreyi hashle
        hashed_password = get_password_hash(setup_request.new_password)
        
        # Kullanıcıyı güncelle: şifre ekle, status'u active yap, invitation_token'ı sil
        await db.users.update_one(
            {"invitation_token": setup_request.token},
            {
                "$set": {
                    "hashed_password": hashed_password,
                    "status": "active"
                },
                "$unset": {
                    "invitation_token": ""
                }
            }
        )
        
        logging.info(f"Personel şifre belirleme tamamlandı: {user.get('username')}")
        return {"message": "Şifreniz başarıyla belirlendi. Giriş yapabilirsiniz."}
    except HTTPException: raise
    except Exception as e:
        logging.error(f"Şifre belirleme hatası: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="Şifre belirlenirken bir hata oluştu")

@api_router.post("/reset-password")
@rate_limit(LIMITS['register'])
async def reset_password(request: Request, reset_request: ResetPasswordRequest, db = Depends(get_db)):
    """Token ile şifreyi sıfırlar."""
    try:
        # Token'ı bul
        token_doc = await db.password_reset_tokens.find_one({
            "token": reset_request.token,
            "used": False
        })
        
        if not token_doc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Geçersiz veya kullanılmış token."
            )
        
        # Token'ın süresi dolmuş mu kontrol et
        expires_at = datetime.fromisoformat(token_doc['expires_at'].replace('Z', '+00:00'))
        if datetime.now(timezone.utc) > expires_at:
            # Süresi dolmuş token'ı işaretle
            await db.password_reset_tokens.update_one(
                {"token": reset_request.token},
                {"$set": {"used": True}}
            )
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Token'ın süresi dolmuş. Lütfen yeni bir şifre sıfırlama talebi oluşturun."
            )
        
        # Kullanıcıyı bul
        username = token_doc['username']
        user = await get_user_from_db(request, username, db=db)
        if not user:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Kullanıcı bulunamadı."
            )
        
        # Yeni şifreyi hashle ve güncelle
        new_hashed_password = get_password_hash(reset_request.new_password)
        await db.users.update_one(
            {"username": username},
            {"$set": {"hashed_password": new_hashed_password}}
        )
        
        # Token'ı kullanıldı olarak işaretle
        await db.password_reset_tokens.update_one(
            {"token": reset_request.token},
            {"$set": {"used": True}}
        )
        
        logging.info(f"Şifre başarıyla sıfırlandı: {username}")
        return {"message": "Şifreniz başarıyla sıfırlandı. Yeni şifrenizle giriş yapabilirsiniz."}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Şifre sıfırlama hatası: {e}")
        import traceback
        logging.error(traceback.format_exc())
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Şifre sıfırlama sırasında bir hata oluştu. Lütfen tekrar deneyin."
        )

# === CUSTOMERS DENORMALIZATION HELPERS ===
# Randevu mutation'larında `customers` collection'daki denormalize sayaçları
# (`total_appointments`, `completed_appointments`, `last_appointment_at`,
# `first_appointment_at`) tutarlı tutmak için tek noktadan yönetilir.
# GET /customers cursor pagination bu sayaçlara dayanır — full appointments
# scan yerine sadece customers collection'ından okur → sub-100ms yanıt.


def _appointment_datetime_iso(date_str: str, time_str: str) -> Optional[str]:
    """`YYYY-MM-DD` + `HH:MM` → ISO datetime string (Europe/Istanbul).

    Bozuk girdide None döner; çağıran taraf min/max operatörü ile atlayabilir.
    """
    try:
        if not date_str:
            return None
        tstr = time_str or "00:00"
        naive = datetime.strptime(f"{date_str} {tstr}", "%Y-%m-%d %H:%M")
        return naive.replace(tzinfo=ZoneInfo("Europe/Istanbul")).isoformat()
    except Exception:
        return None


# Türkçe karakter-tolerant case-insensitive regex builder.
# MongoDB `$regex` + `$options: 'i'` sadece ASCII case-insensitive. Türkçe için
# İ/i/ı/I, Ş/ş/S/s, Ç/ç/C/c, Ğ/ğ/G/g, Ü/ü/U/u, Ö/ö/O/o çiftleri arasında
# match yapmaz. Bu helper girilen string'i karakter-bazlı bir char class'a
# çevirir (örn. "şen" → "[şŞsS][eE][nN]"), böylece kullanıcı hangi harflerle
# yazsa da isim doğru eşleşir.
_TR_CASE_CLASS_MAP = {
    'i': '[iıIİ]', 'ı': '[iıIİ]', 'I': '[iıIİ]', 'İ': '[iıIİ]',
    's': '[sSşŞ]', 'S': '[sSşŞ]', 'ş': '[sSşŞ]', 'Ş': '[sSşŞ]',
    'c': '[cCçÇ]', 'C': '[cCçÇ]', 'ç': '[cCçÇ]', 'Ç': '[cCçÇ]',
    'g': '[gGğĞ]', 'G': '[gGğĞ]', 'ğ': '[gGğĞ]', 'Ğ': '[gGğĞ]',
    'u': '[uUüÜ]', 'U': '[uUüÜ]', 'ü': '[uUüÜ]', 'Ü': '[uUüÜ]',
    'o': '[oOöÖ]', 'O': '[oOöÖ]', 'ö': '[oOöÖ]', 'Ö': '[oOöÖ]',
    'a': '[aAâÂ]', 'A': '[aAâÂ]', 'â': '[aAâÂ]', 'Â': '[aAâÂ]',
    'e': '[eE]', 'E': '[eE]',
    'b': '[bB]', 'B': '[bB]',
    'd': '[dD]', 'D': '[dD]',
    'f': '[fF]', 'F': '[fF]',
    'h': '[hH]', 'H': '[hH]',
    'j': '[jJ]', 'J': '[jJ]',
    'k': '[kK]', 'K': '[kK]',
    'l': '[lL]', 'L': '[lL]',
    'm': '[mM]', 'M': '[mM]',
    'n': '[nN]', 'N': '[nN]',
    'p': '[pP]', 'P': '[pP]',
    'r': '[rR]', 'R': '[rR]',
    't': '[tT]', 'T': '[tT]',
    'v': '[vV]', 'V': '[vV]',
    'y': '[yY]', 'Y': '[yY]',
    'z': '[zZ]', 'Z': '[zZ]',
    'w': '[wW]', 'W': '[wW]',
    'x': '[xX]', 'X': '[xX]',
    'q': '[qQ]', 'Q': '[qQ]',
}


def build_turkish_case_insensitive_pattern(s: str) -> str:
    parts = []
    for ch in s:
        klass = _TR_CASE_CLASS_MAP.get(ch)
        if klass:
            parts.append(klass)
        else:
            # Rakamlar, boşluk, tire vb. — regex için escape et.
            parts.append(re.escape(ch))
    return ''.join(parts)


def _phone_match_variants(phone: str) -> list:
    """Bir telefonun olası saklama formatlarını üretir (05.. / +90.. / 90.. / 10 hane).

    Randevudan otomatik müşteri eklemede, aynı kişinin farklı formatla girilip
    mükerrer kayıt oluşturmasını engellemek için kullanılır. 10 haneden kısa
    (ör. "0" telefonsuz walk-in) değerlerde davranış DEĞİŞMEZ: sadece kendisini döner.
    """
    raw = (phone or "").strip()
    clean = re.sub(r"\D", "", raw)
    if len(clean) < 10:
        return [raw] if raw else []
    variants = {raw, clean}
    core = None
    if clean.startswith("90") and len(clean) == 12:
        core = clean[2:]
    elif clean.startswith("0") and len(clean) == 11:
        core = clean[1:]
    elif len(clean) == 10 and not clean.startswith("0"):
        core = clean
    if core and len(core) == 10:
        variants.update({core, "0" + core, "90" + core, "+90" + core})
    return list(variants)


async def _apply_customer_delta(
    db,
    organization_id: str,
    phone: str,
    name: Optional[str] = None,
    delta_total: int = 0,
    delta_completed: int = 0,
    appointment_date: Optional[str] = None,
    appointment_time: Optional[str] = None,
) -> None:
    """`customers` collection'ında tek atomik upsert ile sayaçları günceller.

    - `delta_total` / `delta_completed` sıfır dahi olsa çağrılabilir (isim/tarih
      güncellemek için); bu durumda sadece `$max`/`$set` uygulanır.
    - Doküman yoksa `$setOnInsert` ile bootstrap yapılır (id, created_at, notes).
    - Sayaçlar sıfırın altına düşmez — geriye doğru inc'den sonra clamp yapılır.
    - Hata loglanır ama swallow edilir (asıl mutation'ı bloklamamak için).
    """
    if not organization_id or not phone:
        return

    try:
        query = {"organization_id": organization_id, "phone": phone}
        now_iso = datetime.now(timezone.utc).isoformat()

        set_on_insert = {
            "id": str(uuid.uuid4()),
            "organization_id": organization_id,
            "phone": phone,
            "created_at": now_iso,
            "notes": "",
        }
        set_ops: dict = {"updated_at": now_iso}
        if name and name.strip():
            set_ops["name"] = name.strip()
        else:
            set_on_insert["name"] = ""

        update_doc: dict = {"$setOnInsert": set_on_insert, "$set": set_ops}

        inc_ops: dict = {}
        if delta_total:
            inc_ops["total_appointments"] = int(delta_total)
        if delta_completed:
            inc_ops["completed_appointments"] = int(delta_completed)
        if inc_ops:
            update_doc["$inc"] = inc_ops

        apt_dt_iso = _appointment_datetime_iso(appointment_date or "", appointment_time or "")
        max_ops: dict = {}
        min_ops: dict = {}
        if apt_dt_iso:
            max_ops["last_appointment_at"] = apt_dt_iso
            min_ops["first_appointment_at"] = apt_dt_iso
        if max_ops:
            update_doc["$max"] = max_ops
        if min_ops:
            update_doc["$min"] = min_ops

        await db.customers.update_one(query, update_doc, upsert=True)

        # Negatif sayaç clamp (delete edilen randevu customers'a hiç yansımadıysa
        # inc -1 → -1'e düşebilir; kullanıcıya yanlış sayı göstermemek için 0'a çek).
        if delta_total < 0 or delta_completed < 0:
            fix: dict = {}
            doc = await db.customers.find_one(
                query,
                {"_id": 0, "total_appointments": 1, "completed_appointments": 1},
            )
            if doc:
                if (doc.get("total_appointments") or 0) < 0:
                    fix["total_appointments"] = 0
                if (doc.get("completed_appointments") or 0) < 0:
                    fix["completed_appointments"] = 0
                if fix:
                    await db.customers.update_one(query, {"$set": fix})
    except Exception as exc:
        logging.warning(
            f"_apply_customer_delta failed org={organization_id[:8] if organization_id else '?'} "
            f"phone={phone} total={delta_total} completed={delta_completed}: {exc}"
        )


# === APPOINTMENTS ROUTES ===
@api_router.delete("/appointments/groups/{session_group_id}")
async def delete_session_group(
    request: Request,
    session_group_id: str,
    current_user: UserInDB = Depends(get_current_user),
):
    """Bir seans paketinin (session_group_id) tamamını ve bağlı transaction'larını siler.

    PLANN v2 — kullanıcı onayıyla, tamamlananlar dahil tüm grup randevuları
    hard-delete edilir. Operasyon geri alınamaz; audit_log tutulur ve
    WebSocket ile bağlı tüm panellerde anlık yansıtılır.
    """
    if current_user.role not in ("admin", "superadmin"):
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok.")

    db = await get_db_from_request(request)
    org_id = current_user.organization_id
    base_query = {"session_group_id": session_group_id, "organization_id": org_id}

    apts = await db.appointments.find(base_query, {"_id": 0}).to_list(None)
    if not apts:
        raise HTTPException(status_code=404, detail="Seans paketi bulunamadı.")

    apt_ids = [a["id"] for a in apts]

    # 1) Randevuları sil
    del_result = await db.appointments.delete_many(base_query)

    # 2) Bağlı transaction'ları sil (legacy "transactions" koleksiyonu)
    try:
        await db.transactions.delete_many({
            "appointment_id": {"$in": apt_ids},
            "organization_id": org_id,
        })
    except Exception as exc:
        logger.warning(f"transactions cleanup failed for group={session_group_id}: {exc}")

    # 3) Audit log
    try:
        await create_audit_log(
            db=db,
            organization_id=org_id,
            user_id=current_user.username,
            user_full_name=current_user.full_name or current_user.username,
            action="DELETE",
            resource_type="SESSION_GROUP",
            resource_id=session_group_id,
            old_value={"deleted_count": del_result.deleted_count, "appointment_ids": apt_ids},
            ip_address=request.client.host if request.client else None,
        )
    except Exception as exc:
        logger.warning(f"audit log failed for session_group_delete={session_group_id}: {exc}")

    # 4) WebSocket — her appointment için ayrı event (frontend zaten dinliyor)
    try:
        for apt_id in apt_ids:
            await emit_to_organization(org_id, "appointment_deleted", {"appointment_id": apt_id})
    except Exception as exc:
        logger.warning(f"emit failed for session_group_delete={session_group_id}: {exc}")

    # 5) Denormalize sayaç güncellemesi — grup tek müşteriye ait olsa da
    #    (phone,name) bazında agregasyon yap; auto-complete'li karma durumlarda güvenli.
    try:
        deltas: dict = {}
        for _a in apts:
            _ph = (_a.get("phone") or "").strip()
            if not _ph:
                continue
            _key = (_ph, _a.get("customer_name") or "")
            if _key not in deltas:
                deltas[_key] = {"total": 0, "completed": 0}
            deltas[_key]["total"] -= 1
            if _a.get("status") == "Tamamlandı":
                deltas[_key]["completed"] -= 1
        for (_ph, _name), _d in deltas.items():
            await _apply_customer_delta(
                db,
                organization_id=org_id,
                phone=_ph,
                name=_name,
                delta_total=_d["total"],
                delta_completed=_d["completed"],
            )
    except Exception as _delta_err:
        logger.warning(f"customer delta (session_group_delete) skipped: {_delta_err}")

    # 6) Cache invalidation — grup silme sonrası müşteri/dashboard listesi tazelensin
    try:
        await invalidate_cache(request, "dashboard_stats", current_user)
        await invalidate_cache(request, "customers_list", current_user)
    except Exception:
        pass

    logger.info(
        f"Session group deleted: group={session_group_id} count={del_result.deleted_count} by={current_user.username}"
    )
    return {"deleted": del_result.deleted_count, "session_group_id": session_group_id}


@api_router.delete("/appointments/{appointment_id}")
async def delete_appointment(request: Request, appointment_id: str, current_user: UserInDB = Depends(get_current_user)):
    db = await get_db_from_request(request); query = {"id": appointment_id, "organization_id": current_user.organization_id}
    
    # Get appointment before deleting (for audit log)
    appointment = await db.appointments.find_one(query, {"_id": 0})
    if not appointment:
        raise HTTPException(status_code=404, detail="Randevu bulunamadı")
    
    result = await db.appointments.delete_one(query)
    
    # Randevuyla ilişkili transaction'ları da sil
    try:
        transaction_query = {
            "appointment_id": appointment_id,
            "organization_id": current_user.organization_id
        }
        transaction_delete_result = await db.transactions.delete_many(transaction_query)
        if transaction_delete_result.deleted_count > 0:
            logger.info(f"Deleted {transaction_delete_result.deleted_count} transaction(s) for appointment {appointment_id}")
    except Exception as e:
        logger.error(f"Error deleting transactions for appointment {appointment_id}: {e}", exc_info=True)

    # Denormalize sayaç güncellemesi: -1 total, was Tamamlandı ise -1 completed
    try:
        _was_completed = appointment.get("status") == "Tamamlandı"
        await _apply_customer_delta(
            db,
            organization_id=current_user.organization_id,
            phone=appointment.get("phone") or "",
            name=appointment.get("customer_name"),
            delta_total=-1,
            delta_completed=-1 if _was_completed else 0,
        )
    except Exception as _delta_err:
        logger.warning(f"customer delta (delete_appointment) skipped: {_delta_err}")

    # Cache invalidation — mevcut mimariyle uyumlu (bkz. POST /appointments)
    try:
        await invalidate_cache(request, "dashboard_stats", current_user)
        await invalidate_cache(request, "customers_list", current_user)
    except Exception:
        pass
    
    # Audit log
    await create_audit_log(
        db=db,
        organization_id=current_user.organization_id,
        user_id=current_user.username,
        user_full_name=current_user.full_name or current_user.username,
        action="DELETE",
        resource_type="APPOINTMENT",
        resource_id=appointment_id,
        old_value=appointment,
        ip_address=request.client.host if request.client else None
    )
    
    # Emit WebSocket event for real-time update
    logger.info(f"About to emit appointment_deleted for org: {current_user.organization_id}")
    try:
        await emit_to_organization(
            current_user.organization_id,
            'appointment_deleted',
            {'appointment_id': appointment_id}
        )
        logger.info(f"Successfully emitted appointment_deleted for org: {current_user.organization_id}")
    except Exception as emit_error:
        logger.error(f"Failed to emit appointment_deleted: {emit_error}", exc_info=True)
    
    return {"message": "Randevu silindi"}

@api_router.put("/appointments/{appointment_id}", response_model=Appointment)
async def update_appointment(request: Request, appointment_id: str, appointment_update: AppointmentUpdate, current_user: UserInDB = Depends(get_current_user)):
    db = await get_db_from_request(request); settings_data = await db.settings.find_one({"organization_id": current_user.organization_id})
    if not settings_data:
        default_settings = Settings(organization_id=current_user.organization_id); settings_data = default_settings.model_dump()
    company_name = settings_data.get("company_name", "İşletmeniz"); support_phone = settings_data.get("support_phone", "Destek Hattı")
    query = {"id": appointment_id, "organization_id": current_user.organization_id}; appointment = await db.appointments.find_one(query, {"_id": 0})
    if not appointment: raise HTTPException(status_code=404, detail="Randevu bulunamadı")
    update_data = {k: v for k, v in appointment_update.model_dump().items() if v is not None}
    # Çoklu hizmet güncellemesi: service_id veya service_ids geldiyse snapshot + toplam değerleri yeniden kur
    if 'service_ids' in update_data or 'service_id' in update_data:
        _settings_flag = await db.settings.find_one({"organization_id": current_user.organization_id}, {"_id": 0, "multi_service_enabled": 1})
        _multi_enabled = bool((_settings_flag or {}).get("multi_service_enabled", False))
        _snap, _tot_dur, _tot_price, _merged, _ids = await resolve_services(
            db, current_user.organization_id, appointment_update, multi_enabled=_multi_enabled
        )
        update_data.pop('service_ids', None)
        update_data['service_id'] = _ids[0]
        update_data['services'] = _snap
        update_data['service_name'] = _merged['name']
        update_data['service_price'] = _merged['price']
        update_data['service_duration'] = _merged['duration']
    # Tarih/saat veya personel değişikliği varsa çakışma kontrolü yap
    if 'appointment_date' in update_data or 'appointment_time' in update_data or 'staff_member_id' in update_data:
        check_date = update_data.get('appointment_date', appointment['appointment_date'])
        check_time = update_data.get('appointment_time', appointment['appointment_time'])
        check_staff = update_data.get('staff_member_id', appointment.get('staff_member_id'))

        # Hizmet süresini bul: çoklu hizmet güncellemesinde update_data'daki toplam süre kullanılır
        if 'service_duration' in update_data:
            service_duration = update_data['service_duration']
        else:
            check_service_id = update_data.get('service_id', appointment.get('service_id'))
            service_duration = appointment.get('service_duration') or 30
            if check_service_id:
                service_doc = await db.services.find_one(
                    {"id": check_service_id, "organization_id": current_user.organization_id},
                    {"_id": 0, "duration": 1}
                )
                if service_doc and service_doc.get('duration'):
                    service_duration = service_doc.get('duration')

        # Yeni randevunun bitiş saatini hesapla
        new_start_hour, new_start_minute = map(int, check_time.split(':'))
        new_end_minute = new_start_minute + int(service_duration)
        new_end_hour = new_start_hour + (new_end_minute // 60)
        new_end_minute = new_end_minute % 60
        new_end_time = f"{str(new_end_hour).zfill(2)}:{str(new_end_minute).zfill(2)}"

        # Aynı personelin aynı günkü randevularında overlap kontrolü
        existing_appointments = await db.appointments.find(
            {
                "organization_id": current_user.organization_id,
                "id": {"$ne": appointment_id},
                "staff_member_id": check_staff,
                "appointment_date": check_date,
                # Müsaitlik ile tutarlı: ödeme bekleyen/iptal slotları çakışma saymaz
                "status": {"$nin": ["İptal", "İptal Edildi", "Cancelled", "Ödeme Bekleniyor"]}
            },
            {"_id": 0, "appointment_time": 1, "service_id": 1, "service_duration": 1}
        ).to_list(200)

        def time_to_minutes(time_str: str) -> int:
            try:
                hour, minute = map(int, time_str.split(':'))
                return hour * 60 + minute
            except Exception:
                return 0

        new_start_min = time_to_minutes(check_time)
        new_end_min = time_to_minutes(new_end_time)

        for existing_appt in existing_appointments:
            existing_start_time = existing_appt.get('appointment_time')
            if not existing_start_time:
                continue
            existing_duration = existing_appt.get('service_duration')
            if not existing_duration and existing_appt.get('service_id'):
                existing_service = await db.services.find_one(
                    {"id": existing_appt.get('service_id'), "organization_id": current_user.organization_id},
                    {"_id": 0, "duration": 1}
                )
                existing_duration = existing_service.get('duration') if existing_service else 30
            existing_duration = int(existing_duration or 30)

            existing_start_min = time_to_minutes(existing_start_time)
            existing_end_min = existing_start_min + existing_duration

            # overlap: new_start < existing_end AND new_end > existing_start
            if (new_start_min < existing_end_min and new_end_min > existing_start_min):
                raise HTTPException(
                    status_code=400,
                    detail=f"Bu personelin {check_date} tarihinde seçilen saat aralığında zaten randevusu var. Lütfen başka bir saat seçin."
                )

        # Time block (break) çakışması kontrolü
        has_break_conflict = await check_break_conflict(
            db, check_staff, check_date, check_time, new_end_time, current_user.organization_id
        )
        if has_break_conflict:
            raise HTTPException(
                status_code=400,
                detail=f"Bu personelin {check_date} tarihinde {check_time} saatinde mola var. Lütfen başka bir saat seçin."
            )
    new_status = update_data.get('status'); old_status = appointment['status']
    if new_status == 'Tamamlandı' and old_status != 'Tamamlandı':
        update_data['completed_at'] = datetime.now(timezone.utc).isoformat()
        transaction = Transaction(organization_id=current_user.organization_id, appointment_id=appointment_id, customer_name=appointment['customer_name'], service_name=appointment['service_name'], amount=appointment['service_price'], date=appointment['appointment_date'])
        trans_doc = transaction.model_dump(); trans_doc['created_at'] = trans_doc['created_at'].isoformat()
        await db.transactions.insert_one(trans_doc)
        # Tamamlanma SMS'i kaldırıldı (maliyet nedeniyle)
    elif new_status == 'İptal' and old_status != 'İptal':
        # Randevu iptal edildiğinde
        try:
            # İptal SMS'i - Default mesaj kullan
            sms_message = build_sms_message(
                company_name, appointment['customer_name'],
                appointment['appointment_date'], appointment['appointment_time'],
                appointment['service_name'], support_phone, sms_type="cancellation"
            )
            send_sms(appointment['phone'], sms_message)
        except Exception as e: logging.error(f"İptal SMS'i gönderilirken hata oluştu: {e}")
    if update_data: await db.appointments.update_one(query, {"$set": update_data})
    updated_appointment = await db.appointments.find_one(query, {"_id": 0})
    if isinstance(updated_appointment['created_at'], str): updated_appointment['created_at'] = datetime.fromisoformat(updated_appointment['created_at'].replace('Z', '+00:00'))
    
    # Audit log
    await create_audit_log(
        db=db,
        organization_id=current_user.organization_id,
        user_id=current_user.username,
        user_full_name=current_user.full_name or current_user.username,
        action="UPDATE",
        resource_type="APPOINTMENT",
        resource_id=appointment_id,
        old_value=appointment,
        new_value=updated_appointment,
        ip_address=request.client.host if request.client else None
    )
    
    # Emit WebSocket event for real-time update
    # Convert datetime to ISO format if it's a datetime object
    # make_json_serializable function will handle this, but we can also do it explicitly
    if isinstance(updated_appointment.get('created_at'), datetime):
        updated_appointment['created_at'] = updated_appointment['created_at'].isoformat()
    logger.info(f"About to emit appointment_updated for org: {current_user.organization_id}")
    try:
        await emit_to_organization(
            current_user.organization_id,
            'appointment_updated',
            {'appointment': updated_appointment}
        )
        logger.info(f"Successfully emitted appointment_updated for org: {current_user.organization_id}")
    except Exception as emit_error:
        logger.error(f"Failed to emit appointment_updated: {emit_error}", exc_info=True)
    
    # Denormalize sayaç güncellemesi + cache invalidation
    try:
        _phone = updated_appointment.get("phone") or appointment.get("phone") or ""
        _name = updated_appointment.get("customer_name") or appointment.get("customer_name")
        _old_st = appointment.get("status")
        _new_st = updated_appointment.get("status")
        _delta_completed = 0
        if _old_st != "Tamamlandı" and _new_st == "Tamamlandı":
            _delta_completed = 1
        elif _old_st == "Tamamlandı" and _new_st != "Tamamlandı":
            _delta_completed = -1
        # Telefon nadiren değişir; buraya ekstra kompleksite eklemiyoruz — customer
        # PUT endpoint'i zaten kendi delta'sını yönetir.
        if _phone and (_delta_completed or _name):
            await _apply_customer_delta(
                db,
                organization_id=current_user.organization_id,
                phone=_phone,
                name=_name,
                delta_total=0,
                delta_completed=_delta_completed,
                appointment_date=updated_appointment.get("appointment_date"),
                appointment_time=updated_appointment.get("appointment_time"),
            )
    except Exception as _delta_err:
        logger.warning(f"customer delta (update_appointment) skipped: {_delta_err}")

    try:
        await invalidate_cache(request, "dashboard_stats", current_user)
        await invalidate_cache(request, "customers_list", current_user)
    except Exception:
        pass

    # Tarih/saat değişince müşteriye geçici olarak CONFIRMATION şablonu gönderiliyor
    try:
        od, ot = appointment.get("appointment_date"), appointment.get("appointment_time")
        nd, nt = updated_appointment.get("appointment_date"), updated_appointment.get("appointment_time")
        
        if (od != nd or ot != nt) and updated_appointment.get("phone"):
            from whatsapp_service import send_whatsapp_template, WHATSAPP_ENABLED
            
            if WHATSAPP_ENABLED:
                ph = updated_appointment.get("phone")
                nm = (updated_appointment.get("customer_name") or "").strip()
                
                send_whatsapp_template(
                    to_number=ph,
                    template_type="CONFIRMATION",
                    customer_name=nm,
                    company_name=company_name, 
                    appointment_date=nd,       
                    appointment_time=nt,       
                    service_name=updated_appointment.get("service_name", ""),
                    support_phone=updated_appointment.get("support_phone", ""),
                    business_lat=updated_appointment.get("lat"), 
                    business_lng=updated_appointment.get("lng"),
                    business_address=updated_appointment.get("address", "")
                )
    except Exception as wa_upd_err:
        logging.warning(f"Reschedule WhatsApp skipped: {wa_upd_err}")
    
    return updated_appointment

@api_router.get("/appointments/{appointment_id}", response_model=Appointment)
async def get_appointment(request: Request, appointment_id: str, current_user: UserInDB = Depends(get_current_user)):
    db = await get_db_from_request(request)
    query = {"id": appointment_id, "organization_id": current_user.organization_id}
    appointment = await db.appointments.find_one(query, {"_id": 0})
    if not appointment: raise HTTPException(status_code=404, detail="Randevu bulunamadı")
    if isinstance(appointment['created_at'], str): appointment['created_at'] = datetime.fromisoformat(appointment['created_at'].replace('Z', '+00:00'))
    return appointment

# Yardımcı fonksiyon: Mola çakışması kontrolü
async def check_break_conflict(db, staff_id: str, date: str, start_time: str, end_time: str, organization_id: str) -> bool:
    """Personelin belirtilen tarih ve saatte mola çakışması var mı kontrol eder"""
    try:
        # Personeli bul ve molalarını al
        staff = await db.users.find_one(
            {"username": staff_id, "organization_id": organization_id},
            {"_id": 0, "breaks": 1}
        )
        
        if not staff:
            return False
        
        staff_breaks = staff.get('breaks', [])
        
        # Zamanları dakika cinsinden sayıya çevir
        def time_to_minutes(time_str):
            try:
                hour, minute = map(int, time_str.split(':'))
                return hour * 60 + minute
            except (ValueError, AttributeError):
                return 0
        
        new_start_min = time_to_minutes(start_time)
        new_end_min = time_to_minutes(end_time)
        
        # O tarihteki molaları kontrol et
        for brk in staff_breaks:
            if brk.get('date') == date:
                break_start = brk.get('start_time')
                break_end = brk.get('end_time')
                
                if break_start and break_end:
                    break_start_min = time_to_minutes(break_start)
                    break_end_min = time_to_minutes(break_end)
                    
                    # Çakışma kontrolü: (yeni_başlangıç < mola_bitiş) VE (yeni_bitiş > mola_başlangıç)
                    if (new_start_min < break_end_min and new_end_min > break_start_min):
                        logging.info(f"⚠️ Break conflict detected: New {start_time}-{end_time} overlaps with break {break_start}-{break_end} for staff {staff_id}")
                        return True
        
        return False
    except Exception as e:
        logging.error(f"Error checking break conflict: {e}", exc_info=True)
        return False

async def _send_appointment_confirmations(db, phone: str, sms_message: str, wa_params: dict, skip_whatsapp: bool):
    """Randevu onay SMS + WhatsApp'ını response DÖNDÜKTEN sonra gönderir (BackgroundTasks).

    send_sms / send_whatsapp_template SENKRON ağ çağrılarıdır; event loop'u bloklamamak
    için asyncio.to_thread ile thread havuzunda çalıştırılır. Response gecikmesini (~1-1.5s)
    ortadan kaldırır — randevu zaten DB'ye yazılmıştır, bu adımlar best-effort'tur.
    """
    import time as _time

    try:
        await asyncio.to_thread(send_sms, phone, sms_message)
    except Exception as sms_err:
        logger.warning(f"⚠️ SMS gönderilemedi (bg): {sms_err}")

    if skip_whatsapp:
        logging.info("WhatsApp CONFIRMATION skipped (skip_confirmation_whatsapp) — paket şablonu bekleniyor")
        return

    _phone = str(phone).lstrip('+')
    try:
        _wa_msg_id = await asyncio.to_thread(lambda: send_whatsapp_template(**wa_params))
        try:
            await db.whatsapp_message_logs.update_one(
                {"message_id": _wa_msg_id},
                {"$set": {
                    "message_id": _wa_msg_id, "recipient": _phone, "status": "sent",
                    "timestamp": int(_time.time()), "recorded_at": datetime.utcnow().isoformat(),
                    "source": "admin_create",
                }},
                upsert=True,
            )
        except Exception:
            pass
    except Exception as whatsapp_error:
        logger.warning(f"⚠️ WhatsApp mesajı gönderilemedi (bg): {whatsapp_error}")
        try:
            await db.whatsapp_message_logs.update_one(
                {"message_id": f"fail_{_phone}_{int(_time.time())}"},
                {"$set": {
                    "recipient": _phone, "status": "failed", "timestamp": int(_time.time()),
                    "recorded_at": datetime.utcnow().isoformat(),
                    "errors": [{"code": "SEND_ERROR", "title": str(whatsapp_error)[:300]}],
                    "source": "admin_create",
                }},
                upsert=True,
            )
        except Exception:
            pass


# ═══ CONTRACT: v1 (FROZEN / REQUEST) ════════════════════════════════════════
# Donmuş mobil build sözleşmesi. AppointmentCreate body'sine YENİ ZORUNLU alan
# EKLEME (eski client göndermez → 400). Yeni alan = Optional + default.
# bkz. OPERATIONS.md §2.1 + .cursor/rules/api-contract.mdc
@api_router.post("/appointments", response_model=Appointment)
async def create_appointment(request: Request, appointment: AppointmentCreate, background_tasks: BackgroundTasks, current_user: UserInDB = Depends(get_current_user)):
    db = await get_db_from_request(request)

    # Otomatik personel: boş / "auto" / "none" → None (tekli randevuda eligible personel döngüsü)
    _sm = getattr(appointment, "staff_member_id", None)
    if _sm is not None:
        _sl = str(_sm).strip().lower()
        if _sl in ("", "auto", "none"):
            object.__setattr__(appointment, "staff_member_id", None)

    # ---------------------------------------------------------
    # 🔒 ADIM 1: REDIS LOCK (ÇİFTE REZERVASYON ENGELLEYİCİ)
    # ---------------------------------------------------------
    # Aynı personelin aynı saatine aynı anda sızılmasını milisaniyeler içinde engeller.
    redis_client = getattr(request.app.state, 'redis_client', None)
    
    if redis_client:
        # Kilit anahtarı: plann:lock:appt:orgID:staffID:date:time
        staff_id = appointment.staff_member_id or "unassigned"
        lock_key = f"plann:lock:appt:{current_user.organization_id}:{staff_id}:{appointment.appointment_date}:{appointment.appointment_time}"
        
        # setnx: Anahtar yoksa oluşturur. Varsa False döner.
        acquired = await redis_client.setnx(lock_key, "locked")
        if not acquired:
            raise HTTPException(
                status_code=409, 
                detail="Bu randevu saati için şu an başka bir işlem yapılıyor. Lütfen 10 saniye sonra tekrar deneyin veya farklı bir saat seçin."
            )
        # Kilidin asılı kalmaması için 5 saniyelik ömür veriyoruz.
        await redis_client.expire(lock_key, 5)
    # ---------------------------------------------------------
    
    # KOTA KONTROLÜ - Randevu oluşturmadan önce kontrol et
    quota_ok, quota_error = await check_quota_and_increment(db, current_user.organization_id)
    if not quota_ok:
        raise HTTPException(status_code=403, detail=quota_error)
    
    # Çoklu hizmet çözümleme (feature flag arkasında). Tek hizmette eski davranış birebir korunur;
    # çoklu hizmette süre/fiyat toplanır ve "merged_service" mevcut mantığı bozmadan kullanılır.
    _settings_flag = await db.settings.find_one({"organization_id": current_user.organization_id}, {"_id": 0, "multi_service_enabled": 1})
    _multi_enabled = bool((_settings_flag or {}).get("multi_service_enabled", False))
    try:
        snapshot_lines, total_duration, total_price, service, resolved_service_ids = await resolve_services(
            db, current_user.organization_id, appointment, multi_enabled=_multi_enabled
        )
    except HTTPException:
        # Kota artırıldı ama hizmet çözümleme başarısız, geri al
        plan_doc = await db.organization_plans.find_one({"organization_id": current_user.organization_id})
        if plan_doc:
            await db.organization_plans.update_one(
                {"organization_id": current_user.organization_id},
                {"$inc": {"quota_usage": -1}}
            )
        raise
    
    # PERSONEL KONTROL: Staff ise sadece kendi hizmetlerine randevu alabilir (can_view_all_appointments yetkisi yoksa)
    if current_user.role == "staff" and not current_user.can_view_all_appointments:
        if any(sid not in current_user.permitted_service_ids for sid in resolved_service_ids):
            # Kota artırıldı ama yetki yok, geri al
            plan_doc = await db.organization_plans.find_one({"organization_id": current_user.organization_id})
            if plan_doc:
                await db.organization_plans.update_one(
                    {"organization_id": current_user.organization_id},
                    {"$inc": {"quota_usage": -1}}
                )
            raise HTTPException(status_code=403, detail="Bu hizmete randevu alma yetkiniz yok")
        
        # Personel kendi adına randevu oluşturuyor, kendi molasını kontrol et
        # Staff role olduğunda, staff_member_id belirtilmese bile personel kendine atanacak
        service_duration = service.get('duration', 30)
        new_start_hour, new_start_minute = map(int, appointment.appointment_time.split(':'))
        new_end_minute = new_start_minute + service_duration
        new_end_hour = new_start_hour + (new_end_minute // 60)
        new_end_minute = new_end_minute % 60
        new_end_time = f"{str(new_end_hour).zfill(2)}:{str(new_end_minute).zfill(2)}"
        
        # Personelin kendi molasını kontrol et
        has_break_conflict = await check_break_conflict(
            db, current_user.username, appointment.appointment_date,
            appointment.appointment_time, new_end_time, current_user.organization_id
        )
        
        if has_break_conflict:
            # Kota artırıldı ama mola çakışması var, geri al
            plan_doc = await db.organization_plans.find_one({"organization_id": current_user.organization_id})
            if plan_doc:
                await db.organization_plans.update_one(
                    {"organization_id": current_user.organization_id},
                    {"$inc": {"quota_usage": -1}}
                )
            raise HTTPException(
                status_code=400,
                detail=f"Bu saatte mola var. Lütfen başka bir saat seçin."
            )
        
        # Personel kendi adına randevu oluşturuyor, otomatik olarak kendine ata
        # Normal akışa devam et, ama staff_member_id kontrolünü atlamak için 
        # appointment.staff_member_id'yi current_user.username olarak set et
        # Pydantic model olduğu için object.__setattr__ kullan
        if not appointment.staff_member_id or appointment.staff_member_id != current_user.username:
            object.__setattr__(appointment, 'staff_member_id', current_user.username)
    
    # Otomatik atama mantığı
    assigned_staff_id = None
    
    if appointment.staff_member_id:
        # Belirli bir personel seçildi - çakışma kontrolü yap (duration'a göre)
        service_duration = service.get('duration', 30)
        
        # Yeni randevunun başlangıç ve bitiş saatlerini hesapla
        new_start_hour, new_start_minute = map(int, appointment.appointment_time.split(':'))
        new_end_minute = new_start_minute + service_duration
        new_end_hour = new_start_hour + (new_end_minute // 60)
        new_end_minute = new_end_minute % 60
        new_end_time = f"{str(new_end_hour).zfill(2)}:{str(new_end_minute).zfill(2)}"
        
        # Bu personelin o tarihteki tüm randevularını çek
        existing_appointments = await db.appointments.find(
            {
                "organization_id": current_user.organization_id,
                "staff_member_id": appointment.staff_member_id,
                "appointment_date": appointment.appointment_date,
                # Müsaitlik ile tutarlı: ödeme bekleyen/iptal slotları çakışma saymaz
                "status": {"$nin": ["İptal", "İptal Edildi", "Cancelled", "Ödeme Bekleniyor"]}
            },
            {"_id": 0, "appointment_time": 1, "service_id": 1, "service_duration": 1}
        ).to_list(100)
        
        # --- PERFORMANS DOKUNUŞU: BULK FETCH ---
        # 1. Mevcut randevulardaki benzersiz hizmet ID'lerini topla
        unique_service_ids = list(set(appt.get('service_id') for appt in existing_appointments if appt.get('service_id')))
        
        # 2. Tüm bu hizmetlerin sürelerini TEK BİR sorguyla çek
        services_data = await db.services.find(
            {"id": {"$in": unique_service_ids}, "organization_id": current_user.organization_id},
            {"id": 1, "duration": 1, "_id": 0}
        ).to_list(len(unique_service_ids))
        
        # 3. Hızlı arama için bir harita (Map) oluştur
        duration_map = {s['id']: s.get('duration', 30) for s in services_data}
        
        # Her randevunun bitiş saatini hesapla ve çakışma kontrolü yap
        has_conflict = False
        for existing_appt in existing_appointments:
            existing_start_time = existing_appt['appointment_time']
            existing_service_id = existing_appt.get('service_id')
            
            # Çoklu hizmette saklı TOPLAM süreyi kullan; yoksa canlı tek-hizmet süresi (hafızadan)
            existing_duration = existing_appt.get('service_duration') or duration_map.get(existing_service_id, 30)
            
            # Mevcut randevunun bitiş saatini hesapla
            existing_start_hour, existing_start_minute = map(int, existing_start_time.split(':'))
            existing_end_minute = existing_start_minute + existing_duration
            existing_end_hour = existing_start_hour + (existing_end_minute // 60)
            existing_end_minute = existing_end_minute % 60
            existing_end_time = f"{str(existing_end_hour).zfill(2)}:{str(existing_end_minute).zfill(2)}"
            
            # Çakışma kontrolü: Zamanları sayısal değerlere çevir (dakika cinsinden)
            def time_to_minutes(time_str):
                """Zaman string'ini (HH:MM) dakika cinsinden sayıya çevir"""
                try:
                    hour, minute = map(int, time_str.split(':'))
                    return hour * 60 + minute
                except (ValueError, AttributeError):
                    return 0
            
            new_start_min = time_to_minutes(appointment.appointment_time)
            new_end_min = time_to_minutes(new_end_time)
            existing_start_min = time_to_minutes(existing_start_time)
            existing_end_min = time_to_minutes(existing_end_time)
            
            # Çakışma kontrolü: (yeni_başlangıç < mevcut_bitiş) VE (yeni_bitiş > mevcut_başlangıç)
            if (new_start_min < existing_end_min and new_end_min > existing_start_min):
                has_conflict = True
                logging.info(f"⚠️ Conflict detected: New {appointment.appointment_time}-{new_end_time} overlaps with existing {existing_start_time}-{existing_end_time}")
                break
        
        if has_conflict:
            # Kota artırıldı ama çakışma var, geri al
            plan_doc = await db.organization_plans.find_one({"organization_id": current_user.organization_id})
            if plan_doc:
                await db.organization_plans.update_one(
                    {"organization_id": current_user.organization_id},
                    {"$inc": {"quota_usage": -1}}
                )
            raise HTTPException(
                status_code=400,
                detail=f"Bu personelin {appointment.appointment_date} tarihinde {appointment.appointment_time} saatinde zaten bir randevusu var. Lütfen başka bir saat seçin."
            )
        
        # Mola çakışması kontrolü
        has_break_conflict = await check_break_conflict(
            db, appointment.staff_member_id, appointment.appointment_date,
            appointment.appointment_time, new_end_time, current_user.organization_id
        )
        
        if has_break_conflict:
            # Kota artırıldı ama mola çakışması var, geri al
            plan_doc = await db.organization_plans.find_one({"organization_id": current_user.organization_id})
            if plan_doc:
                await db.organization_plans.update_one(
                    {"organization_id": current_user.organization_id},
                    {"$inc": {"quota_usage": -1}}
                )
            raise HTTPException(
                status_code=400,
                detail=f"Bu personelin {appointment.appointment_date} tarihinde {appointment.appointment_time} saatinde mola var. Lütfen başka bir saat seçin."
            )
        
        assigned_staff_id = appointment.staff_member_id
    else:
        # Otomatik atama: Bu hizmeti verebilen personellerden boş olanı bul
        # Admin'in de hizmet verip vermediğini kontrol et
        settings_data = await db.settings.find_one({"organization_id": current_user.organization_id})
        admin_provides_service = settings_data.get('admin_provides_service', True) if settings_data else True
        no_assignable_staff = False
        
        # Bu hizmet(ler)i verebilen personelleri bul (çoklu hizmette KESİŞİM: tümünü verebilmeli)
        qualified_staff_query = {
            "organization_id": current_user.organization_id,
            "permitted_service_ids": {"$all": resolved_service_ids}
        }
        
        # Admin hizmet vermiyorsa, admin'i listeden çıkar
        if not admin_provides_service:
            qualified_staff_query["role"] = {"$ne": "admin"}
        
        qualified_staff = await db.users.find(
            qualified_staff_query,
            {"_id": 0, "username": 1, "role": 1}
        ).to_list(1000)
        
        # Eğer admin_provides_service açıksa ve başka personel yoksa, admin'i personel listesine ekle
        if not qualified_staff:
            # Admin'in bu hizmeti verebilip veremediğini kontrol et
            admin_user = await db.users.find_one(
                {"username": current_user.username, "organization_id": current_user.organization_id, "role": "admin"},
                {"_id": 0, "username": 1, "role": 1, "permitted_service_ids": 1}
            )
            if admin_provides_service and admin_user and all(sid in (admin_user.get('permitted_service_ids') or []) for sid in resolved_service_ids):
                qualified_staff = [{"username": admin_user['username'], "role": "admin"}]
                logging.info(f"⚠️ No staff found, but admin can provide service. Using admin: {admin_user['username']}")

        if not qualified_staff:
            # Admin panelinde personel hiç yoksa (ve admin de hizmet vermiyorsa),
            # randevuyu personelsiz (unassigned) kaydetmeye izin ver.
            if current_user.role == "admin":
                no_assignable_staff = True
                assigned_staff_id = None
                logging.info("ℹ️ No assignable staff found; creating appointment as unassigned (admin)")
            else:
                # Kota artırıldı ama personel bulunamadı, geri al
                plan_doc = await db.organization_plans.find_one({"organization_id": current_user.organization_id})
                if plan_doc:
                    await db.organization_plans.update_one(
                        {"organization_id": current_user.organization_id},
                        {"$inc": {"quota_usage": -1}}
                    )
                raise HTTPException(
                    status_code=400,
                    detail="Bu hizmet için uygun personel bulunamadı"
                )

        # Personelsiz kayıt (admin) durumunda, organizasyon bazlı çakışma kontrolü yapılır
        # (personel seçilmediği için herhangi bir randevu ile çakışmayı engelle)
        if no_assignable_staff:
            service_duration = service.get('duration', 30)

            new_start_hour, new_start_minute = map(int, appointment.appointment_time.split(':'))
            new_end_minute = new_start_minute + service_duration
            new_end_hour = new_start_hour + (new_end_minute // 60)
            new_end_minute = new_end_minute % 60
            new_end_time = f"{str(new_end_hour).zfill(2)}:{str(new_end_minute).zfill(2)}"

            org_appointments = await db.appointments.find(
                {
                    "organization_id": current_user.organization_id,
                    "appointment_date": appointment.appointment_date,
                    # Müsaitlik ile tutarlı: ödeme bekleyen/iptal slotları çakışma saymaz
                    "status": {"$nin": ["İptal", "İptal Edildi", "Cancelled", "Ödeme Bekleniyor"]}
                },
                {"_id": 0, "appointment_time": 1, "service_id": 1, "service_duration": 1}
            ).to_list(1000)

            def time_to_minutes(time_str):
                try:
                    hour, minute = map(int, time_str.split(':'))
                    return hour * 60 + minute
                except (ValueError, AttributeError):
                    return 0

            new_start_min = time_to_minutes(appointment.appointment_time)
            new_end_min = time_to_minutes(new_end_time)

            for existing_appt in org_appointments:
                existing_start_time = existing_appt.get('appointment_time')
                if not existing_start_time:
                    continue
                existing_service_id = existing_appt.get('service_id')

                # Çoklu hizmette saklı TOPLAM süreyi kullan; yoksa canlı tek-hizmet süresi
                stored_dur = existing_appt.get('service_duration')
                if stored_dur:
                    existing_duration = int(stored_dur)
                elif existing_service_id:
                    existing_service = await db.services.find_one({"id": existing_service_id}, {"_id": 0, "duration": 1})
                    existing_duration = existing_service.get('duration', 30) if existing_service else 30
                else:
                    existing_duration = 30

                existing_start_hour, existing_start_minute = map(int, existing_start_time.split(':'))
                existing_end_minute = existing_start_minute + existing_duration
                existing_end_hour = existing_start_hour + (existing_end_minute // 60)
                existing_end_minute = existing_end_minute % 60
                existing_end_time = f"{str(existing_end_hour).zfill(2)}:{str(existing_end_minute).zfill(2)}"

                existing_start_min = time_to_minutes(existing_start_time)
                existing_end_min = time_to_minutes(existing_end_time)

                if new_start_min < existing_end_min and new_end_min > existing_start_min:
                    plan_doc = await db.organization_plans.find_one({"organization_id": current_user.organization_id})
                    if plan_doc:
                        await db.organization_plans.update_one(
                            {"organization_id": current_user.organization_id},
                            {"$inc": {"quota_usage": -1}}
                        )
                    raise HTTPException(
                        status_code=400,
                        detail="Bu saat dilimi doludur. Lütfen başka bir saat seçin."
                    )
        else:
            # Otomatik atama: org-geneli "bu saat dolu" YOK — eligible personel içinde ilk müsait (_check_slot_conflict ile tek kaynak)
            service_duration = service.get('duration', 30)
            new_start_hour, new_start_minute = map(int, appointment.appointment_time.split(':'))
            new_end_minute = new_start_minute + service_duration
            new_end_hour = new_start_hour + (new_end_minute // 60)
            new_end_minute = new_end_minute % 60
            new_end_time = f"{str(new_end_hour).zfill(2)}:{str(new_end_minute).zfill(2)}"

            for staff in qualified_staff:
                sid = staff["username"]
                if await _check_slot_conflict(
                    db,
                    current_user.organization_id,
                    sid,
                    appointment.appointment_date,
                    appointment.appointment_time,
                    service_duration,
                ):
                    continue
                if await check_break_conflict(
                    db,
                    sid,
                    appointment.appointment_date,
                    appointment.appointment_time,
                    new_end_time,
                    current_user.organization_id,
                ):
                    continue
                assigned_staff_id = sid
                logging.info(f"✅ Auto-assigned to {sid} for {appointment.appointment_time}")
                break

            if not assigned_staff_id:
                # Kota artırıldı ama personel bulunamadı, geri al
                plan_doc = await db.organization_plans.find_one({"organization_id": current_user.organization_id})
                if plan_doc:
                    await db.organization_plans.update_one(
                        {"organization_id": current_user.organization_id},
                        {"$inc": {"quota_usage": -1}}
                    )
                raise HTTPException(
                    status_code=400,
                    detail="Bu saat dilimi doludur. Lütfen başka bir saat seçin."
                )
    
    appointment_data = appointment.model_dump(); 
    appointment_data['service_id'] = resolved_service_ids[0]  # ilk hizmet (geriye dönük uyum)
    appointment_data['services'] = snapshot_lines  # çoklu hizmet snapshot kalemleri
    appointment_data['service_name'] = service['name']; 
    appointment_data['service_price'] = service['price']
    appointment_data['staff_member_id'] = assigned_staff_id
    appointment_data['service_duration'] = service.get('duration', 30)  # Toplam hizmet süresi
    try:
        turkey_tz = ZoneInfo("Europe/Istanbul"); now = datetime.now(turkey_tz); dt_str = f"{appointment.appointment_date} {appointment.appointment_time}"
        naive_dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M"); appointment_dt = naive_dt.replace(tzinfo=turkey_tz)
        # Randevu bitiş saatini hesapla (başlangıç saati + hizmet süresi)
        service_duration_minutes = service.get('duration', 30)
        completion_threshold = appointment_dt + timedelta(minutes=service_duration_minutes)
        if now >= completion_threshold: appointment_data['status'] = 'Tamamlandı'; appointment_data['completed_at'] = datetime.now(timezone.utc).isoformat()
        else: appointment_data['status'] = 'Bekliyor'
    except (ValueError, TypeError) as e: logging.warning(f"Randevu durumu ayarlanırken tarih hatası: {e}"); appointment_data['status'] = 'Bekliyor'
    appointment_obj = Appointment(**appointment_data, organization_id=current_user.organization_id)
    doc = appointment_obj.model_dump(); doc['created_at'] = doc['created_at'].isoformat()
    await db.appointments.insert_one(doc)
    # ---------------------------------------------------------
    # 🧹 ADIM 2: SÜPÜRGE (Cache Invalidation)
    # ---------------------------------------------------------
    # Veritabanına kayıt başarılı! Şimdi Redis'teki eski verileri temizleme vakti.
    try:
        # Dashboard ve Müşteri listesini siliyoruz ki yeni veri anında görünsün
        await invalidate_cache(request, "dashboard_stats", current_user)
        await invalidate_cache(request, "customers_list", current_user)
        await invalidate_availability_cache(request, current_user.organization_id)
        
        # 🔑 İşimiz bitti, en başta koyduğumuz kilidi (Lock) artık kaldırabiliriz
        redis_client = getattr(request.app.state, 'redis_client', None)
        if redis_client:
            # Fonksiyonun başında tanımladığımız lock_key'i siliyoruz
            await redis_client.delete(lock_key)
            logging.info(f"🔑 Lock released and cache cleared for org: {current_user.organization_id}")
            
    except Exception as e:
        logging.error(f"❌ Süpürge işlemi sırasında hata oluştu: {e}")
    # ---------------------------------------------------------
    # Müşteriyi customers collection'ına ekle (eğer yoksa)
    # Format-dedup: aynı kişi farklı telefon formatıyla (05.. / +90.. / 90..) girilirse
    # yeni müşteri açılmasın, mevcut kayda bağlansın. resolved_customer_phone, sayaç
    # upsert'inin doğru (mevcut) dokümana yazması için aşağıda kullanılır.
    resolved_customer_phone = appointment.phone
    try:
        # Aynı telefonun tüm format varyantlarıyla eşleşen müşterileri bul
        _phone_variants = _phone_match_variants(appointment.phone)
        customers_with_phone = await db.customers.find(
            {
                "organization_id": current_user.organization_id,
                "phone": {"$in": _phone_variants}
            },
            {"_id": 0, "name": 1, "phone": 1}
        ).to_list(100)
        
        # İsim-soyisim kontrolü (büyük-küçük harf duyarsız)
        customer_name_normalized = appointment.customer_name.strip().lower()
        existing_customer = None
        for customer in customers_with_phone:
            if customer.get("name", "").strip().lower() == customer_name_normalized:
                existing_customer = customer
                break
        if existing_customer and existing_customer.get("phone"):
            # Mevcut kaydın saklı telefon formatını kanonik kabul et (sayaç oraya yazılsın)
            resolved_customer_phone = existing_customer.get("phone")
        
        if not existing_customer:
            # Müşteri yoksa ekle
            customer_doc = {
                "id": str(uuid.uuid4()),
                "organization_id": current_user.organization_id,
                "name": appointment.customer_name.strip(),
                "phone": appointment.phone,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "notes": ""
            }
            await db.customers.insert_one(customer_doc)
            logging.info(f"Customer auto-added: {appointment.customer_name} ({appointment.phone}) for org {current_user.organization_id}")
            
            # Cache invalidation SONRASI müşteri eklendi — tekrar temizle (race condition fix)
            try:
                await invalidate_cache(request, "customers_list", current_user)
            except Exception:
                pass
            
            # WebSocket event gönder (müşteriler listesini güncellemek için)
            try:
                await emit_to_organization(
                    current_user.organization_id,
                    'customer_added',
                    {'customer': customer_doc}
                )
            except Exception as emit_error:
                logging.warning(f"Failed to emit customer_added event: {emit_error}")
    except Exception as e:
        logging.warning(f"Error adding customer to collection: {e}")

    # Denormalize sayaç güncellemesi — mevcut customer doc'a +1 total (+1 completed
    # eğer randevu anında Tamamlandı geldiyse). Yeni customer'sa upsert oluşturur.
    try:
        await _apply_customer_delta(
            db,
            organization_id=current_user.organization_id,
            phone=resolved_customer_phone,
            name=appointment.customer_name,
            delta_total=1,
            delta_completed=1 if appointment_obj.status == "Tamamlandı" else 0,
            appointment_date=appointment.appointment_date,
            appointment_time=appointment.appointment_time,
        )
    except Exception as _delta_err:
        logger.warning(f"customer delta (create_appointment) skipped: {_delta_err}")

    if appointment_obj.status == 'Tamamlandı':
        transaction = Transaction(organization_id=current_user.organization_id, appointment_id=appointment_obj.id, customer_name=appointment_obj.customer_name, service_name=appointment_obj.service_name, amount=appointment_obj.service_price, date=appointment_obj.appointment_date)
        trans_doc = transaction.model_dump(); trans_doc['created_at'] = trans_doc['created_at'].isoformat()
        await db.transactions.insert_one(trans_doc)

    settings_data = await db.settings.find_one({"organization_id": current_user.organization_id})
    if not settings_data:
        default_settings = Settings(organization_id=current_user.organization_id); settings_data = default_settings.model_dump()
    company_name = settings_data.get("company_name", "İşletmeniz")
    support_phone = settings_data.get("support_phone", "Destek Hattı")
    business = settings_data or {}
    settings = business.get('settings') if isinstance(business.get('settings'), dict) else business
    location = (settings or {}).get('location') or {}
    coordinates = (location or {}).get('coordinates', {})
    lat = coordinates.get('lat')
    lng = coordinates.get('lng')
    
    # SMS + WhatsApp ONAY — SENKRON AĞ ÇAĞRILARI RESPONSE'U BLOKLAMASIN.
    # Randevu DB'ye yazıldı; onay mesajları BackgroundTasks ile response DÖNDÜKTEN
    # sonra gönderilir (sync sender'lar to_thread ile) → "Oluştur" ~1-1.5s yerine anında yanıtlar.
    sms_message = build_sms_message(
        company_name, appointment.customer_name,
        appointment.appointment_date, appointment.appointment_time,
        service['name'], support_phone, sms_type="confirmation"
    )
    wa_params = {
        "to_number": appointment.phone,
        "template_type": "CONFIRMATION",
        "customer_name": appointment.customer_name,
        "company_name": company_name,
        "appointment_date": appointment.appointment_date,
        "appointment_time": appointment.appointment_time,
        "service_name": service['name'],
        "support_phone": support_phone or "Destek Hattı",
        "business_lat": lat,
        "business_lng": lng,
        "business_address": location.get('address'),
    }
    background_tasks.add_task(
        _send_appointment_confirmations,
        db, appointment.phone, sms_message, wa_params,
        bool(getattr(appointment, "skip_confirmation_whatsapp", False)),
    )
    
    # Audit log
    await create_audit_log(
        db=db,
        organization_id=current_user.organization_id,
        user_id=current_user.username,
        user_full_name=current_user.full_name or current_user.username,
        action="CREATE",
        resource_type="APPOINTMENT",
        resource_id=appointment_obj.id,
        new_value=doc,
        ip_address=request.client.host if request.client else None
    )
    
    # Emit WebSocket event for real-time update
    # Use appointment_obj.model_dump() instead of doc to avoid MongoDB _id issues
    appointment_for_emit = appointment_obj.model_dump()
    appointment_for_emit['created_at'] = appointment_for_emit['created_at'].isoformat()
    logger.info(f"About to emit appointment_created for org: {current_user.organization_id}")
    try:
        await emit_to_organization(
            current_user.organization_id,
            'appointment_created',
            {'appointment': appointment_for_emit}
        )
        logger.info(f"Successfully emitted appointment_created for org: {current_user.organization_id}")
    except Exception as emit_error:
        logger.error(f"Failed to emit appointment_created: {emit_error}", exc_info=True)
    
    return appointment_obj

# === SESSION PACKAGE MODELS ===
class BulkCheckRequest(BaseModel):
    service_id: str
    staff_id: Optional[str] = None
    slots: list[str]
    staff_ids: Optional[list[Optional[str]]] = None


class WavePlanRequest(BaseModel):
    service_id: str
    first_session_date: str
    first_session_time: str
    remaining_count: int = Field(ge=1, le=50)
    staff_member_id: Optional[str] = None


class CancelSelectedRequest(BaseModel):
    appointment_ids: List[str]
    refund: bool = False

IDEMPOTENCY_BULK_SESSION_TTL_SEC = int(os.environ.get("IDEMPOTENCY_BULK_SESSION_TTL_SEC", str(72 * 3600)))
IDEMPOTENCY_BULK_SESSION_LOCK_SEC = int(os.environ.get("IDEMPOTENCY_BULK_SESSION_LOCK_SEC", "120"))

# === SESSION PACKAGE ENDPOINTS ===

def _slugify_service(name: str) -> str:
    """Hizmet adından analytics için stabil slug üretir (TR karakter desteği)."""
    s = (name or "").strip().lower()
    trmap = str.maketrans("çğıöşü", "cgiosu")
    s = s.translate(trmap)
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "service"


async def resolve_services(db, organization_id: str, appointment, *, multi_enabled: bool = True):
    """Çoklu hizmet çözümleme + doğrulama (tek kaynak).

    Returns: (snapshot_lines: List[dict], total_duration: int, total_price: float,
              merged_service: dict, service_ids: List[str])
    merged_service: mevcut tek-hizmet mantığını bozmadan kullanılacak birleşik hizmet
    (id=ilk hizmet, name=birleşik, price=Σ, duration=Σ).
    Geçersiz girdide HTTPException fırlatır.
    """
    sid_single = getattr(appointment, "service_id", None)
    sid_list = getattr(appointment, "service_ids", None)

    if sid_single and sid_list:
        # İç replay (public booking pending payload → /public/verify-code) hem
        # service_id hem service_ids gönderir; tekil hizmette bile service_ids=[id]
        # + service_id=id birlikte gelir. Bu ambiguity DEĞİL — tutarlıysa
        # (service_id, service_ids içinde) service_ids'i otoritatif kabul edip
        # service_id'yi yok say. Yalnızca GERÇEKTEN çelişen girdide (service_id
        # listede yok) 400 dön.
        if sid_single in sid_list:
            sid_single = None
        else:
            raise HTTPException(status_code=400, detail="service_id ve service_ids aynı anda gönderilemez")

    if sid_list is not None:
        if not multi_enabled:
            sid_list = sid_list[:1]
        if len(sid_list) == 0:
            raise HTTPException(status_code=400, detail="En az bir hizmet seçilmelidir")
        if len(set(sid_list)) != len(sid_list):
            raise HTTPException(status_code=400, detail="Aynı hizmet birden fazla seçilemez")
        ordered_ids = list(sid_list)
    elif sid_single:
        ordered_ids = [sid_single]
    else:
        raise HTTPException(status_code=400, detail="Hizmet seçilmedi")

    docs = await db.services.find(
        {"id": {"$in": ordered_ids}, "organization_id": organization_id}, {"_id": 0}
    ).to_list(len(ordered_ids))
    by_id = {d["id"]: d for d in docs}
    for sid in ordered_ids:
        d = by_id.get(sid)
        if not d:
            raise HTTPException(status_code=404, detail="Hizmet bulunamadı")
        if d.get("is_active") is False:
            raise HTTPException(status_code=400, detail="Pasif hizmet seçildi")

    pkg_count = sum(1 for sid in ordered_ids if (by_id[sid].get("session_count") or 1) > 1)
    if pkg_count and len(ordered_ids) > 1:
        raise HTTPException(status_code=400, detail="Seans paketi diğer hizmetlerle birlikte seçilemez")

    snapshot_lines = []
    total_duration = 0
    total_price = 0.0
    for idx, sid in enumerate(ordered_ids):
        d = by_id[sid]
        dur = int(d.get("duration", 30) or 30)
        price = float(d.get("price", 0) or 0)
        total_duration += dur
        total_price += price
        snapshot_lines.append({
            "service_id": sid,
            "sequence": idx,
            "name_snapshot": d.get("name", ""),
            "slug_snapshot": _slugify_service(d.get("name", "")),
            "duration_snapshot": dur,
            "price_snapshot": price,
            "service_version": int(d.get("version", 1) or 1),
        })

    merged_service = dict(by_id[ordered_ids[0]])
    merged_service["duration"] = total_duration
    merged_service["price"] = total_price
    merged_service["name"] = " + ".join(line["name_snapshot"] for line in snapshot_lines)
    # Çoklu hizmette toplam ödeme kuralı: herhangi biri online/deposit ise toplu kapora (Akıllı D).
    if len(ordered_ids) > 1:
        rules = [(by_id[sid].get("payment_rule") or "on_site") for sid in ordered_ids]
        if any(r in ("online", "deposit", "full_online") for r in rules):
            merged_service["payment_rule"] = "deposit"
        else:
            merged_service["payment_rule"] = "on_site"
    return snapshot_lines, total_duration, total_price, merged_service, ordered_ids


def _time_to_minutes(time_str: str) -> int:
    try:
        hour, minute = map(int, time_str.split(':'))
        return hour * 60 + minute
    except (ValueError, AttributeError):
        return 0

def _minutes_to_time(minutes: int) -> str:
    return f"{str(minutes // 60).zfill(2)}:{str(minutes % 60).zfill(2)}"

async def _check_slot_conflict(db, organization_id: str, staff_id: str, date: str, time: str, service_duration: int) -> bool:
    """Belirtilen slotun çakışma durumunu kontrol eder. True = çakışma var.
    staff_id boş/None ise: o gün tüm personellerin randevuları arasında zaman çakışması aranır (işletme takvimi).

    staff_id verildiğinde: o staff'a ait randevular VE staff atanmamış (boş/None)
    randevular birlikte kontrol edilir. Auto-assign veya silinmiş personele ait
    randevular "kim verirse versin slot dolu" prensibiyle çakışma sayılır;
    böylece personelsiz seans paketleri çift booking'e karşı korunur.
    """
    new_start_min = _time_to_minutes(time)
    new_end_min = new_start_min + service_duration
    query = {"organization_id": organization_id, "appointment_date": date, "status": {"$nin": ["İptal", "İptal Edildi", "Cancelled", "Ödeme Bekleniyor"]}}
    if staff_id:
        query["$or"] = [
            {"staff_member_id": staff_id},
            {"staff_member_id": None},
            {"staff_member_id": ""},
            {"staff_member_id": {"$exists": False}},
        ]
    existing = await db.appointments.find(
        query,
        {"_id": 0, "appointment_time": 1, "service_id": 1, "service_duration": 1}
    ).to_list(200)
    for appt in existing:
        ex_start = _time_to_minutes(appt.get('appointment_time', '00:00'))
        ex_dur = appt.get('service_duration') or 30
        ex_end = ex_start + ex_dur
        if new_start_min < ex_end and new_end_min > ex_start:
            return True
    return False

async def _find_alternative_slots(db, organization_id: str, staff_id: str, date: str, preferred_time: str, service_duration: int, max_alternatives: int = 3) -> list[str]:
    """Çakışan slot için aynı günden en yakın müsait saatleri bulur.

    `_check_slot_conflict` ile aynı semantiği uygular: seçili staff'a ait
    randevular VE staff atanmamış (boş/None) randevular birlikte busy sayılır.
    """
    preferred_min = _time_to_minutes(preferred_time)
    alt_query = {
        "organization_id": organization_id,
        "appointment_date": date,
        "status": {"$nin": ["İptal", "İptal Edildi", "Ödeme Bekleniyor"]},
        "$or": [
            {"staff_member_id": staff_id},
            {"staff_member_id": None},
            {"staff_member_id": ""},
            {"staff_member_id": {"$exists": False}},
        ],
    }
    existing = await db.appointments.find(
        alt_query,
        {"_id": 0, "appointment_time": 1, "service_duration": 1}
    ).to_list(200)
    busy_ranges = []
    for appt in existing:
        s = _time_to_minutes(appt.get('appointment_time', '00:00'))
        d = appt.get('service_duration') or 30
        busy_ranges.append((s, s + d))
    # Molalar da kontrol
    if staff_id:
        staff_doc = await db.users.find_one({"username": staff_id, "organization_id": organization_id}, {"_id": 0, "breaks": 1})
        if staff_doc:
            for brk in staff_doc.get('breaks', []):
                if brk.get('date') == date:
                    bs = _time_to_minutes(brk.get('start_time', '00:00'))
                    be = _time_to_minutes(brk.get('end_time', '00:00'))
                    busy_ranges.append((bs, be))
    def is_free(start_min):
        end_min = start_min + service_duration
        for bs, be in busy_ranges:
            if start_min < be and end_min > bs:
                return False
        return True
    alternatives = []
    # Tercih edilen saatten 30'ar dk uzaklaşarak ara
    for offset in range(1, 20):
        for delta in [offset * 30, -offset * 30]:
            candidate = preferred_min + delta
            if candidate < 480 or candidate > 1200:  # 08:00 - 20:00 arası
                continue
            if is_free(candidate):
                alternatives.append(_minutes_to_time(candidate))
                if len(alternatives) >= max_alternatives:
                    return alternatives
    return alternatives

async def _find_available_staff_for_slot(db, organization_id: str, service_id: str, date: str, time: str, service_duration: int, exclude_staff_id: str = None) -> str | None:
    """Belirtilen slotta bu hizmeti verebilen müsait bir personel bulur. Bulamazsa None döner."""
    # Bu hizmeti verebilen tüm personelleri bul
    staff_query = {"organization_id": organization_id, "role": {"$in": ["staff", "admin"]}, "permitted_service_ids": service_id}
    staff_members = await db.users.find(staff_query, {"_id": 0, "username": 1}).to_list(50)
    
    for staff in staff_members:
        sid = staff.get("username")
        if not sid or sid == exclude_staff_id:
            continue
        # Çakışma kontrolü
        has_conflict = await _check_slot_conflict(db, organization_id, sid, date, time, service_duration)
        if has_conflict:
            continue
        # Mola kontrolü
        end_time = _minutes_to_time(_time_to_minutes(time) + service_duration)
        has_break = await check_break_conflict(db, sid, date, time, end_time, organization_id)
        if has_break:
            continue
        return sid
    return None


async def _assert_staff_can_provide_service(
    db,
    organization_id: str,
    service_id: str,
    staff_username: str,
    org_settings: Optional[dict] = None,
) -> str:
    """Açık personel ataması için: kullanıcı var mı, hizmete yetkili mi.

    Not: Admin için `admin_provides_service=False` ayarı, admin'in
    `permitted_service_ids` listesinde bu hizmete AÇIKÇA izin verilmişse
    göz ardı edilir. Bu, `/availability`, `/sessions/wave-plan` ve
    `/public/appointments` fallback mantıklarıyla tutarlı olması için
    zorunludur — aksi halde aynı org'ta bir endpoint admin'e fallback
    yaparken başka endpoint "Yönetici bu hizmeti veremez" hatası
    dönerek tek admin'li org'ların paket planlamasını bozar.
    """
    staff_username = (staff_username or "").strip()
    if not staff_username:
        raise HTTPException(status_code=400, detail="Geçersiz personel")
    user = await db.users.find_one(
        {"username": staff_username, "organization_id": organization_id},
        {"_id": 0, "role": 1, "permitted_service_ids": 1},
    )
    if not user:
        raise HTTPException(status_code=400, detail=f"Personel bulunamadı: {staff_username}")
    permitted = user.get("permitted_service_ids") or []
    if service_id not in permitted:
        raise HTTPException(status_code=400, detail="Seçilen personel bu hizmeti veremez")
    # Admin permitted_service_ids'inde hizmete açıkça izin verdiğinde,
    # admin_provides_service ayarı permissive olarak override edilir.
    return staff_username


@api_router.post("/appointments/bulk-session")
async def create_bulk_session(request: Request, bulk: BulkSessionCreate, current_user: UserInDB = Depends(get_current_user)):
    """Seans paketi oluştur — atomik toplu randevu oluşturma"""
    db = await get_db_from_request(request)
    redis_client = getattr(request.app.state, 'redis_client', None)
    request_id = str(uuid.uuid4())
    assert_capability_for_user(current_user, CAP_SESSION_PLAN, request_id=request_id)
    if not bulk.sessions:
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error_code": "validation_failed",
                "message": "En az bir seans slotu gereklidir",
                "request_id": request_id,
            },
        )
    
    # Hizmet doğrulama
    service = await db.services.find_one({"id": bulk.service_id, "organization_id": current_user.organization_id}, {"_id": 0})
    if not service:
        raise HTTPException(status_code=404, detail="Hizmet bulunamadı")
    service_duration = service.get('duration', 30)
    
    # Personel doğrulama
    assigned_staff_id = bulk.staff_member_id
    if current_user.role == "staff" and not current_user.can_view_all_appointments:
        if service["id"] not in (current_user.permitted_service_ids or []):
            raise HTTPException(status_code=403, detail="Bu hizmete randevu alma yetkiniz yok")
        assigned_staff_id = current_user.username

    payload_hash = _canonical_bulk_session_hash(bulk)
    idem_header = (request.headers.get("Idempotency-Key") or request.headers.get("idempotency-key") or "").strip()
    idem_main_key = None
    idem_proc_key = None
    if idem_header and redis_client:
        safe_idem = hashlib.sha256(f"{current_user.organization_id}:{idem_header}".encode()).hexdigest()[:48]
        idem_main_key = f"plann:idempotency:bulk_session:{current_user.organization_id}:{safe_idem}"
        idem_proc_key = idem_main_key + ":processing"
        raw = await redis_client.get(idem_main_key)
        if raw:
            try:
                stored = json.loads(raw)
                if stored.get("payload_hash") == payload_hash:
                    return JSONResponse(content=stored["response"], status_code=stored.get("http_status", 201))
                raise HTTPException(
                    status_code=422,
                    detail={
                        "success": False,
                        "error_code": "idempotency_key_mismatch",
                        "message": "Idempotency-Key already used with a different request body",
                        "request_id": request_id,
                    },
                )
            except HTTPException:
                raise
            except Exception:
                pass
        got = await redis_client.set(idem_proc_key, payload_hash, nx=True, ex=IDEMPOTENCY_BULK_SESSION_LOCK_SEC)
        if not got:
            raw = await redis_client.get(idem_main_key)
            if raw:
                try:
                    stored = json.loads(raw)
                    if stored.get("payload_hash") == payload_hash:
                        return JSONResponse(content=stored["response"], status_code=stored.get("http_status", 201))
                except Exception:
                    pass
            raise HTTPException(
                status_code=409,
                detail={
                    "success": False,
                    "error_code": "idempotency_in_progress",
                    "message": "A request with this Idempotency-Key is in progress",
                    "request_id": request_id,
                },
            )
    elif idem_header and not redis_client:
        logging.warning("Idempotency-Key set but Redis unavailable; replay disabled for this request")
    
    # Kota kontrolü (toplam seans sayısı kadar)
    total_sessions = len(bulk.sessions)
    for i in range(total_sessions):
        quota_ok, quota_error = await check_quota_and_increment(db, current_user.organization_id)
        if not quota_ok:
            # Artırılan kotaları geri al
            if i > 0:
                await db.organization_plans.update_one(
                    {"organization_id": current_user.organization_id},
                    {"$inc": {"quota_usage": -i}}
                )
            raise HTTPException(
                status_code=403,
                detail={
                    "success": False,
                    "error_code": "validation_failed",
                    "message": str(quota_error),
                    "request_id": request_id,
                },
            )
    
    # Yeniden planlama: aynı session_group_id ile geliyorsa eski seansları iptal et
    if bulk.session_group_id:
        cancel_result = await db.appointments.update_many(
            {
                "session_group_id": bulk.session_group_id,
                "organization_id": current_user.organization_id,
                "session_number": {"$gte": bulk.starting_session_number},
                "status": {"$nin": ["İptal Edildi", "Tamamlandı"]},
            },
            {"$set": {"status": "İptal Edildi"}}
        )
        if cancel_result.modified_count > 0:
            logging.info(f"♻️ Yeniden planlama: {cancel_result.modified_count} eski seans iptal edildi (group={bulk.session_group_id})")
    
    # Her slot için çakışma kontrolü (satır bazlı personel: SessionSlot.staff_member_id)
    org_settings = await db.settings.find_one(
        {"organization_id": current_user.organization_id},
        {"_id": 0, "admin_provides_service": 1},
    )
    lock_keys = []
    staff_assignments = {}
    # Aynı bulk isteğinde iki satırın çakışmasını DB'ye yazmadan yakala
    bulk_reserved_ranges = []  # list of (date_str, start_min, end_min)
    try:
        for idx, session in enumerate(bulk.sessions):
            slot_staff = assigned_staff_id
            if current_user.role == "staff" and not current_user.can_view_all_appointments:
                slot_staff = current_user.username
            else:
                per_slot = getattr(session, "staff_member_id", None)
                if per_slot and str(per_slot).strip():
                    slot_staff = await _assert_staff_can_provide_service(
                        db,
                        current_user.organization_id,
                        bulk.service_id,
                        str(per_slot).strip(),
                        org_settings,
                    )

            new_start_min = _time_to_minutes(session.time)
            new_end_min = new_start_min + service_duration
            for bd, bs, be in bulk_reserved_ranges:
                if bd == session.date and new_start_min < be and new_end_min > bs:
                    raise HTTPException(
                        status_code=409,
                        detail={
                            "success": False,
                            "error_code": "slot_conflict",
                            "message": f"Seans {idx+1}: Aynı istekte bu tarih/saat çakışıyor.",
                            "request_id": request_id,
                            "failed_indexes": [idx],
                        },
                    )

            if redis_client and slot_staff:
                lk = f"plann:lock:appt:{current_user.organization_id}:{slot_staff}:{session.date}:{session.time}"
                acquired = await redis_client.setnx(lk, "locked")
                if not acquired:
                    raise HTTPException(
                        status_code=409,
                        detail={
                            "success": False,
                            "error_code": "slot_conflict",
                            "message": f"Seans {idx+1} ({session.date} {session.time}) için başka bir işlem yapılıyor.",
                            "request_id": request_id,
                            "failed_indexes": [idx],
                        },
                    )
                await redis_client.expire(lk, 10)
                lock_keys.append(lk)

            # Tek doğruluk kaynağı: istemcinin gönderdiği personel + slot — org-geneli veya otomatik personel ataması yok.
            has_conflict = False
            has_break = False
            if slot_staff:
                has_conflict = await _check_slot_conflict(
                    db, current_user.organization_id, slot_staff, session.date, session.time, service_duration
                )
                if not has_conflict:
                    end_time = _minutes_to_time(_time_to_minutes(session.time) + service_duration)
                    has_break = await check_break_conflict(
                        db, slot_staff, session.date, session.time, end_time, current_user.organization_id
                    )
            else:
                logging.warning(
                    "bulk-session row %s: no staff on slot; skipping conflict/break check (legacy)",
                    idx + 1,
                )

            if has_conflict or has_break:
                raise HTTPException(
                    status_code=409,
                    detail={
                        "success": False,
                        "error_code": "slot_conflict",
                        "message": f"Seans {idx+1}: {session.date} {session.time} müsait değil (personel: {slot_staff or '-'}).",
                        "request_id": request_id,
                        "failed_indexes": [idx],
                        "conflicts": [{"row_index": idx, "reason": "slot_or_break_conflict"}],
                    },
                )
            staff_assignments[idx] = slot_staff
            bulk_reserved_ranges.append((session.date, new_start_min, new_end_min))
        
        # Tümü geçti — randevuları oluştur
        group_id = bulk.session_group_id or str(uuid.uuid4())
        session_total = bulk.session_total or total_sessions
        created_appointments = []
        
        for idx, session in enumerate(bulk.sessions):
            session_number = bulk.starting_session_number + idx
            is_first_session = (session_number == 1)
            
            # Tarih/saat ile status hesapla
            turkey_tz = ZoneInfo("Europe/Istanbul")
            now = datetime.now(turkey_tz)
            try:
                naive_dt = datetime.strptime(f"{session.date} {session.time}", "%Y-%m-%d %H:%M")
                appointment_dt = naive_dt.replace(tzinfo=turkey_tz)
                completion_threshold = appointment_dt + timedelta(minutes=service_duration)
                apt_status = 'Tamamlandı' if now >= completion_threshold else 'Bekliyor'
            except (ValueError, TypeError):
                apt_status = 'Bekliyor'
            
            apt_doc = {
                "id": str(uuid.uuid4()),
                "organization_id": current_user.organization_id,
                "customer_name": bulk.customer_name,
                "phone": bulk.phone,
                "service_id": bulk.service_id,
                "service_name": service['name'],
                "service_price": service['price'] if is_first_session else 0,
                "appointment_date": session.date,
                "appointment_time": session.time,
                "notes": bulk.notes,
                "status": apt_status,
                "staff_member_id": staff_assignments.get(idx),
                "service_duration": service_duration,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "session_group_id": group_id,
                "session_number": session_number,
                "session_total": session_total,
                "payment_status": "paid" if is_first_session else "package_included",
                "source": "admin_bulk"
            }
            if apt_status == 'Tamamlandı':
                apt_doc['completed_at'] = datetime.now(timezone.utc).isoformat()
            created_appointments.append(apt_doc)
        
        # Atomik insert — mümkünse transaction (tek commit); aksi halde insert_many
        if created_appointments:
            tx_ok = False
            motor_client = getattr(db, "client", None)
            if motor_client is not None:
                try:
                    async with await motor_client.start_session() as mongo_sess:
                        async with mongo_sess.start_transaction():
                            await db.appointments.insert_many(created_appointments, session=mongo_sess)
                    tx_ok = True
                except Exception as tx_err:
                    logging.warning(
                        "bulk-session: transaction insert not applied (%s); retrying without session",
                        tx_err,
                    )
            if not tx_ok:
                await db.appointments.insert_many(created_appointments)
        
        # Müşteri kaydı (upsert)
        try:
            existing_customer = await db.customers.find_one({"phone": bulk.phone, "organization_id": current_user.organization_id})
            if not existing_customer:
                await db.customers.insert_one({
                    "id": str(uuid.uuid4()), "organization_id": current_user.organization_id,
                    "name": bulk.customer_name, "phone": bulk.phone,
                    "created_at": datetime.now(timezone.utc).isoformat(), "notes": ""
                })
        except Exception as e:
            logging.warning(f"Customer upsert error in bulk-session: {e}")

        # Denormalize sayaç: her bir seans için +1 total ve auto-Tamamlandı gelenler
        # için +1 completed. Tek müşteri (bulk.phone) — birleştirilmiş delta uygula.
        try:
            _bulk_total = len(created_appointments)
            _bulk_completed = sum(1 for a in created_appointments if a.get("status") == "Tamamlandı")
            if _bulk_total:
                _first = created_appointments[0]
                await _apply_customer_delta(
                    db,
                    organization_id=current_user.organization_id,
                    phone=bulk.phone,
                    name=bulk.customer_name,
                    delta_total=_bulk_total,
                    delta_completed=_bulk_completed,
                    appointment_date=_first.get("appointment_date"),
                    appointment_time=_first.get("appointment_time"),
                )
        except Exception as _delta_err:
            logging.warning(f"customer delta (bulk-session) skipped: {_delta_err}")
        
        # Cache invalidation
        try:
            await invalidate_cache(request, "dashboard_stats", current_user)
            await invalidate_cache(request, "customers_list", current_user)
        except Exception as e:
            logging.warning(f"Cache invalidation error: {e}")
        
        # WebSocket emit
        try:
            await emit_to_organization(current_user.organization_id, 'appointments_bulk_created', {
                'session_group_id': group_id, 'count': len(created_appointments)
            })
        except Exception as e:
            logging.warning(f"WebSocket emit error: {e}")
        
        # WhatsApp SESSION_PACKAGE — template ile seans bilgisi
        async def _send_session_wa():
            try:
                from whatsapp_service import format_date_for_display

                wa_settings = await db.settings.find_one({"organization_id": current_user.organization_id})
                wa_company = (wa_settings or {}).get("company_name", "İşletme")
                wa_support = (wa_settings or {}).get("support_phone", "") or (wa_settings or {}).get("phone", "")
                wa_loc = (wa_settings or {}).get("location") or {}
                wa_coords = wa_loc.get("coordinates", {})
                wa_lat = wa_coords.get("lat")
                wa_lng = wa_coords.get("lng")
                wa_address = wa_loc.get("address")

                first_apt = created_appointments[0]
                new_appt_ids = [a["id"] for a in created_appointments]
                # Admin "yeni paket" akışı (AppointmentFormWizard): 1. seans
                # bu bulk'tan hemen önce POST /appointments ile yaratılmıştır,
                # bu yüzden müşteriye gönderilen tek WhatsApp mesajında 1. seans
                # da yer almalı. Bu durumda gruptaki tüm seansları çekiyoruz.
                # Kalanları sonradan planlayan akışlarda (SessionPlanner) flag
                # False kalır ve yalnızca yeni planlananlar listelenir; çünkü
                # müşteri 1. seansı önceki mesajdan zaten biliyor.
                base_query = {
                    "session_group_id": group_id,
                    "organization_id": current_user.organization_id,
                    "status": {"$nin": ["İptal Edildi"]},
                }
                if not getattr(bulk, "notify_include_existing_sessions", False):
                    base_query["id"] = {"$in": new_appt_ids}
                all_sessions = await db.appointments.find(base_query).sort("session_number", 1).to_list(100)

                # Template'e gönderilecek "kapak" tarih/saati: mesajda listelenen
                # en erken seans (session_number bakımından). Admin yeni paket
                # akışında gerçek 1. seansı kullanırız; kalanları planla akışında
                # listedeki ilk yeni seansı kullanırız.
                header_apt = all_sessions[0] if all_sessions else first_apt

                session_total = first_apt.get("session_total", len(created_appointments))

                number_emojis = ["1️⃣","2️⃣","3️⃣","4️⃣","5️⃣","6️⃣","7️⃣","8️⃣","9️⃣","🔟"]
                tr_days = ["Pazartesi","Salı","Çarşamba","Perşembe","Cuma","Cumartesi","Pazar"]
                tr_months = ["","Ocak","Şubat","Mart","Nisan","Mayıs","Haziran","Temmuz","Ağustos","Eylül","Ekim","Kasım","Aralık"]
                parts = []
                for s in all_sessions:
                    idx = (s.get("session_number", 1) - 1)
                    emoji = number_emojis[idx] if idx < len(number_emojis) else f"{idx+1}."
                    try:
                        from datetime import datetime as _dt
                        dt = _dt.strptime(s.get("appointment_date", ""), "%Y-%m-%d")
                        day_name = tr_days[dt.weekday()]
                        month_name = tr_months[dt.month]
                        appt_time = s.get("appointment_time", "")
                        parts.append(f"{emoji} {dt.day} {month_name} {day_name} {appt_time}")
                    except Exception:
                        d = format_date_for_display(s.get("appointment_date", ""))
                        appt_time = s.get("appointment_time", "")
                        parts.append(f"{emoji} {d} {appt_time}")
                session_dates_text = " | ".join(parts)

                try:
                    await asyncio.to_thread(
                        send_whatsapp_template,
                        first_apt["phone"], "SESSION_PACKAGE",
                        first_apt["customer_name"], wa_company,
                        header_apt.get("appointment_date", first_apt["appointment_date"]),
                        header_apt.get("appointment_time", first_apt["appointment_time"]),
                        first_apt["service_name"], wa_support or "Destek Hattı",
                        business_lat=wa_lat, business_lng=wa_lng, business_address=wa_address,
                        session_count=session_total, session_dates_text=session_dates_text,
                    )
                    logging.info(f"✓ SESSION_PACKAGE WhatsApp sent to {first_apt['phone']} ({session_total} sessions)")
                except Exception as wa_err:
                    logging.warning(f"WhatsApp session notification failed: {wa_err}")
            except Exception as wa_outer_err:
                logging.warning(f"WhatsApp session notifications error: {wa_outer_err}")
        asyncio.create_task(_send_session_wa())
        
        logging.info(f"✅ Bulk session created: {len(created_appointments)} appointments, group={group_id}")
        await create_audit_log(
            db=db,
            organization_id=current_user.organization_id,
            user_id=current_user.username,
            user_full_name=current_user.full_name or current_user.username,
            action="CREATE",
            resource_type="APPOINTMENT_BULK",
            resource_id=group_id,
            new_value={
                "kind": "bulk_session",
                "session_group_id": group_id,
                "created_appointment_ids": [a["id"] for a in created_appointments],
                "session_count": len(created_appointments),
                "idempotency_key": idem_header or None,
                "request_id": request_id,
            },
            ip_address=request.client.host if request.client else None,
        )
        await log_operation_audit(
            db,
            request_id=request_id,
            organization_id=current_user.organization_id,
            actor_user_id=current_user.username,
            action="bulk_session",
            target_kind="session_group",
            target_ids=[group_id],
            result="success",
            meta={
                "created_appointment_ids": [a["id"] for a in created_appointments],
                "idempotency_key": idem_header,
            },
        )
        clean_appointments = [{k: v for k, v in apt.items() if k != '_id'} for apt in created_appointments]
        created_ids = [a["id"] for a in created_appointments]
        response_body = {
            "success": True,
            "message": f"{len(created_appointments)} seans başarıyla oluşturuldu",
            "created_appointment_ids": created_ids,
            "session_group_id": group_id,
            "idempotency_key": idem_header or None,
            "request_id": request_id,
            "appointments": clean_appointments,
        }
        if idem_main_key and redis_client:
            try:
                await redis_client.set(
                    idem_main_key,
                    json.dumps({
                        "payload_hash": payload_hash,
                        "http_status": 201,
                        "response": response_body,
                    }, ensure_ascii=False),
                    ex=IDEMPOTENCY_BULK_SESSION_TTL_SEC,
                )
            except Exception as ex:
                logging.warning(f"Idempotency persist failed: {ex}")
        return JSONResponse(content=response_body, status_code=201)
    
    except HTTPException:
        # Kota geri al
        await db.organization_plans.update_one(
            {"organization_id": current_user.organization_id},
            {"$inc": {"quota_usage": -total_sessions}}
        )
        raise
    finally:
        # Lock'ları temizle
        if redis_client and lock_keys:
            for lk in lock_keys:
                try:
                    await redis_client.delete(lk)
                except Exception:
                    pass
        if redis_client and idem_proc_key:
            try:
                await redis_client.delete(idem_proc_key)
            except Exception:
                pass

@api_router.post("/appointments/bulk-package")
async def create_bulk_package(
    request: Request,
    package: BulkPackageCreate,
    current_user: UserInDB = Depends(get_current_user),
):
    """
    Atomik seans paketi yaratımı: 1..N seansı tek MongoDB transaction'da
    yaratır. Hata olursa hiçbir doc DB'ye yazılmaz. Tek SESSION_PACKAGE
    WhatsApp mesajı gönderilir (1..N seansları içerir).

    Plan 12.1 (backward compatibility):
        - /bulk-session geriye uyumlu olarak korunur (kalan planla akışı).
        - /bulk-package yeni "atomic new package" akışı için paralel eklendi.
    Plan 12.2 (atomic transaction safety):
        - Tüm DB yazımları transaction içinde, WhatsApp tetiği commit
          SONRASINDA asyncio.create_task ile (transaction bloğu dışında).
    Plan 12.3 (idempotency):
        - session_group_id tabanlı no-op success: aynı group_id ile yapılan
          retry/double-submit yeni doc yaratmaz, mevcut paketi döndürür.
        - Transaction başarısız olursa kota geri alınır + güvenli abort
          (HTTPException), sistem CRASH ETMEZ.

    Implementation: aynı core logic /bulk-session ile paylaşıldığı için
    bu endpoint wrapper'dır → BulkPackageCreate'i BulkSessionCreate'e
    çevirip create_bulk_session()'ı çağırır.
    """
    db = await get_db_from_request(request)
    request_id = str(uuid.uuid4())

    if not package.sessions:
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error_code": "validation_failed",
                "message": "En az bir seans slotu gereklidir",
                "request_id": request_id,
            },
        )

    # Idempotency check (Plan 12.3): aynı session_group_id ile var olan paketi
    # no-op success olarak döndür (retry / double-submit güvenliği).
    if package.session_group_id:
        existing = await db.appointments.find_one(
            {
                "organization_id": current_user.organization_id,
                "session_group_id": package.session_group_id,
            },
            {"_id": 0},
        )
        if existing:
            existing_apts = (
                await db.appointments.find(
                    {
                        "organization_id": current_user.organization_id,
                        "session_group_id": package.session_group_id,
                    },
                    {"_id": 0},
                )
                .sort("session_number", 1)
                .to_list(200)
            )
            logging.info(
                "♻️ /bulk-package idempotent replay: session_group_id=%s (%d existing appts)",
                package.session_group_id,
                len(existing_apts),
            )
            return JSONResponse(
                content={
                    "success": True,
                    "message": "Paket zaten oluşturulmuş (idempotent replay)",
                    "created_appointment_ids": [a["id"] for a in existing_apts],
                    "session_group_id": package.session_group_id,
                    "appointments": existing_apts,
                    "idempotent_replay": True,
                    "request_id": request_id,
                },
                status_code=200,
            )

    # BulkSessionCreate'e dönüştür (Plan 12.1 wrapper):
    # /bulk-session'ın atomic transaction + conflict check + quota + lock +
    # WhatsApp + audit log + cache invalidation logic'i olduğu gibi kullanılır.
    session_total = package.session_total or len(package.sessions)
    bulk = BulkSessionCreate(
        customer_name=package.customer_name,
        phone=package.phone,
        service_id=package.service_id,
        staff_member_id=package.staff_member_id,
        notes=package.notes or "",
        session_group_id=package.session_group_id or str(uuid.uuid4()),
        starting_session_number=1,  # Yeni paket her zaman 1'den başlar.
        session_total=session_total,
        sessions=package.sessions,
        # SESSION_PACKAGE WhatsApp mesajında 1..N seansları listele.
        # skip_confirmation_whatsapp True ise WhatsApp gönderimini handler
        # tetikler ama biz burada bayrağı bulk yapısına yansıtamayız (eski
        # şema). Bunun yerine handler'ın WhatsApp tetikleyici task'ı zaten
        # asyncio.create_task ile fire-and-forget, transaction blokunu
        # etkilemiyor. skip_confirmation_whatsapp gelişimi sonradan eklenebilir.
        notify_include_existing_sessions=True,
    )

    try:
        return await create_bulk_session(request, bulk, current_user)
    except HTTPException:
        # Plan 12.3: transaction başlatılamazsa güvenli abort (alt katman
        # zaten kota'yı geri alıyor). HTTPException'ı olduğu gibi yukarı sürer.
        raise
    except Exception as exc:
        # Beklenmeyen hata: sistem crash olmasın → log + HTTPException(500).
        logging.error(
            "/bulk-package wrapper unexpected error (session_group_id=%s): %s",
            package.session_group_id,
            exc,
            exc_info=True,
        )
        raise HTTPException(
            status_code=500,
            detail={
                "success": False,
                "error_code": "bulk_package_unexpected_error",
                "message": "Paket oluşturulurken beklenmedik bir hata oluştu.",
                "request_id": request_id,
            },
        )


@api_router.post("/availability/bulk-check")
async def bulk_check_availability(request: Request, check: BulkCheckRequest, current_user: UserInDB = Depends(get_current_user)):
    """Birden fazla tarih/saat slotunun müsaitliğini kontrol eder ve alternatif önerir."""
    db = await get_db_from_request(request)
    logging.info(
        f"🔍 bulk-check payload: service_id={check.service_id}, staff_id={check.staff_id}, "
        f"slots={check.slots}, staff_ids={check.staff_ids}"
    )
    service = await db.services.find_one({"id": check.service_id, "organization_id": current_user.organization_id}, {"_id": 0})
    if not service:
        logging.warning(f"❌ bulk-check: service not found {check.service_id}")
        raise HTTPException(status_code=404, detail="Hizmet bulunamadı")
    service_duration = service.get('duration', 30)
    if check.staff_ids is not None and len(check.staff_ids) != len(check.slots):
        logging.warning(
            f"❌ bulk-check: staff_ids/slots length mismatch slots={len(check.slots)} staff_ids={len(check.staff_ids)}"
        )
        raise HTTPException(status_code=400, detail="staff_ids uzunluğu slots ile aynı olmalıdır")
    org_settings = await db.settings.find_one(
        {"organization_id": current_user.organization_id},
        {"_id": 0, "admin_provides_service": 1},
    )
    results = []
    for i, slot_str in enumerate(check.slots):
        staff_id = check.staff_id
        if current_user.role == "staff" and not current_user.can_view_all_appointments:
            staff_id = current_user.username
        elif check.staff_ids is not None and i < len(check.staff_ids):
            ov = check.staff_ids[i]
            if ov is not None and str(ov).strip():
                try:
                    staff_id = await _assert_staff_can_provide_service(
                        db,
                        current_user.organization_id,
                        check.service_id,
                        str(ov).strip(),
                        org_settings,
                    )
                except HTTPException as exc:
                    logging.warning(
                        f"❌ bulk-check assert failed for staff='{ov}' service={check.service_id}: {exc.detail}"
                    )
                    raise
        try:
            parts = slot_str.split('T')
            date = parts[0]
            time = parts[1] if len(parts) > 1 else "00:00"
        except (IndexError, ValueError):
            results.append({"slot": slot_str, "available": False, "alternatives": [], "error": "Geçersiz format"})
            continue
        conflict = await _check_slot_conflict(db, current_user.organization_id, staff_id, date, time, service_duration)
        break_conflict = False
        if staff_id:
            end_time = _minutes_to_time(_time_to_minutes(time) + service_duration)
            break_conflict = await check_break_conflict(db, staff_id, date, time, end_time, current_user.organization_id)
        available = not conflict and not break_conflict
        alternatives = []
        suggested_staff = None
        suggested_staff_name = None
        if not available:
            # Aynı saatte başka müsait personel var mı?
            alt_staff = await _find_available_staff_for_slot(db, current_user.organization_id, check.service_id, date, time, service_duration, exclude_staff_id=staff_id)
            if alt_staff:
                suggested_staff = alt_staff
                staff_doc = await db.users.find_one({"username": alt_staff, "organization_id": current_user.organization_id}, {"_id": 0, "full_name": 1})
                suggested_staff_name = staff_doc.get("full_name", alt_staff) if staff_doc else alt_staff
                available = True  # Başka personel müsait → slot kullanılabilir
            elif staff_id:
                alternatives = await _find_alternative_slots(db, current_user.organization_id, staff_id, date, time, service_duration)
        result_item = {"date": date, "time": time, "available": available, "alternatives": alternatives}
        if suggested_staff:
            result_item["suggested_staff"] = suggested_staff
            result_item["suggested_staff_name"] = suggested_staff_name
        results.append(result_item)
    return results


@api_router.post("/sessions/wave-plan")
async def wave_plan_package_sessions(
    request: Request,
    body: WavePlanRequest,
    current_user: UserInDB = Depends(get_current_user),
):
    """Kalan paket seansları — haftalık hedef + dalga saat sırası + personel fallback (PLANN sözleşmesi)."""
    db = await get_db_from_request(request)
    assert_capability_for_user(current_user, CAP_SESSION_PLAN, request_id=str(uuid.uuid4()))
    settings = await db.settings.find_one({"organization_id": current_user.organization_id}, {"_id": 0})
    if not settings:
        raise HTTPException(status_code=400, detail="Ayarlar bulunamadı")

    admin_provides = settings.get("admin_provides_service", True)
    staff_query: dict = {
        "organization_id": current_user.organization_id,
        "permitted_service_ids": {"$in": [body.service_id]},
    }
    if not admin_provides:
        staff_query["role"] = {"$ne": "admin"}

    staff_members = await db.users.find(
        staff_query,
        {"_id": 0, "username": 1, "full_name": 1, "days_off": 1, "role": 1, "permitted_service_ids": 1, "breaks": 1},
    ).to_list(200)

    # Fallback: hizmet için yetkili staff yoksa AMA admin bu hizmeti verebiliyorsa
    # admin'i kullan. Bu, "/availability" ve "/public/appointments" akışlarındaki
    # davranışla tutarlıdır — tek admin'li org'larda public'ten gelen seans
    # paketlerinin "Kalan seansları planla" akışında "uygun slot bulunamadı"
    # dönmemesi için kritiktir. admin_provides_service=False olsa bile, randevu
    # zaten admin'e atanmış (public booking aynı fallback'i kullanıyor) ya da
    # staff_member_id=None olarak kayıtlı; her iki durumda da planlama
    # admin üzerinden yürütülmelidir.
    if not staff_members:
        admin_user = await db.users.find_one(
            {"organization_id": current_user.organization_id, "role": "admin"},
            {"_id": 0, "username": 1, "full_name": 1, "days_off": 1, "role": 1, "permitted_service_ids": 1, "breaks": 1},
        )
        if admin_user and body.service_id in (admin_user.get("permitted_service_ids") or []):
            staff_members = [admin_user]
            logging.info(
                f"⚠️ Wave-plan: No staff found for service {body.service_id}, falling back to admin {admin_user['username']}"
            )

    sessions, unplanned = await plan_wave_remaining_sessions(
        db,
        current_user.organization_id,
        body.service_id,
        body.first_session_date,
        body.first_session_time,
        body.remaining_count,
        settings,
        staff_members,
        preferred_staff_username=(body.staff_member_id or "").strip() or None,
    )

    return {
        "sessions": sessions,
        "unplanned_count": unplanned,
        "message": "OK" if unplanned == 0 else "partial",
    }


@api_router.post("/appointments/session-group/{group_id}/cancel-remaining")
async def cancel_remaining_sessions(request: Request, group_id: str, from_session_number: int = Query(1, ge=1), current_user: UserInDB = Depends(get_current_user)):
    """Seans grubundaki seçilen seans numarasından itibaren bekleyen randevuları iptal eder."""
    db = await get_db_from_request(request)
    turkey_tz = ZoneInfo("Europe/Istanbul")
    today_str = datetime.now(turkey_tz).strftime("%Y-%m-%d")
    query = {
        "session_group_id": group_id,
        "organization_id": current_user.organization_id,
        "session_number": {"$gte": from_session_number},
        "status": {"$nin": ["İptal", "İptal Edildi", "Tamamlandı"]}
    }
    if current_user.role == "staff" and not current_user.can_view_all_appointments:
        query["staff_member_id"] = current_user.username
    future_sessions = await db.appointments.find(query, {"_id": 0, "id": 1}).to_list(100)
    if not future_sessions:
        raise HTTPException(status_code=404, detail="İptal edilecek bekleyen seans bulunamadı")
    cancel_count = len(future_sessions)
    ids_to_cancel = [s['id'] for s in future_sessions]
    await db.appointments.update_many(
        {"id": {"$in": ids_to_cancel}, "organization_id": current_user.organization_id},
        {"$set": {"status": "İptal Edildi"}}
    )
    # Kota geri iadesi
    try:
        await db.organization_plans.update_one(
            {"organization_id": current_user.organization_id},
            {"$inc": {"quota_usage": -cancel_count}}
        )
    except Exception as e:
        logging.warning(f"Quota refund error: {e}")
    # Cache + WebSocket
    try:
        await invalidate_cache(request, "dashboard_stats", current_user)
        await invalidate_cache(request, "customers_list", current_user)
        await emit_to_organization(current_user.organization_id, 'appointments_bulk_updated', {'session_group_id': group_id, 'cancelled': cancel_count})
    except Exception as e:
        logging.warning(f"Post-cancel cleanup error: {e}")
    logging.info(f"✅ Cancelled {cancel_count} remaining sessions for group {group_id}")
    return {"message": f"{cancel_count} seans iptal edildi", "cancelled_count": cancel_count, "session_group_id": group_id}

@api_router.post("/appointments/{appointment_id}/refund")
async def refund_single_appointment(request: Request, appointment_id: str, current_user: UserInDB = Depends(get_current_user)):
    """Tekil randevu için Stripe iade başlatır + state machine + wallet günceller."""
    if current_user.role not in ("admin", "superadmin"):
        raise HTTPException(status_code=403, detail="Yalnızca admin iade başlatabilir")
    db = await get_db_from_request(request)
    apt = await db.appointments.find_one({"id": appointment_id, "organization_id": current_user.organization_id})
    if not apt:
        raise HTTPException(status_code=404, detail="Randevu bulunamadı")

    REFUNDABLE_STATES = ("captured", "settled", "available")

    txn = await db.merchant_transactions.find_one({
        "organization_id": current_user.organization_id,
        "appointment_id": appointment_id,
        "state": {"$in": list(REFUNDABLE_STATES)},
    })
    if not txn and apt.get("session_group_id"):
        txn = await db.merchant_transactions.find_one({
            "organization_id": current_user.organization_id,
            "session_group_id": apt["session_group_id"],
            "state": {"$in": list(REFUNDABLE_STATES)},
        })

    if not txn or not txn.get("stripe_payment_intent_id"):
        raise HTTPException(status_code=400, detail="Bu randevu için iade edilebilir ödeme bulunamadı")

    from financial.state_machine import StateMachine

    is_session_partial = bool(apt.get("session_group_id") and apt.get("session_total") and apt["session_total"] > 1)
    refund_amount = txn.get("amount_display_minor", 0)
    if is_session_partial:
        refund_amount = int(refund_amount / apt["session_total"])

        if refund_amount <= 0:
            raise HTTPException(status_code=400, detail="İade tutarı hesaplanamadı")

    try:
        stripe_refund = stripe.Refund.create(
            payment_intent=txn["stripe_payment_intent_id"],
            amount=refund_amount,
            reason="requested_by_customer",
            metadata={
                "organization_id": current_user.organization_id,
                "appointment_id": appointment_id,
                "transaction_id": txn["id"],
            },
        )
    except stripe.error.StripeError as e:
        logging.error(f"Stripe refund failed for appointment {appointment_id}: {e}")
        raise HTTPException(status_code=400, detail=f"Stripe iade hatası: {str(e)}")

    now_iso = datetime.now(timezone.utc).isoformat()

    if is_session_partial:
        prev_refunded = txn.get("partial_refund_total_minor", 0)
        await db.merchant_transactions.update_one(
            {"id": txn["id"]},
            {
                "$set": {"partial_refund_total_minor": prev_refunded + refund_amount, "updated_at": now_iso},
                "$push": {"state_history": {
                    "state": "partial_refund", "timestamp": now_iso,
                    "trigger": "single_session_refund", "refund_amount_minor": refund_amount,
                    "stripe_refund_id": stripe_refund.id, "appointment_id": appointment_id,
                }},
            },
        )
        refund_tx_doc = {
            "id": str(uuid.uuid4()),
            "organization_id": current_user.organization_id,
            "base_currency": txn.get("base_currency", "TRY"),
            "type": "refund",
            "state": "refunded",
            "amount_display_minor": refund_amount,
            "fee_amount_minor": 0,
            "stripe_refund_id": stripe_refund.id,
            "stripe_payment_intent_id": txn.get("stripe_payment_intent_id"),
            "appointment_id": appointment_id,
            "session_group_id": apt.get("session_group_id"),
            "parent_transaction_id": txn["id"],
            "customer_name": txn.get("customer_name"),
            "customer_phone": txn.get("customer_phone"),
            "created_at": now_iso, "updated_at": now_iso,
            "state_history": [{"state": "refunded", "timestamp": now_iso, "trigger": "single_session_refund"}],
        }
        await db.merchant_transactions.insert_one(refund_tx_doc)
        await db.merchant_wallets.update_one(
            {"organization_id": current_user.organization_id},
            {"$inc": {"available_balance_minor": -refund_amount, "total_refunded_minor": refund_amount}, "$set": {"updated_at": now_iso}},
        )
    else:
        sm = StateMachine(db)
        try:
            await sm.transition(
                tx_id=txn["id"],
                new_state="refunded",
                trigger="appointment_refund",
                metadata={
                    "stripe_refund_id": stripe_refund.id,
                    "refund_amount_minor": refund_amount,
                    "refund_appointment_id": appointment_id,
                },
            )
        except Exception as e:
            logging.error(f"State transition after refund failed for tx {txn['id']}: {e}")

        await db.appointments.update_one(
            {"id": appointment_id},
        {"$set": {"payment_status": "refunded", "status": "İptal Edildi"}}
    )

    base_currency = txn.get("base_currency", "TRY")
    logging.info(f"Single refund {stripe_refund.id} for appointment {appointment_id}: {refund_amount} minor ({base_currency}), session_partial={is_session_partial}")
    return {
        "message": "İade başlatıldı",
        "refund_id": stripe_refund.id,
        "amount_minor": refund_amount,
    }

@api_router.post("/appointments/session-group/{group_id}/cancel-and-refund")
async def cancel_remaining_sessions_and_refund(request: Request, group_id: str, from_session_number: int = Query(1, ge=1), current_user: UserInDB = Depends(get_current_user)):
    """Seans grubundaki seçilen seans numarasından itibaren iptal eder ve orantılı iade başlatır."""
    if current_user.role not in ("admin", "superadmin"):
        raise HTTPException(status_code=403, detail="Yalnızca admin iade başlatabilir")
    db = await get_db_from_request(request)
    turkey_tz = ZoneInfo("Europe/Istanbul")
    today_str = datetime.now(turkey_tz).strftime("%Y-%m-%d")
    query = {
        "session_group_id": group_id,
        "organization_id": current_user.organization_id,
        "session_number": {"$gte": from_session_number},
        "status": {"$nin": ["İptal", "İptal Edildi", "Tamamlandı"]}
    }
    future_sessions = await db.appointments.find(query, {"_id": 0, "id": 1}).to_list(100)
    if not future_sessions:
        raise HTTPException(status_code=404, detail="İptal edilecek bekleyen seans bulunamadı")

    # Get total session count for this group
    total_in_group = await db.appointments.count_documents({
        "session_group_id": group_id,
        "organization_id": current_user.organization_id,
    })
    cancel_count = len(future_sessions)
    ids_to_cancel = [s['id'] for s in future_sessions]

    # Cancel
    await db.appointments.update_many(
        {"id": {"$in": ids_to_cancel}, "organization_id": current_user.organization_id},
        {"$set": {"status": "İptal Edildi"}}
    )

    # Quota refund
    try:
        await db.organization_plans.update_one(
            {"organization_id": current_user.organization_id},
            {"$inc": {"quota_usage": -cancel_count}}
        )
    except Exception as e:
        logging.warning(f"Quota refund error: {e}")

    # Proportional refund via financial engine + state machine
    refund_result = None
    try:
        from financial.state_machine import StateMachine
        import stripe as stripe_lib

        REFUNDABLE_STATES = ("captured", "settled", "available")

        # 1) Try session_group_id first
        txn = await db.merchant_transactions.find_one({
            "organization_id": current_user.organization_id,
            "session_group_id": group_id,
            "state": {"$in": list(REFUNDABLE_STATES)},
        })

        # 2) Fallback: find by appointment_id of the first (paid) session
        if not txn:
            first_session = await db.appointments.find_one({
                "session_group_id": group_id,
                "organization_id": current_user.organization_id,
                "session_number": 1,
            })
            if first_session:
                txn = await db.merchant_transactions.find_one({
                    "organization_id": current_user.organization_id,
                    "appointment_id": first_session["id"],
                    "state": {"$in": list(REFUNDABLE_STATES)},
                })
                if txn and not txn.get("session_group_id"):
                    await db.merchant_transactions.update_one(
                        {"id": txn["id"]},
                        {"$set": {"session_group_id": group_id}},
                    )
                    logging.info(f"Backfilled session_group_id on tx {txn['id']}")

        if txn and txn.get("stripe_payment_intent_id"):
            proportion = cancel_count / total_in_group if total_in_group > 0 else 0
            refund_amount = int(txn.get("amount_display_minor", 0) * proportion)
            is_full_refund = cancel_count >= total_in_group

            if refund_amount > 0:
                stripe_refund = stripe_lib.Refund.create(
                    payment_intent=txn["stripe_payment_intent_id"],
                    amount=refund_amount,
                    reason="requested_by_customer",
                    metadata={
                        "organization_id": current_user.organization_id,
                        "session_group_id": group_id,
                        "transaction_id": txn["id"],
                        "sessions_cancelled": str(cancel_count),
                    },
                )

                now_iso = datetime.now(ZoneInfo("UTC")).isoformat()

                if is_full_refund:
                    sm = StateMachine(db)
                    try:
                        await sm.transition(
                            tx_id=txn["id"],
                            new_state="refunded",
                            trigger="session_group_refund",
                            metadata={
                                "stripe_refund_id": stripe_refund.id,
                                "refund_amount_minor": refund_amount,
                                "sessions_cancelled": cancel_count,
                                "total_sessions": total_in_group,
                            },
                        )
                    except Exception as e:
                        logging.error(f"State transition after full session refund failed for tx {txn['id']}: {e}")
                else:
                    # Partial refund: keep original tx, create a visible refund record
                    prev_refunded = txn.get("partial_refund_total_minor", 0)
                    await db.merchant_transactions.update_one(
                        {"id": txn["id"]},
                        {
                            "$set": {
                                "partial_refund_total_minor": prev_refunded + refund_amount,
                                "updated_at": now_iso,
                            },
                            "$push": {
                                "state_history": {
                                    "state": "partial_refund",
                                    "timestamp": now_iso,
                                    "trigger": "session_group_partial_refund",
                                    "refund_amount_minor": refund_amount,
                                    "stripe_refund_id": stripe_refund.id,
                                    "sessions_cancelled": cancel_count,
                                    "total_sessions": total_in_group,
                                },
                            },
                        },
                    )

                    refund_tx_doc = {
                        "id": str(uuid.uuid4()),
                        "organization_id": current_user.organization_id,
                        "base_currency": txn.get("base_currency", "TRY"),
                        "type": "refund",
                        "state": "refunded",
                        "amount_display_minor": refund_amount,
                        "fee_amount_minor": 0,
                        "stripe_refund_id": stripe_refund.id,
                        "stripe_payment_intent_id": txn.get("stripe_payment_intent_id"),
                        "session_group_id": group_id,
                        "parent_transaction_id": txn["id"],
                        "customer_name": txn.get("customer_name"),
                        "customer_phone": txn.get("customer_phone"),
                        "created_at": now_iso,
                        "updated_at": now_iso,
                        "state_history": [{"state": "refunded", "timestamp": now_iso, "trigger": "session_group_partial_refund"}],
                    }
                    await db.merchant_transactions.insert_one(refund_tx_doc)

                    wallet_debit = {
                        "available_balance_minor": -refund_amount,
                        "total_refunded_minor": refund_amount,
                    }
                    await db.merchant_wallets.update_one(
                        {"organization_id": current_user.organization_id},
                        {"$inc": wallet_debit, "$set": {"updated_at": now_iso}},
                    )
                    logging.info(f"Partial session refund: created refund tx {refund_tx_doc['id']}, wallet debit: {wallet_debit}")

                refund_result = {
                    "refund_id": stripe_refund.id,
                    "amount_minor": refund_amount,
                    "sessions_refunded": cancel_count,
                    "is_full_refund": is_full_refund,
                }
                logging.info(f"Session refund {stripe_refund.id}: {refund_amount} minor ({cancel_count}/{total_in_group} sessions, full={is_full_refund})")
        else:
            logging.warning(f"No refundable transaction found for session group {group_id}")
            refund_result = {"error": "İade edilecek ödeme işlemi bulunamadı"}
    except Exception as e:
        logging.warning(f"Refund attempt error: {e}")
        refund_result = {"error": str(e)}

    # Cache + WebSocket
    try:
        await invalidate_cache(request, "dashboard_stats", current_user)
        await invalidate_cache(request, "customers_list", current_user)
        await emit_to_organization(current_user.organization_id, 'appointments_bulk_updated', {'session_group_id': group_id, 'cancelled': cancel_count})
    except Exception as e:
        logging.warning(f"Post-cancel cleanup error: {e}")

    logging.info(f"✅ Cancelled+refund {cancel_count} remaining sessions for group {group_id}")
    return {
        "message": f"{cancel_count} seans iptal edildi" + (" ve iade başlatıldı" if refund_result and "refund_id" in refund_result else ""),
        "cancelled_count": cancel_count,
        "session_group_id": group_id,
        "refund": refund_result,
    }


async def _insert_refund_reconciliation_pending(
    db,
    *,
    organization_id: str,
    request_id: str,
    failed_step: str,
    stripe_refund_id: str,
    stripe_payment_intent_id: str,
    merchant_transaction_id: str,
    session_group_id: str,
    appointment_ids: list,
    refund_amount_minor: int,
    is_full_refund: bool,
    error_message: str,
    snapshot_prev_partial_minor: int | None = None,
) -> None:
    """Stripe iade alındıktan sonra DB/state güncellemesi başarısız — mutabakat kuyruğu."""
    doc = {
        "id": str(uuid.uuid4()),
        "organization_id": organization_id,
        "status": "pending",
        "failed_step": failed_step,
        "stripe_refund_id": stripe_refund_id,
        "stripe_payment_intent_id": stripe_payment_intent_id,
        "merchant_transaction_id": merchant_transaction_id,
        "session_group_id": session_group_id,
        "appointment_ids": list(appointment_ids),
        "refund_amount_minor": int(refund_amount_minor),
        "is_full_refund": bool(is_full_refund),
        "error_message": str(error_message)[:4000],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "request_id": request_id,
        "retry_count": 0,
        "next_retry_at": None,
        "locked_by": None,
        "locked_at": None,
    }
    if snapshot_prev_partial_minor is not None:
        doc["snapshot_prev_partial_minor"] = int(snapshot_prev_partial_minor)
    try:
        await db.refund_reconciliation_pending.insert_one(doc)
    except Exception as e:
        logging.error(f"refund_reconciliation_pending insert failed: {e}", exc_info=True)


def _raise_stripe_refund_db_failed(request_id: str, stripe_refund_id: str) -> None:
    raise HTTPException(
        status_code=503,
        detail={
            "success": False,
            "error_code": "stripe_refund_succeeded_db_failed",
            "refund_status": "pending_reconciliation",
            "message": "Ödeme iadesi alındı; kayıtlar senkronize edilemedi. Mutabakat kuyruğuna eklendi. Gerekirse destek ile iletişime geçin.",
            "request_id": request_id,
            "stripe_refund_id": stripe_refund_id,
        },
    )


@api_router.post("/appointments/cancel-selected")
async def cancel_selected_appointments(
    request: Request,
    body: CancelSelectedRequest,
    current_user: UserInDB = Depends(get_current_user),
):
    """Seçili randevu id\'lerini toplu iptal; refund=true ise (admin) önce Stripe iade, sonra DB güncelle — tek session_group_id paketi."""
    request_id = str(uuid.uuid4())
    assert_capability_for_user(current_user, CAP_APPOINTMENT_CANCEL, request_id=request_id)
    db = await get_db_from_request(request)
    raw_ids = [str(x).strip() for x in (body.appointment_ids or []) if str(x).strip()]
    if not raw_ids:
        raise HTTPException(
            status_code=400,
            detail={"error_code": "validation_failed", "message": "appointment_ids gerekli", "request_id": request_id},
        )
    ids = list(dict.fromkeys(raw_ids))

    appts = await db.appointments.find(
        {"id": {"$in": ids}, "organization_id": current_user.organization_id},
        {"_id": 0},
    ).to_list(500)
    if len(appts) != len(ids):
        found = {a["id"] for a in appts}
        missing = [i for i in ids if i not in found]
        raise HTTPException(
            status_code=404,
            detail={"error_code": "validation_failed", "message": f"Randevu bulunamadı: {missing[:5]}", "request_id": request_id},
        )

    blocked = frozenset({"Tamamlandı", "İptal", "İptal Edildi", "Cancelled"})
    for a in appts:
        if a.get("status") in blocked:
            raise HTTPException(
                status_code=400,
                detail={
                    "error_code": "validation_failed",
                    "message": "Seçilen randevulardan biri iptal edilemez (tamamlanmış veya zaten iptal).",
                    "request_id": request_id,
                },
            )

    if current_user.role == "staff" and not getattr(current_user, "can_view_all_appointments", False):
        if staff_cannot_access_appointments(appts, current_user.username):
            raise HTTPException(
                status_code=403,
                detail={"error_code": "permission_denied", "message": "Bu randevulara erişim yetkiniz yok", "request_id": request_id},
            )

    k = len(ids)

    if not body.refund:
        await db.appointments.update_many(
            {"id": {"$in": ids}, "organization_id": current_user.organization_id},
            {"$set": {"status": "İptal Edildi"}},
        )
        try:
            await db.organization_plans.update_one(
                {"organization_id": current_user.organization_id},
                {"$inc": {"quota_usage": -k}},
            )
        except Exception as e:
            logging.warning(f"Quota refund error (cancel-selected): {e}")
        try:
            await invalidate_cache(request, "dashboard_stats", current_user)
            await invalidate_cache(request, "customers_list", current_user)
            await emit_to_organization(
                current_user.organization_id,
                "appointments_bulk_updated",
                {"cancelled": k, "appointment_ids": ids},
            )
        except Exception as e:
            logging.warning(f"Post-cancel-selected cleanup: {e}")
        await create_audit_log(
            db=db,
            organization_id=current_user.organization_id,
            user_id=current_user.username,
            user_full_name=current_user.full_name or current_user.username,
            action="UPDATE",
            resource_type="APPOINTMENT",
            resource_id=request_id,
            new_value={
                "operation": "cancel_selected",
                "refund": False,
                "appointment_ids": ids,
                "cancelled_count": k,
                "request_id": request_id,
            },
            ip_address=request.client.host if request.client else None,
        )
        await log_operation_audit(
            db,
            request_id=request_id,
            organization_id=current_user.organization_id,
            actor_user_id=current_user.username,
            action="cancel_selected",
            target_kind="appointment",
            target_ids=ids,
            result="success",
            meta={"refund": False},
        )
        return {
            "success": True,
            "cancelled_ids": ids,
            "refund_status": "none",
            "refunded_amount_minor": None,
            "stripe_refund_id": None,
            "error_code": None,
            "request_id": request_id,
        }

    assert_capability_for_user(current_user, CAP_APPOINTMENT_REFUND, request_id=request_id)

    sgids = {a.get("session_group_id") for a in appts if a.get("session_group_id")}
    if len(sgids) != 1:
        raise HTTPException(
            status_code=422,
            detail={
                "error_code": "validation_failed",
                "message": "İade için seçim tek bir seans paketine (aynı session_group_id) ait olmalıdır.",
                "request_id": request_id,
            },
        )
    group_id = sgids.pop()
    session_total = max((int(a.get("session_total") or 0) for a in appts), default=k)
    if session_total < 1:
        session_total = k

    REFUNDABLE_STATES = ("captured", "settled", "available")
    txn = await db.merchant_transactions.find_one({
        "organization_id": current_user.organization_id,
        "session_group_id": group_id,
        "state": {"$in": list(REFUNDABLE_STATES)},
    })
    if not txn:
        fs = await db.appointments.find_one(
            {"session_group_id": group_id, "organization_id": current_user.organization_id, "session_number": 1},
            {"_id": 0, "id": 1},
        )
        if fs:
            txn = await db.merchant_transactions.find_one({
                "organization_id": current_user.organization_id,
                "appointment_id": fs["id"],
                "state": {"$in": list(REFUNDABLE_STATES)},
            })

    if not txn or not txn.get("stripe_payment_intent_id"):
        raise HTTPException(
            status_code=400,
            detail={"error_code": "validation_failed", "message": "İade edilebilir ödeme bulunamadı", "request_id": request_id},
        )

    total_capture = int(txn.get("amount_display_minor") or 0)
    prev_part = int(txn.get("partial_refund_total_minor") or 0)
    per_session = int(total_capture / session_total) if session_total > 0 else 0
    refund_amount = per_session * k
    remaining_cap = max(0, total_capture - prev_part)
    refund_amount = min(refund_amount, remaining_cap)
    if refund_amount <= 0:
        raise HTTPException(
            status_code=400,
            detail={"error_code": "validation_failed", "message": "İade tutarı sıfır veya kalan tutar yetersiz", "request_id": request_id},
        )

    is_full_refund = k >= session_total

    try:
        stripe_refund = stripe.Refund.create(
            payment_intent=txn["stripe_payment_intent_id"],
            amount=refund_amount,
            reason="requested_by_customer",
            metadata={
                "organization_id": current_user.organization_id,
                "session_group_id": group_id,
                "transaction_id": txn["id"],
                "cancel_selected": "1",
            },
        )
    except stripe.error.StripeError as e:
        logging.error(f"cancel-selected Stripe error: {e}")
        err_code = "stripe_insufficient_balance" if "balance" in str(e).lower() else "validation_failed"
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error_code": err_code,
                "message": str(e),
                "request_id": request_id,
                "refund_status": "failed",
            },
        )

    now_iso = datetime.now(timezone.utc).isoformat()
    from financial.state_machine import StateMachine
    if is_full_refund:
        sm = StateMachine(db)
        try:
            await sm.transition(
                tx_id=txn["id"],
                new_state="refunded",
                trigger="cancel_selected_refund",
                metadata={
                    "stripe_refund_id": stripe_refund.id,
                    "refund_amount_minor": refund_amount,
                    "appointment_ids": ids,
                },
            )
        except Exception as e:
            logging.error(f"StateMachine after cancel-selected: {e}", exc_info=True)
            await _insert_refund_reconciliation_pending(
                db,
                organization_id=current_user.organization_id,
                request_id=request_id,
                failed_step="state_machine",
                stripe_refund_id=stripe_refund.id,
                stripe_payment_intent_id=txn.get("stripe_payment_intent_id") or "",
                merchant_transaction_id=txn["id"],
                session_group_id=group_id,
                appointment_ids=ids,
                refund_amount_minor=refund_amount,
                is_full_refund=True,
                error_message=str(e),
            )
            _raise_stripe_refund_db_failed(request_id, stripe_refund.id)
    else:
        try:
            await db.merchant_transactions.update_one(
                {"id": txn["id"]},
                {
                    "$set": {"partial_refund_total_minor": prev_part + refund_amount, "updated_at": now_iso},
                    "$push": {"state_history": {
                        "state": "partial_refund",
                        "timestamp": now_iso,
                        "trigger": "cancel_selected_partial",
                        "refund_amount_minor": refund_amount,
                        "stripe_refund_id": stripe_refund.id,
                        "appointment_ids": ids,
                    }},
                },
            )
            refund_tx_doc = {
                "id": str(uuid.uuid4()),
                "organization_id": current_user.organization_id,
                "base_currency": txn.get("base_currency", "TRY"),
                "type": "refund",
                "state": "refunded",
                "amount_display_minor": refund_amount,
                "fee_amount_minor": 0,
                "stripe_refund_id": stripe_refund.id,
                "stripe_payment_intent_id": txn.get("stripe_payment_intent_id"),
                "session_group_id": group_id,
                "parent_transaction_id": txn["id"],
                "customer_name": txn.get("customer_name"),
                "customer_phone": txn.get("customer_phone"),
                "created_at": now_iso,
                "updated_at": now_iso,
                "state_history": [{"state": "refunded", "timestamp": now_iso, "trigger": "cancel_selected_partial"}],
            }
            await db.merchant_transactions.insert_one(refund_tx_doc)
            await db.merchant_wallets.update_one(
                {"organization_id": current_user.organization_id},
                {"$inc": {"available_balance_minor": -refund_amount, "total_refunded_minor": refund_amount}, "$set": {"updated_at": now_iso}},
            )
        except Exception as e:
            logging.error(f"cancel-selected partial merchant DB failed: {e}", exc_info=True)
            await _insert_refund_reconciliation_pending(
                db,
                organization_id=current_user.organization_id,
                request_id=request_id,
                failed_step="partial_merchant_updates",
                stripe_refund_id=stripe_refund.id,
                stripe_payment_intent_id=txn.get("stripe_payment_intent_id") or "",
                merchant_transaction_id=txn["id"],
                session_group_id=group_id,
                appointment_ids=ids,
                refund_amount_minor=refund_amount,
                is_full_refund=False,
                error_message=str(e),
                snapshot_prev_partial_minor=prev_part,
            )
            _raise_stripe_refund_db_failed(request_id, stripe_refund.id)

    try:
        await db.appointments.update_many(
            {"id": {"$in": ids}, "organization_id": current_user.organization_id},
            {"$set": {"status": "İptal Edildi", "payment_status": "refunded"}},
        )
    except Exception as e:
        logging.error(f"cancel-selected appointments update after Stripe refund: {e}", exc_info=True)
        await _insert_refund_reconciliation_pending(
            db,
            organization_id=current_user.organization_id,
            request_id=request_id,
            failed_step="appointments_cancel",
            stripe_refund_id=stripe_refund.id,
            stripe_payment_intent_id=txn.get("stripe_payment_intent_id") or "",
            merchant_transaction_id=txn["id"],
            session_group_id=group_id,
            appointment_ids=ids,
            refund_amount_minor=refund_amount,
            is_full_refund=is_full_refund,
            error_message=str(e),
        )
        _raise_stripe_refund_db_failed(request_id, stripe_refund.id)
    try:
        await db.organization_plans.update_one(
            {"organization_id": current_user.organization_id},
            {"$inc": {"quota_usage": -k}},
        )
    except Exception as e:
        logging.warning(f"Quota adjust cancel-selected refund: {e}")

    try:
        await invalidate_cache(request, "dashboard_stats", current_user)
        await invalidate_cache(request, "customers_list", current_user)
        await emit_to_organization(
            current_user.organization_id,
            "appointments_bulk_updated",
            {"session_group_id": group_id, "cancelled": k, "appointment_ids": ids},
        )
    except Exception as e:
        logging.warning(f"cancel-selected post cleanup: {e}")

    await create_audit_log(
        db=db,
        organization_id=current_user.organization_id,
        user_id=current_user.username,
        user_full_name=current_user.full_name or current_user.username,
        action="UPDATE",
        resource_type="APPOINTMENT",
        resource_id=request_id,
        new_value={
            "operation": "cancel_selected",
            "refund": True,
            "session_group_id": group_id,
            "appointment_ids": ids,
            "cancelled_count": k,
            "refunded_amount_minor": refund_amount,
            "stripe_refund_id": stripe_refund.id,
            "request_id": request_id,
        },
        ip_address=request.client.host if request.client else None,
    )
    await log_operation_audit(
        db,
        request_id=request_id,
        organization_id=current_user.organization_id,
        actor_user_id=current_user.username,
        action="cancel_selected",
        target_kind="session_group",
        target_ids=[group_id],
        result="success",
        meta={
            "refund": True,
            "appointment_ids": ids,
            "refunded_amount_minor": refund_amount,
            "stripe_refund_id": stripe_refund.id,
        },
    )
    return {
        "success": True,
        "cancelled_ids": ids,
        "refund_status": "succeeded",
        "refunded_amount_minor": refund_amount,
        "stripe_refund_id": stripe_refund.id,
        "error_code": None,
        "request_id": request_id,
    }


# ═══ CONTRACT: v1 (FROZEN / RESPONSE) ═══════════════════════════════════════
# Donmuş mobil build sözleşmesi. response_model=List[Appointment] — alan
# ekleme/çıkarma/rename YASAK; yeni davranış yalnızca YENİ query parametresi
# arkasında. bkz. OPERATIONS.md §2 + .cursor/rules/api-contract.mdc
#
# FAZ 2 (parametre-versiyonlu pagination — ADDITIVE, legacy şekle DOKUNMAZ):
#   • limit YOK  → eskisi gibi TAM ARRAY (List[Appointment]) — donmuş build bunu bekler.
#   • ?limit=&cursor= → YENİ şekil {items, next_cursor, has_more} (JSONResponse ile
#     response_model baypas edilir). items[] her elemanı yine birebir Appointment şekli.
#   Cursor sırası: (appointment_date, appointment_time, id) artan; base64(date|time|id).
@api_router.get("/appointments", response_model=List[Appointment])
async def get_appointments(
    request: Request, 
    date: Optional[str] = None, 
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None, 
    search: Optional[str] = None,
    staff_member_id: Optional[str] = None,
    limit: Optional[int] = None,
    cursor: Optional[str] = None,
    session_only: Optional[bool] = None,
    current_user: UserInDB = Depends(get_current_user)
):
    db = await get_db_from_request(request)
    query = {"organization_id": current_user.organization_id, "status": {"$ne": "Ödeme Bekleniyor"}}

    # Seans paketi ekranı geçmiş seansları da göstermek zorunda (Tamamlanan sekmesi),
    # bu yüzden tarih penceresine sığmıyor. Bu filtre ile yalnızca pakete ait randevular
    # çekilir — küçük ve sınırlı bir küme. Parametre GÖNDERİLMEZSE davranış birebir eskisi.
    if session_only:
        query["session_group_id"] = {"$exists": True, "$ne": None}
    
    # Personel sadece kendine atanan randevuları görebilir (can_view_all_appointments yetkisi yoksa)
    if current_user.role == "staff" and not current_user.can_view_all_appointments:
        query['staff_member_id'] = current_user.username
    elif staff_member_id and (current_user.role == "admin" or current_user.can_view_all_appointments):
        # Admin için personel filtresi
        if staff_member_id != "all" and staff_member_id != "unassigned":
            query['staff_member_id'] = staff_member_id
        elif staff_member_id == "unassigned":
            query['$or'] = [
                {'staff_member_id': {'$exists': False}},
                {'staff_member_id': None},
                {'staff_member_id': ''}
            ]
    
    if date: 
        query['appointment_date'] = date
    elif start_date and end_date:
        # Tarih aralığı sorgusu
        query['appointment_date'] = {
            '$gte': start_date,
            '$lte': end_date
        }
    elif start_date:
        query['appointment_date'] = {'$gte': start_date}
    elif end_date:
        query['appointment_date'] = {'$lte': end_date}
    
    if status: query['status'] = status
    if search:
        query['$or'] = [
            {'customer_name': {'$regex': search, '$options': 'i'}},
            {'phone': {'$regex': search, '$options': 'i'}}
        ]
    
    # ── FAZ 2: parametre-versiyonlu fetch ────────────────────────────────────
    # limit YOK → legacy tam array (donmuş build). limit VAR → cursor pagination.
    use_pagination = limit is not None
    has_more = False
    if use_pagination:
        effective_limit = max(1, int(limit))
        if cursor:
            try:
                _raw = base64.urlsafe_b64decode(cursor.encode()).decode()
                c_date, c_time, c_id = _raw.split("|", 2)
                tie_break = {
                    "$or": [
                        {"appointment_date": {"$gt": c_date}},
                        {"appointment_date": c_date, "appointment_time": {"$gt": c_time}},
                        {"appointment_date": c_date, "appointment_time": c_time, "id": {"$gt": c_id}},
                    ]
                }
                # Mevcut query'de $or olabilir (search/unassigned) → $and ile birleştir.
                base_wo_org = {k: v for k, v in query.items() if k != "organization_id"}
                query = {"organization_id": current_user.organization_id, "$and": [tie_break]}
                if base_wo_org:
                    query["$and"].append(base_wo_org)
            except Exception:
                logging.warning(f"GET /appointments: bozuk cursor atlandı: {cursor}")
        appointments_from_db = await (
            db.appointments.find(query, {"_id": 0})
            .sort([("appointment_date", 1), ("appointment_time", 1), ("id", 1)])
            .limit(effective_limit + 1)
        ).to_list(effective_limit + 1)
        has_more = len(appointments_from_db) > effective_limit
        if has_more:
            appointments_from_db = appointments_from_db[:effective_limit]
    else:
        # Penceresiz dump'ta 1000 tavanı kalsın (tüm geçmişi yanlışlıkla çekmeye karşı).
        # Tarih veya seans filtresi varken aynı tavan eşleşen kayıtları sessizce kesiyordu:
        # yoğun bir işletmede "bugün" / takvim ayı listenin dışına düşüyordu.
        windowed = bool(date or start_date or end_date or session_only)
        appointments_from_db = await (
            db.appointments.find(query, {"_id": 0})
            .sort([("appointment_date", 1), ("appointment_time", 1)])
            .to_list(None if windowed else 1000)
        )
    try:
        turkey_tz = ZoneInfo("Europe/Istanbul"); now = datetime.now(turkey_tz)
    except Exception:
        turkey_tz = timezone(timedelta(hours=3)); now = datetime.now(turkey_tz)
    
    # Tüm servisleri bir kerede çek (performans için) — çoklu hizmette
    # snapshot içindeki tüm service_id'leri de topla ki toplam süre doğru hesaplansın
    service_ids = []
    for appt in appointments_from_db:
        if appt.get('service_id'):
            service_ids.append(appt['service_id'])
        for line in (appt.get('services') or []):
            if line.get('service_id'):
                service_ids.append(line['service_id'])
    services_dict = {}
    if service_ids:
        unique_service_ids = list(set(service_ids))
        logging.info(f"🔍 GET /appointments: {len(unique_service_ids)} unique service_id bulundu: {unique_service_ids[:5]}")
        services = await db.services.find(
            {"id": {"$in": unique_service_ids}, "organization_id": current_user.organization_id},
            {"_id": 0, "id": 1, "duration": 1}
        ).to_list(1000)
        services_dict = {s['id']: s.get('duration', 30) for s in services}
        logging.info(f"✅ GET /appointments: {len(services_dict)} servis bulundu, durations: {list(services_dict.values())[:5]}")
    else:
        logging.warning("⚠️ GET /appointments: Hiç service_id bulunamadı")
    
    ids_to_update = []; transactions_to_create = [] 
    for appt in appointments_from_db:
        if isinstance(appt.get('created_at'), str): appt['created_at'] = datetime.fromisoformat(appt['created_at'].replace('Z', '+00:00'))
        
        # Service duration ekle (bitiş saati hesaplamak için)
        appt_service_id = appt.get('service_id')
        appt_services = appt.get('services')
        if isinstance(appt_services, list) and len(appt_services) > 0:
            # Çoklu hizmet: tüm hizmetlerin süresini topla (canlı süre yoksa snapshot)
            total_dur = 0
            for line in appt_services:
                sid = line.get('service_id')
                total_dur += services_dict.get(sid, line.get('duration_snapshot', 30) or 30)
            appt['service_duration'] = total_dur or 30
        elif appt_service_id and appt_service_id in services_dict:
            appt['service_duration'] = services_dict[appt_service_id]
            logging.debug(f"✅ Randevu {appt.get('id', 'unknown')}: service_duration={appt['service_duration']} (service_id={appt_service_id})")
        else:
            appt['service_duration'] = 30
            if appt_service_id:
                logging.warning(f"⚠️ Randevu {appt.get('id', 'unknown')}: service_id={appt_service_id} services_dict'te bulunamadı, default 30 kullanılıyor")
            else:
                logging.warning(f"⚠️ Randevu {appt.get('id', 'unknown')}: service_id yok, default 30 kullanılıyor")
        
        if appt.get('status') == 'Bekliyor':
            try:
                dt_str = f"{appt['appointment_date']} {appt['appointment_time']}"
                naive_dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M"); appointment_dt = naive_dt.replace(tzinfo=turkey_tz)
                
                # Randevu bitiş saatini hesapla (başlangıç saati + hizmet süresi)
                service_duration_minutes = appt.get('service_duration', 30)
                completion_threshold = appointment_dt + timedelta(minutes=service_duration_minutes)
                
                # Şu anki saat >= bitiş saati ise tamamlandı olarak işaretle
                if now >= completion_threshold:
                    appt['status'] = 'Tamamlandı'; completed_at_iso = datetime.now(timezone.utc).isoformat()
                    appt['completed_at'] = completed_at_iso; ids_to_update.append(appt['id'])
                    logging.info(f"✅ Randevu {appt.get('id', 'unknown')} otomatik tamamlandı: {appt['appointment_time']} + {service_duration_minutes}dk = {completion_threshold.strftime('%H:%M')}, şimdi: {now.strftime('%H:%M')}")
                    transaction = Transaction(organization_id=current_user.organization_id, appointment_id=appt['id'], customer_name=appt['customer_name'], service_name=appt['service_name'], amount=appt['service_price'], date=appt['appointment_date'], staff_member_id=appt.get('staff_member_id'))
                    trans_doc = transaction.model_dump(); trans_doc['created_at'] = trans_doc['created_at'].isoformat()
                    transactions_to_create.append(trans_doc)
            except (ValueError, TypeError) as e: logging.warning(f"Randevu {appt['id']} için tarih ayrıştırılamadı: {e}")
    if ids_to_update:
        await db.appointments.update_many({"organization_id": current_user.organization_id, "id": {"$in": ids_to_update}}, {"$set": {"status": "Tamamlandı", "completed_at": datetime.now(timezone.utc).isoformat()}})
    if transactions_to_create:
        await db.transactions.insert_many(transactions_to_create)
    
    # UI: Stripe üzerinden iade edilebilir paket ödemesi var mı (aynı session_group)
    _gids = list({a.get("session_group_id") for a in appointments_from_db if a.get("session_group_id")})
    _eligible = set()
    if _gids:
        _tx = await db.merchant_transactions.find(
            {
                "organization_id": current_user.organization_id,
                "session_group_id": {"$in": _gids},
                "state": {"$in": list(REFUNDABLE_MERCHANT_STATES)},
                "stripe_payment_intent_id": {"$exists": True, "$nin": [None, ""]},
            },
            {"_id": 0, "session_group_id": 1},
        ).to_list(500)
        _eligible = session_group_ids_from_merchant_transactions(_tx)
    attach_refund_eligible(appointments_from_db, _eligible)
    
    # ── FAZ 2: paginated şekil (limit VAR) — legacy array'e DOKUNMAZ ──────────
    if use_pagination:
        # Her item yine BİREBİR Appointment şekli (donmuş alan seti korunur).
        items = [jsonable_encoder(Appointment(**a)) for a in appointments_from_db]
        next_cursor: Optional[str] = None
        if has_more and appointments_from_db:
            _last = appointments_from_db[-1]
            _raw = f"{_last.get('appointment_date', '')}|{_last.get('appointment_time', '')}|{_last.get('id', '')}"
            next_cursor = base64.urlsafe_b64encode(_raw.encode()).decode()
        return JSONResponse({"items": items, "next_cursor": next_cursor, "has_more": has_more})

    return appointments_from_db

# === SERVICES ROUTES ===
@api_router.post("/services/reorder", status_code=204)
async def reorder_services(request: Request, payload: dict, current_user: UserInDB = Depends(get_current_user)):
    # Sadece admin sıralayabilir
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")

    service_ids = payload.get("service_ids")
    if not isinstance(service_ids, list) or not service_ids:
        raise HTTPException(status_code=400, detail="Invalid service_ids")

    unique_ids = list(dict.fromkeys(service_ids))
    db = await get_db_from_request(request)

    existing = await db.services.find(
        {"organization_id": current_user.organization_id, "id": {"$in": unique_ids}},
        {"_id": 0, "id": 1}
    ).to_list(2000)
    existing_ids = {s.get("id") for s in existing}
    missing = [sid for sid in unique_ids if sid not in existing_ids]
    if missing:
        raise HTTPException(status_code=400, detail="Invalid service_ids")

    now = datetime.now(timezone.utc).isoformat()
    for idx, sid in enumerate(unique_ids):
        await db.services.update_one(
            {"organization_id": current_user.organization_id, "id": sid},
            {"$set": {"order": idx, "updated_at": now}}
        )

    return None

@api_router.delete("/services/{service_id}")
async def delete_service(request: Request, service_id: str, current_user: UserInDB = Depends(get_current_user)):
    # Sadece admin silebilir
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request); query = {"id": service_id, "organization_id": current_user.organization_id}
    result = await db.services.delete_one(query)
    if result.deleted_count == 0: raise HTTPException(status_code=404, detail="Hizmet bulunamadı")
    try:
        await invalidate_service_caches(request, current_user.organization_id)
    except Exception as e:
        logging.error(f"Service cache invalidation failed: {e}")
    return {"message": "Hizmet silindi"}

async def _get_org_base_currency(db, organization_id: str) -> str:
    """Organizasyonun para birimi (GBP / TRY). Ayar yoksa TRY."""
    try:
        settings_doc = await db.settings.find_one(
            {"organization_id": organization_id}, {"_id": 0, "base_currency": 1}
        )
        currency = (settings_doc or {}).get("base_currency") or "TRY"
        return currency.upper()
    except Exception as e:
        logging.error(f"base_currency okunamadı ({organization_id}): {e}")
        return "TRY"


def _validate_service_payment_limits(
    price: float,
    payment_rule: Optional[str],
    deposit_amount: Optional[float],
    base_currency: str = "TRY",
):
    """Minimum online ödeme / kapora limitleri.

    Eşikler para birimine göre değişir ve tek kaynaktan (financial.money) okunur;
    checkout tarafı (merchant_checkout) da aynı sözlükleri kullanıyor. Eskiden burada
    TRY değerleri sabit kodluydu, bu yüzden GBP işletmelerde £300 altındaki her hizmet
    reddediliyordu.
    """
    if not payment_rule or payment_rule == "on_site":
        return

    from financial.money import MIN_ONLINE_PAYMENT_MINOR, MIN_DEPOSIT_MINOR, format_display_short

    currency = (base_currency or "TRY").upper()
    if currency not in MIN_ONLINE_PAYMENT_MINOR:
        currency = "TRY"
    min_online = MIN_ONLINE_PAYMENT_MINOR[currency]
    min_deposit = MIN_DEPOSIT_MINOR[currency]

    price_minor = int(round(price * 100))
    if payment_rule in ("online", "full_online") and price_minor < min_online:
        raise HTTPException(
            status_code=400,
            detail=f"Online ödeme için hizmet fiyatı en az {format_display_short(min_online, currency)} olmalıdır.",
        )
    if payment_rule == "deposit":
        if price_minor < min_online:
            raise HTTPException(
                status_code=400,
                detail=f"Kapora için hizmet fiyatı en az {format_display_short(min_online, currency)} olmalıdır.",
            )
        if deposit_amount is not None:
            dep_minor = int(round(deposit_amount * 100))
            if dep_minor < min_deposit:
                raise HTTPException(
                    status_code=400,
                    detail=f"Kapora tutarı en az {format_display_short(min_deposit, currency)} olmalıdır.",
                )
            if dep_minor > price_minor:
                raise HTTPException(status_code=400, detail="Kapora tutarı, hizmetin toplam fiyatını aşamaz.")

@api_router.put("/services/{service_id}", response_model=Service)
async def update_service(request: Request, service_id: str, service_update: ServiceUpdate, current_user: UserInDB = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request); query = {"id": service_id, "organization_id": current_user.organization_id}
    service = await db.services.find_one(query, {"_id": 0})
    if not service: raise HTTPException(status_code=404, detail="Hizmet bulunamadı")
    effective_price = service_update.price if service_update.price is not None else service.get("price", 0)
    effective_rule = service_update.payment_rule if service_update.payment_rule is not None else service.get("payment_rule")
    effective_deposit = service_update.deposit_amount if service_update.deposit_amount is not None else service.get("deposit_amount")
    _validate_service_payment_limits(
        effective_price, effective_rule, effective_deposit,
        await _get_org_base_currency(db, current_user.organization_id),
    )
    if service_update.category_id:
        cat = await db.categories.find_one({"id": service_update.category_id, "organization_id": current_user.organization_id}, {"_id": 0, "id": 1})
        if not cat:
            raise HTTPException(status_code=400, detail="Geçersiz kategori")
    raw = service_update.model_dump()
    update_data = {k: v for k, v in raw.items() if v is not None}
    unset_data = {}
    for field in ("session_count", "deposit_percentage", "deposit_amount", "category_id"):
        if field in raw and raw[field] is None and field in service:
            unset_data[field] = ""
    # Optimistic lock / versiyonlama: fiyat/süre/isim değişiminde version artır.
    # Böylece açık checkout'lar snapshot'taki service_version ile tutarlı kalır.
    version_changing = any(
        f in update_data and update_data[f] != service.get(f)
        for f in ("price", "duration", "name")
    )
    if version_changing:
        update_data["version"] = int(service.get("version", 1)) + 1
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    ops = {}
    if update_data: ops["$set"] = update_data
    if unset_data: ops["$unset"] = unset_data
    if ops:
        await db.services.update_one(query, ops)
        try:
            await invalidate_service_caches(request, current_user.organization_id)
        except Exception as e:
            logging.error(f"Service cache invalidation failed: {e}")
    updated_service = await db.services.find_one(query, {"_id": 0})
    if isinstance(updated_service['created_at'], str): updated_service['created_at'] = datetime.fromisoformat(updated_service['created_at'].replace('Z', '+00:00'))
    return updated_service

@api_router.get("/services/{service_id}", response_model=Service)
async def get_service(request: Request, service_id: str, current_user: UserInDB = Depends(get_current_user)):
    db = await get_db_from_request(request); query = {"id": service_id, "organization_id": current_user.organization_id}
    service = await db.services.find_one(query, {"_id": 0})
    if not service: raise HTTPException(status_code=404, detail="Hizmet bulunamadı")
    if isinstance(service['created_at'], str): service['created_at'] = datetime.fromisoformat(service['created_at'].replace('Z', '+00:00'))
    return service

@api_router.post("/services", response_model=Service)
async def create_service(request: Request, service: ServiceCreate, current_user: UserInDB = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    _validate_service_payment_limits(
        service.price, service.payment_rule, service.deposit_amount,
        await _get_org_base_currency(db, current_user.organization_id),
    )
    if service.category_id:
        cat = await db.categories.find_one({"id": service.category_id, "organization_id": current_user.organization_id}, {"_id": 0, "id": 1})
        if not cat:
            raise HTTPException(status_code=400, detail="Geçersiz kategori")

    service_data = service.model_dump()
    if service_data.get("order") is None:
        last = await db.services.find(
            {"organization_id": current_user.organization_id},
            {"_id": 0, "order": 1}
        ).sort("order", -1).limit(1).to_list(1)
        last_order = None
        if last:
            last_order = last[0].get("order")
        service_data["order"] = (last_order + 1) if isinstance(last_order, int) else 0

    service_obj = Service(**service_data, organization_id=current_user.organization_id)
    doc = service_obj.model_dump(); doc['created_at'] = doc['created_at'].isoformat()
    await db.services.insert_one(doc)
    try:
        # Yeni hizmet eklenince org'daki tüm admin+staff'a otomatik atanır.
        # Admin sonradan istediğini Personel Yönetimi'nden kaldırabilir.
        await db.users.update_many(
            {
                "organization_id": current_user.organization_id,
                "role": {"$in": ["admin", "staff"]},
            },
            {"$addToSet": {"permitted_service_ids": service_obj.id}},
        )
    except Exception as e:
        logging.error(f"Failed to auto-assign service to staff: {e}")
    try:
        await invalidate_service_caches(request, current_user.organization_id)
    except Exception as e:
        logging.error(f"Service cache invalidation failed: {e}")
    return service_obj

@api_router.get("/services", response_model=List[Service])
async def get_services(request: Request, current_user: UserInDB = Depends(get_current_user)):
    db = await get_db_from_request(request); query = {"organization_id": current_user.organization_id}
    services = await db.services.find(query, {"_id": 0}).to_list(1000)
    for service in services:
        if isinstance(service.get('created_at'), str):
            service['created_at'] = datetime.fromisoformat(service['created_at'].replace('Z', '+00:00'))

    def sort_key(svc: dict):
        order = svc.get("order")
        created = svc.get("created_at")
        created_str = created.isoformat() if isinstance(created, datetime) else (created or "")
        return (
            order if isinstance(order, int) else 1_000_000,
            created_str,
            svc.get("name") or "",
        )

    services_sorted = sorted(services, key=sort_key)
    return services_sorted

# === CATEGORIES ROUTES (Hizmet Kategorileri) ===
def _sort_categories(categories: List[dict]) -> List[dict]:
    def key(cat: dict):
        order = cat.get("order")
        created = cat.get("created_at")
        created_str = created.isoformat() if isinstance(created, datetime) else (created or "")
        return (order if isinstance(order, int) else 1_000_000, created_str, cat.get("name") or "")
    return sorted(categories, key=key)

@api_router.get("/categories", response_model=List[Category])
async def get_categories(request: Request, current_user: UserInDB = Depends(get_current_user)):
    db = await get_db_from_request(request)
    cats = await db.categories.find({"organization_id": current_user.organization_id}, {"_id": 0}).to_list(1000)
    for cat in cats:
        if isinstance(cat.get("created_at"), str):
            cat["created_at"] = datetime.fromisoformat(cat["created_at"].replace("Z", "+00:00"))
    return _sort_categories(cats)

@api_router.post("/categories", response_model=Category)
async def create_category(request: Request, category: CategoryCreate, current_user: UserInDB = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    db = await get_db_from_request(request)
    data = category.model_dump()
    if data.get("order") is None:
        last = await db.categories.find(
            {"organization_id": current_user.organization_id}, {"_id": 0, "order": 1}
        ).sort("order", -1).limit(1).to_list(1)
        last_order = last[0].get("order") if last else None
        data["order"] = (last_order + 1) if isinstance(last_order, int) else 0
    cat_obj = Category(**data, organization_id=current_user.organization_id)
    doc = cat_obj.model_dump(); doc["created_at"] = doc["created_at"].isoformat()
    await db.categories.insert_one(doc)
    return cat_obj

@api_router.put("/categories/{category_id}", response_model=Category)
async def update_category(request: Request, category_id: str, category_update: CategoryUpdate, current_user: UserInDB = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    db = await get_db_from_request(request); query = {"id": category_id, "organization_id": current_user.organization_id}
    cat = await db.categories.find_one(query, {"_id": 0})
    if not cat:
        raise HTTPException(status_code=404, detail="Kategori bulunamadı")
    update_data = {k: v for k, v in category_update.model_dump().items() if v is not None}
    if update_data:
        await db.categories.update_one(query, {"$set": update_data})
    updated = await db.categories.find_one(query, {"_id": 0})
    if isinstance(updated.get("created_at"), str):
        updated["created_at"] = datetime.fromisoformat(updated["created_at"].replace("Z", "+00:00"))
    return updated

@api_router.delete("/categories/{category_id}")
async def delete_category(request: Request, category_id: str, current_user: UserInDB = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    db = await get_db_from_request(request); query = {"id": category_id, "organization_id": current_user.organization_id}
    cat = await db.categories.find_one(query, {"_id": 0})
    if not cat:
        raise HTTPException(status_code=404, detail="Kategori bulunamadı")
    # Önce bağlı hizmetlerin category_id'sini atomik şekilde temizle, sonra kategoriyi sil.
    # Yarım kalmayı önlemek için update_many başarısız olursa kategori silinmez.
    await db.services.update_many(
        {"organization_id": current_user.organization_id, "category_id": category_id},
        {"$unset": {"category_id": ""}}
    )
    await db.categories.delete_one(query)
    try:
        await invalidate_service_caches(request, current_user.organization_id)
    except Exception as e:
        logging.error(f"Category cache invalidation failed: {e}")
    return {"message": "Kategori silindi"}

@api_router.post("/categories/{category_id}/services")
async def assign_services_to_category(request: Request, category_id: str, payload: dict, current_user: UserInDB = Depends(get_current_user)):
    """Bir kategorinin hizmet üyeliğini topluca ayarlar.

    `service_ids` kategorinin YENİ tam üye listesidir: listede olanlara `category_id`
    atanır, kategoride olup listede olmayanların ataması kaldırılır. Boş liste
    göndermek kategoriyi boşaltır. Bu semantik, arayüzdeki çoklu seçim kutusunun
    "kaydet" davranışıyla birebir örtüşüyor.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")

    service_ids = payload.get("service_ids")
    if not isinstance(service_ids, list):
        raise HTTPException(status_code=400, detail="service_ids listesi gerekli")
    unique_ids = [sid for sid in dict.fromkeys(service_ids) if isinstance(sid, str) and sid]

    db = await get_db_from_request(request)
    org_id = current_user.organization_id

    cat = await db.categories.find_one({"id": category_id, "organization_id": org_id}, {"_id": 0, "id": 1})
    if not cat:
        raise HTTPException(status_code=404, detail="Kategori bulunamadı")

    # Başka organizasyonun hizmeti sızmasın: gönderilen her id bu org'da var mı?
    if unique_ids:
        existing = await db.services.find(
            {"organization_id": org_id, "id": {"$in": unique_ids}}, {"_id": 0, "id": 1}
        ).to_list(2000)
        existing_ids = {s.get("id") for s in existing}
        missing = [sid for sid in unique_ids if sid not in existing_ids]
        if missing:
            raise HTTPException(status_code=400, detail="Geçersiz hizmet listesi")

    assigned = 0
    if unique_ids:
        result = await db.services.update_many(
            {"organization_id": org_id, "id": {"$in": unique_ids}},
            {"$set": {"category_id": category_id}}
        )
        assigned = result.modified_count

    removed_result = await db.services.update_many(
        {"organization_id": org_id, "category_id": category_id, "id": {"$nin": unique_ids}},
        {"$unset": {"category_id": ""}}
    )

    try:
        await invalidate_service_caches(request, org_id)
    except Exception as e:
        logging.error(f"Category assignment cache invalidation failed: {e}")

    return {
        "category_id": category_id,
        "assigned_count": assigned,
        "removed_count": removed_result.modified_count,
        "total_in_category": len(unique_ids),
    }

@api_router.post("/categories/reorder", status_code=204)
async def reorder_categories(request: Request, payload: dict, current_user: UserInDB = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    category_ids = payload.get("category_ids")
    if not isinstance(category_ids, list) or not category_ids:
        raise HTTPException(status_code=400, detail="Invalid category_ids")
    unique_ids = list(dict.fromkeys(category_ids))
    db = await get_db_from_request(request)
    existing = await db.categories.find(
        {"organization_id": current_user.organization_id, "id": {"$in": unique_ids}}, {"_id": 0, "id": 1}
    ).to_list(2000)
    existing_ids = {c.get("id") for c in existing}
    if any(cid not in existing_ids for cid in unique_ids):
        raise HTTPException(status_code=400, detail="Invalid category_ids")
    for idx, cid in enumerate(unique_ids):
        await db.categories.update_one(
            {"organization_id": current_user.organization_id, "id": cid},
            {"$set": {"order": idx}}
        )
    return None

# === TRANSACTIONS ROUTES ===
@api_router.get("/transactions", response_model=List[Transaction])
async def get_transactions(request: Request, start_date: Optional[str] = None, end_date: Optional[str] = None, current_user: UserInDB = Depends(get_current_user)):
    # Sadece admin görebilir
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request); query = {"organization_id": current_user.organization_id}
    if start_date and end_date: query['date'] = {'$gte': start_date, '$lte': end_date}
    elif start_date: query['date'] = {'$gte': start_date}
    elif end_date: query['date'] = {'$lte': end_date}
    transactions = await db.transactions.find(query, {"_id": 0}).to_list(1000)
    for transaction in transactions:
        if isinstance(transaction['created_at'], str): transaction['created_at'] = datetime.fromisoformat(transaction['created_at'].replace('Z', '+00:00'))
    return transactions

@api_router.put("/transactions/{transaction_id}", response_model=Transaction)
async def update_transaction(request: Request, transaction_id: str, transaction_update: TransactionUpdate, current_user: UserInDB = Depends(get_current_user)):
    # Sadece admin güncelleyebilir
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request); query = {"id": transaction_id, "organization_id": current_user.organization_id}
    transaction = await db.transactions.find_one(query, {"_id": 0})
    if not transaction: raise HTTPException(status_code=404, detail="İşlem bulunamadı")
    await db.transactions.update_one(query, {"$set": {"amount": transaction_update.amount}})
    updated_transaction = await db.transactions.find_one(query, {"_id": 0})
    if isinstance(updated_transaction['created_at'], str): updated_transaction['created_at'] = datetime.fromisoformat(updated_transaction['created_at'].replace('Z', '+00:00'))
    
    # Emit WebSocket event for real-time update
    logger.info(f"About to emit transaction_updated for org: {current_user.organization_id}")
    try:
        await emit_to_organization(
            current_user.organization_id,
            'transaction_updated',
            updated_transaction
        )
        logger.info(f"Successfully emitted transaction_updated for org: {current_user.organization_id}")
    except Exception as emit_error:
        logger.error(f"Failed to emit transaction_updated: {emit_error}", exc_info=True)
    
    return updated_transaction

@api_router.delete("/transactions/{transaction_id}")
async def delete_transaction(request: Request, transaction_id: str, current_user: UserInDB = Depends(get_current_user)):
    # Sadece admin silebilir
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request); query = {"id": transaction_id, "organization_id": current_user.organization_id}
    result = await db.transactions.delete_one(query)
    if result.deleted_count == 0: raise HTTPException(status_code=404, detail="İşlem bulunamadı")
    
    # Emit WebSocket event for real-time update
    logger.info(f"About to emit transaction_deleted for org: {current_user.organization_id}")
    try:
        await emit_to_organization(
            current_user.organization_id,
            'transaction_deleted',
            {'id': transaction_id}
        )
        logger.info(f"Successfully emitted transaction_deleted for org: {current_user.organization_id}")
    except Exception as emit_error:
        logger.error(f"Failed to emit transaction_deleted: {emit_error}", exc_info=True)
    
    return {"message": "İşlem silindi"}

# === DASHBOARD STATS ===
# === PLAN ENDPOINT'LERİ ===
@api_router.get("/plans")
async def get_plans(request: Request):
    """Tüm planları getir (herkese açık) - Para birimine göre Stripe fiyatlarını çeker"""
    # Dil ve para birimi algılama (Accept-Language header'ından)
    accept_language = request.headers.get('Accept-Language', 'tr')
    lang = 'en' if accept_language.startswith('en') else 'tr'
    currency = 'gbp' if lang == 'en' else 'try'
    
    # Features çeviri mapping'i
    features_translation = {
        'tr': {
            '50 Randevu veya 7 Gün (Hangisi önce)': '50 Randevu veya 7 Gün (Hangisi önce)',
            'Randevu Hatırlatma Dahil': 'Randevu Hatırlatma Dahil',
            'Sınırsız Personel': 'Sınırsız Personel',
            'Sınırsız Müşteri': 'Sınırsız Müşteri',
            'Online Randevu': 'Online Randevu',
            'İstatistikler': 'İstatistikler',
            'Yapay Zeka Akıllı Asistan (Test)': 'Yapay Zeka Akıllı Asistan (Test)',
            '100 Randevu/Ay': '100 Randevu/Ay',
            '300 Randevu/Ay': '300 Randevu/Ay',
            '600 Randevu/Ay': '600 Randevu/Ay',
            '900 Randevu/Ay': '900 Randevu/Ay',
            '1.200 Randevu/Ay': '1.200 Randevu/Ay',
            'Sınırsız Randevu/Ay': 'Sınırsız Randevu/Ay',
            'Yapay Zeka Akıllı Asistan (Standart Kullanım)': 'Yapay Zeka Akıllı Asistan (Standart Kullanım)',
            'Yapay Zeka Akıllı Asistan (Gelişmiş Kullanım)': 'Yapay Zeka Akıllı Asistan (Gelişmiş Kullanım)',
            'Yapay Zeka Akıllı Asistan (Limitsiz)': 'Yapay Zeka Akıllı Asistan (Limitsiz)',
            'Akıllı Pazarlama Otopilotu': 'Akıllı Pazarlama Otopilotu',
        },
        'en': {
            '50 Randevu veya 7 Gün (Hangisi önce)': '50 Appointments or 7 Days (Whichever comes first)',
            'Randevu Hatırlatma Dahil': 'Appointment Reminders Included',
            'Sınırsız Personel': 'Unlimited Staff',
            'Sınırsız Müşteri': 'Unlimited Customers',
            'Online Randevu': 'Online Appointment',
            'İstatistikler': 'Statistics',
            'Yapay Zeka Akıllı Asistan (Test)': 'AI Smart Assistant (Test)',
            '100 Randevu/Ay': '100 Appointments / Monthly',
            '300 Randevu/Ay': '300 Appointments / Monthly',
            '600 Randevu/Ay': '600 Appointments / Monthly',
            '900 Randevu/Ay': '900 Appointments / Monthly',
            '1.200 Randevu/Ay': '1,200 Appointments / Monthly',
            'Sınırsız Randevu/Ay': 'Unlimited Appointments / Monthly',
            'Yapay Zeka Akıllı Asistan (Standart Kullanım)': 'AI Smart Assistant (Standard Use)',
            'Yapay Zeka Akıllı Asistan (Gelişmiş Kullanım)': 'AI Smart Assistant (Advanced Use)',
            'Yapay Zeka Akıllı Asistan (Limitsiz)': 'AI Smart Assistant (Unlimited)',
            'Akıllı Pazarlama Otopilotu': 'Smart Marketing Autopilot',
        }
    }
    
    # İlk ay %25 indirimli fiyatları hesapla
    plans_with_discount = []
    for plan in PLANS:
        if plan['id'] == 'tier_trial':
            plan_copy = plan.copy()
            # Features'ları çevir
            if 'features' in plan_copy and lang == 'en':
                plan_copy['features'] = [features_translation['en'].get(f, f) for f in plan_copy['features']]
            plans_with_discount.append(plan_copy)
            continue
        
        plan_copy = plan.copy()
        
        # Stripe'dan fiyatları çek (GBP için)
        if currency == 'gbp':
            try:
                # Monthly fiyat
                monthly_lookup_key = get_stripe_lookup_key(plan['id'], 'monthly', 'gbp')
                monthly_price = get_stripe_price_by_lookup_key(monthly_lookup_key)
                if monthly_price:
                    plan_copy['price_monthly'] = monthly_price.unit_amount / 100  # Stripe cent'ten çevir
                    logger.info(f"✅ Stripe'dan GBP monthly fiyat çekildi: {plan['id']} = {plan_copy['price_monthly']} GBP")
                else:
                    logger.warning(f"⚠️ Stripe'dan GBP monthly fiyat bulunamadı: {plan['id']}, TRY fiyatı kullanılıyor")
                
                # Yearly fiyat
                yearly_lookup_key = get_stripe_lookup_key(plan['id'], 'yearly', 'gbp')
                yearly_price = get_stripe_price_by_lookup_key(yearly_lookup_key)
                if yearly_price:
                    plan_copy['price_yearly'] = yearly_price.unit_amount / 100  # Stripe cent'ten çevir
                    logger.info(f"✅ Stripe'dan GBP yearly fiyat çekildi: {plan['id']} = {plan_copy['price_yearly']} GBP")
                else:
                    # Yearly bulunamazsa monthly'den hesapla (veya TRY fiyatını kullan)
                    if 'price_monthly' in plan_copy:
                        plan_copy['price_yearly'] = plan_copy['price_monthly'] * 10
                    else:
                        logger.warning(f"⚠️ Stripe'dan GBP yearly fiyat bulunamadı: {plan['id']}, TRY fiyatı kullanılıyor")
            except Exception as e:
                logger.warning(f"⚠️ Stripe'dan GBP fiyat çekilemedi, TRY fiyatları kullanılıyor: {e}")
                # Hata durumunda TRY fiyatlarını kullan (plan_copy zaten plan.copy() ile oluşturuldu)
        
        # Features'ları çevir
        if 'features' in plan_copy:
            plan_copy['features'] = [features_translation.get(lang, features_translation['tr']).get(f, f) for f in plan_copy['features']]
        
        # İlk ay indirimli fiyatları hesapla
        plan_copy['price_monthly_original'] = plan_copy.get('price_monthly', plan['price_monthly'])
        plan_copy['price_monthly_discounted'] = int(plan_copy['price_monthly_original'] * 0.75)  # %25 indirim
        plan_copy['discount_percentage'] = 25
        plans_with_discount.append(plan_copy)
    
    return {"plans": plans_with_discount}

@api_router.get("/plan/current")
async def get_current_plan(request: Request, current_user: UserInDB = Depends(get_current_user)):
    """Mevcut plan bilgisini getir"""
    db = await get_db_from_request(request)
    plan_doc = await get_organization_plan(db, current_user.organization_id)
    if not plan_doc:
        raise HTTPException(status_code=404, detail="Plan bilgisi bulunamadı")
    
    plan_id = plan_doc.get('plan_id', 'tier_trial')
    plan_info = await get_plan_info(plan_id)
    if not plan_info:
        raise HTTPException(status_code=404, detail="Plan bilgisi geçersiz")
    
    # Datetime'ları string'e çevir
    billing_cycle = plan_doc.get('billing_cycle', 'monthly')
    
    # quota_limit zaten database'de aylık olarak tutuluyor (lazy quota reset mekanizması ile)
    # Yıllık plan için de aylık gösterilecek, sadece badge'de "Yıllık Plan" yazacak
    quota_limit = plan_doc.get('quota_limit')
    if not quota_limit:
        # Fallback: Eğer database'de quota_limit yoksa plan_info'dan al (aylık)
        quota_limit = plan_info.get('quota_monthly_appointments', 50)
    # Plan tanımı sınırsız ise (-1) DB kaydından bağımsız olarak override et
    if plan_info.get('quota_monthly_appointments') == -1:
        quota_limit = -1
    
    result = {
        "plan_id": plan_id,
        "plan_name": plan_info.get('name'),
        "quota_usage": plan_doc.get('quota_usage', 0),
        "quota_limit": quota_limit,
        "quota_reset_date": plan_doc.get('quota_reset_date').isoformat() if isinstance(plan_doc.get('quota_reset_date'), datetime) else plan_doc.get('quota_reset_date'),
        "is_first_month": plan_doc.get('is_first_month', False),
        "trial_start_date": plan_doc.get('trial_start_date').isoformat() if plan_doc.get('trial_start_date') and isinstance(plan_doc.get('trial_start_date'), datetime) else plan_doc.get('trial_start_date'),
        "trial_end_date": plan_doc.get('trial_end_date').isoformat() if plan_doc.get('trial_end_date') and isinstance(plan_doc.get('trial_end_date'), datetime) else plan_doc.get('trial_end_date'),
        "is_trial": plan_id == 'tier_trial',
        "billing_cycle": billing_cycle,
        "is_yearly": billing_cycle == 'yearly'
    }
    
    # Trial kontrolü
    if plan_id == 'tier_trial':
        trial_end = plan_doc.get('trial_end_date')
        if isinstance(trial_end, str):
            trial_end = datetime.fromisoformat(trial_end.replace('Z', '+00:00'))
        if trial_end:
            result['trial_days_remaining'] = max(0, (trial_end - datetime.now(timezone.utc)).days)
    
    return result

@api_router.put("/plan/update")
async def update_plan(request: Request, plan_update: dict, current_user: UserInDB = Depends(get_current_user)):
    """Plan güncelle (şimdilik sadece plan_id değişikliği)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    new_plan_id = plan_update.get('plan_id')
    if not new_plan_id:
        raise HTTPException(status_code=400, detail="plan_id gerekli")
    
    plan_info = await get_plan_info(new_plan_id)
    if not plan_info:
        raise HTTPException(status_code=400, detail="Geçersiz plan_id")
    
    db = await get_db_from_request(request)
    
    # Mevcut plan bilgisini al
    plan_doc = await get_organization_plan(db, current_user.organization_id)
    
    # Yeni plana geç
    quota_reset = datetime.now(timezone.utc) + timedelta(days=30)
    # Trial dışına çıkarken is_first_month tekrar TRUE olmamalı
    is_first_month = False if new_plan_id != 'tier_trial' else False
    
    update_data = {
        "plan_id": new_plan_id,
        "quota_usage": 0,  # Yeni plana geçince sıfırla
        "quota_reset_date": quota_reset.isoformat(),
        "is_first_month": is_first_month,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Trial paketine geçiliyorsa trial tarihlerini ayarla
    if new_plan_id == 'tier_trial':
        trial_start = datetime.now(timezone.utc)
        trial_end = trial_start + timedelta(days=7)
        update_data['trial_start_date'] = trial_start.isoformat()
        update_data['trial_end_date'] = trial_end.isoformat()
    else:
        # Trial'dan çıkıyorsa trial tarihlerini temizle
        update_data['trial_start_date'] = None
        update_data['trial_end_date'] = None
    
    await db.organization_plans.update_one(
        {"organization_id": current_user.organization_id},
        {"$set": update_data}
    )
    
    return {"message": "Plan güncellendi", "plan_id": new_plan_id}

@api_router.post("/subscription/portal")
async def create_customer_portal_session(
    request: Request,
    current_user: UserInDB = Depends(get_current_user)
):
    """Stripe Customer Portal session oluştur (iptal, fatura görüntüleme vb.)"""
    try:
        if current_user.role != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        db = await get_db_from_request(request)
        
        # Mevcut plan bilgisini al
        plan_doc = await get_organization_plan(db, current_user.organization_id)
        if not plan_doc:
            raise HTTPException(status_code=404, detail="Plan bilgisi bulunamadı")
        
        current_plan_id = plan_doc.get('plan_id', 'tier_trial')
        
        # Trial'daysa portal'a gerek yok
        if current_plan_id == 'tier_trial':
            raise HTTPException(status_code=400, detail="Trial paketinde Stripe portal kullanılamaz")
        
        # Stripe customer_id'yi bul
        customer_id = None
        
        # Önce organization_plans'dan stripe_customer_id'yi kontrol et
        stripe_customer_id = plan_doc.get('stripe_customer_id')
        if stripe_customer_id:
            customer_id = stripe_customer_id
            logger.info(f"Customer ID organization_plans'dan alındı: {customer_id}")
        else:
            # Payment logs'dan bul
            try:
                payment_log = await db.payment_logs.find_one(
                    {
                        "organization_id": current_user.organization_id,
                        "status": {"$in": ["active", "completed"]}
                    },
                    sort=[("created_at", -1)]
                )
                
                if payment_log and payment_log.get('session_id'):
                    # Stripe session'dan customer_id'yi al
                    session = stripe.checkout.Session.retrieve(payment_log['session_id'])
                    customer_id = session.customer
                    logger.info(f"Customer ID payment_logs'dan alındı: {customer_id}")
            except Exception as e:
                logger.error(f"Customer ID bulunamadı: {e}")
        
        if not customer_id:
            logger.warning(f"Customer ID bulunamadı - org: {current_user.organization_id}, plan: {current_plan_id}")
            raise HTTPException(status_code=404, detail="Stripe müşteri bilgisi bulunamadı. Henüz bir ödeme yapmamış olabilirsiniz.")
        
        # Customer Portal Session oluştur
        portal_session = stripe.billing_portal.Session.create(
            customer=customer_id,
            return_url=PAYMENT_SUCCESS_URL  # Portal'dan döndükten sonra gidilecek sayfa
        )
        
        logger.info(f"Stripe Customer Portal oluşturuldu: customer={customer_id}, org={current_user.organization_id}")
        
        return {
            "portal_url": portal_session.url
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Customer Portal oluşturma hatası: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Portal oluşturulamadı: {str(e)}")

@api_router.post("/payments/create-checkout-session")
async def create_checkout_session(
    request: Request,
    plan_request: PlanUpdateRequest,
    current_user: UserInDB = Depends(get_current_user)
):
    """PayTR ödeme oturumu oluştur"""
    try:
        # Maps normalized 'lookup_key' to the ACTUAL Stripe Coupon ID
        COUPON_MAP = {
            # --- MONTHLY TRY COUPONS (Turkey) ---
            'standard_monthly_try':     'qcrphAHF',  # LAUNCH_STD_TRY (-200 TL)
            'professional_monthly_try': 'LhtiDQfU',  # LAUNCH_PRO_TRY (-250 TL)
            'premium_monthly_try':      'aMUqlGOn',  # LAUNCH_PRE_TRY (-400 TL)
            'business_monthly_try':     'SJSEK6LZ',  # LAUNCH_BUS_TRY (-500 TL)
            'enterprise_monthly_try':   'EgiWhRFI',  # LAUNCH_ENT_TRY (-600 TL)
            'corporate_monthly_try':    'z7n9rmxE',  # LAUNCH_COR_TRY (-900 TL)
            # --- MONTHLY GBP COUPONS (UK) ---
            'standard_monthly_gbp':     'JqOc3ta6',  # LAUNCH_STD_GBP (-£4)
            'professional_monthly_gbp': 'drRoM3Kc',  # LAUNCH_PRO_GBP (-£6)
            'premium_monthly_gbp':      'NFnpf2Og',  # LAUNCH_PRE_GBP (-£8)
            'business_monthly_gbp':     'XOnrysiP',  # LAUNCH_BUS_GBP (-£10)
            'enterprise_monthly_gbp':   'C6g1vI2G',  # LAUNCH_ENT_GBP (-£14)
            'corporate_monthly_gbp':    'u4jV6TAJ',  # LAUNCH_COR_GBP (-£20)
        }
        
        # Log mesajını hem console'a hem de file'a yaz
        billing_cycle = plan_request.billing_cycle or "monthly"
        log_msg = f"Payment checkout session başlatılıyor: user={current_user.username}, plan_id={plan_request.plan_id}, billing_cycle={billing_cycle}"
        logger.info(log_msg)
        print(f"[PAYMENT] {log_msg}")  # Console'a da yaz
        
        if current_user.role != "admin":
            logger.warning(f"Payment endpoint: Yetkisiz erişim denemesi - user={current_user.username}, role={current_user.role}")
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        # Stripe ayarlarını kontrol et
        if not STRIPE_SECRET_KEY:
            logger.error("STRIPE_SECRET_KEY tanımlı değil!")
            raise HTTPException(status_code=500, detail="Ödeme sistemi yapılandırılmamış")
        
        # 1. İstenen planı bul ve fiyatını al
        plan = await get_plan_info(plan_request.plan_id)
        if not plan:
            logger.error(f"Plan bulunamadı: plan_id={plan_request.plan_id}")
            raise HTTPException(status_code=404, detail="Plan bulunamadı")
        
        # Plan dict'inin gerekli alanlarını kontrol et
        if 'price_monthly' not in plan or 'name' not in plan:
            logger.error(f"Plan eksik alanlar içeriyor: plan={plan}, plan_id={plan_request.plan_id}")
            raise HTTPException(status_code=500, detail="Plan verisi eksik veya geçersiz")
        
        # Yıllık/Aylık fiyat belirleme
        is_yearly = billing_cycle == "yearly"
        price_yearly = plan.get('price_yearly', plan.get('price_monthly', 0) * 10)
        
        # Trial paketini satın alınamaz
        if plan_request.plan_id == 'tier_trial':
            raise HTTPException(status_code=400, detail="Trial paketi satın alınamaz")
        
        db = await get_db_from_request(request)

        # 2. İndirimi uygula (organizasyon bazında sadece 1 kere)
        # NOT: is_first_month flag'ine güvenme; direkt payment_logs geçmişine bak.
        old_payment = await db.payment_logs.find_one({
            "organization_id": current_user.organization_id,
            "status": {"$in": ["completed", "active"]}
        })
        is_old_customer = bool(old_payment)
        is_first_month = not is_old_customer
        
        # price_monthly değerini güvenli şekilde al
        price_monthly = plan.get('price_monthly', 0)
        if not isinstance(price_monthly, (int, float)) or price_monthly < 0:
            logger.error(f"Geçersiz price_monthly değeri: {price_monthly}, plan_id={plan_request.plan_id}")
            raise HTTPException(status_code=500, detail="Plan fiyatı geçersiz")
        
        # E-posta kontrolü
        user_email = (current_user.username or "").strip().lower()
        if not user_email or "@" not in user_email:
            logger.error(f"Geçersiz email (kullanıcı: {current_user.username}): {user_email}")
            raise HTTPException(status_code=400, detail="Geçerli bir e-posta adresi gerekli")
        
        # Para birimi algılama
        # Settings'den telefon numarasını al
        settings_doc = await db.settings.find_one({"organization_id": current_user.organization_id})
        user_phone = settings_doc.get("support_phone") if settings_doc else None
        
        currency = plan_request.currency or detect_currency(request, user_phone)
        currency_upper = currency.upper()  # 'GBP' veya 'TRY'
        
        logger.info(f"💱 Para birimi algılandı: {currency_upper} (plan_id={plan_request.plan_id}, phone={user_phone})")
        
        # Stripe Checkout Session oluştur
        try:
            # Yıllık veya aylık fiyat belirleme
            if is_yearly:
                # Yıllık abonelik - tek seferlik yıllık ödeme
                price_amount = price_yearly
                recurring_interval = 'year'
                plan_suffix = "(Yıllık)"
                # Yıllık ödemede ilk ay indirimi yok
                is_first_month = False
            else:
                # Aylık abonelik
                price_amount = price_monthly
                recurring_interval = 'month'
                plan_suffix = "(Aylık)"
            
            # Verbose logging: Raw payload
            logger.info(f"🔍 DEBUG - Raw Payload:")
            logger.info(f"   plan_id: '{plan_request.plan_id}'")
            logger.info(f"   currency: '{currency}'")
            logger.info(f"   billing_cycle: '{billing_cycle}'")
            
            # Lookup key ile Stripe'dan fiyat çekmeyi dene
            # Not: Lookup key currency içerir (örn: standard_monthly_gbp, professional_monthly_try)
            lookup_key = get_stripe_lookup_key(plan_request.plan_id, billing_cycle, currency)
            logger.info(f"🔑 Generated Lookup Key: '{lookup_key}'")
            
            stripe_price = get_stripe_price_by_lookup_key(lookup_key)
            
            # COUPON_MAP için lookup_key'i normalize et (standart -> standard)
            # Normalize: lowercase, strip whitespace, standart -> standard
            coupon_lookup_key = lookup_key.lower().strip()
            coupon_lookup_key = coupon_lookup_key.replace('standart_', 'standard_') if 'standart_' in coupon_lookup_key else coupon_lookup_key
            logger.info(f"🔑 Normalized Coupon Lookup Key: '{coupon_lookup_key}'")
            
            if stripe_price:
                # Lookup key ile fiyat bulundu, kullan
                logger.info(f"✅ Stripe lookup key ile fiyat bulundu: {lookup_key} (currency: {currency})")
                price = stripe_price
            else:
                # Lookup key ile bulunamadı, dinamik oluştur (fallback)
                logger.warning(f"⚠️ Stripe lookup key bulunamadı: {lookup_key}, dinamik fiyat oluşturuluyor")
                
                # Para birimine göre fiyat dönüşümü - Gerçek kur API kullan
                if currency == 'gbp':
                    # GBP fiyatları (plan'dan alınan TRY fiyatlarını GBP'ye çevir)
                    # Gerçek kur API'sinden kur bilgisi al
                    exchange_rate = await get_exchange_rate('TRY', 'GBP')
                    price_amount_gbp = price_amount * exchange_rate
                    price_kurus = int(price_amount_gbp * 100)  # GBP için pence
                    logger.info(f"💱 Kur dönüşümü: {price_amount} TRY × {exchange_rate:.6f} = {price_amount_gbp:.2f} GBP")
                else:
                    # TRY fiyatları
                    price_kurus = int(price_amount * 100)  # TRY için kuruş
                
                # Stripe Price objesi oluştur (fallback - lookup key kullanmıyoruz)
                # Not: lookup_key kullanmıyoruz çünkü Stripe'da zaten o key'e sahip fiyatlar olabilir
                # Bu dinamik fiyatlar geçici/fallback amaçlı, lookup key'e gerek yok
                price = stripe.Price.create(
                    currency=currency,
                    unit_amount=price_kurus,
                    recurring={'interval': recurring_interval},
                    product_data={
                        'name': f'PLANN {plan.get("name", "Plan")} {plan_suffix}'
                    }
                    # lookup_key kullanmıyoruz - çakışma önlemek için
                )
                logger.info(f"✅ Dinamik Stripe fiyat oluşturuldu (fallback): currency={currency}, amount={price_kurus}, lookup_key={lookup_key} (kullanılmadı)")
            
            logger.info(f"💰 Fiyat Hesaplama - Plan: {plan_request.plan_id}")
            logger.info(f"💰 Billing Cycle: {billing_cycle}")
            logger.info(f"💰 Currency: {currency_upper}")
            logger.info(f"💰 Fiyat: {price_amount} ({currency_upper}) ({recurring_interval})")
            logger.info(f"💰 is_first_month: {is_first_month}")
            
            # COUPON_MAP'ten coupon ID al (sadece aylık planlar için)
            coupon_id = None
            if (not is_yearly) and (not is_old_customer):
                # Explicit key check with verbose logging
                logger.info(f"🔍 Checking COUPON_MAP for key: '{coupon_lookup_key}'")
                logger.info(f"📋 Available Keys in COUPON_MAP: {list(COUPON_MAP.keys())}")
                
                if coupon_lookup_key in COUPON_MAP:
                    coupon_id = COUPON_MAP[coupon_lookup_key]
                    logger.info(f"✅ Coupon found: '{coupon_id}' for key '{coupon_lookup_key}'")
                else:
                    error_msg = f"❌ CRITICAL MISMATCH: Key '{coupon_lookup_key}' is NOT in COUPON_MAP."
                    logger.error(error_msg)
                    logger.error(f"📋 Available Keys in Map: {list(COUPON_MAP.keys())}")
                    logger.error(f"🔑 Original Lookup Key: '{lookup_key}'")
                    logger.error(f"🔑 Normalized Key: '{coupon_lookup_key}'")
                    raise ValueError(f"{error_msg} Available keys: {list(COUPON_MAP.keys())}")
            
            # Checkout Session oluştur
            # Native app için deep link, web için normal URL
            is_native = plan_request.platform in ['android', 'ios']
            
            # Domain algılama: Host/Origin/Referer üzerinden domain'i çıkar
            domain = _get_domain_from_request(request)
            
            success_url = PAYMENT_SUCCESS_URL_NATIVE if is_native else f"https://{domain}/?session_id={{CHECKOUT_SESSION_ID}}#/payment-success"
            cancel_url = PAYMENT_CANCEL_URL_NATIVE if is_native else f"https://{domain}/#/subscribe"
            
            logger.info(f"🌐 Domain algılandı: {domain}, cancel_url: {cancel_url}, success_url: {success_url}")
            
            session_params = {
                'payment_method_types': ['card'],
                'mode': 'subscription',
                'line_items': [{
                    'price': price.id,
                    'quantity': 1,
                }],
                'customer_email': user_email,
                'metadata': {
                    'user_id': current_user.organization_id,
                    'plan_id': plan_request.plan_id,
                    'organization_id': current_user.organization_id,
                    'is_first_month': str(is_first_month),
                    'billing_cycle': billing_cycle,
                    'currency': currency,
                    'appointment_limit': str(plan.get('quota_monthly_appointments', 0))  # Metadata'ya appointment_limit ekle
                },
                'success_url': success_url,
                'cancel_url': cancel_url,
                'billing_address_collection': 'required',
            }
            
            # COUPON_MAP'ten gelen coupon'u uygula (sadece monthly planlar için)
            if coupon_id:
                session_params['discounts'] = [{'coupon': coupon_id}]
                logger.info(f"🎁 COUPON_MAP coupon uygulandı: {coupon_id} (lookup_key: {coupon_lookup_key})")
            else:
                # Coupon yoksa promotion code'a izin ver
                session_params['allow_promotion_codes'] = True
            
            try:
                session = stripe.checkout.Session.create(**session_params)
            except stripe.error.StripeError as stripe_err:
                # Stripe hatası durumunda detaylı log
                logger.error(f"❌ Stripe Checkout Session Creation Failed:")
                logger.error(f"   Error Type: {type(stripe_err).__name__}")
                logger.error(f"   Error Message: {str(stripe_err)}")
                if coupon_id:
                    logger.error(f"   Attempted Coupon ID: '{coupon_id}'")
                    logger.error(f"   ⚠️ Coupon '{coupon_id}' might not exist in Stripe!")
                logger.error(f"   Session Params: {session_params}")
                # Re-raise the error so it's properly handled
                raise
            
            # Payment log oluştur
            if is_yearly:
                actual_amount = price_yearly
                original_amount = price_yearly
            else:
                actual_amount = price_monthly * 0.75 if is_first_month else price_monthly
                original_amount = price_monthly
            
            await db.payment_logs.insert_one({
                "session_id": session.id,
                "organization_id": current_user.organization_id,
                "user_id": current_user.username,
                "plan_id": plan_request.plan_id,
                "status": "pending",
                "amount": actual_amount,
                "original_amount": original_amount,
                "currency": currency_upper,
                "is_first_month": is_first_month,
                "discount_applied": is_first_month and not is_yearly,
                "billing_cycle": billing_cycle,
                "payment_provider": "stripe",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            
            logger.info(f"Stripe checkout session oluşturuldu: {session.id} - {plan_request.plan_id}")
            
            return {
                "checkout_url": session.url,
                "session_id": session.id
            }
            
        except stripe.error.StripeError as e:
            logger.error(f"Stripe hatası: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Ödeme oturumu oluşturulamadı: {str(e)}")
        except Exception as e:
            logger.error(f"Stripe checkout session oluşturma hatası: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Sunucu hatası: {str(e)}")
        
        # 3. Kullanıcı IP Adresini Al
        user_ip = request.client.host if request.client else "127.0.0.1"
        
        # 4. Benzersiz Sipariş ID'si Oluştur (Sadece alfanumerik karakterler)
        # PayTR özel karakter kabul etmez, organization_id'deki tireleri kaldır
        org_id_clean = current_user.organization_id.replace('-', '')
        timestamp_str = str(int(datetime.now(timezone.utc).timestamp()))
        merchant_oid = f"PLANN{org_id_clean}{timestamp_str}"
        
        # 5. BASİTLEŞTİRİLMİŞ KULLANICI BİLGİLERİ (None-Safe)
        
        # E-posta (Temel Kontrol)
        user_email = (current_user.username or "").strip().lower()
        if not user_email or "@" not in user_email or not re.match(r'^[a-zA-Z0-9._+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', user_email):
            logger.error(f"Geçersiz email (kullanıcı: {current_user.username}): {user_email}")
            raise HTTPException(status_code=400, detail="Geçerli bir e-posta adresi gerekli")
        
        user_name = (current_user.full_name or user_email.split('@')[0]).strip()[:50]
        
        # Adres ve Telefon (None-Safe Kontrol)
        settings = await db.settings.find_one({"organization_id": current_user.organization_id})
        
        user_address = "Adres Bilgisi Yok"  # PayTR boş kabul etmez
        user_phone = "05000000000"  # PayTR için geçerli bir varsayılan
        
        if settings:
            # DB'den gelen veriyi al ve 'str' olduğuna emin ol
            user_address_raw = str(settings.get("address", "")).strip()
            user_phone_raw = str(settings.get("support_phone", "")).strip()
            
            # Placeholder (TODO) metinlerini ve boşlukları temizle
            if user_address_raw and "Müşteri Adresi" not in user_address_raw:
                user_address = user_address_raw
            if user_phone_raw and "Müşteri Telefonu" not in user_phone_raw and len(user_phone_raw) >= 10:
                user_phone = user_phone_raw
        
        # 6. Sepet Bilgisi
        plan_name = plan.get('name', 'Plan')
        user_basket = base64.b64encode(json.dumps([
            [plan_name, str(price_to_pay), 1]
        ]).encode('utf-8')).decode('utf-8')  # UTF-8 kullan
        
        # 7. PayTR Ek Parametreleri
        no_installment = '1'  # Taksit yok
        max_installment = '0'  # Max taksit sayısı
        currency = 'TL'
        test_mode = '0'  # Production mode (1 = test mode)
        debug_on = '1'  # Debug açık
        timeout_limit = '30'  # Timeout süresi (dakika)
        store_card = '1'  # Kart tokenize et (recurring payment için)
        
        # 8. PAYTR TOKEN (HASH) OLUŞTURMA
        # Doğru sıra: merchant_id + user_ip + merchant_oid + email + payment_amount + user_basket + no_installment + max_installment + currency + test_mode
        # NOT: store_card hash'e dahil DEĞİL, sadece post_data'ya eklenir
        hash_str = f"{PAYTR_MERCHANT_ID}{user_ip}{merchant_oid}{user_email}{payment_amount_kurus}{user_basket}{no_installment}{max_installment}{currency}{test_mode}"
        
        paytr_token = base64.b64encode(hmac.new(
            PAYTR_MERCHANT_KEY.encode('utf-8'), 
            hash_str.encode('utf-8') + PAYTR_MERCHANT_SALT.encode('utf-8'), 
            hashlib.sha256
        ).digest()).decode('utf-8')
        
        # 9. PAYTR API'SİNE İSTEK GÖNDERME
        logger.info(f"PayTR için email: '{user_email}' (len: {len(user_email)}), user_ip: '{user_ip}'")
        
        post_data = {
            'merchant_id': PAYTR_MERCHANT_ID,
            'user_ip': user_ip,  # Zorunlu parametre
            'merchant_oid': merchant_oid,
            'email': user_email,  # 'user_email' değil, 'email' kullanılmalı!
            'payment_amount': payment_amount_kurus,
            'paytr_token': paytr_token,
            'user_basket': user_basket,
            'debug_on': debug_on,
            'no_installment': no_installment,
            'max_installment': max_installment,
            'user_name': user_name,
            'user_address': user_address[:400],  # Limiti aşmadığından emin ol
            'user_phone': user_phone[:20],
            'merchant_ok_url': PAYTR_SUCCESS_URL,
            'merchant_fail_url': PAYTR_FAIL_URL,
            'timeout_limit': timeout_limit,
            'currency': currency,
            'test_mode': test_mode,
            'store_card': store_card  # Kart saklama
        }
        
        # PayTR için kritik alanları logla
        logger.info(f"PayTR post_data hazırlandı: email='{post_data['email']}', user_ip='{post_data['user_ip']}', user_address='{post_data['user_address'][:50]}...', user_phone='{post_data['user_phone']}'")
        
        try:
            # PayTR'a isteği gönder - data parametresi form-urlencoded formatında gönderir
            response = requests.post(PAYTR_API_URL, data=post_data, timeout=10)
            logger.info(f"PayTR API HTTP Status: {response.status_code}")
            logger.info(f"PayTR API Response Text: {response.text}")
            
            # Debug: Gönderilen request body'yi logla
            if hasattr(response, 'request') and hasattr(response.request, 'body'):
                logger.info(f"PayTR Request Body (ilk 500 karakter): {str(response.request.body)[:500]}")
            
            # HTTP hata kontrolü
            try:
                response.raise_for_status()
            except requests.exceptions.HTTPError as e:
                logger.error(f"PayTR API HTTP Hatası: {e}")
                logger.error(f"Response text: {response.text}")
                raise HTTPException(status_code=503, detail="Ödeme servisi şu an hizmet veremiyor.")
            
            # PayTR'den gelen yanıtı parse et
            try:
                res_data = response.json()
            except Exception as e:
                logger.error(f"PayTR yanıtı JSON parse edilemedi: {e}")
                logger.error(f"Response text: {response.text}")
                raise HTTPException(status_code=500, detail="Ödeme servisi yanıtı işlenemedi")
            
            logger.info(f"PayTR API yanıtı (JSON): {res_data}")
            
            # PayTR'den gelen hata detaylarını logla (status: 'failed' durumu)
            if res_data.get('status') != 'success':
                error_reason = res_data.get('reason', 'Bilinmeyen hata')
                error_details = res_data.get('errors', {})
                
                logger.error(f"PayTR HATA DETAYLARI: reason={error_reason}, errors={error_details}")
                logger.error(f"PayTR Full Response: {res_data}")
                
                if 'email' in str(error_details).lower() or 'user_email' in str(error_details).lower() or 'email' in str(error_reason).lower():
                    logger.error(f"EMAIL HATASI TESPİT EDİLDİ!")
                    logger.error(f"Gönderilen email: '{user_email}'")
                    logger.error(f"Email uzunluk: {len(user_email)}")
                    logger.error(f"Email bytes: {user_email.encode('utf-8')}")
                    logger.error(f"Email repr: {repr(user_email)}")
                    logger.error(f"Email karakterler (ilk 20): {[ord(c) for c in user_email[:20]]}")
                    
                    # Email'i hex formatında göster
                    logger.error(f"Email hex: {user_email.encode('utf-8').hex()}")
                    
                    # PayTR post_data'daki email'i göster
                    logger.error(f"post_data['user_email']: '{post_data.get('user_email')}'")
                    logger.error(f"post_data['user_email'] type: {type(post_data.get('user_email'))}")
                    
                    # Hata mesajını oluştur
                    if error_details:
                        error_details_list = []
                        for field, message in error_details.items():
                            error_details_list.append(f"{field}: {message}")
                        error_msg = f"{error_reason} - {', '.join(error_details_list)}"
                    else:
                        error_msg = error_reason
                    
                    logger.error(f"PayTR hatası nedeniyle HTTPException fırlatılıyor: {error_msg}")
                    raise HTTPException(status_code=500, detail=f"Ödeme başlatılamadı: {error_msg}")
            
            if res_data.get('status') == 'success':
                # 7. BAŞARILI: ÖDEME LİNKİNİ (TOKEN) AL
                paytr_iframe_token = res_data.get('token')
                
                if not paytr_iframe_token:
                    logger.error(f"PayTR başarılı ama token yok: {res_data}")
                    raise HTTPException(status_code=500, detail="PayTR token alınamadı")
                
                # Bu link, frontend'i PayTR ödeme sayfasına yönlendirecek
                checkout_url = f"https://www.paytr.com/odeme/guvenli/{paytr_iframe_token}"
                
                # 8. 'merchant_oid'yi veritabanına 'pending' olarak kaydet
                await db.payment_logs.insert_one({
                    "merchant_oid": merchant_oid,
                    "organization_id": current_user.organization_id,
                    "user_id": current_user.username,
                    "plan_id": plan_request.plan_id,
                    "status": "pending",
                    "amount": price_to_pay,
                    "amount_kurus": payment_amount_kurus,
                    "is_first_month": is_first_month,
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
                
                logger.info(f"PayTR checkout session oluşturuldu: {merchant_oid} - {plan_request.plan_id}")
                return {"checkout_url": checkout_url}
            else:
                # PayTR token oluşturamadı
                # PayTR hata mesajını düzgün çıkar
                error_reason = res_data.get('reason', 'Bilinmeyen hata')
                errors = res_data.get('errors', {})
                
                # Eğer errors dict'i varsa, içindeki hataları birleştir
                if errors:
                    error_details = []
                    for field, message in errors.items():
                        error_details.append(f"{field}: {message}")
                    error_msg = f"{error_reason} - {', '.join(error_details)}"
                else:
                    error_msg = error_reason
                
                logger.error(f"PayTR token alma hatası: reason={error_reason}, errors={errors}, full_response={res_data}")
                raise HTTPException(status_code=500, detail=f"Ödeme başlatılamadı: {error_msg}")
        except requests.exceptions.RequestException as e:
            logger.error(f"PayTR API isteği hatası: {e}", exc_info=True)
            raise HTTPException(status_code=503, detail="Ödeme servisi şu an hizmet veremiyor.")
    
    except HTTPException:
        # HTTPException'ları tekrar fırlat (zaten doğru şekilde işlenmiş)
        raise
    except AttributeError as e:
        # NoneType hatası veya eksik attribute hatası
        logger.error(f"create_checkout_session İÇİNDE AttributeError: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Veri hatası: {str(e)}")
    except Exception as e:
        # BU, TÜM DİĞER (AttributeError vb.) ÇÖKMELERİ YAKALAR
        logger.error(f"create_checkout_session İÇİNDE BEKLENMEDİK HATA: {e}", exc_info=True)
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Sunucu hatası: {str(e)}")


def _stripe_obj_get(obj, key, default=None):
    """StripeObject [] destekler; .get() yok. Dict ile uyumlu erişim."""
    if obj is None:
        return default
    if isinstance(obj, dict):
        return obj.get(key, default)
    try:
        val = obj[key]
        return default if val is None else val
    except (KeyError, TypeError):
        pass
    try:
        val = getattr(obj, key, None)
        return default if val is None else val
    except Exception:
        return default


def _stripe_metadata_dict(stripe_obj):
    """Session/Price/Product vb. üzerindeki metadata alanını düz dict yap."""
    meta = _stripe_obj_get(stripe_obj, "metadata", None)
    if meta is None:
        return {}
    if isinstance(meta, dict):
        return meta
    try:
        return dict(meta)
    except Exception:
        try:
            keys = meta.keys() if hasattr(meta, "keys") else []
            return {k: meta[k] for k in keys}
        except Exception:
            return {}


def _stripe_line_items_data(obj):
    """Checkout Session line_items → data listesi."""
    raw = _stripe_obj_get(obj, "line_items", None)
    if raw is None:
        return []
    if isinstance(raw, list):
        return raw
    inner = _stripe_obj_get(raw, "data", None)
    if isinstance(inner, list):
        return inner
    if isinstance(raw, dict) and "data" in raw:
        d = raw["data"]
        return d if isinstance(d, list) else []
    return []


def _stripe_invoice_lines(inv):
    """Invoice lines.data listesi (StripeObject/dict uyumlu)."""
    lines = _stripe_obj_get(inv, "lines", None)
    data = _stripe_obj_get(lines, "data", None) if lines is not None else None
    return data if isinstance(data, list) else []


def _stripe_invoice_subscription_id(inv):
    """invoice → subscription id.

    Stripe eski API'de `invoice.subscription` (top-level) vardı. 2025-11 sonrası
    API sürümlerinde bu alan KALDIRILDI; subscription bilgisi artık
    `parent.subscription_details.subscription` veya
    `lines.data[].parent.subscription_item_details.subscription` altında.
    Tüm sürümlerle uyum için sırayla dener.
    """
    sub_id = _stripe_obj_get(inv, "subscription", None)
    if sub_id:
        return sub_id
    parent = _stripe_obj_get(inv, "parent", None)
    if parent is not None:
        sd = _stripe_obj_get(parent, "subscription_details", None)
        sub_id = _stripe_obj_get(sd, "subscription", None) if sd is not None else None
        if sub_id:
            return sub_id
    for ln in _stripe_invoice_lines(inv):
        lp = _stripe_obj_get(ln, "parent", None)
        sid = _stripe_obj_get(lp, "subscription_item_details", None) if lp is not None else None
        sub_id = _stripe_obj_get(sid, "subscription", None) if sid is not None else None
        if sub_id:
            return sub_id
    return None


def _stripe_invoice_period_end(inv):
    """invoice → yeni dönem sonu (unix ts).

    Yeni API'de güvenilir dönem sonu `lines.data[].period.end` içinde; birden
    fazla satır varsa (proration vb.) en ileri tarih alınır. Bulunamazsa
    top-level `period_end` (eski API) fallback.
    """
    best = None
    for ln in _stripe_invoice_lines(inv):
        per = _stripe_obj_get(ln, "period", None)
        end = _stripe_obj_get(per, "end", None) if per is not None else None
        if end:
            try:
                end_int = int(end)
                best = end_int if best is None else max(best, end_int)
            except (TypeError, ValueError):
                pass
    if best is not None:
        return best
    return _stripe_obj_get(inv, "period_end", None)


class StripeConfirmCheckoutRequest(BaseModel):
    session_id: str


@api_router.post("/payments/confirm-checkout-session")
async def confirm_stripe_checkout_session(
    request: Request,
    body: StripeConfirmCheckoutRequest,
    current_user: UserInDB = Depends(get_current_user)
):
    try:
        if current_user.role != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")

        if not STRIPE_SECRET_KEY:
            logger.error("STRIPE_SECRET_KEY tanımlı değil!")
            raise HTTPException(status_code=500, detail="Ödeme sistemi yapılandırılmamış")

        session_id = (body.session_id or "").strip()
        if not session_id:
            raise HTTPException(status_code=400, detail="session_id gerekli")

        logger.info(f"🔁 Confirm checkout session requested: session_id={session_id}, org={current_user.organization_id}")

        session = stripe.checkout.Session.retrieve(
            session_id,
            expand=['line_items.data.price.product']
        )

        payment_status = _stripe_obj_get(session, 'payment_status')
        status_value = _stripe_obj_get(session, 'status')
        if payment_status not in ('paid', 'no_payment_required') and status_value != 'complete':
            raise HTTPException(status_code=400, detail=f"Ödeme tamamlanmamış (payment_status={payment_status}, status={status_value})")

        metadata = _stripe_metadata_dict(session)
        organization_id = metadata.get('organization_id') or metadata.get('user_id')
        if organization_id and organization_id != current_user.organization_id:
            raise HTTPException(status_code=403, detail="session organization uyumsuz")
        organization_id = current_user.organization_id

        db = await get_db_from_request(request)

        payment_log = await db.payment_logs.find_one({"session_id": session_id})
        if payment_log and payment_log.get("status") == "active":
            try:
                plan_id_existing = payment_log.get('plan_id') or (metadata.get('plan_id') if isinstance(metadata, dict) else None)
                plan_data_existing = await get_plan_info(plan_id_existing) if plan_id_existing else None
                plan_name_existing = (plan_data_existing or {}).get('name') or plan_id_existing or ''
                billing_cycle_existing = payment_log.get('billing_cycle') or (metadata.get('billing_cycle', 'monthly') if isinstance(metadata, dict) else 'monthly')
                await _send_subscription_email_once(
                    db=db,
                    request=request,
                    organization_id=organization_id,
                    session_id=session_id,
                    plan_name=plan_name_existing,
                    billing_cycle=billing_cycle_existing,
                    payment_log=payment_log,
                )
            except Exception as e:
                logger.warning(f"Subscription email (already processed) gönderilemedi: {e}")
            return {"status": "ok", "already_processed": True}

        plan_id = None
        if payment_log:
            plan_id = payment_log.get('plan_id')
        if not plan_id:
            plan_id = metadata.get('plan_id')
        if not plan_id:
            raise HTTPException(status_code=400, detail="Plan bilgisi bulunamadı")

        plan_data = await get_plan_info(plan_id)
        if not plan_data:
            raise HTTPException(status_code=400, detail="Plan bulunamadı")

        billing_cycle = metadata.get('billing_cycle', 'monthly')

        appointment_limit = None
        quota_limit = None
        try:
            line_items = _stripe_line_items_data(session)

            if line_items and len(line_items) > 0:
                price_obj = _stripe_obj_get(line_items[0], 'price')
                if price_obj:
                    price_metadata = _stripe_metadata_dict(price_obj)
                    appointment_limit = price_metadata.get('appointment_limit')
                    if not appointment_limit:
                        product_obj = _stripe_obj_get(price_obj, 'product')
                        if isinstance(product_obj, dict):
                            product_metadata = product_obj.get('metadata', {})
                            appointment_limit = product_metadata.get('appointment_limit')
                        elif isinstance(product_obj, str):
                            try:
                                product_retrieved = stripe.Product.retrieve(product_obj)
                                product_metadata = _stripe_metadata_dict(product_retrieved)
                                appointment_limit = product_metadata.get('appointment_limit')
                            except Exception as e:
                                logger.warning(f"⚠️ Product retrieve hatası: {e}")

                    if appointment_limit:
                        try:
                            quota_limit = int(appointment_limit)
                        except (ValueError, TypeError):
                            quota_limit = None

            if not quota_limit:
                appointment_limit = metadata.get('appointment_limit')
                if appointment_limit:
                    try:
                        quota_limit = int(appointment_limit)
                    except (ValueError, TypeError):
                        quota_limit = None
            if not quota_limit:
                quota_limit = plan_data.get('quota_monthly_appointments', 100)
        except Exception as e:
            logger.error(f"❌ Confirm: appointment_limit alınırken hata: {e}")
            quota_limit = plan_data.get('quota_monthly_appointments', 100)

        now = datetime.now(timezone.utc)
        if billing_cycle == 'yearly':
            quota_reset = now + timedelta(days=365)
        else:
            quota_reset = now + timedelta(days=30)

        update_data = {
            "plan_id": plan_id,
            "quota_usage": 0,
            "quota_limit": quota_limit,
            "quota_reset_date": quota_reset.isoformat(),
            "quota_last_reset_date": now.isoformat(),
            "is_first_month": False,
            "billing_cycle": billing_cycle,
            "updated_at": now.isoformat(),
            "trial_start_date": None,
            "trial_end_date": None,
        }

        subscription_id = _stripe_obj_get(session, 'subscription')
        customer_id = _stripe_obj_get(session, 'customer')
        if subscription_id:
            update_data['stripe_subscription_id'] = subscription_id
            update_data['stripe_customer_id'] = customer_id
            update_data['next_billing_date'] = quota_reset.isoformat()

        await db.organization_plans.update_one(
            {"organization_id": organization_id},
            {"$set": update_data},
            upsert=True
        )

        # Gerçek Stripe tahsilat tutarını al (kuruş/cent → TL/GBP)
        stripe_amount_total = _stripe_obj_get(session, 'amount_total')
        payment_log_update = {
                "status": "active",
                "completed_at": now.isoformat(),
                "subscription_id": subscription_id,
                "stripe_customer_id": customer_id,
                "stripe_subscription_id": subscription_id
        }
        if stripe_amount_total is not None:
            try:
                payment_log_update["amount"] = round(int(stripe_amount_total) / 100.0, 2)
            except (TypeError, ValueError):
                pass

        await db.payment_logs.update_one(
            {"session_id": session_id},
            {"$set": payment_log_update},
            upsert=True
        )

        try:
            updated_payment_log = await db.payment_logs.find_one({"session_id": session_id})
            await _send_subscription_email_once(
                db=db,
                request=request,
                organization_id=organization_id,
                session_id=session_id,
                plan_name=plan_data.get('name') if isinstance(plan_data, dict) else plan_id,
                billing_cycle=billing_cycle,
                payment_log=updated_payment_log,
            )
        except Exception as e:
            logger.warning(f"Subscription email gönderilemedi: {e}")

        return {"status": "ok", "plan_id": plan_id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Confirm checkout session hatası: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Confirm işlemi başarısız")

@api_router.post("/webhook/stripe")
async def handle_stripe_webhook(request: Request):
    """Stripe webhook - Ödeme başarılı olduğunda çağrılır"""
    # Stripe webhook secret kontrolü
    if not STRIPE_WEBHOOK_SECRET:
        logger.error("STRIPE_WEBHOOK_SECRET tanımlı değil!")
        return Response(content="ERROR", status_code=500)
    
    try:
        # Stripe webhook signature doğrulaması
        payload = await request.body()
        sig_header = request.headers.get('stripe-signature')
        
        if not sig_header:
            logger.warning("Stripe webhook: stripe-signature header eksik")
            return Response(content="ERROR", status_code=400)
        
        try:
            event = stripe.Webhook.construct_event(
                payload, sig_header, STRIPE_WEBHOOK_SECRET
            )
        except ValueError as e:
            logger.error(f"Stripe webhook: Invalid payload - {e}")
            return Response(content="ERROR", status_code=400)
        except stripe.error.SignatureVerificationError as e:
            logger.error(f"Stripe webhook: Invalid signature - {e}")
            return Response(content="ERROR", status_code=400)
        
        # === SIGNATURE DOĞRULANDI, WEBHOOK GÜVENLİ ===
        logger.info(f"✅ Stripe webhook event: {event['type']}")
        
        db = await get_db_from_request(request)
        
        # checkout.session.completed - Ödeme başarılı
        if event['type'] == 'checkout.session.completed':
            session = event['data']['object']
            session_id = session['id']
            
            logger.info(f"💳 Ödeme başarılı: session_id={session_id}")
            # SaaS aboneliği değil; business (merchant) ödemesi ayrı webhook'ta işlenir.
            # payment_type=business VEYA legacy is_merchant_payment VEYA mode==payment → atla.
            # (Abonelik akışı mode==subscription + payment_type=subscription ile buradan devam eder.)
            _sess_meta = _stripe_metadata_dict(session)
            if (
                _sess_meta.get("payment_type") == "business"
                or _sess_meta.get("is_merchant_payment") == "true"
                or _stripe_obj_get(session, "mode") == "payment"
            ):
                return Response(content="OK", status_code=200)
            
            # Payment log'u bul
            payment_log = await db.payment_logs.find_one({"session_id": session_id})
            
            if not payment_log:
                metadata = _stripe_metadata_dict(session)
                plan_id_meta = metadata.get('plan_id')
                organization_id_meta = metadata.get('organization_id') or metadata.get('user_id')
                if not plan_id_meta or not organization_id_meta:
                    logger.error(f"Webhook hatası: session_id={session_id} payment_logs kaydı yok ve metadata eksik (plan_id/org_id).")
                    return Response(content="OK", status_code=200)

                billing_cycle_meta = metadata.get('billing_cycle', 'monthly')
                customer_email = _stripe_obj_get(session, 'customer_email')
                now_iso = datetime.now(timezone.utc).isoformat()
                await db.payment_logs.update_one(
                    {"session_id": session_id},
                    {
                        "$setOnInsert": {
                            "session_id": session_id,
                            "organization_id": organization_id_meta,
                            "user_id": customer_email,
                            "plan_id": plan_id_meta,
                            "status": "pending",
                            "payment_provider": "stripe",
                            "created_at": now_iso,
                        },
                        "$set": {
                            "stripe_customer_id": _stripe_obj_get(session, 'customer'),
                            "stripe_subscription_id": _stripe_obj_get(session, 'subscription'),
                            "billing_cycle": billing_cycle_meta,
                            "currency": (_stripe_obj_get(session, 'currency') or '').upper() or None,
                            "updated_at": now_iso,
                        }
                    },
                    upsert=True
                )
                payment_log = await db.payment_logs.find_one({"session_id": session_id})
                logger.warning(f"Webhook: payment_logs kaydı yoktu, metadata ile upsert edildi. session_id={session_id}, org={organization_id_meta}, plan={plan_id_meta}")
            
            if payment_log.get("status") == "active":
                logger.info(f"Webhook: {session_id} zaten işlenmiş.")
                return Response(content="OK", status_code=200)
            
            # 4. ABONELİĞİ GÜNCELLE (Kritik Mantık)
            plan_id = payment_log.get("plan_id")
            organization_id = payment_log.get("organization_id")
            
            plan_data = await get_plan_info(plan_id)
            if not plan_data:
                logger.error(f"Webhook hatası: Plan {plan_id} bulunamadı.")
                return Response(content="OK", status_code=200)
            
            # Mevcut plan bilgisini al
            plan_doc = await get_organization_plan(db, organization_id)
            
            # Billing cycle'ı metadata'dan al
            metadata = _stripe_metadata_dict(session)
            billing_cycle = metadata.get('billing_cycle', 'monthly')
            
            # TASK 1: appointment_limit'i Price Object'in metadata'sından al (Single Source of Truth)
            # Stripe session'da line_items var, onun içindeki price'ın metadata'sına bakmalıyız
            appointment_limit = None
            quota_limit = None
            
            try:
                # TASK 1: Price Object'in metadata'sından appointment_limit almak için
                # Stripe session'da line_items expand edilmemiş olabilir, retrieve et
                # Product metadata fallback için product'ı da expand et
                session_expanded = stripe.checkout.Session.retrieve(
                    session_id,
                    expand=['line_items.data.price.product']
                )
                
                line_items = _stripe_line_items_data(session_expanded)
                
                # İlk line item'dan price'ı al
                if line_items and len(line_items) > 0:
                    price_obj = _stripe_obj_get(line_items[0], 'price')
                    if price_obj:
                        # FALLBACK STRATEGY: First check price.metadata, then product.metadata
                        # Step 1: Check Price metadata
                        price_metadata = _stripe_metadata_dict(price_obj)
                        appointment_limit = price_metadata.get('appointment_limit')
                        
                        # Step 2: If not found in price, check product metadata
                        if not appointment_limit:
                            product_obj = _stripe_obj_get(price_obj, 'product')
                            if product_obj:
                                # product_obj can be a string (ID) or dict (expanded)
                                if isinstance(product_obj, dict):
                                    product_metadata = product_obj.get('metadata', {})
                                    appointment_limit = product_metadata.get('appointment_limit')
                                    if appointment_limit:
                                        logger.info(f"✅ Appointment limit Product metadata'dan alındı (fallback): {appointment_limit}")
                                elif isinstance(product_obj, str):
                                    # Product not expanded, retrieve it
                                    try:
                                        product_retrieved = stripe.Product.retrieve(product_obj)
                                        product_metadata = _stripe_metadata_dict(product_retrieved)
                                        appointment_limit = product_metadata.get('appointment_limit')
                                        if appointment_limit:
                                            logger.info(f"✅ Appointment limit Product metadata'dan alındı (fallback, retrieved): {appointment_limit}")
                                    except Exception as e:
                                        logger.warning(f"⚠️ Product retrieve hatası: {e}")
                        
                        if appointment_limit:
                            try:
                                quota_limit = int(appointment_limit)
                                source = "Price metadata" if price_metadata.get('appointment_limit') else "Product metadata"
                                logger.info(f"✅ Appointment limit {source}'dan alındı: {quota_limit} (Monthly Capacity)")
                            except (ValueError, TypeError):
                                logger.warning(f"⚠️ Metadata'daki appointment_limit geçersiz: {appointment_limit}")
                        else:
                            logger.warning(f"⚠️ Price ve Product metadata'da appointment_limit bulunamadı. Price ID: {_stripe_obj_get(price_obj, 'id')}")
                
                if not quota_limit:
                    # Fallback: Session metadata'dan dene
                    appointment_limit = metadata.get('appointment_limit')
                    if appointment_limit:
                        try:
                            quota_limit = int(appointment_limit)
                            logger.info(f"✅ Appointment limit session metadata'dan alındı: {quota_limit}")
                        except (ValueError, TypeError):
                            pass
                
                # Son fallback: Plan'dan al
                if not quota_limit:
                    quota_limit = plan_data.get('quota_monthly_appointments', 100)
                    logger.warning(f"⚠️ Appointment limit Price metadata'dan alınamadı, plan'dan alındı: {quota_limit}")
                    
            except Exception as e:
                logger.error(f"❌ Price metadata'dan appointment_limit alınırken hata: {e}")
                # Fallback: Plan'dan al
                quota_limit = plan_data.get('quota_monthly_appointments', 100)
                logger.warning(f"⚠️ Hata nedeniyle appointment limit plan'dan alındı: {quota_limit}")
            
            # CRITICAL RULE: quota_limit metadata'daki değer aynen kullanılır (Monthly Capacity)
            # Yearly planlar için 12 ile çarpılmaz - her ay aynı limit yenilenir
            logger.info(f"📊 Quota limit kaydediliyor: {quota_limit} (Monthly Capacity, billing_cycle={billing_cycle})")
            
            # Yeni plana geç - yıllık ise 365 gün, aylık ise 30 gün
            if billing_cycle == 'yearly':
                quota_reset = datetime.now(timezone.utc) + timedelta(days=365)
            else:
                quota_reset = datetime.now(timezone.utc) + timedelta(days=30)
            
            # quota_last_reset_date'i şimdi olarak ayarla (lazy reset için)
            now = datetime.now(timezone.utc)
            
            update_data = {
                "plan_id": plan_id,
                "quota_usage": 0,  # Yeni plana geçince sıfırla
                "quota_limit": quota_limit,  # Price metadata'dan alınan appointment_limit (Monthly Capacity)
                "quota_reset_date": quota_reset.isoformat(),
                "quota_last_reset_date": now.isoformat(),  # Lazy reset için son reset tarihi
                "is_first_month": False,  # Artık indirim kullanamaz
                "billing_cycle": billing_cycle,  # Yıllık/Aylık bilgisi
                "updated_at": now.isoformat()
            }
            
            # Trial tarihlerini temizle
            update_data['trial_start_date'] = None
            update_data['trial_end_date'] = None
            
            # Stripe subscription bilgilerini kaydet
            subscription_id = _stripe_obj_get(session, 'subscription')
            customer_id = _stripe_obj_get(session, 'customer')
            
            if subscription_id:
                update_data['stripe_subscription_id'] = subscription_id
                update_data['stripe_customer_id'] = customer_id
                update_data['next_billing_date'] = quota_reset.isoformat()
                logger.info(f"Stripe subscription kaydedildi: organization_id={organization_id}, subscription_id={subscription_id}, billing_cycle={billing_cycle}")
            
            await db.organization_plans.update_one(
                {"organization_id": organization_id},
                {"$set": update_data},
                upsert=True
            )
            
            # 5. Ödeme kaydını 'active' yap + gerçek Stripe tutarını güncelle
            wh_payment_update = {
                    "status": "active",
                    "completed_at": datetime.now(timezone.utc).isoformat(),
                    "subscription_id": subscription_id
            }
            wh_amount_total = _stripe_obj_get(session, 'amount_total')
            if wh_amount_total is not None:
                try:
                    wh_payment_update["amount"] = round(int(wh_amount_total) / 100.0, 2)
                except (TypeError, ValueError):
                    pass
            await db.payment_logs.update_one(
                {"session_id": session_id},
                {"$set": wh_payment_update}
            )
            
            logger.info(f"✅ STRIPE BAŞARILI: {session_id} - Plan güncellendi. Organization: {organization_id}, Plan: {plan_id}")

            try:
                updated_payment_log = await db.payment_logs.find_one({"session_id": session_id})
                await _send_subscription_email_once(
                    db=db,
                    request=request,
                    organization_id=organization_id,
                    session_id=session_id,
                    plan_name=plan_data.get('name') if isinstance(plan_data, dict) else plan_id,
                    billing_cycle=billing_cycle,
                    payment_log=updated_payment_log,
                )
            except Exception as e:
                logger.warning(f"Subscription email (webhook) gönderilemedi: {e}")

            # Funnel: subscription_completed — webhook-owned (backend sole emitter)
            # Stripe event.id'si ledger insert_id olarak kullanılır → idempotent retry'de dup'lenmez.
            try:
                owner_user = await db.users.find_one(
                    {"organization_id": organization_id, "role": "admin"},
                    {"_id": 0, "user_id": 1, "username": 1},
                    sort=[("created_at", 1)],
                )
                distinct_id = (owner_user or {}).get("user_id") or (owner_user or {}).get("username") or organization_id
                await track_webhook_event(
                    db,
                    event_name="subscription_completed",
                    distinct_id=distinct_id,
                    organization_id=organization_id,
                    user_id=(owner_user or {}).get("user_id"),
                    insert_id=f"stripe_{event['id']}",
                    properties={
                        "plan_id": plan_id,
                        "billing_cycle": billing_cycle,
                        "session_id": session_id,
                        "stripe_subscription_id": subscription_id,
                    },
                )
            except Exception as e:
                logger.warning(f"subscription_completed funnel event emit failed: {e}")

        # invoice.paid — yenileme dönemleri + Stripe'dan gerçek tahsilat tutarı (subscription_cycle)
        elif event['type'] == 'invoice.paid':
            inv = event['data']['object']
            inv_id = _stripe_obj_get(inv, 'id')
            # NOT: Stripe 2025-11+ API'de invoice.subscription/period_end alanları
            # taşındı; helper'lar hem eski hem yeni yapıyı okur.
            sub_id = _stripe_invoice_subscription_id(inv)
            if sub_id and inv_id:
                plan_doc = await db.organization_plans.find_one({"stripe_subscription_id": sub_id})
                if plan_doc:
                    org_id = plan_doc.get('organization_id')
                    period_end = _stripe_invoice_period_end(inv)
                    if period_end:
                        ned = datetime.fromtimestamp(int(period_end), tz=timezone.utc)
                        await db.organization_plans.update_one(
                            {"organization_id": org_id},
                            {"$set": {
                                "next_billing_date": ned.isoformat(),
                                "quota_reset_date": ned.isoformat(),
                                "updated_at": datetime.now(timezone.utc).isoformat(),
                            }},
                        )
                    br = _stripe_obj_get(inv, 'billing_reason') or ''
                    if br in ('subscription_cycle', 'subscription_update'):
                        dup = await db.payment_logs.find_one({"stripe_invoice_id": inv_id})
                        if not dup:
                            amount_paid = int(_stripe_obj_get(inv, 'amount_paid') or 0)
                            cur = (_stripe_obj_get(inv, 'currency') or 'try').upper()
                            amt = amount_paid / 100.0
                            await db.payment_logs.insert_one({
                                "stripe_invoice_id": inv_id,
                                "organization_id": org_id,
                                "plan_id": plan_doc.get('plan_id'),
                                "status": "completed",
                                "amount": amt,
                                "currency": cur,
                                "payment_provider": "stripe",
                                "billing_reason": br,
                                "created_at": datetime.now(timezone.utc).isoformat(),
                            })
                            logger.info(f"✅ invoice.paid kaydı eklendi: org={org_id} invoice={inv_id} amount={amt} {cur}")
        
        # customer.subscription.updated - Abonelik güncellendi (iptal planlandı)
        elif event['type'] == 'customer.subscription.updated':
            subscription = event['data']['object']
            subscription_id = subscription['id']
            cancel_at_period_end = _stripe_obj_get(subscription, 'cancel_at_period_end', False)
            
            if cancel_at_period_end:
                logger.info(f"⚠️ Abonelik iptal planlandı (dönem sonunda): subscription_id={subscription_id}")
                
                # Subscription ID ile organization'ı bul
                plan_doc = await db.organization_plans.find_one({"stripe_subscription_id": subscription_id})
                
                if plan_doc:
                    organization_id = plan_doc.get('organization_id')
                    # İsteğe bağlı: Organization'a bildirim gönderebilirsiniz
                    logger.info(f"ℹ️ Organization {organization_id} aboneliğini dönem sonunda iptal edecek")
        
        # customer.subscription.deleted - Abonelik iptal edildi
        elif event['type'] == 'customer.subscription.deleted':
            subscription = event['data']['object']
            subscription_id = subscription['id']
            customer_id = subscription['customer']
            
            logger.info(f"🚫 Abonelik iptal edildi: subscription_id={subscription_id}, customer_id={customer_id}")
            
            # Subscription ID ile organization'ı bul
            plan_doc = await db.organization_plans.find_one({"stripe_subscription_id": subscription_id})
            
            if not plan_doc:
                logger.warning(f"Webhook: subscription_id={subscription_id} için organization bulunamadı")
                return Response(content="OK", status_code=200)
            
            organization_id = plan_doc.get('organization_id')
            
            # Trial'a döndür
            trial_start = datetime.now(timezone.utc)
            trial_end = trial_start + timedelta(days=7)
            
            update_data = {
                "plan_id": "tier_trial",
                "quota_usage": 0,
                "quota_reset_date": trial_end.isoformat(),
                "trial_start_date": trial_start.isoformat(),
                "trial_end_date": trial_end.isoformat(),
                "is_first_month": False,
                "stripe_subscription_id": None,
                "stripe_customer_id": None,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.organization_plans.update_one(
                {"organization_id": organization_id},
                {"$set": update_data}
            )
            
            logger.info(f"✅ Abonelik iptal edildi ve trial'a döndürüldü: organization_id={organization_id}")
        
        # Diğer event'ler için
        logger.info(f"ℹ️ Stripe webhook event işlendi: {event['type']}")
        return Response(content="OK", status_code=200)
        
    except Exception as e:
        logger.error(f"Stripe webhook işleme hatası: {e}", exc_info=True)
        return Response(content="ERROR", status_code=500)


@api_router.post("/payments/webhook/stripe")
async def handle_stripe_webhook_compat(request: Request):
    return await handle_stripe_webhook(request)


@app.post("/webhook/stripe")
async def handle_stripe_webhook_root(request: Request):
    return await handle_stripe_webhook(request)


@app.post("/payments/webhook/stripe")
async def handle_stripe_webhook_payments_root(request: Request):
    return await handle_stripe_webhook(request)

@api_router.post("/payments/process-recurring")
async def process_recurring_payment(request: Request, organization_id: str, current_user: UserInDB = Depends(get_current_user)):
    """Kayıtlı kart ile recurring payment çek (Internal API - Cron job için)"""
    try:
        # Sadece superadmin veya sistem çağrısı yapabilir
        if current_user.role != "superadmin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        db = await get_db_from_request(request)
        
        # Organization plan bilgilerini al
        plan_doc = await get_organization_plan(db, organization_id)
        if not plan_doc:
            logger.error(f"Recurring payment: Plan bulunamadı - organization_id={organization_id}")
            return {"status": "error", "message": "Plan bulunamadı"}
        
        # Kart token bilgilerini kontrol et
        if not plan_doc.get('card_saved') or not plan_doc.get('payment_utoken') or not plan_doc.get('payment_ctoken'):
            logger.warning(f"Recurring payment: Kayıtlı kart yok - organization_id={organization_id}")
            return {"status": "error", "message": "Kayıtlı kart bulunamadı"}
        
        utoken = plan_doc.get('payment_utoken')
        ctoken = plan_doc.get('payment_ctoken')
        plan_id = plan_doc.get('plan_id')
        
        # Plan bilgilerini al
        plan_info = await get_plan_info(plan_id)
        if not plan_info:
            logger.error(f"Recurring payment: Plan info bulunamadı - plan_id={plan_id}")
            return {"status": "error", "message": "Plan bilgisi bulunamadı"}
        
        # Ödeme tutarı (recurring'de indirim yok)
        price_monthly = plan_info.get('price_monthly', 0)
        payment_amount_kurus = int(price_monthly * 100)
        
        # Organization bilgilerini al
        user = await db.users.find_one({"organization_id": organization_id, "role": "admin"})
        if not user:
            logger.error(f"Recurring payment: Admin user bulunamadı - organization_id={organization_id}")
            return {"status": "error", "message": "Admin kullanıcı bulunamadı"}
        
        user_email = user.get('username', 'noreply@royalpremiumcare.com')
        user_name = user.get('full_name', 'Kullanıcı')
        
        # Settings'den adres ve telefon al
        settings = await db.settings.find_one({"organization_id": organization_id})
        user_address = settings.get('address', 'Adres Bilgisi Yok') if settings else 'Adres Bilgisi Yok'
        user_phone = settings.get('support_phone', '05000000000') if settings else '05000000000'
        
        # Merchant OID oluştur
        org_id_clean = organization_id.replace('-', '')
        timestamp_str = str(int(datetime.now(timezone.utc).timestamp()))
        merchant_oid = f"RECUR{org_id_clean}{timestamp_str}"
        
        # User IP (sistem iç çağrısı için)
        user_ip = request.client.host if request.client else "127.0.0.1"
        
        # Sepet bilgisi
        plan_name = plan_info.get('name', 'Plan')
        user_basket = base64.b64encode(json.dumps([
            [plan_name, str(price_monthly), 1]
        ]).encode('utf-8')).decode('utf-8')
        
        # PayTR parametreleri
        no_installment = '1'
        max_installment = '0'
        currency = 'TL'
        test_mode = '0'
        non_3d = '1'  # Recurring payment için Non-3D zorunlu
        payment_type = 'card'
        
        # Hash oluştur (recurring payment için farklı sıra)
        hash_str = f"{PAYTR_MERCHANT_ID}{user_ip}{merchant_oid}{user_email}{payment_amount_kurus}{payment_type}0{currency}{test_mode}{non_3d}"
        paytr_token = base64.b64encode(hmac.new(
            PAYTR_MERCHANT_KEY.encode('utf-8'), 
            hash_str.encode('utf-8') + PAYTR_MERCHANT_SALT.encode('utf-8'), 
            hashlib.sha256
        ).digest()).decode('utf-8')
        
        # PayTR API'sine recurring payment isteği gönder
        post_data = {
            'merchant_id': PAYTR_MERCHANT_ID,
            'user_ip': user_ip,
            'merchant_oid': merchant_oid,
            'email': user_email,
            'payment_type': payment_type,
            'payment_amount': payment_amount_kurus,
            'currency': currency,
            'test_mode': test_mode,
            'non_3d': non_3d,
            'merchant_ok_url': PAYTR_SUCCESS_URL,
            'merchant_fail_url': PAYTR_FAIL_URL,
            'user_name': user_name,
            'user_address': user_address[:400],
            'user_phone': user_phone[:20],
            'user_basket': user_basket,
            'debug_on': '1',
            'paytr_token': paytr_token,
            'utoken': utoken,
            'ctoken': ctoken,
            'installment_count': '0'
        }
        
        # Payment log oluştur
        payment_log = {
            "merchant_oid": merchant_oid,
            "organization_id": organization_id,
            "plan_id": plan_id,
            "amount": price_monthly,
            "status": "pending",
            "payment_type": "recurring",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.payment_logs.insert_one(payment_log)
        
        # PayTR'a istek gönder
        logger.info(f"Recurring payment başlatılıyor: organization_id={organization_id}, merchant_oid={merchant_oid}")
        response = requests.post("https://www.paytr.com/odeme", data=post_data, timeout=15)
        
        logger.info(f"PayTR recurring response: {response.status_code} - {response.text}")
        
        if response.status_code == 200:
            res_data = response.json() if response.headers.get('content-type', '').startswith('application/json') else {}
            
            if res_data.get('status') == 'success':
                logger.info(f"Recurring payment başarılı: organization_id={organization_id}")
                return {"status": "success", "message": "Ödeme başarılı", "merchant_oid": merchant_oid}
            else:
                error_msg = res_data.get('reason', 'Bilinmeyen hata')
                logger.error(f"Recurring payment başarısız: {error_msg}")
                
                # Başarısız ödeme kaydını güncelle
                await db.payment_logs.update_one(
                    {"merchant_oid": merchant_oid},
                    {"$set": {"status": "failed", "failed_reason": error_msg}}
                )
                
                return {"status": "failed", "message": error_msg}
        else:
            logger.error(f"PayTR recurring HTTP hatası: {response.status_code}")
            return {"status": "error", "message": "Ödeme servisi hatası"}
            
    except Exception as e:
        logger.error(f"Recurring payment işleme hatası: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}

async def _auto_complete_pending_for_org(db, organization_id: str) -> bool:
    """Bugünkü süresi geçmiş `Bekliyor` randevuları toplu `Tamamlandı`ya alır.

    Dashboard endpoint her çağrıda cache'in DIŞINDA bu helper'ı çağırır; yan etki
    tespit edilirse dashboard cache invalidate edilir → HIT'ler stale kalmaz.
    Denormalize edilen `customers.completed_appointments` sayaçları da burada
    güncellenir (delta_completed=+1 per (phone,name)).

    Return: True → en az bir randevu tamamlandı; False → değişiklik yok.
    """
    turkey_tz = ZoneInfo("Europe/Istanbul")
    now = datetime.now(turkey_tz)
    today = now.date().isoformat()
    base_query = {"organization_id": organization_id}

    today_waiting_appointments = await db.appointments.find(
        {**base_query, "appointment_date": today, "status": "Bekliyor"},
        {"_id": 0, "id": 1, "appointment_date": 1, "appointment_time": 1, "service_price": 1, "customer_name": 1, "service_name": 1, "service_id": 1, "service_duration": 1, "staff_member_id": 1, "phone": 1}
    ).to_list(1000)

    if not today_waiting_appointments:
        return False

    service_ids = list({appt.get('service_id') for appt in today_waiting_appointments if appt.get('service_id')})
    services_dict: dict = {}
    if service_ids:
        services = await db.services.find(
            {"id": {"$in": service_ids}, "organization_id": organization_id},
            {"_id": 0, "id": 1, "duration": 1}
        ).to_list(100)
        services_dict = {s['id']: s.get('duration', 30) for s in services}

    ids_to_update: list = []
    transactions_to_create: list = []
    completed_deltas: list = []
    for appt in today_waiting_appointments:
        try:
            dt_str = f"{appt['appointment_date']} {appt['appointment_time']}"
            appointment_dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M").replace(tzinfo=turkey_tz)
            duration = appt.get('service_duration') or services_dict.get(appt.get('service_id'), 30)
            if now >= (appointment_dt + timedelta(minutes=duration)):
                ids_to_update.append(appt['id'])
                transactions_to_create.append(Transaction(
                    organization_id=organization_id,
                    appointment_id=appt['id'],
                    customer_name=appt['customer_name'],
                    service_name=appt['service_name'],
                    amount=appt.get('service_price', 0) or 0,
                    date=appt['appointment_date'],
                    staff_member_id=appt.get('staff_member_id')
                ).model_dump(mode='json'))
                if appt.get('phone'):
                    completed_deltas.append({
                        "phone": appt['phone'],
                        "name": appt.get('customer_name') or "",
                        "appointment_date": appt.get('appointment_date'),
                        "appointment_time": appt.get('appointment_time'),
                    })
        except Exception as e:
            logging.warning(f"Randevu {appt.get('id')} işlenirken hata: {e}")

    if not ids_to_update:
        return False

    await db.appointments.update_many(
        {"organization_id": organization_id, "id": {"$in": ids_to_update}},
        {"$set": {"status": "Tamamlandı", "completed_at": datetime.now(timezone.utc).isoformat()}}
    )
    if transactions_to_create:
        for t in transactions_to_create:
            if 'created_at' not in t:
                t['created_at'] = datetime.now(timezone.utc).isoformat()
        await db.transactions.insert_many(transactions_to_create)

    # Denormalize sayaçları güncelle (async'te sırayla — hata tek customer'ı atlar)
    for d in completed_deltas:
        try:
            await _apply_customer_delta(
                db,
                organization_id=organization_id,
                phone=d["phone"],
                name=d["name"],
                delta_total=0,
                delta_completed=1,
                appointment_date=d.get("appointment_date"),
                appointment_time=d.get("appointment_time"),
            )
        except Exception:
            pass

    logger.info(f"⏱️ auto-complete: {len(ids_to_update)} randevu Tamamlandı'ya alındı (org={organization_id[:8]})")
    return True


@cache_result(prefix="dashboard_stats", ttl=60)
async def _compute_dashboard_stats(request: Request, current_user: UserInDB):
    """Cache'lenebilir saf dashboard hesaplaması — yan etki YOK.

    Auto-complete side-effect'i dışarıda `_auto_complete_pending_for_org` ile
    çalıştırılır; bu fonksiyon sadece MongoDB read'lerini yapar.
    """
    db = await get_db_from_request(request)
    turkey_tz = ZoneInfo("Europe/Istanbul")
    now = datetime.now(turkey_tz)
    today = now.date().isoformat()
    month_start = now.date().replace(day=1).isoformat()
    org = current_user.organization_id
    base_query = {"organization_id": org}

    # Bağımsız read'ler PARALEL + sunucu-taraflı $group toplama (doc transferi yerine).
    # Response şekli DEĞİŞMEZ; sadece I/O ve ağ yükü düşer (cache-miss ~3-4s → ~sub-sn hedef).
    async def _count_today_appts():
        return await db.appointments.count_documents(
            {**base_query, "appointment_date": today, "status": {"$nin": ["İptal", "İptal Edildi", "Ödeme Bekleniyor"]}}
        )

    async def _today_completed_agg():
        cur = db.appointments.aggregate([
            {"$match": {**base_query, "appointment_date": today, "status": "Tamamlandı"}},
            {"$group": {"_id": None, "count": {"$sum": 1}, "total": {"$sum": {"$ifNull": ["$service_price", 0]}}}},
        ])
        docs = await cur.to_list(1)
        if docs:
            return int(docs[0].get("count", 0) or 0), (docs[0].get("total", 0) or 0)
        return 0, 0

    async def _sum_income(date_filter: dict):
        cur = db.transactions.aggregate([
            {"$match": {**base_query, **date_filter}},
            {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$amount", 0]}}}},
        ])
        docs = await cur.to_list(1)
        return (docs[0].get("total", 0) or 0) if docs else 0

    (
        today_appointments,
        (today_completed, bugunku_toplam_hizmet_tutari),
        today_income,
        month_income,
    ) = await asyncio.gather(
        _count_today_appts(),
        _today_completed_agg(),
        _sum_income({"date": today}),
        _sum_income({"date": {"$gte": month_start}}),
    )

    plan_doc = await get_organization_plan(db, org)
    quota_info = None
    if plan_doc:
        plan_id = plan_doc.get('plan_id', 'tier_trial')
        plan_info = await get_plan_info(plan_id)
        if plan_info:
            quota_usage = plan_doc.get('quota_usage', 0)
            quota_limit = plan_doc.get('quota_limit') or plan_info.get('quota_monthly_appointments', 50)
            quota_remaining = -1 if quota_limit == -1 else max(0, quota_limit - quota_usage)
            quota_percentage = (quota_usage / quota_limit * 100) if quota_limit > 0 else 0
            quota_info = {
                "plan_id": plan_id, "plan_name": plan_info.get('name'),
                "quota_usage": quota_usage, "quota_limit": quota_limit,
                "quota_remaining": quota_remaining, "quota_percentage": round(quota_percentage, 2),
                "is_trial": plan_id == 'tier_trial', "is_low_quota": quota_percentage >= 90
            }

    return {
        "today_appointments": today_appointments,
        "today_completed": today_completed,
        "today_income": today_income,
        "bugunku_toplam_hizmet_tutari": bugunku_toplam_hizmet_tutari,
        "month_income": month_income,
        "quota": quota_info
    }


@api_router.get("/stats/dashboard")
async def get_dashboard_stats(request: Request, current_user: UserInDB = Depends(get_current_user)):
    if current_user.role not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")

    if current_user.role == "superadmin":
        return {
            "today_appointments": 0, "today_completed": 0, "today_income": 0,
            "bugunku_toplam_hizmet_tutari": 0, "month_income": 0, "quota": None
        }

    logger.info(f"📊 Stats endpoint çağrıldı - Organization: {current_user.organization_id}")
    db = await get_db_from_request(request)

    # 1) Yan etkiyi cache'in DIŞINDA çalıştır → cache'i asla kirletmez
    try:
        did_change = await _auto_complete_pending_for_org(db, current_user.organization_id)
    except Exception as exc:
        logger.warning(f"auto-complete pending atlandı: {exc}")
        did_change = False

    # 2) Auto-complete yeni Tamamlandı yarattıysa cache'i uçur → fresh hesap
    if did_change:
        try:
            await invalidate_cache(request, "dashboard_stats", current_user)
        except Exception:
            pass

    # 3) Cache'li hesabı çağır (HIT → milisaniye, MISS → hesapla + Redis'e yaz)
    return await _compute_dashboard_stats(request=request, current_user=current_user)

@api_router.get("/stats/personnel")
async def get_personnel_stats(request: Request, current_user: UserInDB = Depends(get_current_user)):
    # Sadece personel görebilir
    if current_user.role != "staff":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    logger.info(f"👤 Personel stats endpoint çağrıldı - Staff: {current_user.username}")
    db = await get_db_from_request(request)
    turkey_tz = ZoneInfo("Europe/Istanbul")
    today = datetime.now(turkey_tz).date().isoformat()
    now = datetime.now(turkey_tz)
    base_query = {"organization_id": current_user.organization_id}
    
    # ÖNCE: Bugünkü "Bekliyor" status'ündeki personelin randevularını otomatik tamamla
    today_waiting_appointments = await db.appointments.find(
        {**base_query, "appointment_date": today, "status": "Bekliyor", "staff_member_id": current_user.username},
        {"_id": 0, "id": 1, "appointment_date": 1, "appointment_time": 1, "service_price": 1, "customer_name": 1, "service_name": 1, "service_id": 1, "service_duration": 1}
    ).to_list(1000)
    
    # Servisleri çek (duration bilgisi için)
    service_ids = [appt.get('service_id') for appt in today_waiting_appointments if appt.get('service_id')]
    services_dict = {}
    if service_ids:
        services = await db.services.find(
            {"id": {"$in": list(set(service_ids))}, "organization_id": current_user.organization_id},
            {"_id": 0, "id": 1, "duration": 1}
        ).to_list(1000)
        services_dict = {s['id']: s.get('duration', 30) for s in services}
    
    ids_to_update = []
    transactions_to_create = []
    for appt in today_waiting_appointments:
        try:
            dt_str = f"{appt['appointment_date']} {appt['appointment_time']}"
            naive_dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
            appointment_dt = naive_dt.replace(tzinfo=turkey_tz)
            # Randevu bitiş saatini hesapla — çoklu hizmette saklı toplam süre
            service_duration_minutes = appt.get('service_duration') or services_dict.get(appt.get('service_id'), 30)
            completion_threshold = appointment_dt + timedelta(minutes=service_duration_minutes)
            if now >= completion_threshold:
                ids_to_update.append(appt['id'])
                transaction = Transaction(
                    organization_id=current_user.organization_id,
                    appointment_id=appt['id'],
                    customer_name=appt['customer_name'],
                    service_name=appt['service_name'],
                    amount=appt.get('service_price', 0),
                    date=appt['appointment_date'],
                    staff_member_id=appt.get('staff_member_id')
                )
                trans_doc = transaction.model_dump()
                trans_doc['created_at'] = trans_doc['created_at'].isoformat()
                transactions_to_create.append(trans_doc)
        except (ValueError, TypeError) as e:
            logging.warning(f"Randevu {appt['id']} için tarih ayrıştırılamadı: {e}")
    
    # Otomatik tamamlanan randevuları güncelle
    if ids_to_update:
        logger.info(f"✅ Personel için {len(ids_to_update)} randevu otomatik tamamlanacak")
        await db.appointments.update_many(
            {"organization_id": current_user.organization_id, "id": {"$in": ids_to_update}},
            {"$set": {"status": "Tamamlandı", "completed_at": datetime.now(timezone.utc).isoformat()}}
        )
    # Otomatik tamamlanan randevular için transaction oluştur
    if transactions_to_create:
        await db.transactions.insert_many(transactions_to_create)
    
    # Personelin bugünkü tamamlanan randevularını bul
    today_completed_appointments = await db.appointments.find(
        {**base_query, "appointment_date": today, "status": "Tamamlandı", "staff_member_id": current_user.username},
        {"_id": 0, "service_price": 1, "id": 1}
    ).to_list(1000)
    
    # Toplam hizmet tutarı ve randevu sayısı
    total_revenue_generated = sum(apt.get('service_price', 0) or 0 for apt in today_completed_appointments)
    completed_appointments_count = len(today_completed_appointments)
    
    logger.info(f"👤 Personel {current_user.username}: {completed_appointments_count} tamamlanan randevu, Toplam: {total_revenue_generated} ₺")
    
    return {
        "total_revenue_generated": total_revenue_generated,
        "completed_appointments_count": completed_appointments_count
    }

# ═══════════════════════════════════════════════════════════════════════════
# İSTATİSTİKLER SAYFASI — kapsamlı agregasyon (Randevular/Müşteriler/Kasa/Personel/Hizmetler)
# ═══════════════════════════════════════════════════════════════════════════

def _stats_canon_phone(phone: Optional[str]) -> str:
    """Telefonu kanonik anahtara indirger (son 10 hane) → mükerrer müşteri sayımını azaltır."""
    digits = re.sub(r"\D", "", phone or "")
    if len(digits) > 10:
        digits = digits[-10:]
    return digits


def _resolve_stats_range(range_key: str, start: Optional[str], end: Optional[str], tz):
    """range_key → (start_date, end_date) date objeleri. Europe/Istanbul bazlı."""
    today = datetime.now(tz).date()
    if range_key == "today":
        return today, today
    if range_key == "this_week":
        return today - timedelta(days=today.weekday()), today
    if range_key == "this_month":
        return today.replace(day=1), today
    if range_key == "last_month":
        first_this = today.replace(day=1)
        last_prev = first_this - timedelta(days=1)
        return last_prev.replace(day=1), last_prev
    if range_key == "this_year":
        return today.replace(month=1, day=1), today
    if range_key == "custom" and start and end:
        try:
            s = datetime.strptime(start, "%Y-%m-%d").date()
            e = datetime.strptime(end, "%Y-%m-%d").date()
            return (e, s) if e < s else (s, e)
        except ValueError:
            pass
    # varsayılan: bu ay
    return today.replace(day=1), today


@api_router.get("/stats/analytics")
async def get_stats_analytics(
    request: Request,
    range_key: str = Query("this_month", alias="range"),
    start: Optional[str] = None,
    end: Optional[str] = None,
    current_user: UserInDB = Depends(get_current_user),
):
    """İstatistikler sayfası için tek agregasyon ucu (sadece admin, org-scoped).

    Yeni uç — mevcut hiçbir response şekli değişmez (API sözleşmesi güvenli).
    Gelir yalnız 'Tamamlandı' randevulardan; çoklu hizmette services[].price_snapshot,
    tekli hizmette service_price. Tahsilat modu (Yerinde/Online/Kapora) hizmetin
    payment_rule'undan türetilir.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")

    from collections import Counter, defaultdict

    db = await get_db_from_request(request)
    tz = ZoneInfo("Europe/Istanbul")
    org = current_user.organization_id
    start_d, end_d = _resolve_stats_range(range_key, start, end, tz)
    start_s, end_s = start_d.isoformat(), end_d.isoformat()
    span_days = (end_d - start_d).days + 1
    granularity = "month" if span_days > 92 else "day"
    base = {"organization_id": org}

    COMPLETED = "Tamamlandı"
    CANCELLED = {"İptal", "İptal Edildi"}
    NOSHOW = "Gelmedi"

    appt_fields = {
        "_id": 0, "status": 1, "source": 1, "appointment_date": 1, "appointment_time": 1,
        "staff_member_id": 1, "service_id": 1, "service_name": 1, "service_price": 1,
        "services": 1, "customer_name": 1, "phone": 1, "payment_status": 1,
    }

    async def _appts_in_range():
        return await db.appointments.find(
            {**base, "appointment_date": {"$gte": start_s, "$lte": end_s}}, appt_fields
        ).to_list(100000)

    async def _expenses_in_range():
        return await db.expenses.find(
            {**base, "date": {"$gte": start_s, "$lte": end_s}},
            {"_id": 0, "amount": 1, "category": 1, "date": 1},
        ).to_list(100000)

    async def _services_all():
        return await db.services.find(
            base, {"_id": 0, "id": 1, "name": 1, "payment_rule": 1}
        ).to_list(100000)

    async def _staff_all():
        return await db.users.find(
            {**base, "role": {"$in": ["admin", "staff"]}},
            {"_id": 0, "username": 1, "full_name": 1},
        ).to_list(10000)

    async def _first_appt_agg():
        cur = db.appointments.aggregate([
            {"$match": base},
            {"$group": {"_id": "$phone", "first_date": {"$min": "$appointment_date"}}},
        ])
        return await cur.to_list(100000)

    async def _customers_all():
        return await db.customers.find(
            base, {"_id": 0, "phone": 1, "created_at": 1}
        ).to_list(100000)

    appts, expenses, services, staff_users, first_appt, customers_docs = await asyncio.gather(
        _appts_in_range(), _expenses_in_range(), _services_all(),
        _staff_all(), _first_appt_agg(), _customers_all(),
    )

    svc_rule = {s["id"]: (s.get("payment_rule") or "on_site") for s in services}
    staff_name = {u["username"]: (u.get("full_name") or u["username"]) for u in staff_users}

    def _amount(a):
        svcs = a.get("services")
        if svcs:
            return sum((x.get("price_snapshot") or 0) for x in svcs)
        return a.get("service_price") or 0

    def _mode(a):
        svcs = a.get("services")
        ids = [x.get("service_id") for x in svcs] if svcs else [a.get("service_id")]
        rules = []
        for sid in ids:
            r = svc_rule.get(sid, "on_site")
            rules.append("online" if r in ("online", "full_online") else r)
        if "deposit" in rules:
            mode = "deposit"
        elif "online" in rules:
            mode = "online"
        else:
            mode = "on_site"
        if mode == "on_site" and a.get("payment_status") == "paid":
            mode = "online"
        return mode

    completed = [a for a in appts if a.get("status") == COMPLETED]
    total_income = sum(_amount(a) for a in completed)
    total_expense = sum(float(e.get("amount") or 0) for e in expenses)
    completed_count = len(completed)
    total_appts = len(appts)

    # ── Zaman serisi (gelir / gider / net / randevu) ──
    def _bucket(d_str):
        return d_str[:7] if granularity == "month" else d_str

    buckets = {}

    def _b(key):
        return buckets.setdefault(key, {"income": 0.0, "expense": 0.0, "appointments": 0})

    ordered = []
    if granularity == "month":
        y, m = start_d.year, start_d.month
        while (y, m) <= (end_d.year, end_d.month):
            key = f"{y:04d}-{m:02d}"
            ordered.append(key)
            _b(key)
            m += 1
            if m > 12:
                m = 1
                y += 1
    else:
        d = start_d
        while d <= end_d:
            key = d.isoformat()
            ordered.append(key)
            _b(key)
            d += timedelta(days=1)

    for a in appts:
        ad = a.get("appointment_date")
        if ad:
            _b(_bucket(ad))["appointments"] += 1
    for a in completed:
        ad = a.get("appointment_date")
        if ad:
            _b(_bucket(ad))["income"] += _amount(a)
    for e in expenses:
        dd = e.get("date")
        if dd:
            _b(_bucket(dd))["expense"] += float(e.get("amount") or 0)

    timeseries = [
        {
            "label": k,
            "income": round(buckets[k]["income"], 2),
            "expense": round(buckets[k]["expense"], 2),
            "net": round(buckets[k]["income"] - buckets[k]["expense"], 2),
            "appointments": buckets[k]["appointments"],
        }
        for k in ordered
    ]

    # ── Randevu kırılımları ──
    status_counter = Counter(a.get("status") or "—" for a in appts)
    source_counter = Counter(
        "public_booking" if (a.get("source") == "public_booking") else "manual" for a in appts
    )
    weekday_counter = Counter()
    hour_counter = Counter()
    for a in appts:
        ad = a.get("appointment_date")
        if ad:
            try:
                weekday_counter[datetime.strptime(ad, "%Y-%m-%d").weekday()] += 1
            except ValueError:
                pass
        tm = a.get("appointment_time") or ""
        if ":" in tm:
            try:
                hour_counter[int(tm.split(":")[0])] += 1
            except ValueError:
                pass
    cancelled = sum(status_counter.get(s, 0) for s in CANCELLED)
    noshow = status_counter.get(NOSHOW, 0)

    # ── Müşteriler ──
    first_date_map = {}
    for row in first_appt:
        ph = _stats_canon_phone(row.get("_id"))
        fd = row.get("first_date")
        if ph and fd:
            if ph not in first_date_map or fd < first_date_map[ph]:
                first_date_map[ph] = fd
    for c in customers_docs:
        ph = _stats_canon_phone(c.get("phone"))
        created = c.get("created_at")
        cd = created[:10] if isinstance(created, str) and len(created) >= 10 else None
        if ph and cd:
            if ph not in first_date_map or cd < first_date_map[ph]:
                first_date_map[ph] = cd
    total_customers = len(first_date_map)
    new_customers = sum(1 for fd in first_date_map.values() if start_s <= fd <= end_s)

    new_by_bucket = Counter()
    for fd in first_date_map.values():
        if start_s <= fd <= end_s:
            new_by_bucket[_bucket(fd)] += 1
    new_series = [{"label": k, "count": new_by_bucket.get(k, 0)} for k in ordered]

    range_customer = defaultdict(lambda: {"name": "", "visits": 0, "spent": 0.0})
    for a in appts:
        ph = _stats_canon_phone(a.get("phone"))
        if not ph:
            continue
        rc = range_customer[ph]
        rc["visits"] += 1
        if not rc["name"]:
            rc["name"] = a.get("customer_name") or ""
    for a in completed:
        ph = _stats_canon_phone(a.get("phone"))
        if ph:
            range_customer[ph]["spent"] += _amount(a)
    new_in_range = returning_in_range = 0
    for ph in range_customer:
        fd = first_date_map.get(ph)
        if fd is not None and start_s <= fd <= end_s:
            new_in_range += 1
        else:
            returning_in_range += 1
    # Tüm müşteriler, harcamaya göre yüksekten düşüğe (frontend listeyi scroll eder)
    top_customers = sorted(
        (
            {"name": v["name"] or "—", "visits": v["visits"], "spent": round(v["spent"], 2)}
            for v in range_customer.values()
        ),
        key=lambda x: x["spent"],
        reverse=True,
    )

    # ── Kasa: tahsilat modu / gider kategori / hizmete göre gelir ──
    mode_agg = defaultdict(lambda: {"amount": 0.0, "count": 0})
    for a in completed:
        md = _mode(a)
        mode_agg[md]["amount"] += _amount(a)
        mode_agg[md]["count"] += 1
    income_by_mode = [
        {"mode": k, "amount": round(v["amount"], 2), "count": v["count"]}
        for k, v in mode_agg.items()
    ]

    cat_agg = defaultdict(float)
    for e in expenses:
        cat_agg[e.get("category") or "Diğer"] += float(e.get("amount") or 0)
    expense_by_category = sorted(
        ({"category": k, "amount": round(v, 2)} for k, v in cat_agg.items()),
        key=lambda x: x["amount"],
        reverse=True,
    )

    # ── Hizmetler + Personel (tamamlanan randevulardan) ──
    svc_agg = defaultdict(lambda: {"name": "", "count": 0, "revenue": 0.0})
    staff_agg = defaultdict(lambda: {"completed": 0, "revenue": 0.0})
    for a in completed:
        amt = _amount(a)
        sid = a.get("staff_member_id")
        skey = sid if sid else "__unassigned__"
        staff_agg[skey]["completed"] += 1
        staff_agg[skey]["revenue"] += amt
        svcs = a.get("services")
        if svcs:
            for x in svcs:
                k = x.get("service_id") or x.get("name_snapshot") or "—"
                row = svc_agg[k]
                row["name"] = x.get("name_snapshot") or row["name"] or "—"
                row["count"] += 1
                row["revenue"] += (x.get("price_snapshot") or 0)
        else:
            k = a.get("service_id") or a.get("service_name") or "—"
            row = svc_agg[k]
            row["name"] = a.get("service_name") or row["name"] or "—"
            row["count"] += 1
            row["revenue"] += amt

    services_stats = sorted(
        (
            {
                "name": v["name"] or "—",
                "count": v["count"],
                "revenue": round(v["revenue"], 2),
                "avg_price": round(v["revenue"] / v["count"], 2) if v["count"] else 0,
            }
            for v in svc_agg.values()
        ),
        key=lambda x: x["count"],
        reverse=True,
    )
    income_by_service = sorted(
        ({"name": s["name"], "amount": s["revenue"]} for s in services_stats),
        key=lambda x: x["amount"],
        reverse=True,
    )[:8]

    personnel = []
    for skey, v in staff_agg.items():
        if skey == "__unassigned__":
            username, full_name = "__unassigned__", None
        else:
            username, full_name = skey, staff_name.get(skey, skey)
        personnel.append({
            "username": username,
            "full_name": full_name,
            "completed": v["completed"],
            "revenue": round(v["revenue"], 2),
            "avg_ticket": round(v["revenue"] / v["completed"], 2) if v["completed"] else 0,
        })
    personnel.sort(key=lambda x: x["revenue"], reverse=True)

    return {
        "range": {"key": range_key, "start": start_s, "end": end_s, "granularity": granularity},
        "summary": {
            "total_income": round(total_income, 2),
            "total_expense": round(total_expense, 2),
            "net_profit": round(total_income - total_expense, 2),
            "total_appointments": total_appts,
            "completed_appointments": completed_count,
            "total_customers": total_customers,
            "new_customers": new_customers,
            "avg_ticket": round(total_income / completed_count, 2) if completed_count else 0,
        },
        "timeseries": timeseries,
        "appointments": {
            "by_status": [{"status": k, "count": v} for k, v in status_counter.items()],
            "by_source": [{"source": k, "count": v} for k, v in source_counter.items()],
            "by_weekday": [{"weekday": wd, "count": weekday_counter.get(wd, 0)} for wd in range(7)],
            "by_hour": [{"hour": h, "count": hour_counter.get(h, 0)} for h in sorted(hour_counter)],
            "cancellation_rate": round(cancelled / total_appts * 100, 1) if total_appts else 0,
            "noshow_rate": round(noshow / total_appts * 100, 1) if total_appts else 0,
        },
        "customers": {
            "new_series": new_series,
            "new_vs_returning": {"new": new_in_range, "returning": returning_in_range},
            "top": top_customers,
        },
        "kasa": {
            "income_by_mode": income_by_mode,
            "expense_by_category": expense_by_category,
            "income_by_service": income_by_service,
        },
        "personnel": personnel,
        "services": services_stats,
    }

# === CHATWOOT IDENTITY (HMAC) ===
@api_router.get("/me/chatwoot-identity")
async def get_chatwoot_identity(current_user: UserInDB = Depends(get_current_user)):
    """
    Chatwoot User Identity Validation için identifier_hash döner.
    Frontend bunu `$chatwoot.setUser(identifier, { identifier_hash })` ile gönderir.
    Chatwoot panelinde User Identity Validation açıkken HMAC zorunludur; aksi halde
    setUser çağrısı sessizce reddedilir ve "small-dew" gibi anonymous rumuz oluşur.
    CHATWOOT_HMAC_TOKEN env tanımlı değilse hash None döner (validation kapalı modu).
    """
    import hmac as _hmac
    import hashlib as _hashlib
    hmac_token = os.environ.get("CHATWOOT_HMAC_TOKEN", "").strip()
    identifier = str(current_user.user_id)
    identifier_hash = None
    if hmac_token:
        identifier_hash = _hmac.new(
            hmac_token.encode("utf-8"),
            identifier.encode("utf-8"),
            _hashlib.sha256,
        ).hexdigest()
    return {"identifier": identifier, "identifier_hash": identifier_hash}

# === PUSH NOTIFICATION ROUTES ===
@api_router.get("/push/vapid-key")
async def get_vapid_public_key():
    """VAPID public key'i döndür - subscription için gerekli"""
    return {"publicKey": VAPID_PUBLIC_KEY}

@api_router.post("/push/subscribe")
async def subscribe_push(request: Request, subscription_data: PushSubscriptionCreate, current_user: UserInDB = Depends(get_current_user)):
    """Push notification aboneliği oluştur"""
    db = await get_db_from_request(request)
    
    subscription = subscription_data.subscription
    endpoint = subscription["endpoint"]
    
    # ÖNEMLİ: Aynı endpoint'e sahip ama farklı kullanıcıya/organization'a ait subscription varsa temizle
    # Bu, cross-account bildirim karışıklığını önler
    existing_any = await db.push_subscriptions.find_one({
        "endpoint": endpoint
    })
    
    if existing_any:
        # Eğer mevcut subscription farklı bir kullanıcıya/organization'a aitse, onu sil
        if (existing_any.get("organization_id") != current_user.organization_id or 
            existing_any.get("user_id") != current_user.username):
            logger.warning(
                f"⚠️ Found existing subscription for endpoint {endpoint[:50]}... "
                f"belonging to different user/org: {existing_any.get('user_id')}/{existing_any.get('organization_id')} "
                f"(current: {current_user.username}/{current_user.organization_id}). Deleting old subscription."
            )
            await db.push_subscriptions.delete_one({"_id": existing_any["_id"]})
            existing_any = None
    
    platform = subscription.get("platform", "web")
    
    # Native platformlar (ios/android) için: aynı kullanıcı+platform için eski token'ı güncelle
    # Token her app restart'ta değişebilir, endpoint bazlı değil user+platform bazlı eşleştir
    if platform in ('ios', 'android'):
        existing_native = await db.push_subscriptions.find_one({
            "organization_id": current_user.organization_id,
            "user_id": current_user.username,
            "platform": platform
        })
        if existing_native:
            await db.push_subscriptions.update_one(
                {"_id": existing_native["_id"]},
                {"$set": {
                    "endpoint": endpoint,
                    "keys": subscription["keys"],
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            logger.info(f"✅ Push subscription updated (token refresh) for user: {current_user.username}, platform: {platform}")
            return {"message": "Push subscription updated"}
    
    # Web için: aynı endpoint bazlı eşleştir
    existing = await db.push_subscriptions.find_one({
        "organization_id": current_user.organization_id,
        "user_id": current_user.username,
        "endpoint": endpoint
    })
    
    if existing:
        await db.push_subscriptions.update_one(
            {"_id": existing["_id"]},
            {"$set": {
                "keys": subscription["keys"], 
                "platform": platform,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        logger.info(f"✅ Push subscription updated for user: {current_user.username}, platform: {platform}")
        return {"message": "Push subscription updated"}
    
    # Yeni abonelik oluştur
    push_doc = {
        "id": str(uuid.uuid4()),
        "organization_id": current_user.organization_id,
        "user_id": current_user.username,
        "user_role": current_user.role,
        "endpoint": endpoint,
        "keys": subscription["keys"],
        "platform": subscription.get("platform", "web"),  # android, ios, web
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.push_subscriptions.insert_one(push_doc)
    logger.info(f"✅ Push subscription created for user: {current_user.username}, platform: {subscription.get('platform', 'web')}")
    
    return {"message": "Push subscription created"}

@api_router.delete("/push/unsubscribe")
async def unsubscribe_push(request: Request, current_user: UserInDB = Depends(get_current_user)):
    """Push notification aboneliğini iptal et"""
    db = await get_db_from_request(request)
    
    result = await db.push_subscriptions.delete_many({
        "organization_id": current_user.organization_id,
        "user_id": current_user.username
    })
    
    return {"message": f"Deleted {result.deleted_count} subscriptions"}


@api_router.get("/push/status")
async def push_subscription_status(request: Request, current_user: UserInDB = Depends(get_current_user)):
    """Kullanıcının backend'de kayıtlı push aboneliği var mı? (OS izni ≠ kayıtlı cihaz)"""
    db = await get_db_from_request(request)
    count = await db.push_subscriptions.count_documents({
        "organization_id": current_user.organization_id,
        "user_id": current_user.username,
    })
    subs = await db.push_subscriptions.find(
        {"organization_id": current_user.organization_id, "user_id": current_user.username},
        {"platform": 1, "_id": 0},
    ).to_list(10)
    platforms = list({s.get("platform") or "web" for s in subs})
    return {"subscribed": count > 0, "count": count, "platforms": platforms}


@api_router.post("/push/debug")
async def push_debug(request: Request):
    """iOS push debug - her adımı logla"""
    try:
        body = await request.json()
        step = body.get("step", "unknown")
        data = body.get("data", {})
        error = body.get("error", None)
        platform = body.get("platform", "unknown")
        logger.info(f"📲 PUSH DEBUG [{platform}] Step: {step} | Data: {data} | Error: {error}")
        return {"ok": True}
    except Exception as e:
        logger.error(f"Push debug error: {e}")
        return {"ok": False}

@api_router.post("/push/test")
async def test_push_notification(request: Request, current_user: UserInDB = Depends(get_current_user)):
    """Test push notification - sadece admin için"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    try:
        await send_push_notification(
            db=db,
            organization_id=current_user.organization_id,
            title="🧪 Test Bildirimi",
            body="Push notification sistemi çalışıyor!",
            data={"type": "test", "url": "/"}
        )
        return {"message": "Test bildirimi gönderildi", "success": True}
    except Exception as e:
        logger.error(f"Test push notification error: {e}")
        return {"message": f"Hata: {str(e)}", "success": False}

# === SETTINGS ROUTES ===
@api_router.get("/settings", response_model=Settings)
async def get_settings(request: Request, current_user: UserInDB = Depends(get_current_user)):
    # Personel okuyabilir, ama sadece admin güncelleyebilir
    db = await get_db_from_request(request); query = {"organization_id": current_user.organization_id}
    settings = await db.settings.find_one(query, {"_id": 0})
    if not settings:
        default_settings = Settings(organization_id=current_user.organization_id); await db.settings.insert_one(default_settings.model_dump())
        return default_settings
    return Settings(**settings)

@api_router.put("/settings", response_model=Settings)
async def update_settings(request: Request, settings: Settings, current_user: UserInDB = Depends(get_current_user)):
    # Sadece admin güncelleyebilir
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    query = {"organization_id": current_user.organization_id}
    
    # Mevcut ayarları al
    current_settings = await db.settings.find_one(query, {"_id": 0})
    
    update_data = settings.model_dump(exclude_unset=True)
    update_data["organization_id"] = current_user.organization_id
    
    # Eğer company_name değiştiyse, yeni slug oluştur
    if current_settings and "company_name" in update_data and current_settings.get('company_name') != settings.company_name:
        base_slug = slugify(settings.company_name)
        if not base_slug:
            raise HTTPException(status_code=400, detail="İşletme adı geçerli karakter içermelidir.")
        
        existing_slug = await db.users.find_one({"slug": base_slug, "username": {"$ne": current_user.username}})
        if existing_slug:
            raise HTTPException(
                status_code=400,
                detail="Bu işletme adı zaten kullanılıyor veya mevcut bir işletmeyle çok benzer. Lütfen farklı bir işletme adı seçin."
            )
        unique_slug = base_slug
        
        await db.users.update_one(
            {"username": current_user.username},
            {"$set": {"slug": unique_slug}}
        )
        
        update_data["slug"] = unique_slug
        logging.info(f"Company name changed. New slug: {unique_slug}")
    
    await db.settings.update_one(query, {"$set": update_data}, upsert=True)
    updated_settings = await db.settings.find_one(query, {"_id": 0})
    
    # business_hours değiştiyse, kapalı günleri tüm personelin days_off'una senkronize et
    if "business_hours" in update_data and update_data.get("business_hours"):
        # Eski (update öncesi) kapalı günleri bul (fallback olarak kullanılacak)
        prev_closed_days = []
        prev_business_hours = (current_settings or {}).get('business_hours', {}) or {}
        prev_day_names = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
        for day in prev_day_names:
            prev_day_settings = prev_business_hours.get(day, {}) or {}
            if not prev_day_settings.get('is_open', True):
                prev_closed_days.append(day)

        # Kapalı günleri bul
        closed_days = []
        day_names = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
        for day in day_names:
            day_settings = settings.business_hours.get(day, {})
            if not day_settings.get('is_open', True):
                closed_days.append(day)

        # Tüm personelin days_off'unu güncelle:
        # - Kapalı günleri days_off'a ekle
        # - Daha önce otomatik senkronize edilip artık açık olan günleri days_off'tan çıkar
        # Not: Bilinçli olarak eklenmiş days_off değerlerini silmemek için, kullanıcı bazında synced_closed_days takip edilir.
        all_users = await db.users.find(
            {"organization_id": current_user.organization_id}
        ).to_list(1000)

        for user in all_users:
            current_days_off = user.get('days_off', []) or []
            # Eğer eski sistemde synced_closed_days alanı yoksa, previous business_hours'tan gelen kapalı günleri fallback olarak kullan
            previously_synced = user.get('synced_closed_days')
            if previously_synced is None:
                previously_synced = prev_closed_days
            previously_synced = previously_synced or []

            # Önceden otomatik eklenmiş ama artık kapalı olmayan günleri kaldır
            to_remove = set(previously_synced) - set(closed_days)
            updated_days_off_set = set(current_days_off) - to_remove

            # Kapalı günleri ekle
            updated_days_off_set |= set(closed_days)

            await db.users.update_one(
                {"username": user['username']},
                {"$set": {"days_off": list(updated_days_off_set), "synced_closed_days": closed_days}}
            )

        logging.info(
            f"✅ Synced business closed days to all staff days_off for org {current_user.organization_id}. closed_days={closed_days}"
        )
    
    # Audit log
    await create_audit_log(
        db=db,
        organization_id=current_user.organization_id,
        user_id=current_user.username,
        user_full_name=current_user.full_name or current_user.username,
        action="UPDATE",
        resource_type="SETTINGS",
        resource_id=current_user.organization_id,
        old_value=current_settings,
        new_value=updated_settings,
        ip_address=request.client.host if request.client else None
    )
    
    return Settings(**updated_settings)

class SettingsLocationUpdate(BaseModel):
    address: str
    coordinates: LocationCoordinates


@api_router.put("/settings/location")
async def update_settings_location(request: Request, data: SettingsLocationUpdate, current_user: UserInDB = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")

    db = await get_db_from_request(request)
    query = {"organization_id": current_user.organization_id}

    current_settings = await db.settings.find_one(query, {"_id": 0})
    old_location = current_settings.get("location") if current_settings else None

    await db.settings.update_one(
        query,
        {"$set": {"location": data.model_dump()}},
        upsert=True
    )

    updated_settings = await db.settings.find_one(query, {"_id": 0})
    new_location = updated_settings.get("location") if updated_settings else None

    await create_audit_log(
        db=db,
        organization_id=current_user.organization_id,
        user_id=current_user.username,
        user_full_name=current_user.full_name or current_user.username,
        action="UPDATE",
        resource_type="SETTINGS_LOCATION",
        resource_id=current_user.organization_id,
        old_value=old_location,
        new_value=new_location,
        ip_address=request.client.host if request.client else None
    )

    return {"location": new_location}


@api_router.delete("/settings/location")
async def delete_settings_location(request: Request, current_user: UserInDB = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")

    db = await get_db_from_request(request)
    query = {"organization_id": current_user.organization_id}

    current_settings = await db.settings.find_one(query, {"_id": 0})
    old_location = current_settings.get("location") if current_settings else None

    await db.settings.update_one(
        query,
        {"$unset": {"location": ""}},
        upsert=False
    )

    await create_audit_log(
        db=db,
        organization_id=current_user.organization_id,
        user_id=current_user.username,
        user_full_name=current_user.full_name or current_user.username,
        action="DELETE",
        resource_type="SETTINGS_LOCATION",
        resource_id=current_user.organization_id,
        old_value=old_location,
        new_value=None,
        ip_address=request.client.host if request.client else None
    )

    return {"message": "Konum kaldırıldı"}

# === ONBOARDING ENDPOINTS ===
@api_router.get("/onboarding/info")
async def get_onboarding_info(request: Request, current_user: UserInDB = Depends(get_current_user)):
    """Onboarding için gerekli bilgileri döndürür (sector, default services, vb.)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    # Settings'den sector bilgisini al
    settings = await db.settings.find_one({"organization_id": current_user.organization_id}, {"_id": 0})
    sector = settings.get("sector") if settings else None
    
    # Mevcut hizmetleri al
    services = await db.services.find(
        {"organization_id": current_user.organization_id}
    ).to_list(100)
    
    # Services'i temizle
    services_clean = []
    for service in services:
        services_clean.append({
            "id": service.get("id"),
            "name": service.get("name"),
            "price": service.get("price", 0),
            "duration": service.get("duration", 30)
        })
    
    # Dil belirleme - request header'dan
    lang = get_request_language(request)
    
    # Sector bazlı default hizmet önerileri (frontend için) - dil bazlı
    default_services = get_sector_default_services(sector, lang) if sector else []
    
    return {
        "user": {
            "username": current_user.username,
            "full_name": current_user.full_name,
            "onboarding_completed": current_user.onboarding_completed
        },
        "sector": sector,
        "existing_services": services_clean,
        "default_services": default_services,
        "business_hours": settings.get("business_hours") if settings else {}
    }

class OnboardingServiceUpdate(BaseModel):
    services: List[dict]  # [{"id": "...", "price": 100, "duration": 30}, ...]

@api_router.post("/onboarding/update-services")
async def update_onboarding_services(request: Request, data: OnboardingServiceUpdate, current_user: UserInDB = Depends(get_current_user)):
    """Onboarding sırasında hizmetlerin fiyat ve sürelerini güncelle"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    updated_services = []
    for service_data in data.services:
        service_id = service_data.get("id")
        price = service_data.get("price", 0)
        duration = service_data.get("duration", 30)
        
        if service_id:
            # Mevcut hizmeti güncelle
            result = await db.services.update_one(
                {"id": service_id, "organization_id": current_user.organization_id},
                {"$set": {"price": price, "duration": duration}}
            )
            
            if result.modified_count > 0:
                updated_service = await db.services.find_one(
                    {"id": service_id, "organization_id": current_user.organization_id},
                    {"_id": 0}
                )
                updated_services.append(updated_service)
    
    return {"message": f"{len(updated_services)} hizmet güncellendi", "services": updated_services}

class OnboardingNewService(BaseModel):
    name: str
    price: float
    duration: int

@api_router.post("/onboarding/add-service")
async def add_onboarding_service(request: Request, data: OnboardingNewService, current_user: UserInDB = Depends(get_current_user)):
    """Onboarding sırasında yeni hizmet ekle"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    service_id = str(uuid.uuid4())
    new_service = Service(
        id=service_id,
        name=data.name,
        price=data.price,
        duration=data.duration,
        organization_id=current_user.organization_id
    )
    
    await db.services.insert_one(new_service.model_dump())

    try:
        await db.users.update_many(
            {"organization_id": current_user.organization_id, "role": "admin"},
            {"$addToSet": {"permitted_service_ids": service_id}}
        )
    except Exception as e:
        logging.error(f"Failed to auto-assign service to admins (onboarding): {e}")
    
    return {"message": "Hizmet eklendi", "service": new_service.model_dump()}

class OnboardingHoursUpdate(BaseModel):
    business_hours: dict

@api_router.post("/onboarding/update-hours")
async def update_onboarding_hours(request: Request, data: OnboardingHoursUpdate, current_user: UserInDB = Depends(get_current_user)):
    """Onboarding sırasında çalışma saatlerini güncelle"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    await db.settings.update_one(
        {"organization_id": current_user.organization_id},
        {"$set": {"business_hours": data.business_hours}},
        upsert=True
    )
    
    return {"message": "Çalışma saatleri güncellendi", "business_hours": data.business_hours}

class OnboardingComplete(BaseModel):
    admin_days_off: Optional[List[str]] = []
    staff_invites: Optional[List[dict]] = []  # [{"username": "...", "full_name": "...", "phone": "..."}]

@api_router.post("/onboarding/complete")
async def complete_onboarding(request: Request, data: OnboardingComplete, current_user: UserInDB = Depends(get_current_user)):
    """Onboarding sihirbazını tamamla - Admin'in tatil günlerini ayarla ve personel davet et"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    # Admin'in tatil günlerini güncelle
    if data.admin_days_off is not None:
        await db.users.update_one(
            {"username": current_user.username},
            {"$set": {"days_off": data.admin_days_off}}
        )
    
    # Personel davetlerini gönder (eğer varsa)
    invited_staff = []
    if data.staff_invites:
        # Organization name'i settings'den al
        org_settings = await db.settings.find_one({"organization_id": current_user.organization_id})
        organization_name = org_settings.get("company_name", "İşletme") if org_settings else "İşletme"
        
        # Admin'in tatil günlerini al (personele de aynısı uygulanacak)
        admin_days_off = data.admin_days_off if data.admin_days_off is not None else []
        
        for staff_data in data.staff_invites:
            username = staff_data.get("username")
            full_name = staff_data.get("full_name")
            
            # Kullanıcı zaten var mı kontrol et
            existing_user = await db.users.find_one({"username": username})
            if existing_user:
                logging.warning(f"⚠️ Kullanıcı zaten mevcut: {username}")
                invited_staff.append({"username": username, "full_name": full_name, "status": "already_exists"})
                continue
            
            # Davet token'ı oluştur
            invitation_token = str(uuid.uuid4())
            
            # Staff için unique slug oluştur (username'den)
            staff_slug = username.split('@')[0] + '-' + str(uuid.uuid4())[:8]
            
            # Yeni staff kullanıcısı oluştur - Admin'in tatil günleriyle
            invite_lang = getattr(current_user, 'language', None) or get_request_language(request)
            if invite_lang not in ['tr', 'en']:
                invite_lang = 'tr'
            staff_user = UserInDB(
                username=username,
                full_name=full_name,
                organization_id=current_user.organization_id,
                role="staff",
                slug=staff_slug,  # Unique slug ekle
                status="pending",
                invitation_token=invitation_token,
                days_off=admin_days_off,  # Admin'in tatil günlerini uygula
                language=invite_lang,
                hashed_password=None  # Şifre davet linkinden belirlenecek
            )
            
            await db.users.insert_one(staff_user.model_dump())
            logging.info(f"✅ Staff kullanıcısı oluşturuldu: {username}")
            
            # Davet e-postası gönder
            try:
                result = await send_personnel_invitation_email(
                    recipient_email=username,
                    recipient_name=full_name,
                    admin_name=current_user.full_name or current_user.username,
                    organization_name=organization_name,
                    invitation_token=invitation_token,
                    lang=invite_lang
                )
                
                if result:
                    invited_staff.append({"username": username, "full_name": full_name, "status": "invited"})
                else:
                    invited_staff.append({"username": username, "full_name": full_name, "status": "email_failed"})
            except Exception as e:
                logging.error(f"❌ Davet e-postası gönderilemedi ({username}): {e}")
                import traceback
                logging.error(traceback.format_exc())
                invited_staff.append({"username": username, "full_name": full_name, "status": "email_failed"})
    
    # Kullanıcının onboarding_completed flag'ini True yap
    await db.users.update_one(
        {"username": current_user.username},
        {"$set": {"onboarding_completed": True}}
    )
    
    # Audit log
    await create_audit_log(
        db=db,
        organization_id=current_user.organization_id,
        user_id=current_user.username,
        user_full_name=current_user.full_name or current_user.username,
        action="UPDATE",
        resource_type="USER",
        resource_id=current_user.username,
        old_value={"onboarding_completed": False},
        new_value={"onboarding_completed": True},
        ip_address=request.client.host if request.client else None
    )
    
    return {
        "message": "Onboarding tamamlandı",
        "onboarding_completed": True,
        "admin_days_off": data.admin_days_off,
        "invited_staff": invited_staff
    }

@api_router.post("/settings/logo")
async def upload_logo(request: Request, file: UploadFile = File(...), current_user: UserInDB = Depends(get_current_user)):
    """Logo upload endpoint"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    # File validation
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="Sadece resim dosyaları yüklenebilir")
    
    # File size check (5MB)
    file_content = await file.read()
    if len(file_content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Dosya boyutu 5MB'dan büyük olamaz")
    
    # Save file to static directory
    static_dir = ROOT_DIR / "static" / "logos"
    static_dir.mkdir(parents=True, exist_ok=True)
    
    file_extension = file.filename.split('.')[-1]
    unique_filename = f"{current_user.organization_id}_{str(uuid.uuid4())[:8]}.{file_extension}"
    file_path = static_dir / unique_filename
    
    with open(file_path, "wb") as f:
        f.write(file_content)
    
    # Update settings with logo URL (with /api prefix for ingress routing)
    logo_url = f"/api/static/logos/{unique_filename}"
    
    db = await get_db_from_request(request)
    query = {"organization_id": current_user.organization_id}
    await db.settings.update_one(query, {"$set": {"logo_url": logo_url}}, upsert=True)
    
    return {"logo_url": logo_url, "message": "Logo başarıyla yüklendi"}


@api_router.post("/upload/image")
async def upload_image(request: Request, file: UploadFile = File(...), current_user: UserInDB = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")

    if not file.content_type or not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="Sadece resim dosyaları yüklenebilir")

    file_content = await file.read()
    if len(file_content) > 8 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Dosya boyutu 8MB'dan büyük olamaz")

    static_dir = ROOT_DIR / "static" / "gallery"
    static_dir.mkdir(parents=True, exist_ok=True)

    unique_filename = f"{current_user.organization_id}_{str(uuid.uuid4())[:8]}.jpg"
    file_path = static_dir / unique_filename

    try:
        img = PilImage.open(io.BytesIO(file_content))
        img = img.convert("RGB")
        img.thumbnail((900, 600), PilImage.LANCZOS)
        img.save(file_path, "JPEG", quality=82, optimize=True)
    except Exception:
        with open(file_path, "wb") as f:
            f.write(file_content)

    url = f"/api/static/gallery/{unique_filename}"
    return {"url": url}

# === USERS/PERSONEL LİSTESİ (Model D) ===
@api_router.get("/users")
async def get_users(request: Request, current_user: UserInDB = Depends(get_current_user)):
    """Aynı organization'daki tüm kullanıcıları listele (şifreler hariç)"""
    db = await get_db_from_request(request)
    
    users = await db.users.find(
        {"organization_id": current_user.organization_id},
        {"_id": 0, "hashed_password": 0}  # Şifreleri gizle
    ).to_list(1000)
    
    return users

# === STAFF/PERSONEL YÖNETİMİ (Model D) ===
class PaymentUpdate(BaseModel):
    payment_type: str
    payment_amount: Optional[float] = None
    days_off: Optional[List[str]] = None

class StaffCreate(BaseModel):
    username: str
    password: Optional[str] = None  # Artık optional, e-posta daveti kullanılıyor
    full_name: Optional[str] = None  # Opsiyonel, e-posta'dan çıkarılabilir
    payment_type: Optional[str] = None  # null olabilir (admin için)
    payment_amount: Optional[float] = None  # null olabilir (admin için)
    role: Optional[str] = "staff"  # "staff" veya "admin" - admin davet için

class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    password: Optional[str] = None
    days_off: Optional[List[str]] = None

@api_router.post("/staff/add")
async def add_staff(request: Request, staff_data: StaffCreate, current_user: UserInDB = Depends(get_current_user)):
    """Admin, yeni personel veya admin ekleyebilir (E-posta daveti ile)"""
    # Sadece admin ekleyebilir
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    # Role validasyonu
    invited_role = staff_data.role or "staff"
    if invited_role not in ["staff", "admin"]:
        raise HTTPException(status_code=400, detail="Geçersiz rol. 'staff' veya 'admin' olmalı")
    
    db = await get_db_from_request(request)
    
    # payment_amount'u float'a çevir (eğer string ise)
    payment_amount = staff_data.payment_amount
    if payment_amount is not None:
        if isinstance(payment_amount, str):
            try:
                payment_amount = float(payment_amount)
            except (ValueError, TypeError):
                payment_amount = 0.0
        elif not isinstance(payment_amount, (int, float)):
            payment_amount = 0.0
    else:
        payment_amount = 0.0
    
    # Kullanıcı adı zaten var mı kontrol et
    existing = await db.users.find_one({"username": staff_data.username})
    if existing:
        raise HTTPException(status_code=400, detail="Bu e-posta adresi zaten kayıtlı")
    
    # Invitation token oluştur
    invitation_token = str(uuid.uuid4())
    
    # İşletme adını al
    settings = await db.settings.find_one({"organization_id": current_user.organization_id})
    organization_name = settings.get("company_name", "İşletme") if settings else "İşletme"

    # İşletmenin kapalı günlerini yeni personele varsayılan olarak uygula
    closed_days = []
    business_hours = (settings or {}).get("business_hours") or {}
    day_names = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    for day in day_names:
        day_settings = business_hours.get(day, {}) or {}
        if not day_settings.get("is_open", True):
            closed_days.append(day)
    
    try:
        # Yeni personel oluştur (password olmadan, pending status ile)
        # Eğer full_name yoksa, email'den isim çıkar
        full_name = staff_data.full_name
        if not full_name:
            # Email'den isim çıkar (örn: john.doe@example.com -> John Doe)
            email_local = staff_data.username.split('@')[0]
            name_parts = email_local.split('.')
            if len(name_parts) > 1:
                full_name = ' '.join([part.capitalize() for part in name_parts])
            else:
                full_name = email_local.capitalize()
        
        # Yeni eklenen kullanıcıya (admin veya staff) org'daki tüm aktif hizmetleri otomatik ata.
        # Admin sonradan Personel Yönetimi'nden istemediğini kaldırabilir.
        services = await db.services.find({"organization_id": current_user.organization_id}).to_list(None)
        all_service_ids = [s.get("id") for s in services if s.get("id")]

        new_user = UserInDB(
            username=staff_data.username,
            full_name=full_name,
            hashed_password=None,  # Şifre henüz belirlenmedi
            organization_id=current_user.organization_id,
            role=invited_role,  # staff veya admin rolü
            slug=None,  # Admin/Personellerin slug'ı yok
            permitted_service_ids=all_service_ids,  # Hem admin hem staff tüm hizmetleri otomatik alır
            payment_type=staff_data.payment_type or "salary" if invited_role == "staff" else None,
            payment_amount=payment_amount if invited_role == "staff" else None,
            status="pending",  # Bekliyor durumu
            invitation_token=invitation_token,
            days_off=closed_days if invited_role == "staff" else [],
            language=getattr(current_user, 'language', None) or get_request_language(request)
        )
        
        user_dict = new_user.model_dump()
        # Personel için slug field'ını kaldır (MongoDB unique index hatası önlemek için)
        user_dict.pop('slug', None)
        if invited_role == "staff":
            user_dict["synced_closed_days"] = closed_days
        # Admin için onboarding tamamlanmış olarak işaretle (organization zaten kurulu)
        if invited_role == "admin":
            user_dict["onboarding_completed"] = True

        # Stable user_id (UUID) — PostHog identity'nin sabit kaynağı
        user_dict["user_id"] = str(uuid.uuid4())

        # Aha activation_state:
        # - staff: hiç görmez ("not_applicable")
        # - davet edilen admin: mevcut org'a katılır, kuruluma gerek yok → "aha_skipped"
        if invited_role == "admin":
            user_dict["activation_state"] = "aha_skipped"
            user_dict["aha_skipped_reason"] = "invited_admin"
        else:
            user_dict["activation_state"] = "not_applicable"

        # MongoDB'ye ekle
        await db.users.insert_one(user_dict)
        
        # E-posta daveti gönder
        email_sent = await send_personnel_invitation_email(
            recipient_email=staff_data.username,
            recipient_name=full_name,
            admin_name=current_user.full_name or current_user.username,
            organization_name=organization_name,
            invitation_token=invitation_token,
            lang=getattr(current_user, 'language', None) or get_request_language(request)
        )
        
        if not email_sent:
            logging.warning(f"{'Admin' if invited_role == 'admin' else 'Personel'} eklendi ancak e-posta gönderilemedi: {staff_data.username}")
        
    except Exception as e:
        logging.error(f"{'Admin' if invited_role == 'admin' else 'Personel'} ekleme hatası: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"{'Admin' if invited_role == 'admin' else 'Personel'} eklenirken bir hata oluştu: {str(e)}")
    
    role_text = "Admin" if invited_role == "admin" else "Personel"
    return {"message": f"{role_text} başarıyla eklendi ve davet e-postası gönderildi", "username": staff_data.username, "full_name": staff_data.full_name, "role": invited_role}

@api_router.put("/users/me")
async def update_current_user(request: Request, user_update: UserUpdate, current_user: UserInDB = Depends(get_current_user), db = Depends(get_db)):
    """Mevcut kullanıcının kendi bilgilerini güncelle"""
    try:
        update_data = {}
        
        if user_update.full_name is not None:
            update_data["full_name"] = user_update.full_name
        
        if user_update.password is not None:
            if len(user_update.password) < 6:
                raise HTTPException(status_code=400, detail="Şifre en az 6 karakter olmalıdır")
            update_data["hashed_password"] = get_password_hash(user_update.password)
        
        if user_update.days_off is not None:
            update_data["days_off"] = user_update.days_off
        
        if not update_data:
            raise HTTPException(status_code=400, detail="Güncellenecek alan belirtilmedi")
        
        await db.users.update_one(
            {"username": current_user.username, "organization_id": current_user.organization_id},
            {"$set": update_data}
        )
        
        # Audit log (hata olursa skip edilir)
        try:
            await create_audit_log(
                db=db,
                organization_id=current_user.organization_id,
                user_id=current_user.username,
                user_full_name=current_user.full_name or current_user.username,
                action="update",
                resource_type="user",
                resource_id=current_user.username,
                new_value={"updated_fields": list(update_data.keys())}
            )
        except Exception as audit_error:
            logger.warning(f"Audit log oluşturulamadı: {audit_error}")
        
        return {"message": "Profil bilgileri güncellendi"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Kullanıcı güncelleme hatası: {e}")
        raise HTTPException(status_code=500, detail="Profil güncellenirken hata oluştu")

@api_router.put("/staff/{staff_id}/payment")
async def update_staff_payment(request: Request, staff_id: str, payment_data: PaymentUpdate, current_user: UserInDB = Depends(get_current_user)):
    """Admin, personelin ödeme ayarlarını (maaş/prim) güncelleyebilir"""
    try:
        # Sadece admin güncelleyebilir
        if current_user.role != "admin":
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        db = await get_db_from_request(request)
        
        # URL decode staff_id (email olabilir)
        import urllib.parse
        staff_id_decoded = urllib.parse.unquote(staff_id)
        
        # Personelin aynı organization'da olduğunu kontrol et
        staff = await db.users.find_one({"username": staff_id_decoded, "organization_id": current_user.organization_id})
        if not staff:
            raise HTTPException(status_code=404, detail="Personel bulunamadı veya erişim yok")
        
        # payment_amount'u float'a çevir (eğer string ise)
        payment_amount = payment_data.payment_amount
        if payment_amount is not None:
            if isinstance(payment_amount, str):
                try:
                    payment_amount = float(payment_amount)
                except (ValueError, TypeError):
                    payment_amount = 0.0
            elif not isinstance(payment_amount, (int, float)):
                payment_amount = 0.0
        else:
            payment_amount = 0.0
        
        # Payment bilgilerini güncelle
        update_fields = {
            "payment_type": payment_data.payment_type,
            "payment_amount": payment_amount
        }
        
        # days_off varsa ekle (PaymentUpdate modelinde artık optional)
        if hasattr(payment_data, 'days_off') and payment_data.days_off is not None:
            update_fields["days_off"] = payment_data.days_off
        
        await db.users.update_one(
            {"username": staff_id_decoded, "organization_id": current_user.organization_id},
            {"$set": update_fields}
        )
        
        logging.info(f"Personel ödeme ayarları güncellendi: {staff_id_decoded}, payment_type={payment_data.payment_type}, payment_amount={payment_amount}")
        
        return {"message": "Personel ödeme ayarları güncellendi", "staff_id": staff_id_decoded, "payment_type": payment_data.payment_type, "payment_amount": payment_amount}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Personel ödeme ayarı güncelleme hatası: {type(e).__name__}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Ödeme ayarları güncellenirken hata oluştu: {str(e)}")

@api_router.put("/staff/{staff_id}/days-off")
async def update_staff_days_off(request: Request, staff_id: str, days_off_data: dict, current_user: UserInDB = Depends(get_current_user)):
    """Admin, personelin tatil günlerini güncelleyebilir"""
    # Sadece admin güncelleyebilir
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    # Personelin aynı organization'da olduğunu kontrol et
    staff = await db.users.find_one({"username": staff_id, "organization_id": current_user.organization_id})
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı veya erişim yok")
    
    # days_off'u al
    days_off = days_off_data.get('days_off', [])
    if not isinstance(days_off, list):
        raise HTTPException(status_code=400, detail="days_off bir liste olmalıdır")
    
    # days_off'u güncelle
    await db.users.update_one(
        {"username": staff_id, "organization_id": current_user.organization_id},
        {"$set": {"days_off": days_off}}
    )
    
    logging.info(f"Personel tatil günleri güncellendi: {staff_id}, days_off={days_off}")
    
    return {"message": "Personel tatil günleri güncellendi", "staff_id": staff_id, "days_off": days_off}

@api_router.put("/staff/{staff_id}/toggle-permission")
async def toggle_staff_permission(request: Request, staff_id: str, permission_data: dict, current_user: UserInDB = Depends(get_current_user)):
    """Admin, personelin tüm randevuları görme yetkisini açıp kapatabilir"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    staff = await db.users.find_one({"username": staff_id, "organization_id": current_user.organization_id})
    if not staff:
        raise HTTPException(status_code=404, detail="Personel bulunamadı veya erişim yok")
    
    can_view_all = permission_data.get('can_view_all_appointments', False)
    
    await db.users.update_one(
        {"username": staff_id, "organization_id": current_user.organization_id},
        {"$set": {"can_view_all_appointments": bool(can_view_all)}}
    )
    
    logging.info(f"Personel yetki güncellendi: {staff_id}, can_view_all_appointments={can_view_all}")
    
    return {"message": "Personel yetkisi güncellendi", "staff_id": staff_id, "can_view_all_appointments": can_view_all}

@api_router.delete("/staff/{staff_id}")
async def delete_staff(request: Request, staff_id: str, current_user: UserInDB = Depends(get_current_user)):
    """Admin, personel/yönetici silebilir (kurucu admin diğer yöneticileri silebilir)"""
    # Sadece admin silebilir
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    # Silinecek kullanıcıyı bul
    staff = await db.users.find_one({"username": staff_id, "organization_id": current_user.organization_id})
    if not staff:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı veya erişim yok")
    
    # Kendini silemez
    if staff.get("username") == current_user.username:
        raise HTTPException(status_code=400, detail="Kendinizi silemezsiniz")
    
    # Eğer silinecek kullanıcı admin ise, kurucu admin kontrolü yap
    if staff.get("role") == "admin":
        # Kurucu admin'i bul (is_founder flag'i veya en eski admin)
        founder = await db.users.find_one({
            "organization_id": current_user.organization_id,
            "role": "admin",
            "is_founder": True
        })
        
        # Eğer is_founder flag yoksa, en eski admin'i kurucu kabul et
        if not founder:
            oldest_admin = await db.users.find_one(
                {"organization_id": current_user.organization_id, "role": "admin"},
                sort=[("_id", 1)]  # MongoDB ObjectId sıralı, ilk oluşturulan en eski
            )
            founder = oldest_admin
        
        # Silinecek kişi kurucu admin ise silinemez
        if staff.get("username") == founder.get("username"):
            raise HTTPException(status_code=400, detail="Kurucu yönetici silinemez")
        
        # Silmeye çalışan kişi kurucu admin değilse, diğer adminleri silemez
        if current_user.username != founder.get("username"):
            raise HTTPException(status_code=403, detail="Sadece kurucu yönetici diğer yöneticileri silebilir")
    
    # Kullanıcıyı sil
    await db.users.delete_one({"username": staff_id, "organization_id": current_user.organization_id})
    
    role_text = "Yönetici" if staff.get("role") == "admin" else "Personel"
    return {"message": f"{role_text} başarıyla silindi"}

@api_router.put("/staff/{staff_id}/services")
async def update_staff_services(request: Request, staff_id: str, service_ids: List[str], current_user: UserInDB = Depends(get_current_user)):
    """Admin, personelin verebileceği hizmetleri güncelleyebilir (sadece kurucu admin diğer adminleri düzenleyebilir)"""
    # Sadece admin güncelleyebilir
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    # Düzenlenecek kullanıcıyı bul
    staff = await db.users.find_one({"username": staff_id, "organization_id": current_user.organization_id})
    if not staff:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı veya erişim yok")
    
    # Eğer düzenlenecek kullanıcı admin ise, kurucu admin kontrolü yap
    if staff.get("role") == "admin":
        # Kurucu admin'i bul
        founder = await db.users.find_one({
            "organization_id": current_user.organization_id,
            "role": "admin",
            "is_founder": True
        })
        if not founder:
            founder = await db.users.find_one(
                {"organization_id": current_user.organization_id, "role": "admin"},
                sort=[("_id", 1)]
            )
        
        # Admin kendi hizmetlerini düzenleyebilir, diğer adminlerin hizmetlerini sadece kurucu düzenleyebilir
        if staff_id != current_user.username and current_user.username != founder.get("username"):
            raise HTTPException(status_code=403, detail="Sadece kurucu yönetici diğer yöneticilerin hizmetlerini düzenleyebilir")
    
    # Kullanıcının permitted_service_ids'ini güncelle
    await db.users.update_one(
        {"username": staff_id, "organization_id": current_user.organization_id},
        {"$set": {"permitted_service_ids": service_ids}}
    )
    
    return {"message": "Kullanıcı hizmetleri güncellendi", "staff_id": staff_id, "permitted_service_ids": service_ids}

# === STAFF BREAKS ROUTES ===
class BreakCreate(BaseModel):
    date: str  # YYYY-MM-DD
    start_time: str  # HH:MM
    end_time: str  # HH:MM

@api_router.get("/staff/breaks")
async def get_my_breaks(request: Request, date: Optional[str] = None, current_user: UserInDB = Depends(get_current_user)):
    """Personel kendi molalarını listeler"""
    db = await get_db_from_request(request)
    user = await db.users.find_one(
        {"username": current_user.username, "organization_id": current_user.organization_id},
        {"_id": 0, "breaks": 1}
    )
    breaks = user.get("breaks", []) if user else []
    
    # Tarih filtresi varsa uygula
    if date:
        breaks = [b for b in breaks if b.get("date") == date]
    
    return {"breaks": breaks}

@api_router.post("/staff/breaks")
async def add_break(request: Request, break_data: BreakCreate, current_user: UserInDB = Depends(get_current_user)):
    """Personel mola ekler"""
    db = await get_db_from_request(request)

    # Mevcut molaları al
    user = await db.users.find_one(
        {"username": current_user.username, "organization_id": current_user.organization_id},
        {"_id": 0, "breaks": 1}
    )
    breaks = user.get("breaks", []) if user else []
    
    # Yeni mola süresini hesapla
    start_parts = break_data.start_time.split(":")
    end_parts = break_data.end_time.split(":")
    start_minutes = int(start_parts[0]) * 60 + int(start_parts[1])
    end_minutes = int(end_parts[0]) * 60 + int(end_parts[1])
    new_break_duration = end_minutes - start_minutes
    
    if new_break_duration <= 0:
        raise HTTPException(status_code=400, detail="Bitiş saati başlangıç saatinden sonra olmalı")
    
    if new_break_duration < 15:
        raise HTTPException(status_code=400, detail="Mola en az 15 dakika olmalı")
    
    # Randevu çakışması kontrolü
    appointments = await db.appointments.find(
        {
            "organization_id": current_user.organization_id,
            "appointment_date": break_data.date,
            "staff_member_id": current_user.username,
            "status": {"$ne": "İptal"}
        },
        {"_id": 0, "appointment_time": 1, "service_id": 1, "service_duration": 1, "customer_name": 1}
    ).to_list(100)
    
    # Her randevunun bitiş saatini hesapla ve çakışma kontrol et
    conflicting_appointments = []
    for appt in appointments:
        # Randevu süresini al — çoklu hizmette saklı TOPLAM süre
        stored_dur = appt.get("service_duration")
        if stored_dur:
            appt_duration = int(stored_dur)
        else:
            appt_service = await db.services.find_one({"id": appt.get("service_id")}, {"_id": 0, "duration": 1})
            appt_duration = appt_service.get("duration", 30) if appt_service else 30
        
        appt_start_time = appt["appointment_time"]
        appt_start_parts = appt_start_time.split(":")
        appt_start_minutes = int(appt_start_parts[0]) * 60 + int(appt_start_parts[1])
        appt_end_minutes = appt_start_minutes + appt_duration
        
        # Çakışma kontrolü: Mola ile randevu saatleri çakışıyor mu?
        if not (end_minutes <= appt_start_minutes or start_minutes >= appt_end_minutes):
            appt_end_hour = appt_end_minutes // 60
            appt_end_minute = appt_end_minutes % 60
            appt_end_time = f"{str(appt_end_hour).zfill(2)}:{str(appt_end_minute).zfill(2)}"
            conflicting_appointments.append({
                "customer_name": appt.get("customer_name", ""),
                "time": f"{appt_start_time} - {appt_end_time}"
            })
    
    if conflicting_appointments:
        conflict_details = ", ".join([f"{a['customer_name']} ({a['time']})" for a in conflicting_appointments])
        raise HTTPException(
            status_code=400, 
            detail=f"Bu saatte randevunuz var: {conflict_details}. Lütfen farklı bir saat seçin."
        )
    
    # Yeni mola oluştur
    new_break = {
        "id": str(uuid.uuid4()),
        "date": break_data.date,
        "start_time": break_data.start_time,
        "end_time": break_data.end_time,
        "created_at": datetime.utcnow().isoformat()
    }
    
    # Veritabanına ekle
    await db.users.update_one(
        {"username": current_user.username, "organization_id": current_user.organization_id},
        {"$push": {"breaks": new_break}}
    )
    
    return {"message": "Mola eklendi", "break": new_break}

@api_router.delete("/staff/breaks/{break_id}")
async def delete_break(request: Request, break_id: str, current_user: UserInDB = Depends(get_current_user)):
    """Personel molasını siler"""
    db = await get_db_from_request(request)
    
    result = await db.users.update_one(
        {"username": current_user.username, "organization_id": current_user.organization_id},
        {"$pull": {"breaks": {"id": break_id}}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Mola bulunamadı")
    
    return {"message": "Mola silindi"}


# === ADMIN TIME BLOCK ROUTES ===
@api_router.get("/admin/staff/{staff_id}/breaks")
async def admin_get_staff_breaks(
    request: Request,
    staff_id: str,
    date: Optional[str] = None,
    current_user: UserInDB = Depends(get_current_user)
):
    """Admin, personelin zaman bloklarını listeler"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    db = await get_db_from_request(request)
    user = await db.users.find_one(
        {"username": staff_id, "organization_id": current_user.organization_id},
        {"_id": 0, "breaks": 1}
    )
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    breaks = user.get("breaks", [])
    if date:
        breaks = [b for b in breaks if b.get("date") == date]
    return {"breaks": breaks}


@api_router.post("/admin/staff/{staff_id}/breaks")
async def admin_add_staff_break(
    request: Request,
    staff_id: str,
    break_data: BreakCreate,
    current_user: UserInDB = Depends(get_current_user)
):
    """Admin, personel için zaman kapatma (blok) ekler"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    db = await get_db_from_request(request)

    staff_user = await db.users.find_one(
        {"username": staff_id, "organization_id": current_user.organization_id},
        {"_id": 0, "username": 1}
    )
    if not staff_user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")

    # Süre hesapla
    start_parts = break_data.start_time.split(":")
    end_parts = break_data.end_time.split(":")
    start_minutes = int(start_parts[0]) * 60 + int(start_parts[1])
    end_minutes = int(end_parts[0]) * 60 + int(end_parts[1])
    new_break_duration = end_minutes - start_minutes

    if new_break_duration <= 0:
        raise HTTPException(status_code=400, detail="Bitiş saati başlangıç saatinden sonra olmalı")
    if new_break_duration < 15:
        raise HTTPException(status_code=400, detail="Mola en az 15 dakika olmalı")

    # Randevu çakışması kontrolü
    appointments = await db.appointments.find(
        {
            "organization_id": current_user.organization_id,
            "appointment_date": break_data.date,
            "staff_member_id": staff_id,
            "status": {"$ne": "İptal"}
        },
        {"_id": 0, "appointment_time": 1, "service_id": 1, "service_duration": 1, "customer_name": 1}
    ).to_list(100)

    conflicting_appointments = []
    for appt in appointments:
        # Çoklu hizmette saklı TOPLAM süre
        stored_dur = appt.get("service_duration")
        if stored_dur:
            appt_duration = int(stored_dur)
        else:
            appt_service = await db.services.find_one({"id": appt.get("service_id")}, {"_id": 0, "duration": 1})
            appt_duration = appt_service.get("duration", 30) if appt_service else 30

        appt_start_time = appt["appointment_time"]
        appt_start_parts = appt_start_time.split(":")
        appt_start_minutes = int(appt_start_parts[0]) * 60 + int(appt_start_parts[1])
        appt_end_minutes = appt_start_minutes + appt_duration

        if not (end_minutes <= appt_start_minutes or start_minutes >= appt_end_minutes):
            appt_end_hour = appt_end_minutes // 60
            appt_end_minute = appt_end_minutes % 60
            appt_end_time = f"{str(appt_end_hour).zfill(2)}:{str(appt_end_minute).zfill(2)}"
            conflicting_appointments.append({
                "customer_name": appt.get("customer_name", ""),
                "time": f"{appt_start_time} - {appt_end_time}"
            })

    if conflicting_appointments:
        conflict_details = ", ".join([f"{a['customer_name']} ({a['time']})" for a in conflicting_appointments])
        raise HTTPException(
            status_code=400,
            detail=f"Bu saatte randevunuz var: {conflict_details}. Lütfen farklı bir saat seçin."
        )

    new_break = {
        "id": str(uuid.uuid4()),
        "date": break_data.date,
        "start_time": break_data.start_time,
        "end_time": break_data.end_time,
        "created_at": datetime.utcnow().isoformat()
    }

    await db.users.update_one(
        {"username": staff_id, "organization_id": current_user.organization_id},
        {"$push": {"breaks": new_break}}
    )

    return {"message": "Mola eklendi", "break": new_break}


@api_router.delete("/admin/staff/{staff_id}/breaks/{break_id}")
async def admin_delete_staff_break(
    request: Request,
    staff_id: str,
    break_id: str,
    current_user: UserInDB = Depends(get_current_user)
):
    """Admin, personelin zaman bloğunu siler"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    db = await get_db_from_request(request)
    result = await db.users.update_one(
        {"username": staff_id, "organization_id": current_user.organization_id},
        {"$pull": {"breaks": {"id": break_id}}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Mola bulunamadı")
    return {"message": "Mola silindi"}

# === CUSTOMERS ROUTES ===
# ═══ CONTRACT: v1 (FROZEN / RESPONSE) — legacy path (limit YOK) ══════════════
# limit YOKKEN flat array döner; randevusuz müşteri is_pending:true, randevulu
# müşteri is_pending İÇERMEZ. Bu şekli değiştirme (donmuş build parse ediyor).
# Yeni davranış = ?limit=&cursor= (paginated) arkasında. bkz. OPERATIONS.md §2.
@api_router.get("/customers")
async def get_customers(
    request: Request,
    limit: Optional[int] = Query(None, ge=1, le=100),
    cursor: Optional[str] = None,
    search: Optional[str] = None,
    current_user: UserInDB = Depends(get_current_user),
):
    """Müşteri listesi.

    Backward-compat kritik: mobil iOS build'ler ve diğer eski client'lar
    `limit` param'ı GÖNDERMEZ → BİREBİR eski shape (appointments+customers
    merge, is_pending sadece randevusuz için) döner. Ekstra alan (`id`,
    `notes`, `last_appointment_at`, `first_appointment_at`) EKLENMEZ →
    eski client hâlâ doğru davranır.

    Yeni web (Customers.js) ise `limit` gönderirse cursor pagination
    aktif olur — `{items, next_cursor, has_more}` formatı döner.
    """
    db = await get_db_from_request(request)
    org_id = current_user.organization_id
    use_pagination = limit is not None

    # ------------------------------------------------------------------
    # BACKWARD-COMPAT PATH (limit YOK) — ESKİ BUILD'LERİN ÇALIŞTIĞI VERSİYON
    # ------------------------------------------------------------------
    # ÖNEMLİ: Bu path DELİBERATLY appointments + customers MERGE yapıyor.
    # Denormalize-only optimizasyonu (appointments taramasını atlamak) mobil
    # eski build'lerde müşteri detay ekranında hatalı liste render'ına yol
    # açtı → geri alındı. Eski build'lerin beklediği birebir shape:
    #   - Randevusu OLAN müşteri: {name(customer_name), phone, total, completed}
    #     → `is_pending` ANAHTARI YOK
    #   - Randevusu OLMAYAN (manuel eklenmiş) müşteri: yukarıdakiler + is_pending:True
    # App Store onayı ~2 gün sürdüğü için bu geriye-uyum kritik; performans
    # (yeni web zaten limit=30 paginated path'i kullanıyor) ikincil.
    if not use_pagination:
        try:
            appointments = await db.appointments.find(
                {"organization_id": org_id},
                {"_id": 0, "phone": 1, "customer_name": 1, "status": 1}
            ).to_list(10000)
        except Exception as exc:
            logging.error(f"GET /customers (legacy) appointments query failed: {exc}", exc_info=True)
            appointments = []

        customer_map: dict = {}
        for apt in appointments:
            phone = apt.get('phone')
            if not phone:
                continue
            if phone not in customer_map:
                customer_map[phone] = {
                    "name": apt.get('customer_name', ''),
                    "phone": phone,
                    "total_appointments": 0,
                    "completed_appointments": 0,
                }
            customer_map[phone]['total_appointments'] += 1
            if apt.get('status') == 'Tamamlandı':
                customer_map[phone]['completed_appointments'] += 1

        try:
            db_customers = await db.customers.find(
                {"organization_id": org_id},
                {"_id": 0, "name": 1, "phone": 1}
            ).to_list(50000)
            for dc in db_customers:
                ph = dc.get('phone')
                if ph and ph not in customer_map:
                    customer_map[ph] = {
                        "name": dc.get('name', ''),
                        "phone": ph,
                        "total_appointments": 0,
                        "completed_appointments": 0,
                        "is_pending": True,  # SADECE randevusuz — eski client anlaşması
                    }
        except Exception as exc:
            logging.warning(f"GET /customers (legacy) customers collection query failed: {exc}")

        result = list(customer_map.values())
        result.sort(key=lambda x: (x.get('name') or '').lower())
        logging.info(f"📋 GET /customers (legacy merge): {len(result)} müşteri (org={org_id[:8]})")
        return result

    # ------------------------------------------------------------------
    # NEW CURSOR PAGINATION PATH (limit VAR) — yeni web Customers.js için
    # ------------------------------------------------------------------
    query: dict = {"organization_id": org_id}
    effective_limit = int(limit)

    if search and search.strip():
        _s = search.strip()
        # Türkçe karakter-tolerant SUBSTRING match — kullanıcı "şenyüz" veya
        # "senyuz" ya da "Şen" yazsın, "Fatih Şenyüz" bulunur.
        name_pattern = build_turkish_case_insensitive_pattern(_s)
        # Phone tarafı: rakam-only karşılaştırma; escape ile substring.
        phone_pattern = re.escape(_s)
        query["$or"] = [
            {"name": {"$regex": name_pattern}},
            {"phone": {"$regex": phone_pattern}},
        ]

    cursor_name: Optional[str] = None
    cursor_id: Optional[str] = None
    if cursor:
        try:
            raw = base64.urlsafe_b64decode(cursor.encode()).decode()
            cursor_name, cursor_id = raw.split("|", 1)
        except Exception:
            logging.warning(f"Bozuk cursor atlandı: {cursor}")

    if cursor_name is not None and cursor_id is not None:
        try:
            from bson import ObjectId
            _cid: object = ObjectId(cursor_id)
        except Exception:
            _cid = cursor_id
        tie_break = {
            "$or": [
                {"name": {"$gt": cursor_name}},
                {"name": cursor_name, "_id": {"$gt": _cid}},
            ]
        }
        existing = {k: v for k, v in query.items() if k != "organization_id"}
        new_q: dict = {"organization_id": org_id, "$and": [tie_break]}
        if existing:
            new_q["$and"].append(existing)
        query = new_q

    projection = {
        "_id": 1, "id": 1, "name": 1, "phone": 1, "notes": 1,
        "total_appointments": 1, "completed_appointments": 1,
        "last_appointment_at": 1, "first_appointment_at": 1,
    }

    try:
        docs = await (
            db.customers.find(query, projection)
            .collation({"locale": "tr", "strength": 2})
            .sort([("name", 1), ("_id", 1)])
            .limit(effective_limit + 1)
        ).to_list(effective_limit + 1)
    except Exception as exc:
        logging.error(f"GET /customers (paginated) query failed: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail="Müşteri listesi getirilemedi")

    has_more = len(docs) > effective_limit
    if has_more:
        docs = docs[:effective_limit]

    items = []
    for d in docs:
        total = int(d.get("total_appointments") or 0)
        completed = int(d.get("completed_appointments") or 0)
        items.append({
            "id": d.get("id") or (str(d.get("_id")) if d.get("_id") else None),
            "name": d.get("name") or "",
            "phone": d.get("phone") or "",
            "notes": d.get("notes") or "",
            "total_appointments": total,
            "completed_appointments": completed,
            "last_appointment_at": d.get("last_appointment_at"),
            "first_appointment_at": d.get("first_appointment_at"),
            "is_pending": total == 0,
        })

    next_cursor: Optional[str] = None
    if has_more and docs:
        last_doc = docs[-1]
        try:
            raw = f"{last_doc.get('name') or ''}|{str(last_doc.get('_id'))}"
            next_cursor = base64.urlsafe_b64encode(raw.encode()).decode()
        except Exception:
            next_cursor = None

    logging.info(
        f"📋 GET /customers (paginated): {len(items)} müşteri (org={org_id[:8]}, "
        f"cursor={'✓' if cursor else '✗'}, search={'✓' if search else '✗'}, has_more={has_more})"
    )

    return {"items": items, "next_cursor": next_cursor, "has_more": has_more}

class CustomerCreate(BaseModel):
    name: str = Field(..., min_length=1, description="Müşteri adı")
    phone: str = Field(..., min_length=10, description="Telefon numarası")

@api_router.post("/customers")
async def create_customer(request: Request, customer_data: CustomerCreate, current_user: UserInDB = Depends(get_current_user)):
    """Yeni müşteri ekle (Sadece admin)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    name = customer_data.name.strip()
    phone = customer_data.phone.strip()
    
    if not name or not phone:
        raise HTTPException(status_code=400, detail="Müşteri adı ve telefon numarası gereklidir")
    
    # Telefon numarasını normalize et
    clean_phone = re.sub(r'\D', '', phone)
    if len(clean_phone) < 10:
        raise HTTPException(status_code=400, detail="Geçerli bir telefon numarası girin")
    
    # Telefon varyasyonlarını oluştur (farklı formatlarla kaydedilmiş olabilir)
    phone_variants = [clean_phone]
    if clean_phone.startswith('90') and len(clean_phone) == 12:
        phone_variants.append(clean_phone[2:])           # 5362231743
        phone_variants.append('0' + clean_phone[2:])     # 05362231743
    elif len(clean_phone) == 10 and not clean_phone.startswith('0'):
        phone_variants.append('90' + clean_phone)        # 905362231743
        phone_variants.append('0' + clean_phone)         # 05362231743
    
    # Aynı telefon numarasına sahip müşteri var mı kontrol et (randevulardan ve customers collection'ından)
    existing_appointment = await db.appointments.find_one(
        {"organization_id": current_user.organization_id, "phone": {"$in": phone_variants}},
        {"_id": 0, "id": 1}
    )
    
    if existing_appointment:
        raise HTTPException(status_code=400, detail="Bu telefon numarasına sahip bir müşteri zaten var")
    
    # customers collection'ında da kontrol et
    existing_customer = await db.customers.find_one(
        {"organization_id": current_user.organization_id, "phone": {"$in": phone_variants}},
        {"_id": 0, "id": 1}
    )
    
    if existing_customer:
        raise HTTPException(status_code=400, detail="Bu telefon numarasına sahip bir müşteri zaten var")
    
    # Müşteriyi customers collection'ına kaydet (eğer collection yoksa otomatik oluşturulur)
    customer_doc = {
        "id": str(uuid.uuid4()),
        "organization_id": current_user.organization_id,
        "name": name,
        "phone": clean_phone,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "notes": ""
    }
    
    # customers collection'ına kaydet
    await db.customers.insert_one(customer_doc)
    
    # Audit log
    await create_audit_log(
        db=db,
        organization_id=current_user.organization_id,
        user_id=current_user.username,
        user_full_name=current_user.full_name or current_user.username,
        action="CREATE",
        resource_type="CUSTOMER",
        resource_id=customer_doc["id"],
        new_value=customer_doc,
        ip_address=request.client.host if request.client else None
    )
    
    logging.info(f"New customer created: {name} ({phone}) for org {current_user.organization_id}")

    try:
        await invalidate_cache(request, "customers_list", current_user)
    except Exception as e:
        logging.warning(f"Customer cache invalidation failed: {e}")
    
    try:
        await emit_to_organization(current_user.organization_id, 'customer_added', {'customer': customer_doc})
    except Exception as e:
        logging.warning(f"Failed to emit customer_added event: {e}")
    
    return {
        "id": customer_doc["id"],
        "name": name,
        "phone": phone,
        "message": "Müşteri başarıyla eklendi"
    }

class CustomerUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None

@api_router.put("/customers/{phone}")
async def update_customer(request: Request, phone: str, update_data: CustomerUpdate, current_user: UserInDB = Depends(get_current_user)):
    """Müşteri bilgilerini güncelle (ad, telefon). Admin ve personel kullanabilir."""
    db = await get_db_from_request(request)
    customer = await db.customers.find_one({"phone": phone, "organization_id": current_user.organization_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    
    updates = {}
    if update_data.name and update_data.name.strip():
        updates["name"] = update_data.name.strip()
    
    new_phone = None
    if update_data.phone and update_data.phone.strip():
        new_phone = re.sub(r'\D', '', update_data.phone.strip())
        if len(new_phone) < 10:
            raise HTTPException(status_code=400, detail="Geçerli bir telefon numarası girin")
        if new_phone != phone:
            existing = await db.customers.find_one({"phone": new_phone, "organization_id": current_user.organization_id})
            if existing:
                raise HTTPException(status_code=400, detail="Bu telefon numarasına sahip başka bir müşteri zaten var")
            updates["phone"] = new_phone
    
    if not updates:
        raise HTTPException(status_code=400, detail="Güncellenecek bir bilgi yok")
    
    await db.customers.update_one({"phone": phone, "organization_id": current_user.organization_id}, {"$set": updates})
    
    # Telefon değiştiyse randevuları da güncelle
    if new_phone and new_phone != phone:
        await db.appointments.update_many(
            {"phone": phone, "organization_id": current_user.organization_id},
            {"$set": {"phone": new_phone}}
        )
    # İsim değiştiyse randevulardaki müşteri adını da güncelle
    if "name" in updates:
        phone_to_match = new_phone if new_phone and new_phone != phone else phone
        await db.appointments.update_many(
            {"phone": phone_to_match, "organization_id": current_user.organization_id},
            {"$set": {"customer_name": updates["name"]}}
        )
    
    try:
        await invalidate_cache(request, "customers_list", current_user)
        await invalidate_cache(request, "dashboard_stats", current_user)
    except Exception:
        pass
    
    try:
        await emit_to_organization(current_user.organization_id, 'customer_updated', {'phone': phone})
    except Exception:
        pass
    
    logging.info(f"Customer updated: {phone} -> {updates} for org {current_user.organization_id}")
    return {"message": "Müşteri bilgileri güncellendi", "updates": updates}

@api_router.delete("/customers/{phone}")
async def delete_customer(request: Request, phone: str, current_user: UserInDB = Depends(get_current_user)):
    """Müşteriyi ve TÜM randevularını sil"""
    # Sadece admin silebilir
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    # Önce customers collection'ından müşteriyi bul
    customer_query = {"phone": phone, "organization_id": current_user.organization_id}
    customer = await db.customers.find_one(customer_query, {"_id": 0})
    
    # Randevuları sil (varsa)
    appointment_query = {"phone": phone, "organization_id": current_user.organization_id}
    appointments_to_delete = await db.appointments.find(appointment_query, {"_id": 0}).to_list(1000)
    appointment_result = await db.appointments.delete_many(appointment_query)
    
    # Transaction'ları da sil (eğer varsa)
    transaction_result = await db.transactions.delete_many(appointment_query)
    
    # Customers collection'ından müşteriyi sil
    customer_result = await db.customers.delete_many(customer_query)
    
    # Eğer ne müşteri ne de randevu bulunamadıysa hata ver
    if customer_result.deleted_count == 0 and appointment_result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bu telefon numarasına ait müşteri veya randevu bulunamadı")

    try:
        await invalidate_cache(request, "customers_list", current_user)
        await invalidate_cache(request, "dashboard_stats", current_user)
        
        if hasattr(request.app.state, 'redis') and request.app.state.redis:
            redis_client = request.app.state.redis
            cache_key = f"plann:org_{current_user.organization_id}:customers_list"
            await redis_client.delete(cache_key)
            logger.info(f"Cache deleted: {cache_key}")
            
    except Exception as e:
        logger.warning(f"Customer cache invalidation failed: {e}")
    
    # Audit log
    await create_audit_log(
        db=db,
        organization_id=current_user.organization_id,
        user_id=current_user.username,
        user_full_name=current_user.full_name or current_user.username,
        action="DELETE",
        resource_type="CUSTOMER",
        resource_id=phone,
        old_value={
            "phone": phone, 
            "customer_deleted": customer_result.deleted_count > 0,
            "appointments": appointments_to_delete, 
            "appointments_count": appointment_result.deleted_count,
            "transactions_count": transaction_result.deleted_count
        },
        ip_address=request.client.host if request.client else None
    )
    
    # Emit WebSocket event for real-time update
    logger.info(f"About to emit customer_deleted for org: {current_user.organization_id}")
    try:
        await emit_to_organization(
            current_user.organization_id,
            'customer_deleted',
            {
                'phone': phone, 
                'deleted_appointments': appointment_result.deleted_count,
                'deleted_customer': customer_result.deleted_count > 0
            }
        )
        logger.info(f"Successfully emitted customer_deleted for org: {current_user.organization_id}")
    except Exception as emit_error:
        logger.error(f"Failed to emit customer_deleted: {emit_error}", exc_info=True)

    try:
        if appointment_result.deleted_count > 0:
            await emit_to_organization(
                current_user.organization_id,
                'appointment_deleted',
                {
                    'deleted_count': appointment_result.deleted_count,
                    'phone': phone,
                }
            )
    except Exception as emit_error:
        logger.error(f"Failed to emit appointment_deleted after customer delete: {emit_error}", exc_info=True)
    
    # Mesaj oluştur
    messages = []
    if customer_result.deleted_count > 0:
        messages.append("Müşteri")
    if appointment_result.deleted_count > 0:
        messages.append(f"{appointment_result.deleted_count} randevu")
    if transaction_result.deleted_count > 0:
        messages.append(f"{transaction_result.deleted_count} işlem")
    
    message = " ve ".join(messages) + " silindi"
    
    return {
        "message": message,
        "deleted_customer": customer_result.deleted_count > 0,
        "deleted_appointments": appointment_result.deleted_count,
        "deleted_transactions": transaction_result.deleted_count
    }

# === AUDIT LOGS ROUTES ===
@api_router.get("/audit-logs", response_model=List[AuditLog])
async def get_audit_logs(
    request: Request,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user_id: Optional[str] = None,
    action: Optional[str] = None,
    resource_type: Optional[str] = None,
    current_user: UserInDB = Depends(get_current_user)
):
    """Denetim günlüklerini getir - Sadece admin"""
    if not AUDIT_LOGS_ENABLED:
        raise HTTPException(status_code=404, detail="Not Found")
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    query = {"organization_id": current_user.organization_id}
    
    # Filters
    if user_id:
        query["user_id"] = user_id
    if action:
        query["action"] = action
    if resource_type:
        query["resource_type"] = resource_type
    if start_date or end_date:
        query["timestamp"] = {}
        if start_date:
            query["timestamp"]["$gte"] = start_date
        if end_date:
            query["timestamp"]["$lte"] = end_date
    
    # Get logs, sorted by timestamp descending
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).to_list(500)
    
    # Convert timestamp strings back to datetime
    for log in logs:
        if isinstance(log.get('timestamp'), str):
            log['timestamp'] = datetime.fromisoformat(log['timestamp'])
    
    return logs

@api_router.get("/export/appointments")
async def export_appointments(request: Request, current_user: UserInDB = Depends(get_current_user)):
    """Randevuları CSV formatında export et"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    appointments = await db.appointments.find(
        {"organization_id": current_user.organization_id},
        {"_id": 0}
    ).sort("appointment_date", -1).to_list(10000)
    
    # CSV formatında hazırla
    import csv
    import io
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        "Randevu ID", "Müşteri Adı", "Telefon", "Tarih", "Saat",
        "Hizmet", "Personel", "Durum", "Fiyat", "Notlar", "Oluşturma Tarihi"
    ])
    
    # Data
    for apt in appointments:
        writer.writerow([
            apt.get('id', ''),
            apt.get('customer_name', ''),
            apt.get('phone', ''),
            apt.get('appointment_date', ''),
            apt.get('appointment_time', ''),
            apt.get('service_name', ''),
            apt.get('staff_member_name', 'Atanmadı'),
            apt.get('status', ''),
            apt.get('price', 0),
            apt.get('notes', ''),
            str(apt.get('created_at', ''))
        ])
    
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=randevular.csv"}
    )

@api_router.get("/export/customers")
async def export_customers(request: Request, current_user: UserInDB = Depends(get_current_user)):
    """Müşterileri CSV formatında export et"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    # Tüm randevuları çek
    appointments = await db.appointments.find(
        {"organization_id": current_user.organization_id},
        {"_id": 0}
    ).to_list(10000)
    
    # Unique müşterileri grupla
    customer_map = {}
    for apt in appointments:
        phone = apt.get('phone')
        if phone and phone not in customer_map:
            customer_map[phone] = {
                "name": apt.get('customer_name', ''),
                "phone": phone,
                "total_appointments": 0,
                "completed_appointments": 0,
                "last_appointment_date": apt.get('appointment_date', '')
            }
        
        if phone:
            customer_map[phone]['total_appointments'] += 1
            if apt.get('status') == 'Tamamlandı':
                customer_map[phone]['completed_appointments'] += 1
            # En son randevu tarihini güncelle
            if apt.get('appointment_date', '') > customer_map[phone]['last_appointment_date']:
                customer_map[phone]['last_appointment_date'] = apt.get('appointment_date', '')
    
    # CSV formatında hazırla
    import csv
    import io
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        "Müşteri Adı", "Telefon", "Toplam Randevu",
        "Tamamlanan Randevu", "Son Randevu Tarihi"
    ])
    
    # Data
    customers = sorted(customer_map.values(), key=lambda x: x['total_appointments'], reverse=True)
    for customer in customers:
        writer.writerow([
            customer['name'],
            customer['phone'],
            customer['total_appointments'],
            customer['completed_appointments'],
            customer['last_appointment_date']
        ])
    
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=musteriler.csv"}
    )

@api_router.get("/customers/{phone}/history")
async def get_customer_history(request: Request, phone: str, current_user: UserInDB = Depends(get_current_user)):
    db = await get_db_from_request(request); query = {"phone": phone, "organization_id": current_user.organization_id, "status": {"$ne": "Ödeme Bekleniyor"}}
    appointments = await db.appointments.find(query, {"_id": 0}).sort("appointment_date", -1).to_list(1000)
    for appointment in appointments:
        if isinstance(appointment['created_at'], str): appointment['created_at'] = datetime.fromisoformat(appointment['created_at'].replace('Z', '+00:00'))
    total_completed = len([a for a in appointments if a['status'] == 'Tamamlandı'])
    
    # Müşteri notlarını getir
    customer_note = await db.customer_notes.find_one(
        {"phone": phone, "organization_id": current_user.organization_id},
        {"_id": 0, "notes": 1}
    )
    notes = customer_note.get("notes", "") if customer_note else ""
    
    return {"phone": phone, "total_appointments": len(appointments), "completed_appointments": total_completed, "appointments": appointments, "notes": notes}

@api_router.put("/customers/{phone}/notes")
async def update_customer_notes(request: Request, phone: str, notes_data: dict, current_user: UserInDB = Depends(get_current_user)):
    """Müşteri notlarını güncelle (Admin ve Personel - sadece kendi müşterileri)"""
    db = await get_db_from_request(request)
    notes = notes_data.get("notes", "")
    
    # Personel için: Bu müşterinin kendisiyle randevusu var mı kontrol et
    if current_user.role == "staff":
        staff_appointments = await db.appointments.find_one(
            {"phone": phone, "organization_id": current_user.organization_id, "staff_member_id": current_user.username},
            {"_id": 1}
        )
        if not staff_appointments:
            raise HTTPException(status_code=403, detail="Bu müşteriye not ekleme yetkiniz yok")
    
    # Müşteri notlarını güncelle veya oluştur
    await db.customer_notes.update_one(
        {"phone": phone, "organization_id": current_user.organization_id},
        {"$set": {"notes": notes, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )

    try:
        await invalidate_cache(request, "customers_list", current_user)
    except Exception as e:
        logging.warning(f"Customer cache invalidation failed: {e}")
    
    return {"message": "Notlar güncellendi", "notes": notes}

# === FINANCE & EXPENSES ROUTES ===
class ExpenseCreate(BaseModel):
    title: str
    amount: float
    category: str
    date: str

class ExpenseUpdate(BaseModel):
    title: Optional[str] = None
    amount: Optional[float] = None
    category: Optional[str] = None
    date: Optional[str] = None

@api_router.get("/finance/summary")
async def get_finance_summary(request: Request, period: str = "this_month", current_user: UserInDB = Depends(get_current_user)):
    """Finans özeti: Gelir, Gider, Net Kâr (Sadece admin)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    # Tarih aralığını hesapla
    from datetime import datetime, timedelta
    # UTC timezone kullan (tutarlılık için)
    now = datetime.now()
    today_str = now.strftime("%Y-%m-%d")
    
    if period == "today":
        start_date = today_str
        end_date = today_str
    elif period == "this_month":
        start_date = now.replace(day=1).strftime("%Y-%m-%d")
        end_date = today_str
    elif period == "last_month":
        first_day_this_month = now.replace(day=1)
        last_day_last_month = first_day_this_month - timedelta(days=1)
        start_date = last_day_last_month.replace(day=1).strftime("%Y-%m-%d")
        end_date = last_day_last_month.strftime("%Y-%m-%d")
    elif period == "this_year":
        start_date = now.replace(month=1, day=1).strftime("%Y-%m-%d")
        end_date = today_str
    else:
        start_date = now.replace(day=1).strftime("%Y-%m-%d")
        end_date = today_str
    
    logging.info(f"Finance summary - Date range calculation: period={period}, start_date={start_date}, end_date={end_date}, today={today_str}, now={now.strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Toplam Gelir: Tamamlanan randevuların toplam hizmet bedeli
    completed_appointments = await db.appointments.find(
        {
            "organization_id": current_user.organization_id,
            "status": "Tamamlandı",
            "appointment_date": {"$gte": start_date, "$lte": end_date}
        },
        {"_id": 0, "service_price": 1, "staff_member_id": 1}
    ).to_list(10000)
    
    total_revenue = sum(apt.get("service_price", 0) or 0 for apt in completed_appointments)
    
    # Personel bazlı kazançları hesapla
    staff_earnings_dict = {}
    for apt in completed_appointments:
        staff_id = apt.get("staff_member_id")
        if staff_id:
            price = apt.get("service_price", 0) or 0
            if staff_id not in staff_earnings_dict:
                staff_earnings_dict[staff_id] = 0
            staff_earnings_dict[staff_id] += price
    
    # Personel isimlerini al
    staff_users = await db.users.find(
        {"organization_id": current_user.organization_id, "username": {"$in": list(staff_earnings_dict.keys())}},
        {"_id": 0, "username": 1, "full_name": 1}
    ).to_list(100)
    staff_name_map = {u["username"]: u.get("full_name", u["username"]) for u in staff_users}
    
    staff_earnings = [
        {"username": staff_id, "full_name": staff_name_map.get(staff_id, staff_id), "total": amount}
        for staff_id, amount in staff_earnings_dict.items()
    ]
    staff_earnings.sort(key=lambda x: x["total"], reverse=True)
    
    # Toplam Gider: Expenses tablosundaki kayıtların toplamı
    # MongoDB sorgusu ile doğrudan filtreleme yapıyoruz (this_month için ay filtresi dahil)
    expense_query = {"organization_id": current_user.organization_id}
    
    # Tarih filtresini MongoDB sorgusuna ekle
    date_conditions = [{"date": {"$exists": True}}, {"date": {"$ne": ""}}]  # Sadece tarihi olan expense'leri al
    
    if period == "this_month":
        # Bu ay için: ayın ilk gününden son gününe kadar
        first_day_of_month = now.replace(day=1).strftime("%Y-%m-%d")
        # Ayın son gününü hesapla
        if now.month == 12:
            last_day_of_month = now.replace(year=now.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last_day_of_month = now.replace(month=now.month + 1, day=1) - timedelta(days=1)
        last_day_str = last_day_of_month.strftime("%Y-%m-%d")
        date_conditions.append({"date": {"$gte": first_day_of_month, "$lte": last_day_str}})
        logging.info(f"Finance summary - this_month filter: {first_day_of_month} to {last_day_str}")
    elif period == "today":
        date_conditions.append({"date": {"$gte": start_date, "$lte": end_date}})
    elif period == "last_month":
        date_conditions.append({"date": {"$gte": start_date, "$lte": end_date}})
    elif period == "this_year":
        first_day_of_year = now.replace(month=1, day=1).strftime("%Y-%m-%d")
        date_conditions.append({"date": {"$gte": first_day_of_year, "$lte": today_str}})
    else:
        # Varsayılan olarak bu ay
        first_day_of_month = now.replace(day=1).strftime("%Y-%m-%d")
        if now.month == 12:
            last_day_of_month = now.replace(year=now.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last_day_of_month = now.replace(month=now.month + 1, day=1) - timedelta(days=1)
        last_day_str = last_day_of_month.strftime("%Y-%m-%d")
        date_conditions.append({"date": {"$gte": first_day_of_month, "$lte": last_day_str}})
    
    if len(date_conditions) > 0:
        expense_query["$and"] = date_conditions
    
    expenses = await db.expenses.find(
        expense_query,
        {"_id": 0, "amount": 1, "date": 1, "title": 1}
    ).to_list(10000)
    
    logging.info(f"Finance summary - period: {period}, start_date: {start_date}, end_date: {end_date}")
    logging.info(f"Finance summary - FILTERED expenses count: {len(expenses)}")
    for exp in expenses[:5]:  # İlk 5 filtrelenmiş expense'i logla
        logging.info(f"  FILTERED Expense: {exp.get('title')}, date: {exp.get('date')}, amount: {exp.get('amount')}")
    
    total_expenses = sum(float(exp.get("amount", 0) or 0) for exp in expenses)
    logging.info(f"Finance summary - total_expenses: {total_expenses}")
    
    # Net Kâr
    net_profit = total_revenue - total_expenses
    
    return {
        "period": period,
        "start_date": start_date,
        "end_date": end_date,
        "total_revenue": total_revenue,
        "total_expenses": total_expenses,
        "net_profit": net_profit,
        "staff_earnings": staff_earnings
    }

@api_router.get("/expenses")
async def get_expenses(request: Request, current_user: UserInDB = Depends(get_current_user)):
    """Giderleri listele (Sadece admin)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    expenses = await db.expenses.find(
        {"organization_id": current_user.organization_id},
        {"_id": 0}
    ).sort("date", -1).to_list(1000)
    
    return expenses

@api_router.post("/expenses")
async def create_expense(request: Request, expense: ExpenseCreate, current_user: UserInDB = Depends(get_current_user)):
    """Yeni gider ekle (Sadece admin)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    expense_data = {
        "id": str(uuid.uuid4()),
        "organization_id": current_user.organization_id,
        "title": expense.title,
        "amount": expense.amount,
        "category": expense.category,
        "date": expense.date,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    try:
        await db.expenses.insert_one(expense_data)
        logging.info(f"Gider başarıyla kaydedildi: {expense_data.get('title')}, tutar: {expense_data.get('amount')}")
    except Exception as insert_error:
        logging.error(f"Gider kaydetme hatası: {type(insert_error).__name__}: {str(insert_error)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Gider kaydedilirken bir hata oluştu: {str(insert_error)}")
    
    # Response için expense_data'yı temizle (MongoDB ObjectId gibi serialize edilemeyen alanları kaldır)
    response_expense = {
        "id": expense_data["id"],
        "organization_id": expense_data["organization_id"],
        "title": expense_data["title"],
        "amount": float(expense_data["amount"]),
        "category": expense_data["category"],
        "date": expense_data["date"],
        "created_at": expense_data["created_at"]
    }
    
    return {"message": "Gider eklendi", "expense": response_expense}

@api_router.put("/expenses/{expense_id}")
async def update_expense(request: Request, expense_id: str, expense_update: ExpenseUpdate, current_user: UserInDB = Depends(get_current_user)):
    """Gider güncelle (Sadece admin)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    update_data = {}
    if expense_update.title is not None:
        update_data["title"] = expense_update.title
    if expense_update.amount is not None:
        update_data["amount"] = expense_update.amount
    if expense_update.category is not None:
        update_data["category"] = expense_update.category
    if expense_update.date is not None:
        update_data["date"] = expense_update.date
    
    result = await db.expenses.update_one(
        {"id": expense_id, "organization_id": current_user.organization_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Gider bulunamadı")
    
    return {"message": "Gider güncellendi"}

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(request: Request, expense_id: str, current_user: UserInDB = Depends(get_current_user)):
    """Gider sil (Sadece admin)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    result = await db.expenses.delete_one(
        {"id": expense_id, "organization_id": current_user.organization_id}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Gider bulunamadı")
    
    return {"message": "Gider silindi"}

@api_router.get("/finance/payroll")
async def get_payroll(request: Request, period: str = "this_month", current_user: UserInDB = Depends(get_current_user)):
    """Personel hakedişlerini hesapla (Sadece admin)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
    
    db = await get_db_from_request(request)
    
    # Tarih aralığını hesapla
    from datetime import datetime, timedelta
    now = datetime.now()
    today_str = now.strftime("%Y-%m-%d")
    
    if period == "this_month":
        start_date = now.replace(day=1).strftime("%Y-%m-%d")
        end_date = today_str
    elif period == "last_month":
        first_day_this_month = now.replace(day=1)
        last_day_last_month = first_day_this_month - timedelta(days=1)
        start_date = last_day_last_month.replace(day=1).strftime("%Y-%m-%d")
        end_date = last_day_last_month.strftime("%Y-%m-%d")
    else:
        start_date = now.replace(day=1).strftime("%Y-%m-%d")
        end_date = today_str
    
    logging.info(f"Payroll - period: {period}, start_date: {start_date}, end_date: {end_date}, today: {today_str}")
    
    # Tüm personelleri getir
    staff_members = await db.users.find(
        {"organization_id": current_user.organization_id, "role": "staff"},
        {"_id": 0, "username": 1, "full_name": 1, "payment_type": 1, "payment_amount": 1}
    ).to_list(1000)
    
    payroll_list = []
    
    for staff in staff_members:
        username = staff.get("username")
        full_name = staff.get("full_name") or username
        payment_type = staff.get("payment_type", "salary")
        payment_amount = staff.get("payment_amount", 0.0) or 0.0
        
        # Hakediş hesapla
        if payment_type == "salary":
            # Sabit maaş
            earned = float(payment_amount) if payment_amount else 0.0
        else:
            # Komisyon: Personelin o ay tamamladığı toplam ciro * (payment_amount / 100)
            completed_appointments = await db.appointments.find(
                {
                    "organization_id": current_user.organization_id,
                    "staff_member_id": username,
                    "status": "Tamamlandı",
                    "appointment_date": {"$gte": start_date, "$lte": end_date}
                },
                {"_id": 0, "service_price": 1}
            ).to_list(10000)
            
            total_revenue = sum(apt.get("service_price", 0) or 0 for apt in completed_appointments)
            commission_rate = float(payment_amount) if payment_amount else 0.0
            earned = total_revenue * (commission_rate / 100.0)
            
            # Debug logging
            logging.info(f"Payroll calculation for {username}: total_revenue={total_revenue}, commission_rate={commission_rate}%, earned={earned}")
        
        # Ödenen tutarı hesapla (Personel Ödemesi kategorisindeki giderler)
        # staff_username field'ı varsa direkt eşleştir, yoksa title'dan kontrol et
        # Tüm personel ödemelerini al (tarih filtresi olmadan) - debug için
        all_staff_payments = await db.expenses.find(
            {
                "organization_id": current_user.organization_id,
                "category": "Personel Ödemesi",
                "$or": [
                    {"staff_username": username},  # Yeni format: staff_username field'ı
                    {"title": {"$regex": username, "$options": "i"}},  # Eski format: title'da username
                    {"title": {"$regex": full_name, "$options": "i"}}  # Eski format: title'da isim
                ]
            },
            {"_id": 0, "amount": 1, "date": 1, "title": 1, "staff_username": 1}
        ).to_list(1000)
        
        logging.info(f"Payroll - {username}: Found {len(all_staff_payments)} total staff payments")
        for payment in all_staff_payments[:5]:
            logging.info(f"  Payment: {payment.get('title')}, date: {payment.get('date')}, amount: {payment.get('amount')}, staff_username: {payment.get('staff_username')}")
        
        # Tarih aralığına göre filtrele
        if period == "this_month":
            # Bu ay içindeki tüm ödemeleri dahil et (ay kontrolü)
            staff_payments = []
            for payment in all_staff_payments:
                payment_date = payment.get('date', '')
                if payment_date:
                    try:
                        payment_date_obj = datetime.strptime(payment_date, "%Y-%m-%d")
                        if payment_date_obj.year == now.year and payment_date_obj.month == now.month:
                            staff_payments.append(payment)
                            logging.info(f"  INCLUDED Payment: {payment.get('title')}, date: {payment_date}, amount: {payment.get('amount')}")
                        else:
                            logging.info(f"  EXCLUDED Payment (wrong month): {payment.get('title')}, date: {payment_date}")
                    except (ValueError, TypeError) as e:
                        logging.warning(f"  EXCLUDED Payment (invalid date): {payment.get('title')}, date: {payment_date}, error: {str(e)}")
        else:
            # Diğer period'lar için normal tarih aralığı kontrolü
            staff_payments = [p for p in all_staff_payments if p.get('date') and start_date <= p.get('date') <= end_date]
        
        logging.info(f"Payroll - {username}: Filtered {len(staff_payments)} payments for period {period}")
        
        paid = sum(float(payment.get("amount", 0) or 0) for payment in staff_payments)
        logging.info(f"Payroll - {username}: paid = {paid}")
        balance = earned - paid
        
        payroll_list.append({
            "username": username,
            "full_name": full_name,
            "payment_type": payment_type,
            "payment_amount": payment_amount,
            "earned": earned,
            "paid": paid,
            "balance": balance
        })
    
    return {
        "period": period,
        "start_date": start_date,
        "end_date": end_date,
        "payroll": payroll_list
    }

class PayrollPaymentRequest(BaseModel):
    staff_username: str
    amount: float
    date: Optional[str] = None

@api_router.post("/finance/payroll/payment")
async def make_payroll_payment(request: Request, payment_data: PayrollPaymentRequest, current_user: UserInDB = Depends(get_current_user)):
    """Personel ödemesi yap (Gider olarak kaydet) (Sadece admin)"""
    logging.info("=== PAYROLL PAYMENT ENDPOINT ÇAĞRILDI ===")
    try:
        logging.info(f"Request body: staff_username={payment_data.staff_username}, amount={payment_data.amount}, date={payment_data.date}")
        logging.info(f"Current user: {current_user.username if current_user else 'None'}, role: {current_user.role if current_user else 'None'}")
    except Exception as log_error:
        logging.error(f"Logging hatası: {type(log_error).__name__}: {str(log_error)}", exc_info=True)
    
    logging.info(f"Personel ödemesi isteği alındı: {payment_data.staff_username}, tutar: {payment_data.amount}")
    try:
        if current_user.role != "admin":
            logging.warning(f"Yetkisiz erişim denemesi: {current_user.username}")
            raise HTTPException(status_code=403, detail="Bu işlem için yetkiniz yok")
        
        db = await get_db_from_request(request)
        logging.info(f"Database bağlantısı başarılı: {current_user.organization_id}")
        
        staff_username = payment_data.staff_username
        amount = float(payment_data.amount)  # Pydantic zaten float olarak validate ediyor ama emin olmak için
        date = payment_data.date or datetime.now().strftime("%Y-%m-%d")
        
        logging.info(f"Personel aranıyor: {staff_username}, organization: {current_user.organization_id}")
        
        # Personel bilgisini al
        staff = await db.users.find_one(
            {"username": staff_username, "organization_id": current_user.organization_id, "role": "staff"},
            {"_id": 0, "full_name": 1}
        )
        
        if not staff:
            logging.error(f"Personel bulunamadı: {staff_username}")
            raise HTTPException(status_code=404, detail="Personel bulunamadı")
        
        staff_name = staff.get("full_name") or staff_username
        logging.info(f"Personel bulundu: {staff_name}")
        
        # Gider olarak kaydet
        expense_data = {
            "id": str(uuid.uuid4()),
            "organization_id": current_user.organization_id,
            "title": f"{staff_name} - Personel Ödemesi",
            "amount": float(amount),
            "category": "Personel Ödemesi",
            "date": date,
            "staff_username": staff_username,  # Personel takibi için
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        logging.info(f"Expense data hazırlandı: {expense_data}")
        logging.info(f"Expenses collection'a kaydediliyor...")
        
        try:
            result = await db.expenses.insert_one(expense_data)
            logging.info(f"MongoDB insert result: inserted_id={result.inserted_id}")
        except Exception as insert_error:
            logging.error(f"Expenses collection'a yazma hatası: {type(insert_error).__name__}: {str(insert_error)}")
            raise
        
        logging.info(f"Personel ödemesi başarıyla kaydedildi: {staff_username}, tutar: {amount}, tarih: {date}")
        
        # Response için expense_data'yı temizle (MongoDB ObjectId gibi serialize edilemeyen alanları kaldır)
        response_expense = {
            "id": expense_data["id"],
            "organization_id": expense_data["organization_id"],
            "title": expense_data["title"],
            "amount": float(expense_data["amount"]),
            "category": expense_data["category"],
            "date": expense_data["date"],
            "staff_username": expense_data.get("staff_username"),
            "created_at": expense_data["created_at"]
        }
        
        try:
            response_data = {"message": "Ödeme kaydedildi", "expense": response_expense}
            logging.info(f"Response hazırlandı: {response_data}")
            return response_data
        except Exception as response_error:
            logging.error(f"Response oluşturma hatası: {type(response_error).__name__}: {str(response_error)}", exc_info=True)
            # Yine de başarılı response dön
            return {"message": "Ödeme kaydedildi", "expense_id": expense_data.get("id")}
    except HTTPException as e:
        logging.error(f"HTTPException: {e.status_code} - {e.detail}")
        raise
    except Exception as e:
        error_type = type(e).__name__
        error_message = str(e)
        import traceback
        full_traceback = traceback.format_exc()
        logging.error(f"Personel ödemesi kaydetme hatası: {error_type}: {error_message}")
        logging.error(f"Full traceback:\n{full_traceback}")
        raise HTTPException(status_code=500, detail=f"Ödeme kaydedilirken bir hata oluştu: {error_message}")

# === PUBLIC API ROUTES (TOKEN GEREKTİRMEZ) ===

# ═══ CONTRACT: v1 (FROZEN / RESPONSE) — GET /api/app/config ═══
# Mobil force-update SİNYAL kaynağı. Backend YALNIZ sinyal verir; isteği ASLA
# hard-reject etmez (kilidi frontend UX katmanı uygular — fail-open felsefesi).
# Response alanı SİLME/RENAME yasak; EKLEME güvenli. Değerler DB'den okunur
# (app_config koleksiyonu, _id="global") → deploy'suz bumplanabilir. Doküman
# yoksa güvenli defaults döner (min_supported=0.0.0 → hiç kimseyi kilitleme).
APP_CONFIG_DEFAULTS = {
    "min_supported_version": {"ios": "0.0.0", "android": "0.0.0"},
    "latest_version": {"ios": "6.1", "android": "2.0"},
    "store_urls": {
        "ios": "https://apps.apple.com/app/id6759719891",
        "android": "https://play.google.com/store/apps/details?id=co.plannapp.app",
    },
}


def _merge_app_config(doc):
    """DB dokümanını defaults ile birleştirir (eksik/bozuk alanlar için güvenli fallback)."""
    cfg = {
        "min_supported_version": dict(APP_CONFIG_DEFAULTS["min_supported_version"]),
        "latest_version": dict(APP_CONFIG_DEFAULTS["latest_version"]),
        "store_urls": dict(APP_CONFIG_DEFAULTS["store_urls"]),
    }
    if isinstance(doc, dict):
        for key in ("min_supported_version", "latest_version", "store_urls"):
            val = doc.get(key)
            if isinstance(val, dict):
                for plat in ("ios", "android"):
                    pv = val.get(plat)
                    if isinstance(pv, str) and pv:
                        cfg[key][plat] = pv
    return cfg


@api_router.get("/app/config")
async def get_app_config(request: Request, response: Response):
    """
    Mobil uygulama config'i (force-update sinyali). PUBLIC — auth gerektirmez ki
    login öncesi de kontrol edilebilsin. DB'den okunur; yoksa güvenli defaults.
    Fail-open: her durumda 200 + geçerli config döner, ASLA kullanıcıyı kilitlemez.

    Cache-Control: no-store — force-update sinyali HER ZAMAN taze olmalı; WebView/
    tarayıcı cache'i eski eşiği döndürürse modal yanlış (kaybolur/kalır) davranır.
    """
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    doc = None
    try:
        db = await get_db_from_request(request)
        doc = await db.app_config.find_one({"_id": "global"}, {"_id": 0})
    except Exception as e:
        logging.warning(f"/app/config okunamadı, defaults dönülüyor: {e}")
    return _merge_app_config(doc)


# ═══ Akıllı uygulama açıcı (smart app launcher) ═══
# E-posta butonları buraya link verir. Cihaz tespiti + "yüklüyse aç, değilse mağaza".
#   - Android: intent:// (scheme=plannapp, package=co.plannapp.app) → app varsa açar,
#     yoksa browser_fallback_url ile Play Store'a düşer (native fallback, tek adım).
#   - iOS: plannapp:// denenir; app açılırsa sayfa arka plana düşer (visibilitychange),
#     açılmazsa kısa timeout sonrası App Store'a yönlendirilir.
#   - Masaüstü/diğer: web paneline (plannapp.co / plannapp.co.uk) gider.
_APP_ANDROID_PACKAGE = "co.plannapp.app"
_APP_URL_SCHEME = "plannapp"
_APP_WEB_ALLOWED = ("https://plannapp.co", "https://plannapp.co.uk")


def _build_app_launcher_html(*, store_ios: str, store_android: str, web_url: str) -> str:
    """Cihaza göre uygulamayı açan / mağazaya yönlendiren minimal HTML sayfası."""
    from urllib.parse import quote
    scheme_url = f"{_APP_URL_SCHEME}://open"
    android_intent = (
        f"intent://open#Intent;scheme={_APP_URL_SCHEME};package={_APP_ANDROID_PACKAGE};"
        f"S.browser_fallback_url={quote(store_android, safe='')};end"
    )
    return f"""<!DOCTYPE html>
<html lang="tr">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>PLANN</title>
  <style>
    html,body{{margin:0;height:100%;background:#fbfbfa;color:#1a1a1a;
      font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;}}
    .wrap{{height:100%;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center;padding:24px;}}
    h1{{margin:0 0 16px;font-weight:300;letter-spacing:10px;font-size:26px;text-transform:uppercase;}}
    p{{margin:0 0 22px;color:#8c8c88;font-size:14px;letter-spacing:1px;}}
    a.btn{{border:1px solid #1a1a1a;color:#1a1a1a;text-decoration:none;padding:14px 34px;
      font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:3px;display:inline-block;}}
  </style>
</head>
<body>
  <div class="wrap">
    <h1>P L A N N</h1>
    <p id="msg">Yönlendiriliyorsunuz…</p>
    <a class="btn" id="fallback" href="{web_url}">Devam Et</a>
  </div>
  <script>
    (function() {{
      var ua = navigator.userAgent || navigator.vendor || '';
      var isAndroid = /android/i.test(ua);
      var isIOS = /iphone|ipad|ipod/i.test(ua) ||
                  (/(Macintosh)/i.test(ua) && 'ontouchend' in document);
      var STORE_IOS = {json.dumps(store_ios)};
      var STORE_ANDROID = {json.dumps(store_android)};
      var WEB = {json.dumps(web_url)};
      var SCHEME = {json.dumps(scheme_url)};
      var ANDROID_INTENT = {json.dumps(android_intent)};
      var fb = document.getElementById('fallback');

      if (isAndroid) {{
        fb.href = STORE_ANDROID;
        window.location.href = ANDROID_INTENT;
        return;
      }}
      if (isIOS) {{
        fb.href = STORE_IOS;
        var opened = false;
        var onHide = function() {{ opened = true; }};
        document.addEventListener('visibilitychange', function() {{
          if (document.visibilityState === 'hidden') onHide();
        }});
        window.addEventListener('pagehide', onHide);
        window.addEventListener('blur', onHide);
        setTimeout(function() {{ window.location.href = SCHEME; }}, 50);
        setTimeout(function() {{
          if (!opened && document.visibilityState === 'visible') {{
            window.location.href = STORE_IOS;
          }}
        }}, 1400);
        return;
      }}
      window.location.href = WEB;
    }})();
  </script>
</body>
</html>"""


@api_router.get("/app/open")
async def app_open(request: Request, web: str = "", path: str = ""):
    """
    Akıllı uygulama açıcı. PUBLIC. Cihaza göre uygulamayı açar; yüklü değilse
    mağazaya (App Store / Play Store), masaüstünde web paneline yönlendirir.
    E-posta butonları bu endpoint'e link verir (`?web=` ile web fallback seçilir).
    """
    doc = None
    try:
        db = await get_db_from_request(request)
        doc = await db.app_config.find_one({"_id": "global"}, {"_id": 0})
    except Exception as e:
        logging.warning(f"/app/open config okunamadı, defaults: {e}")
    cfg = _merge_app_config(doc)
    store_ios = cfg["store_urls"]["ios"]
    store_android = cfg["store_urls"]["android"]

    web_url = web if web in _APP_WEB_ALLOWED else "https://plannapp.co"

    html = _build_app_launcher_html(store_ios=store_ios, store_android=store_android, web_url=web_url)
    return HTMLResponse(
        content=html,
        headers={"Cache-Control": "no-store, no-cache, must-revalidate", "Pragma": "no-cache"},
    )


@api_router.put("/superadmin/app-config")
async def update_app_config(request: Request, current_user: UserInDB = Depends(get_superadmin_user)):
    """
    Force-update eşiklerini deploy'suz güncelle (superadmin). Kısmi güncelleme:
    yalnız gönderilen platform alanları yazılır. Body örneği:
      { "min_supported_version": {"ios": "6.0"}, "latest_version": {"ios": "6.1"} }
    """
    db = await get_db_from_request(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Geçersiz JSON body")
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="Body bir obje olmalı")

    update = {}
    for key in ("min_supported_version", "latest_version", "store_urls"):
        val = body.get(key)
        if isinstance(val, dict):
            for plat in ("ios", "android"):
                pv = val.get(plat)
                if isinstance(pv, str) and pv:
                    update[f"{key}.{plat}"] = pv
    if not update:
        raise HTTPException(status_code=400, detail="Güncellenecek geçerli alan yok")
    update["updated_at"] = datetime.utcnow()
    update["updated_by"] = current_user.username
    await db.app_config.update_one({"_id": "global"}, {"$set": update}, upsert=True)
    doc = await db.app_config.find_one({"_id": "global"}, {"_id": 0})
    return _merge_app_config(doc)


@api_router.get("/public/business/{slug}")
async def get_public_business(request: Request, slug: str):
    """Slug ile işletme bilgilerini, hizmetlerini, personellerini ve ayarlarını getir (Model D)"""
    db = await get_db_from_request(request)
    
    # İlk önce user'dan slug'ı bul (admin kullanıcısı)
    admin_user = await db.users.find_one({"slug": slug}, {"_id": 0})
    if not admin_user:
        raise HTTPException(status_code=404, detail="İşletme bulunamadı")
    
    organization_id = admin_user.get('organization_id')
    
    # Hizmetleri çek
    services = await db.services.find({"organization_id": organization_id}, {"_id": 0}).to_list(1000)

    def sort_key(svc: dict):
        order = svc.get("order")
        return (
            order if isinstance(order, int) else 1_000_000,
            svc.get("created_at") or "",
            svc.get("name") or "",
        )

    services = sorted(services, key=sort_key)

    # Kategorileri çek (sıralı)
    categories = await db.categories.find({"organization_id": organization_id}, {"_id": 0}).to_list(1000)
    categories = _sort_categories(categories)
    
    # Ayarları çek
    settings = await db.settings.find_one({"organization_id": organization_id}, {"_id": 0})
    if not settings:
        # Varsayılan ayarlar
        settings = {
            "customer_can_choose_staff": False,
            "work_start_hour": 9,
            "work_end_hour": 18,
            "appointment_interval": 30,
            "company_name": admin_user.get('full_name', 'İşletme'),
            "admin_provides_service": True,
            "show_service_duration_on_public": True,
            "show_service_price_on_public": True
        }
    else:
        # Yeni alanlar yoksa varsayılan değerleri ekle
        if "show_service_duration_on_public" not in settings:
            settings["show_service_duration_on_public"] = True
        if "show_service_price_on_public" not in settings:
            settings["show_service_price_on_public"] = True
    
    # Tüm personelleri çek (SADECE gerekli alanlar - ŞİFRELER HARİÇ!)
    staff_members = await db.users.find(
        {"organization_id": organization_id}, 
        {"_id": 0, "full_name": 1, "id": 1, "permitted_service_ids": 1, "username": 1, "role": 1}
    ).to_list(1000)
    
    # Admin hizmet vermiyorsa admin'i listeden çıkar
    if not settings.get('admin_provides_service', True):
        staff_members = [staff for staff in staff_members if staff.get('role') != 'admin']
    
    return {
        "business_name": settings.get('company_name', admin_user.get('full_name', 'İşletme')),
        "logo_url": settings.get('logo_url'),
        "images": settings.get('images', []) or [],
        "location": settings.get('location'),
        "organization_id": organization_id,
        "services": services,
        "categories": categories,
        "staff_members": staff_members,
        "settings": {
            "customer_can_choose_staff": settings.get('customer_can_choose_staff', False),
            "work_start_hour": settings.get('work_start_hour', 9),
            "work_end_hour": settings.get('work_end_hour', 18),
            "appointment_interval": settings.get('appointment_interval', 30),
            "show_service_duration_on_public": settings.get('show_service_duration_on_public', True),
            "show_service_price_on_public": settings.get('show_service_price_on_public', True),
            "multi_service_enabled": settings.get('multi_service_enabled', False)
        }
    }

@api_router.get("/public/availability/{organization_id}")
async def get_availability(request: Request, organization_id: str, service_id: Optional[str] = None, service_ids: Optional[str] = None, date: str = "", staff_id: Optional[str] = None, exclude_appointment_id: Optional[str] = None):
    """Model D: Personel bazlı akıllı müsaitlik kontrolü - business_hours ve days_off kullanır.

    Çoklu hizmet: service_ids (virgülle ayrılmış) verilirse toplam süre kadar kesintisiz
    blok aranır ve personel KESİŞİMİ (tüm hizmetleri verebilen) uygulanır.
    """
    db = await get_db_from_request(request)

    # Çoklu/tekli hizmet parametre normalizasyonu
    service_id_list = [s.strip() for s in service_ids.split(",") if s.strip()] if service_ids else ([service_id] if service_id else [])
    if not service_id_list:
        raise HTTPException(status_code=400, detail="Hizmet seçilmedi")
    service_id = service_id_list[0]

    # Availability cache (org+staff+date+service_ids) — randevu create/update/delete'te invalidate edilir
    redis_client = getattr(request.app.state, 'redis_client', None) if hasattr(request, 'app') else None
    _avail_cache_key = f"plann:avail:{organization_id}:{date}:{staff_id or 'auto'}:{exclude_appointment_id or '-'}:{'-'.join(sorted(service_id_list))}"
    if redis_client:
        try:
            _cached = await redis_client.get(_avail_cache_key)
            if _cached:
                return json.loads(_cached)
        except Exception:
            pass
    
    # Ayarları al (admin_provides_service ve business_hours için)
    settings = await db.settings.find_one({"organization_id": organization_id}, {"_id": 0})
    if not settings:
        settings = {
            "work_start_hour": 9, 
            "work_end_hour": 18, 
            "appointment_interval": 30, 
            "admin_provides_service": True,
            "business_hours": {
                "monday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
                "tuesday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
                "wednesday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
                "thursday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
                "friday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
                "saturday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
                "sunday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"}
            }
        }
    
    admin_provides_service = settings.get('admin_provides_service', True)
    business_hours = settings.get('business_hours', {})
    
    # Tarihin hangi güne denk geldiğini bul (0=Monday, 6=Sunday)
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        weekday = date_obj.weekday()  # 0=Monday, 6=Sunday
        day_names = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
        day_name = day_names[weekday]
    except (ValueError, TypeError):
        day_name = "monday"  # Varsayılan
    
    # Eğer müşteri belirli bir personel seçtiyse, sadece o personele bak
    selected_staff = None
    if staff_id:
        staff_query = {
            "organization_id": organization_id,
            "username": staff_id,
            "permitted_service_ids": {"$all": service_id_list}
        }
        
        # Admin'in "hizmet verir" ayarı kapalıysa, admin'i personel listesinden çıkar
        if not admin_provides_service:
            staff_query["role"] = {"$ne": "admin"}
        
        staff_members = await db.users.find(
            staff_query,
            {"_id": 0, "id": 1, "username": 1, "role": 1, "days_off": 1, "permitted_service_ids": 1, "breaks": 1}
        ).to_list(1000)
        
        # Eğer seçilen personel bulunamadıysa ve admin_provides_service kapalıysa bile, admin'in bu hizmeti verebilip veremediğini kontrol et
        if not staff_members:
            # Admin'in bu hizmeti verebilip veremediğini kontrol et (eğer seçilen staff_id admin ise)
            admin_user = await db.users.find_one(
                {"organization_id": organization_id, "role": "admin", "username": staff_id},
                {"_id": 0, "id": 1, "username": 1, "role": 1, "days_off": 1, "permitted_service_ids": 1, "breaks": 1}
            )
            if admin_user and all(s in (admin_user.get('permitted_service_ids') or []) for s in service_id_list):
                # Admin bu hizmeti verebiliyorsa, admin'i personel listesine ekle
                staff_members = [admin_user]
                logging.info(f"⚠️ Availability: Selected staff not found, but admin can provide service. Using admin: {admin_user['username']}")
        
        if not staff_members:
            return {"available_slots": [], "message": "Seçilen personel bu hizmeti veremiyor"}
        
        selected_staff = staff_members[0]
        
        # Personelin days_off kontrolü
        staff_days_off = selected_staff.get('days_off') or []
        if day_name in staff_days_off:
            # Personel bu gün izinli, müsaitlik yok
            logging.info(f"Staff {staff_id} is off on {day_name}")
            return {
                "available_slots": [],
                "all_slots": [],
                "busy_slots": [],
                "message": "Seçili personel bu gün izinli"
            }
    else:
        # O hizmet(ler)i verebilen TÜM personelleri bul (çoklu hizmette KESİŞİM: tümünü verebilmeli)
        staff_query = {
            "organization_id": organization_id,
            "permitted_service_ids": {"$all": service_id_list}
        }
        
        # Admin'in "hizmet verir" ayarı kapalıysa, admin'i personel listesinden çıkar
        if not admin_provides_service:
            staff_query["role"] = {"$ne": "admin"}
        
        staff_members = await db.users.find(
            staff_query,
            {"_id": 0, "id": 1, "username": 1, "role": 1, "days_off": 1, "permitted_service_ids": 1, "breaks": 1}
        ).to_list(1000)
        
        # Eğer başka personel yoksa ve admin_provides_service kapalıysa bile, admin'in bu hizmeti verebilip veremediğini kontrol et
        if not staff_members:
            logging.info(f"⚠️ Availability: No staff found for service_ids={service_id_list}, admin_provides_service={admin_provides_service}. Checking admin...")
            # Admin'in bu hizmeti verebilip veremediğini kontrol et
            admin_user = await db.users.find_one(
                {"organization_id": organization_id, "role": "admin"},
                {"_id": 0, "id": 1, "username": 1, "role": 1, "days_off": 1, "permitted_service_ids": 1, "breaks": 1}
            )
            if admin_user:
                admin_permitted_services = admin_user.get('permitted_service_ids') or []
                logging.info(f"⚠️ Availability: Admin found: {admin_user['username']}, permitted_service_ids: {admin_permitted_services}, checking service_ids: {service_id_list}")
                if all(s in admin_permitted_services for s in service_id_list):
                    # Admin bu hizmeti verebiliyorsa, admin'i personel listesine ekle
                    # Admin'in tüm gerekli alanlarını ekle (breaks dahil — slot meşgul listesine doğru eklenmesi için)
                    admin_staff_member = {
                        "id": admin_user.get('id', admin_user.get('username')),
                        "username": admin_user['username'],
                        "role": "admin",
                        "days_off": admin_user.get('days_off', []),
                        "permitted_service_ids": admin_permitted_services,
                        "breaks": admin_user.get('breaks', []),
                    }
                    staff_members = [admin_staff_member]
                    logging.info(f"✅ Availability: Admin can provide service. Using admin: {admin_user['username']}, staff_member: {admin_staff_member}")
                else:
                    logging.warning(f"❌ Availability: Admin found but service_id {service_id} not in permitted_service_ids: {admin_permitted_services}")
            else:
                logging.warning(f"❌ Availability: No admin user found for organization_id: {organization_id}")
        
        if not staff_members:
            # Hiç personel yoksa veya hiçbiri bu hizmeti vermiyorsa boş dön
            logging.warning(f"❌ Availability: No staff members found after admin check. Returning empty slots.")
            return {"available_slots": [], "message": "Bu hizmet için uygun personel bulunamadı"}
    
        logging.info(f"✅ Availability: Found {len(staff_members)} staff member(s): {[s.get('username') for s in staff_members]}")
    
        # Tüm personellerin bu gün izinli olup olmadığını kontrol et
        all_staff_off = all(
            day_name in (staff.get('days_off') or [])
            for staff in staff_members
        )
        if all_staff_off:
            logging.info(f"❌ Availability: All staff are off on {day_name}. Staff: {[s.get('username') for s in staff_members]}, days_off: {[s.get('days_off') for s in staff_members]}")
            return {
                "available_slots": [],
                "all_slots": [],
                "busy_slots": [],
                "message": "Tüm personeller bu gün izinli"
            }
    
    # business_hours'dan o günün saatlerini al
    day_hours = business_hours.get(day_name, {})
    logging.info(f"🔍 Availability Check - Date: {date}, Day: {day_name}, day_hours: {day_hours}")
    is_open_value = day_hours.get('is_open', True)
    logging.info(f"🔍 is_open value: {is_open_value}, type: {type(is_open_value)}")
    
    if not is_open_value:
        # İşletme bu gün kapalı
        logging.info(f"❌ Business is closed on {day_name}")
        return {
            "available_slots": [],
            "all_slots": [],
            "busy_slots": [],
            "message": "İşletme bu gün kapalı"
        }
    
    logging.info(f"✅ Business is open on {day_name}")
    
    # Açılış ve kapanış saatlerini parse et
    open_time_str = day_hours.get('open_time', '09:00')
    close_time_str = day_hours.get('close_time', '18:00')
    
    try:
        open_hour, open_minute = map(int, open_time_str.split(':'))
        close_hour, close_minute = map(int, close_time_str.split(':'))
        # "00:00" gece yarısı demek — 24:00 olarak işle (günün sonu)
        if close_hour == 0 and close_minute == 0:
            close_hour = 24
    except (ValueError, AttributeError):
        # Varsayılan saatler
        open_hour, open_minute = 9, 0
        close_hour, close_minute = 18, 0
    
    # Hizmet süresini al — çoklu hizmette TOPLAM süre (kesintisiz blok)
    _svc_docs = await db.services.find({"id": {"$in": service_id_list}, "organization_id": organization_id}, {"_id": 0, "id": 1, "duration": 1}).to_list(len(service_id_list))
    _dur_map = {d["id"]: int(d.get("duration", 30) or 30) for d in _svc_docs}
    if any(sid not in _dur_map for sid in service_id_list):
        return {"available_slots": [], "all_slots": [], "busy_slots": [], "message": "Hizmet bulunamadı"}
    service_duration = sum(_dur_map[sid] for sid in service_id_list)  # Dakika cinsinden (toplam)
    
    # KRİTİK: Personel aynı anda sadece 1 müşteriye hizmet verebilir
    # O hizmeti verebilen personellerin o tarihteki TÜM randevularını çek (hangi hizmet olursa olsun)
    staff_ids = [staff['username'] for staff in staff_members]
    logging.info(f"📋 Availability: staff_members count: {len(staff_members)}, staff_ids: {staff_ids}, staff_id param: {staff_id}")
    
    # Personellerin o gün için TÜM randevularını al (tüm hizmetler dahil) - başlangıç ve bitiş saatleriyle
    # staff_member_id=None (atanmamış) randevuları da dahil et — tek kişilik org'larda çakışma kaçmasın
    appt_query = {
        "organization_id": organization_id,
        "appointment_date": date,
        "status": {"$nin": ["İptal", "İptal Edildi", "Cancelled", "Ödeme Bekleniyor"]},
        "$or": [
            {"staff_member_id": {"$in": staff_ids}},
            {"staff_member_id": None}
        ]
    }
    if exclude_appointment_id:
        appt_query["id"] = {"$ne": exclude_appointment_id}
    all_staff_appointments = await db.appointments.find(
        appt_query,
        {"_id": 0, "appointment_time": 1, "staff_member_id": 1, "service_name": 1, "service_id": 1, "service_duration": 1, "services": 1}
    ).to_list(1000)
    
    logging.info(f"📋 Found {len(all_staff_appointments)} appointments for staff_ids: {staff_ids}")
    if len(all_staff_appointments) > 0:
        for appt in all_staff_appointments[:5]:  # İlk 5 randevuyu logla
            logging.info(f"   - {appt.get('appointment_time')} - Staff: {appt.get('staff_member_id')} - Service: {appt.get('service_id')}")
    else:
        logging.info(f"   ⚠️ No appointments found for date {date} and staff_ids {staff_ids}")
        # Tüm randevuları kontrol et (debug için)
        all_appts_debug = await db.appointments.find(
            {
                "organization_id": organization_id,
                "appointment_date": date,
                "status": {"$nin": ["İptal", "İptal Edildi", "Ödeme Bekleniyor"]}
            },
            {"_id": 0, "appointment_time": 1, "staff_member_id": 1, "service_id": 1}
        ).to_list(10)
        logging.info(f"   🔍 Debug: Total appointments for date {date} (all staff): {len(all_appts_debug)}")
        for appt_debug in all_appts_debug:
            logging.info(f"      - {appt_debug.get('appointment_time')} - Staff: {appt_debug.get('staff_member_id')} - Service: {appt_debug.get('service_id')}")
    
    # Her randevunun bitiş saatini hesapla (hizmet süresine göre)
    appointments_with_end_time = []
    for appt in all_staff_appointments:
        # Çoklu hizmette saklı TOPLAM süreyi kullan; yoksa canlı tek-hizmet süresi
        stored_dur = appt.get('service_duration')
        if stored_dur:
            appt_duration = int(stored_dur)
        else:
            appt_service_id = appt.get('service_id')
            if appt_service_id:
                appt_service = await db.services.find_one({"id": appt_service_id}, {"_id": 0, "duration": 1})
                appt_duration = appt_service.get('duration', 30) if appt_service else 30
            else:
                appt_duration = 30  # Varsayılan
        
        start_time_str = appt['appointment_time']
        start_hour, start_minute = map(int, start_time_str.split(':'))
        end_minute = start_minute + appt_duration
        end_hour = start_hour + (end_minute // 60)
        end_minute = end_minute % 60
        end_time_str = f"{str(end_hour).zfill(2)}:{str(end_minute).zfill(2)}"
        
        appointments_with_end_time.append({
            "start_time": start_time_str,
            "end_time": end_time_str,
            "staff_member_id": appt['staff_member_id']
        })
    
    # Personel molalarını da meşgul slotlara ekle
    for staff in staff_members:
        staff_breaks = staff.get('breaks', [])
        for brk in staff_breaks:
            if brk.get('date') == date:
                appointments_with_end_time.append({
                    "start_time": brk['start_time'],
                    "end_time": brk['end_time'],
                    "staff_member_id": staff['username'],
                    "is_break": True
                })
    
    # Gizli adım aralığı (15 dakika)
    STEP_INTERVAL = 15  # Dakika
    
    # Bugünün saatini al (geçmiş saat kontrolü için)
    now = datetime.now(timezone.utc)
    turkey_tz = timezone(timedelta(hours=3))
    now_turkey = now.astimezone(turkey_tz)
    is_today = date == now_turkey.strftime("%Y-%m-%d")
    current_hour_now = now_turkey.hour
    current_minute_now = now_turkey.minute
    
    # Potansiyel slotları oluştur (15 dakikalık adımlarla)
    potential_slots = []
    current_hour = open_hour
    current_minute = open_minute
    
    while True:
        # Kapanış saatine ulaştık mı kontrol et
        if current_hour > close_hour or (current_hour == close_hour and current_minute >= close_minute):
            break
        
        time_str = f"{str(current_hour).zfill(2)}:{str(current_minute).zfill(2)}"
        potential_slots.append(time_str)
        
        # 15 dakika ilerle
        current_minute += STEP_INTERVAL
        if current_minute >= 60:
            current_minute = current_minute % 60
            current_hour += 1
    
    # Müsait saatleri hesapla
    final_available_slots = []
    busy_slots = []
    
    for potential_start_time in potential_slots:
        # Bitiş saatini hesapla
        start_hour, start_minute = map(int, potential_start_time.split(':'))
        end_minute = start_minute + service_duration
        end_hour = start_hour + (end_minute // 60)
        end_minute = end_minute % 60
        potential_end_time = f"{str(end_hour).zfill(2)}:{str(end_minute).zfill(2)}"
        
        # Geçmiş saat kontrolü
        if is_today:
            if start_hour < current_hour_now or (start_hour == current_hour_now and start_minute < current_minute_now):
                # Geçmiş saat - busy slot olarak işaretle (gösterilecek ama seçilemeyecek)
                busy_slots.append(potential_start_time)
                continue  # Bu slotu atla
        
        # Genel saat kontrolü (bitiş saati kapanış saatini aşıyor mu?)
        if end_hour > close_hour or (end_hour == close_hour and end_minute > close_minute):
            continue  # Bu slotu atla
        
        # Randevu çakışma kontrolü
        if staff_id:
            # Belirli bir personel seçildi - sadece onun randevularını kontrol et
            has_conflict = False
            for appt in appointments_with_end_time:
                appt_staff = appt.get('staff_member_id')
                # Boş/bilinmeyen staff_member_id → kim verirse versin slot dolu say
                if appt_staff and appt_staff != staff_id:
                    continue

                appt_start = appt['start_time']
                appt_end = appt['end_time']
                
                # Çakışma kontrolü: Zamanları sayısal değerlere çevir (dakika cinsinden)
                def time_to_minutes(time_str):
                    """Zaman string'ini (HH:MM) dakika cinsinden sayıya çevir"""
                    try:
                        hour, minute = map(int, time_str.split(':'))
                        return hour * 60 + minute
                    except (ValueError, AttributeError):
                        return 0
                
                potential_start_min = time_to_minutes(potential_start_time)
                potential_end_min = time_to_minutes(potential_end_time)
                appt_start_min = time_to_minutes(appt_start)
                appt_end_min = time_to_minutes(appt_end)
                
                # Çakışma kontrolü: (yeni_başlangıç < mevcut_bitiş) VE (yeni_bitiş > mevcut_başlangıç)
                if (potential_start_min < appt_end_min and potential_end_min > appt_start_min):
                    has_conflict = True
                    logging.info(f"   ⚠️ Conflict: Slot {potential_start_time}-{potential_end_time} overlaps with {appt_start}-{appt_end} (Staff: {staff_id})")
                    break
            
            if has_conflict:
                # Seçili personel için busy slot
                busy_slots.append(potential_start_time)
                continue
            else:
                # Seçili personel için müsait (çakışma yok)
                final_available_slots.append(potential_start_time)
                logging.debug(f"   ✅ Available slot: {potential_start_time}-{potential_end_time} (Staff: {staff_id})")
        else:
            # Otomatik atama - Tüm personeller için kontrol et
            # En az bir personel müsait olmalı
            busy_staff_at_slot = []
            
            # Zamanları sayısal değerlere çevir (dakika cinsinden)
            def time_to_minutes(time_str):
                """Zaman string'ini (HH:MM) dakika cinsinden sayıya çevir"""
                try:
                    hour, minute = map(int, time_str.split(':'))
                    return hour * 60 + minute
                except (ValueError, AttributeError):
                    return 0
            
            potential_start_min = time_to_minutes(potential_start_time)
            potential_end_min = time_to_minutes(potential_end_time)
            
            for appt in appointments_with_end_time:
                appt_start = appt['start_time']
                appt_end = appt['end_time']
                
                appt_start_min = time_to_minutes(appt_start)
                appt_end_min = time_to_minutes(appt_end)
                
                # Bu personel bu slotta dolu mu?
                if (potential_start_min < appt_end_min and potential_end_min > appt_start_min):
                    busy_staff_at_slot.append(appt['staff_member_id'])
            
            busy_staff_unique = list(set(busy_staff_at_slot))
            
            # Eğer tüm personeller doluysa busy slot
            if len(busy_staff_unique) >= len(staff_members):
                busy_slots.append(potential_start_time)
                logging.debug(f"   🚫 All staff busy at {potential_start_time}-{potential_end_time}")
            else:
                # En az bir personel müsait - slot müsait
                final_available_slots.append(potential_start_time)
                available_count = len(staff_members) - len(busy_staff_unique)
                logging.debug(f"   ✅ Available slot: {potential_start_time}-{potential_end_time} ({available_count}/{len(staff_members)} staff available)")
    
    logging.info(f"🔍 Services: {service_id_list}, Date: {date}, Duration: {service_duration}min")
    logging.info(f"👥 Qualified staff: {len(staff_members)} - {staff_ids}")
    logging.info(f"📅 Total appointments: {len(appointments_with_end_time)}")
    logging.info(f"✅ Available slots: {len(final_available_slots)}")
    logging.info(f"🚫 Busy slots: {len(busy_slots)}")
    _avail_result = {
        "available_slots": final_available_slots,
        "busy_slots": busy_slots,
        "all_slots": potential_slots,
        "message": "OK"
    }
    if redis_client:
        try:
            await redis_client.setex(_avail_cache_key, 45, json.dumps(_avail_result))
        except Exception:
            pass
    return _avail_result


@api_router.get("/availability")
async def get_internal_availability(
    request: Request,
    date: str,
    service_id: Optional[str] = None,
    service_ids: Optional[str] = None,
    staff_id: Optional[str] = None,
    exclude_appointment_id: Optional[str] = None,
    current_user: UserInDB = Depends(get_current_user)
):
    """İç kullanım müsaitlik endpoint'i.

    Admin panelindeki randevu oluşturma akışında, personel seçilmediyse (auto-assign)
    slotlar personelin haftalık izin günlerine (days_off) bağlı olmamalıdır.
    staff_id verilirse, seçilen personelin days_off kontrolü uygulanır.
    Çoklu hizmet: service_ids (virgülle ayrılmış) verilirse toplam süre + personel kesişimi.
    """
    db = await get_db_from_request(request)
    organization_id = current_user.organization_id

    service_id_list = [s.strip() for s in service_ids.split(",") if s.strip()] if service_ids else ([service_id] if service_id else [])
    if not service_id_list:
        raise HTTPException(status_code=400, detail="Hizmet seçilmedi")
    service_id = service_id_list[0]

    if staff_id is not None:
        _sid = str(staff_id).strip().lower()
        if _sid in ("", "auto", "none"):
            staff_id = None

    settings = await db.settings.find_one({"organization_id": organization_id}, {"_id": 0})
    if not settings:
        settings = {
            "work_start_hour": 9,
            "work_end_hour": 18,
            "appointment_interval": 30,
            "admin_provides_service": True,
            "business_hours": {
                "monday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
                "tuesday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
                "wednesday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
                "thursday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
                "friday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
                "saturday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"},
                "sunday": {"is_open": True, "open_time": "09:00", "close_time": "18:00"}
            }
        }

    admin_provides_service = settings.get('admin_provides_service', True)
    business_hours = settings.get('business_hours', {})

    # Tarihin hangi güne denk geldiğini bul (0=Monday, 6=Sunday)
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        weekday = date_obj.weekday()  # 0=Monday, 6=Sunday
        day_names = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
        day_name = day_names[weekday]
    except (ValueError, TypeError):
        day_name = "monday"  # Varsayılan

    if staff_id:
        staff_query = {
            "organization_id": organization_id,
            "username": staff_id,
            "permitted_service_ids": {"$all": service_id_list}
        }

        if not admin_provides_service:
            staff_query["role"] = {"$ne": "admin"}

        staff_members = await db.users.find(
            staff_query,
            {"_id": 0, "id": 1, "username": 1, "role": 1, "days_off": 1, "permitted_service_ids": 1, "breaks": 1}
        ).to_list(1000)

        # Fallback: admin_provides_service=False filtre admin'i elediğinde,
        # AMA seçilen staff_id gerçekten admin ve admin permitted_service_ids'inde
        # bu hizmete açıkça izin vermişse, kullan. Bu, public /availability,
        # /sessions/wave-plan ve /availability/bulk-check ile tutarlıdır.
        if not staff_members:
            admin_user = await db.users.find_one(
                {"organization_id": organization_id, "role": "admin", "username": staff_id},
                {"_id": 0, "id": 1, "username": 1, "role": 1, "days_off": 1, "permitted_service_ids": 1, "breaks": 1}
            )
            if admin_user and all(s in (admin_user.get('permitted_service_ids') or []) for s in service_id_list):
                staff_members = [admin_user]
                logging.info(f"⚠️ Internal /availability: admin_provides_service=False ama admin {staff_id} permitted'da, fallback aktif")

        if not staff_members:
            return {"available_slots": [], "all_slots": [], "busy_slots": [], "message": "Seçilen personel bu hizmeti veremiyor"}

        selected_staff = staff_members[0]
        staff_days_off = selected_staff.get('days_off') or []
        if day_name in staff_days_off:
            return {
                "available_slots": [],
                "all_slots": [],
                "busy_slots": [],
                "message": "Seçili personel bu gün izinli"
            }
    else:
        staff_query = {
            "organization_id": organization_id,
            "permitted_service_ids": {"$all": service_id_list}
        }

        if not admin_provides_service:
            staff_query["role"] = {"$ne": "admin"}

        staff_members = await db.users.find(
            staff_query,
            {"_id": 0, "id": 1, "username": 1, "role": 1, "days_off": 1, "permitted_service_ids": 1, "breaks": 1}
        ).to_list(1000)

        # Fallback: hizmet için yetkili staff yoksa AMA admin permitted'da varsa, admin'i kullan
        # (public /availability ile tutarlı)
        if not staff_members:
            admin_user = await db.users.find_one(
                {"organization_id": organization_id, "role": "admin"},
                {"_id": 0, "id": 1, "username": 1, "role": 1, "days_off": 1, "permitted_service_ids": 1, "breaks": 1}
            )
            if admin_user and all(s in (admin_user.get('permitted_service_ids') or []) for s in service_id_list):
                staff_members = [admin_user]
                logging.info(f"⚠️ Internal /availability (no staff_id): falling back to admin {admin_user['username']}")

        if not staff_members:
            # Admin panelinde hiç uygun personel yoksa (ve staff_id seçilmemişse),
            # admin unassigned randevu yazabilsin diye slotları işletme saatlerinden üret.
            # Not: staff_id seçilmişse zaten yukarıda "Seçilen personel..." mesajı dönüyoruz.
            if current_user.role != "admin":
                return {"available_slots": [], "all_slots": [], "busy_slots": [], "message": "Bu hizmet için uygun personel bulunamadı"}

            # business_hours'dan o günün saatlerini al
            day_hours = business_hours.get(day_name, {})
            is_open_value = day_hours.get('is_open', True)
            if not is_open_value:
                return {
                    "available_slots": [],
                    "all_slots": [],
                    "busy_slots": [],
                    "message": "İşletme bu gün kapalı"
                }

            open_time_str = day_hours.get('open_time', '09:00')
            close_time_str = day_hours.get('close_time', '18:00')
            try:
                open_hour, open_minute = map(int, open_time_str.split(':'))
                close_hour, close_minute = map(int, close_time_str.split(':'))
                # "00:00" gece yarısı demek — 24:00 olarak işle (günün sonu)
                if close_hour == 0 and close_minute == 0:
                    close_hour = 24
            except (ValueError, AttributeError):
                open_hour, open_minute = 9, 0
                close_hour, close_minute = 18, 0

            _svc_docs = await db.services.find({"id": {"$in": service_id_list}, "organization_id": organization_id}, {"_id": 0, "id": 1, "duration": 1}).to_list(len(service_id_list))
            _dur_map = {d["id"]: int(d.get("duration", 30) or 30) for d in _svc_docs}
            if any(sid not in _dur_map for sid in service_id_list):
                return {"available_slots": [], "all_slots": [], "busy_slots": [], "message": "Hizmet bulunamadı"}
            service_duration = sum(_dur_map[sid] for sid in service_id_list)

            # Organizasyondaki tüm randevular: aynı gün/saatte çakışma varsa slotu "busy" say.
            int_appt_query = {
                "organization_id": organization_id,
                "appointment_date": date,
                "status": {"$nin": ["İptal", "İptal Edildi", "Ödeme Bekleniyor"]},
            }
            if exclude_appointment_id:
                int_appt_query["id"] = {"$ne": exclude_appointment_id}
            org_appointments = await db.appointments.find(
                int_appt_query,
                {"_id": 0, "appointment_time": 1, "service_id": 1, "service_duration": 1}
            ).to_list(1000)

            appointments_with_end_time = []
            for appt in org_appointments:
                # Çoklu hizmette saklı TOPLAM süreyi kullan; yoksa canlı tek-hizmet süresi
                stored_dur = appt.get('service_duration')
                if stored_dur:
                    appt_duration = int(stored_dur)
                else:
                    appt_service_id = appt.get('service_id')
                    if appt_service_id:
                        appt_service = await db.services.find_one(
                            {"id": appt_service_id, "organization_id": organization_id},
                            {"_id": 0, "duration": 1}
                        )
                        appt_duration = appt_service.get('duration', 30) if appt_service else 30
                    else:
                        appt_duration = 30

                start_time_str = appt.get('appointment_time')
                if not start_time_str:
                    continue
                start_hour, start_minute = map(int, start_time_str.split(':'))
                end_minute = start_minute + appt_duration
                end_hour = start_hour + (end_minute // 60)
                end_minute = end_minute % 60
                end_time_str = f"{str(end_hour).zfill(2)}:{str(end_minute).zfill(2)}"
                appointments_with_end_time.append({
                    "start_time": start_time_str,
                    "end_time": end_time_str,
                })

            STEP_INTERVAL = 15

            now = datetime.now(timezone.utc)
            turkey_tz = timezone(timedelta(hours=3))
            now_turkey = now.astimezone(turkey_tz)
            is_today = date == now_turkey.strftime("%Y-%m-%d")
            current_now_minutes = now_turkey.hour * 60 + now_turkey.minute

            potential_slots = []
            current_hour = open_hour
            current_minute = open_minute
            while True:
                if current_hour > close_hour or (current_hour == close_hour and current_minute >= close_minute):
                    break
                potential_slots.append(f"{str(current_hour).zfill(2)}:{str(current_minute).zfill(2)}")
                current_minute += STEP_INTERVAL
                if current_minute >= 60:
                    current_minute = current_minute % 60
                    current_hour += 1

            def time_to_minutes(time_str: str) -> int:
                try:
                    h, m = map(int, time_str.split(':'))
                    return h * 60 + m
                except Exception:
                    return 0

            available_slots = []
            busy_slots = []
            for start_time in potential_slots:
                start_minutes = time_to_minutes(start_time)
                end_minutes = start_minutes + service_duration

                # Kapanış saatini aşan başlangıçları engelle
                close_minutes = close_hour * 60 + close_minute
                if end_minutes > close_minutes:
                    continue

                if is_today and start_minutes <= current_now_minutes:
                    continue

                has_conflict = False
                for ex in appointments_with_end_time:
                    ex_start = time_to_minutes(ex['start_time'])
                    ex_end = time_to_minutes(ex['end_time'])
                    if start_minutes < ex_end and end_minutes > ex_start:
                        has_conflict = True
                        break

                if has_conflict:
                    busy_slots.append(start_time)
                else:
                    available_slots.append(start_time)

            return {
                "available_slots": available_slots,
                "busy_slots": busy_slots,
                "all_slots": potential_slots,
                "message": "OK"
            }

    # business_hours'dan o günün saatlerini al
    day_hours = business_hours.get(day_name, {})
    is_open_value = day_hours.get('is_open', True)
    if not is_open_value:
        return {
            "available_slots": [],
            "all_slots": [],
            "busy_slots": [],
            "message": "İşletme bu gün kapalı"
        }

    open_time_str = day_hours.get('open_time', '09:00')
    close_time_str = day_hours.get('close_time', '18:00')
    try:
        open_hour, open_minute = map(int, open_time_str.split(':'))
        close_hour, close_minute = map(int, close_time_str.split(':'))
        # "00:00" gece yarısı demek — 24:00 olarak işle (günün sonu)
        if close_hour == 0 and close_minute == 0:
            close_hour = 24
    except (ValueError, AttributeError):
        open_hour, open_minute = 9, 0
        close_hour, close_minute = 18, 0

    _svc_docs = await db.services.find({"id": {"$in": service_id_list}, "organization_id": organization_id}, {"_id": 0, "id": 1, "duration": 1}).to_list(len(service_id_list))
    _dur_map = {d["id"]: int(d.get("duration", 30) or 30) for d in _svc_docs}
    if any(sid not in _dur_map for sid in service_id_list):
        return {"available_slots": [], "all_slots": [], "busy_slots": [], "message": "Hizmet bulunamadı"}
    service_duration = sum(_dur_map[sid] for sid in service_id_list)

    staff_ids = [staff['username'] for staff in staff_members]
    # Unassigned randevuları (staff_member_id None/"" /eksik) da dahil et — admin
    # panelden personelsiz oluşturulan randevular busy_slots'a girmeli. Public
    # booking endpoint'iyle tutarlı davranış (Sorun 2).
    int_staff_query = {
        "organization_id": organization_id,
        "appointment_date": date,
        "status": {"$nin": ["İptal", "İptal Edildi", "Ödeme Bekleniyor"]},
        "$or": [
            {"staff_member_id": {"$in": staff_ids}},
            {"staff_member_id": None},
            {"staff_member_id": ""},
            {"staff_member_id": {"$exists": False}},
        ],
    }
    if exclude_appointment_id:
        int_staff_query["id"] = {"$ne": exclude_appointment_id}
    all_staff_appointments = await db.appointments.find(
        int_staff_query,
        {"_id": 0, "appointment_time": 1, "staff_member_id": 1, "service_id": 1, "service_duration": 1}
    ).to_list(1000)

    appointments_with_end_time = []
    for appt in all_staff_appointments:
        # Çoklu hizmette saklı TOPLAM süreyi kullan; yoksa canlı tek-hizmet süresi
        stored_dur = appt.get('service_duration')
        if stored_dur:
            appt_duration = int(stored_dur)
        else:
            appt_service_id = appt.get('service_id')
            if appt_service_id:
                appt_service = await db.services.find_one(
                    {"id": appt_service_id, "organization_id": organization_id},
                    {"_id": 0, "duration": 1}
                )
                appt_duration = appt_service.get('duration', 30) if appt_service else 30
            else:
                appt_duration = 30

        start_time_str = appt['appointment_time']
        start_hour, start_minute = map(int, start_time_str.split(':'))
        end_minute = start_minute + appt_duration
        end_hour = start_hour + (end_minute // 60)
        end_minute = end_minute % 60
        end_time_str = f"{str(end_hour).zfill(2)}:{str(end_minute).zfill(2)}"

        appointments_with_end_time.append({
            "start_time": start_time_str,
            "end_time": end_time_str,
            "staff_member_id": appt['staff_member_id']
        })

    for staff in staff_members:
        staff_breaks = staff.get('breaks', [])
        for brk in staff_breaks:
            if brk.get('date') == date:
                appointments_with_end_time.append({
                    "start_time": brk['start_time'],
                    "end_time": brk['end_time'],
                    "staff_member_id": staff['username'],
                    "is_break": True
                })

    STEP_INTERVAL = 15

    now = datetime.now(timezone.utc)
    turkey_tz = timezone(timedelta(hours=3))
    now_turkey = now.astimezone(turkey_tz)
    is_today = date == now_turkey.strftime("%Y-%m-%d")
    current_hour_now = now_turkey.hour
    current_minute_now = now_turkey.minute

    potential_slots = []
    current_hour = open_hour
    current_minute = open_minute
    while True:
        if current_hour > close_hour or (current_hour == close_hour and current_minute >= close_minute):
            break
        potential_slots.append(f"{str(current_hour).zfill(2)}:{str(current_minute).zfill(2)}")
        current_minute += STEP_INTERVAL
        if current_minute >= 60:
            current_minute = current_minute % 60
            current_hour += 1

    final_available_slots = []
    busy_slots = []

    def time_to_minutes(time_str):
        try:
            hour, minute = map(int, time_str.split(':'))
            return hour * 60 + minute
        except (ValueError, AttributeError):
            return 0

    for potential_start_time in potential_slots:
        start_hour, start_minute = map(int, potential_start_time.split(':'))
        end_minute_total = start_minute + service_duration
        end_hour = start_hour + (end_minute_total // 60)
        end_minute = end_minute_total % 60
        potential_end_time = f"{str(end_hour).zfill(2)}:{str(end_minute).zfill(2)}"

        if is_today:
            if start_hour < current_hour_now or (start_hour == current_hour_now and start_minute < current_minute_now):
                busy_slots.append(potential_start_time)
                continue

        if end_hour > close_hour or (end_hour == close_hour and end_minute > close_minute):
            continue

        potential_start_min = time_to_minutes(potential_start_time)
        potential_end_min = time_to_minutes(potential_end_time)

        if staff_id:
            has_conflict = False
            for appt in appointments_with_end_time:
                appt_staff = appt.get('staff_member_id')
                # Boş/bilinmeyen staff_member_id → slot kim verirse versin dolu say
                # (auto-assign edilmiş ya da silinmiş personele ait randevular için güvenli davranış)
                if appt_staff and appt_staff != staff_id:
                    continue
                appt_start_min = time_to_minutes(appt['start_time'])
                appt_end_min = time_to_minutes(appt['end_time'])
                if (potential_start_min < appt_end_min and potential_end_min > appt_start_min):
                    has_conflict = True
                    break
            if has_conflict:
                busy_slots.append(potential_start_time)
                continue
            final_available_slots.append(potential_start_time)
        else:
            # Auto-assign: en az bir personelde çakışma olmamalı (days_off dikkate alınmaz)
            any_available = False
            for staff_username in staff_ids:
                has_conflict = False
                for appt in appointments_with_end_time:
                    appt_staff = appt.get('staff_member_id')
                    # Boş/bilinmeyen staff'lı randevular her staff için çakışma sayılır
                    if appt_staff and appt_staff != staff_username:
                        continue
                    appt_start_min = time_to_minutes(appt['start_time'])
                    appt_end_min = time_to_minutes(appt['end_time'])
                    if (potential_start_min < appt_end_min and potential_end_min > appt_start_min):
                        has_conflict = True
                        break
                if not has_conflict:
                    any_available = True
                    break
            if any_available:
                final_available_slots.append(potential_start_time)
            else:
                busy_slots.append(potential_start_time)

    return {
        "available_slots": final_available_slots,
        "busy_slots": busy_slots,
        "all_slots": potential_slots,
        "message": "OK"
    }

# ═══════════════════════════════════════════════════════════════════════
# 🔒 PLANN GÜVENLİK — Yardımcı Fonksiyonlar (5 Katmanlı Savunma)
# ═══════════════════════════════════════════════════════════════════════

async def verify_turnstile(token: str, client_ip: str = "") -> bool:
    """Cloudflare Turnstile token doğrulama — Katman 1"""
    secret = os.getenv("TURNSTILE_SECRET_KEY", "")
    if not secret:
        logging.warning("TURNSTILE_SECRET_KEY eksik — Turnstile atlanıyor (dev mode)")
        return True
    if not token:
        logging.warning(f"🔒 Turnstile token boş geldi — IP: {client_ip}")
        return False
    try:
        def _check():
            resp = requests.post(
                "https://challenges.cloudflare.com/turnstile/v0/siteverify",
                data={"secret": secret, "response": token, "remoteip": client_ip},
                timeout=5.0,
            )
            return resp.json().get("success", False)
        return await asyncio.to_thread(_check)
    except Exception as e:
        logging.error(f"Turnstile doğrulama hatası: {e}")
        return False

_PHONE_COUNTRY_MAP = {
    "90": "TR", "44": "GB", "1": "US", "49": "DE",
    "33": "FR", "39": "IT", "34": "ES", "31": "NL",
    "7": "RU", "86": "CN", "81": "JP", "82": "KR",
    "61": "AU", "55": "BR", "91": "IN",
}

def get_phone_country(phone: str) -> str:
    """Telefon numarasından ülke kodu tahmin et — Katman 4"""
    digits = re.sub(r"\D", "", phone)
    if digits.startswith("0") and len(digits) == 11:
        return "TR"
    for prefix_len in (3, 2, 1):
        country = _PHONE_COUNTRY_MAP.get(digits[:prefix_len])
        if country:
            return country
    return ""

async def get_geo_country(redis_client, ip: str) -> str:
    """IP → ülke kodu (Redis 24h cache, ip-api.com) — Katman 4"""
    if not ip or ip in ("127.0.0.1", "::1", ""):
        return ""
    cache_key = f"plann:geo:{ip}"
    try:
        if redis_client:
            cached = await redis_client.get(cache_key)
            if cached:
                return cached.decode() if isinstance(cached, bytes) else cached
        def _fetch():
            resp = requests.get(
                f"http://ip-api.com/json/{ip}?fields=countryCode",
                timeout=3.0,
            )
            return resp.json().get("countryCode", "")
        country = await asyncio.to_thread(_fetch)
        if redis_client and country:
            await redis_client.setex(cache_key, 86400, country)
        return country
    except Exception:
        return ""

async def create_pending_verification(redis_client, phone: str, org_id: str, appointment_data: dict) -> str:
    """
    WhatsApp doğrulama için Redis'e kayıt yaz (TTL 15 dk).
    Aynı numara için eski pending varsa EZİLİR — upsert mantığı.
    APScheduler gerekmez, Redis TTL otomatik temizler.
    """
    code = str(secrets.randbelow(900000) + 100000)
    phone_map_key = f"plann:pending:phone:{phone}:{org_id}"
    ttl = 900  # 15 dakika
    if redis_client:
        old_code = await redis_client.get(phone_map_key)
        if old_code:
            old_str = old_code.decode() if isinstance(old_code, bytes) else old_code
            await redis_client.delete(f"plann:pending:{old_str}")
        code_key = f"plann:pending:{code}"
        await redis_client.hset(code_key, mapping={
            "phone": phone,
            "org_id": org_id,
            "appointment_data": json.dumps(appointment_data, ensure_ascii=False),
        })
        await redis_client.expire(code_key, ttl)
        await redis_client.setex(phone_map_key, ttl, code)
    return code


# ═══ CONTRACT: v1 (FROZEN / REQUEST) — public booking ═══════════════════════
# Donmuş mobil + web public build'ler bu body'yi gönderir. AppointmentCreate'e
# YENİ ZORUNLU alan EKLEME (eski client 400 alır). bkz. OPERATIONS.md §2.1
@api_router.post("/public/appointments")
async def create_public_appointment(request: Request, appointment: AppointmentCreate, organization_id: str):
    """Model D: Public randevu oluştur — 5 Katmanlı Güvenlik + Akıllı personel atama"""
    db = await get_db_from_request(request)
    redis_client = getattr(request.app.state, 'redis_client', None)

    # Gerçek istemci IP'si (Cloudflare proxy arkasında da doğru çalışır)
    client_ip = (
        request.headers.get("CF-Connecting-IP") or
        request.headers.get("X-Forwarded-For", "").split(",")[0].strip() or
        (request.client.host if request.client else "")
    )

    # ═══════════════════════════════════════════════════════════
    # 🔒 KATMAN 1: CLOUDFLARE TURNSTILE (soft — diğer katmanlar korur)
    # ═══════════════════════════════════════════════════════════
    turnstile_ok = await verify_turnstile(appointment.turnstile_token or "", client_ip)
    if not turnstile_ok:
        logging.warning(f"⚠️ Turnstile doğrulama başarısız — IP: {client_ip}, devam ediliyor (soft mode)")

    # ═══════════════════════════════════════════════════════════
    # 🔒 KATMAN 2: IP RATE LIMIT (15 istek / 12 saat)
    # ═══════════════════════════════════════════════════════════
    if redis_client and client_ip and client_ip not in ("127.0.0.1", "::1"):
        ip_key = f"plann:rate:ip:{client_ip}"
        ip_count = await redis_client.incr(ip_key)
        if ip_count == 1:
            await redis_client.expire(ip_key, 43200)  # 12 saat
        if ip_count > 50:
            raise HTTPException(
                status_code=429,
                detail="Bu ağdan çok fazla randevu talebi geldi. Lütfen daha sonra tekrar deneyin.",
            )

    # ═══════════════════════════════════════════════════════════
    # 🔒 KATMAN 3: TANINIK MÜŞTERİ BYPASS (MongoDB customers)
    # ═══════════════════════════════════════════════════════════
    # Normalize phone: strip +, leading 0, country code → last 10 digits
    _raw_phone = re.sub(r'\D', '', appointment.phone or "")
    _phone_suffix = _raw_phone[-10:] if len(_raw_phone) >= 10 else _raw_phone
    # Separator-tolerant suffix match: stored phone may include spaces, dashes,
    # plus signs or parentheses (e.g. "+90 555 123 45 67"). Build a regex that
    # allows non-digit chars between each of the last-10 digits so the suffix
    # still matches regardless of formatting.
    if _phone_suffix:
        _flex_regex = r"\D*" + r"\D*".join(re.escape(d) for d in _phone_suffix) + r"$"
    else:
        _flex_regex = "^$"
    existing_customer = await db.customers.find_one({
        "organization_id": organization_id,
        "phone": {"$regex": _flex_regex},
    })

    # ═══════════════════════════════════════════════════════════
    # 🔒 KATMAN 4: COĞRAFİ RİSK SKORU (ip-api.com, Redis 24h cache)
    # ═══════════════════════════════════════════════════════════
    high_risk = False
    if not existing_customer:
        geo_country = await get_geo_country(redis_client, client_ip)
        phone_country = get_phone_country(appointment.phone)
        if geo_country and phone_country and geo_country != phone_country:
            high_risk = True
            logging.info(f"🌍 Geo uyuşmazlık: IP={geo_country}, tel={phone_country} → yüksek risk")

    # Çoklu hizmet çözümleme (feature flag arkasında) + toplam ödeme kuralı.
    # Geçersiz girdide burada 400 döner — kota/lock'a hiç dokunulmaz.
    _settings_flag = await db.settings.find_one({"organization_id": organization_id}, {"_id": 0, "multi_service_enabled": 1})
    _multi_enabled = bool((_settings_flag or {}).get("multi_service_enabled", False))
    snapshot_lines, total_duration, total_price, service, resolved_service_ids = await resolve_services(
        db, organization_id, appointment, multi_enabled=_multi_enabled
    )
    # Geriye dönük uyum: appointment.service_id'yi ilk hizmete normalize et
    object.__setattr__(appointment, "service_id", resolved_service_ids[0])
    payment_rule = service.get("payment_rule", "on_site")

    # KATMAN 3 KARAR: VIP bypass mı, WhatsApp doğrulama mı?
    skip_verification = False

    # Online ödeme hizmetlerinde doğrulama atla — ödeme = kimlik doğrulama
    if payment_rule in ("online", "deposit"):
        skip_verification = True
        logging.info(f"✅ Online ödeme hizmeti → doğrulama atlanıyor (payment_rule={payment_rule})")
    elif existing_customer and not high_risk:
        if redis_client:
            known_key = f"plann:rate:known:{appointment.phone}"
            known_count = await redis_client.incr(known_key)
            if known_count == 1:
                await redis_client.expire(known_key, 86400)  # 24 saat
            # Anti-spam: tanıdık müşteri günde en fazla 3 randevuyu WhatsApp
            # doğrulamasız alabilir; sonrası WhatsApp zorunlu (spam'i kısar).
            if known_count <= 3:
                skip_verification = True
                logging.info(f"✅ Tanıdık müşteri bypass: {appointment.phone} ({known_count}/3 bugün)")
            else:
                logging.info(f"⚠️ Tanıdık müşteri günlük limit ({known_count}/3), WhatsApp zorunlu")
        else:
            skip_verification = True  # Redis yoksa bypass ver

    # ═══════════════════════════════════════════════════════════
    # 🔒 KATMAN 5: WHATSAPP NUMARA DOĞRULAMA (Kod BİZ göndeririz)
    # ═══════════════════════════════════════════════════════════
    if not skip_verification:
        # WA flood kontrolü (5 kod / 1 saat)
        if redis_client:
            wa_key = f"plann:rate:wa_req:{appointment.phone}"
            wa_count = await redis_client.incr(wa_key)
            if wa_count == 1:
                await redis_client.expire(wa_key, 3600)
            if wa_count > 5:
                raise HTTPException(
                    status_code=429,
                    detail="Çok fazla doğrulama kodu istediniz. Lütfen 1 saat sonra tekrar deneyin.",
                )

        appointment_payload = {
            "customer_name": appointment.customer_name,
            "phone": appointment.phone,
            "service_id": resolved_service_ids[0],
            "service_ids": resolved_service_ids,
            "appointment_date": appointment.appointment_date,
            "appointment_time": appointment.appointment_time,
            "notes": appointment.notes,
            "staff_member_id": appointment.staff_member_id,
        }
        code = await create_pending_verification(
            redis_client, appointment.phone, organization_id, appointment_payload
        )

        # Doğrulama kodunu müşteriye WhatsApp Authentication Template ile gönder
        try:
            from whatsapp_service import format_phone_number, detect_language_from_phone, _post_with_retry
            phone_clean = format_phone_number(appointment.phone)
            wa_lang = detect_language_from_phone(appointment.phone)
            wa_template_name = "randevu_dogrulama" if wa_lang == "TR" else "randevu_dogrulama_en"
            wa_lang_code = "tr" if wa_lang == "TR" else "en"

            wa_api_url = f"https://graph.facebook.com/{os.getenv('META_API_VERSION', 'v21.0')}/{os.getenv('META_PHONE_NUMBER_ID')}/messages"
            wa_headers = {
                "Authorization": f"Bearer {os.getenv('META_ACCESS_TOKEN')}",
                "Content-Type": "application/json",
            }
            wa_payload = {
                "messaging_product": "whatsapp",
                "to": phone_clean,
                "type": "template",
                "template": {
                    "name": wa_template_name,
                    "language": {"code": wa_lang_code},
                    "components": [
                        {
                            "type": "body",
                            "parameters": [
                                {"type": "text", "text": code}
                            ]
                        },
                        {
                            "type": "button",
                            "sub_type": "url",
                            "index": "0",
                            "parameters": [
                                {"type": "text", "text": code}
                            ]
                        }
                    ]
                }
            }
            await asyncio.to_thread(
                lambda: _post_with_retry(wa_api_url, wa_headers, wa_payload, context="verification_code")
            )
            logging.info(f"📱 Doğrulama kodu müşteriye gönderildi: {appointment.phone} → {code}")
        except Exception as wa_err:
            logging.error(f"❌ WA doğrulama kodu gönderim hatası: {wa_err}")

        return {
            "needs_verification": True,
            "code": code,
            "message": "Doğrulama kodunuz WhatsApp'a gönderildi.",
        }

    # ═══════════════════════════════════════════════════════════
    # ✅ GÜVENLİK KATMANLARI ONAYLADI — RANDEVU OLUŞTURMA
    # ═══════════════════════════════════════════════════════════

    # ---------------------------------------------------------
    # 🔒 ADIM 1: REDIS LOCK (ÇİFTE REZERVASYON ENGELLEYİCİ)
    # ---------------------------------------------------------
    lock_key = None # Hata durumunda silmek için dışarıda tanımlıyoruz

    if redis_client:
        # Eğer personel seçilmemişse "auto" kullanıyoruz
        staff_id_lock = appointment.staff_member_id or "auto"
        lock_key = f"plann:lock:appt:{organization_id}:{staff_id_lock}:{appointment.appointment_date}:{appointment.appointment_time}"
        
        acquired = await redis_client.setnx(lock_key, "locked")
        if not acquired:
            # Milisaniyelik farkla başkası tıkladı!
            raise HTTPException(
                status_code=409, 
                detail="Bu saat dilimi şu an başka bir müşteri tarafından dolduruluyor. Lütfen 10 saniye sonra tekrar deneyin."
            )
        await redis_client.expire(lock_key, 15) # Public tarafta işlem süresi (WhatsApp vs.) uzun sürebilir, 15sn ideal.
    
    # DEBUG: Frontend'ten gelen veriyi logla
    logging.info(f"🔍 PUBLIC APPOINTMENT REQUEST - staff_member_id: {appointment.staff_member_id}, service_id: {appointment.service_id}")
    
    # KOTA KONTROLÜ - Randevu oluşturmadan önce kontrol et
    quota_ok, quota_error = await check_quota_and_increment(db, organization_id)
    if not quota_ok:
        raise HTTPException(status_code=403, detail=quota_error)
    
    # Service zaten yukarıda resolve_services ile çözümlendi (merged_service = toplam süre/fiyat/isim)
    assigned_staff_id = None
    
    # AKILLI ATAMA MANTIĞI
    service_duration = service.get('duration', 30)
    
    # Yeni randevunun başlangıç ve bitiş saatlerini hesapla
    new_start_hour, new_start_minute = map(int, appointment.appointment_time.split(':'))
    new_end_minute = new_start_minute + service_duration
    new_end_hour = new_start_hour + (new_end_minute // 60)
    new_end_minute = new_end_minute % 60
    new_end_time = f"{str(new_end_hour).zfill(2)}:{str(new_end_minute).zfill(2)}"
    
    if appointment.staff_member_id:
        # Müşteri belirli bir personel seçti - çakışma kontrolü yap (duration'a göre)
        # Bu personelin o tarihteki tüm randevularını çek
        existing_appointments = await db.appointments.find(
            {
            "organization_id": organization_id,
            "staff_member_id": appointment.staff_member_id,
            "appointment_date": appointment.appointment_date,
            "status": {"$nin": ["İptal", "İptal Edildi", "Ödeme Bekleniyor"]}
            },
            {"_id": 0, "appointment_time": 1, "service_id": 1, "service_duration": 1}
        ).to_list(100)
        
        # Her randevunun bitiş saatini hesapla ve çakışma kontrolü yap
        has_conflict = False
        for existing_appt in existing_appointments:
            existing_start_time = existing_appt['appointment_time']
            existing_service_id = existing_appt.get('service_id')
            
            # Mevcut randevunun hizmet süresini bul — çoklu hizmette saklı TOPLAM süre
            stored_dur = existing_appt.get('service_duration')
            if stored_dur:
                existing_duration = int(stored_dur)
            elif existing_service_id:
                existing_service = await db.services.find_one({"id": existing_service_id}, {"_id": 0, "duration": 1})
                existing_duration = existing_service.get('duration', 30) if existing_service else 30
            else:
                existing_duration = 30
            
            # Mevcut randevunun bitiş saatini hesapla
            existing_start_hour, existing_start_minute = map(int, existing_start_time.split(':'))
            existing_end_minute = existing_start_minute + existing_duration
            existing_end_hour = existing_start_hour + (existing_end_minute // 60)
            existing_end_minute = existing_end_minute % 60
            existing_end_time = f"{str(existing_end_hour).zfill(2)}:{str(existing_end_minute).zfill(2)}"
            
            # Çakışma kontrolü
            if (appointment.appointment_time < existing_end_time and new_end_time > existing_start_time):
                has_conflict = True
                logging.info(f"⚠️ Public booking conflict: New {appointment.appointment_time}-{new_end_time} overlaps with existing {existing_start_time}-{existing_end_time}")
                break
        
        if has_conflict:
            # Kota artırıldı ama personel dolu, geri al
            plan_doc = await db.organization_plans.find_one({"organization_id": organization_id})
            if plan_doc:
                await db.organization_plans.update_one(
                    {"organization_id": organization_id},
                    {"$inc": {"quota_usage": -1}}
                )
            raise HTTPException(
                status_code=400,
                detail="Seçtiğiniz personel bu saatte dolu. Lütfen başka bir saat veya personel seçin."
            )
        
        # Mola çakışması kontrolü
        has_break_conflict = await check_break_conflict(
            db, appointment.staff_member_id, appointment.appointment_date,
            appointment.appointment_time, new_end_time, organization_id
        )
        
        if has_break_conflict:
            # Kota artırıldı ama mola çakışması var, geri al
            plan_doc = await db.organization_plans.find_one({"organization_id": organization_id})
            if plan_doc:
                await db.organization_plans.update_one(
                    {"organization_id": organization_id},
                    {"$inc": {"quota_usage": -1}}
                )
            raise HTTPException(
                status_code=400,
                detail="Seçtiğiniz personel bu saatte mola. Lütfen başka bir saat veya personel seçin."
            )
        
        assigned_staff_id = appointment.staff_member_id
    else:
        # Müşteri "Farketmez" seçti veya personel seçimi yok
        # Önce customer_can_choose_staff ve admin_provides_service ayarlarını kontrol et
        settings_data = await db.settings.find_one({"organization_id": organization_id})
        customer_can_choose_staff = settings_data.get('customer_can_choose_staff', False) if settings_data else False
        admin_provides_service = settings_data.get('admin_provides_service', True) if settings_data else True
        
        # Eğer customer_can_choose_staff kapalıysa ama admin_provides_service açıksa, otomatik atama yap
        # customer_can_choose_staff açıksa da otomatik atama yap
        # Her ikisi de kapalıysa bile, admin hizmet verebiliyorsa otomatik atama yap
        if not customer_can_choose_staff and not admin_provides_service:
            # Her ikisi de kapalıysa, önce admin'i kontrol et
            admin_user = await db.users.find_one(
                {"organization_id": organization_id, "role": "admin"},
                {"_id": 0, "username": 1, "role": 1, "permitted_service_ids": 1}
            )
            if admin_user and all(sid in (admin_user.get('permitted_service_ids') or []) for sid in resolved_service_ids):
                # Çakışma kontrolü — admin'in VE atanmamış randevuları kontrol et
                admin_username = admin_user['username']
                conflict_appts = await db.appointments.find(
                    {
                        "organization_id": organization_id,
                        "$or": [
                            {"staff_member_id": admin_username},
                            {"staff_member_id": None}
                        ],
                        "appointment_date": appointment.appointment_date,
                        "status": {"$nin": ["İptal", "İptal Edildi", "Ödeme Bekleniyor"]}
                    },
                    {"_id": 0, "appointment_time": 1, "service_id": 1}
                ).to_list(100)

                has_conflict = False
                for ca in conflict_appts:
                    ca_start = ca.get('appointment_time', '')
                    if not ca_start:
                        continue
                    ca_svc = await db.services.find_one({"id": ca.get('service_id', '')}, {"_id": 0, "duration": 1})
                    ca_dur = ca_svc.get('duration', 30) if ca_svc else 30
                    ca_sh, ca_sm = map(int, ca_start.split(':'))
                    ca_em = ca_sm + ca_dur
                    ca_eh = ca_sh + ca_em // 60
                    ca_em = ca_em % 60
                    ca_end = f"{str(ca_eh).zfill(2)}:{str(ca_em).zfill(2)}"
                    if appointment.appointment_time < ca_end and new_end_time > ca_start:
                        has_conflict = True
                        logging.info(f"⚠️ Public conflict (no-staff-setting): {appointment.appointment_time}-{new_end_time} overlaps {ca_start}-{ca_end}")
                        break

                if has_conflict:
                    plan_doc = await db.organization_plans.find_one({"organization_id": organization_id})
                    if plan_doc:
                        await db.organization_plans.update_one(
                            {"organization_id": organization_id},
                            {"$inc": {"quota_usage": -1}}
                        )
                    raise HTTPException(status_code=400, detail="Bu saat dilimi doludur. Lütfen başka bir saat seçin.")

                assigned_staff_id = admin_username
                logging.info(f"ℹ️ customer_can_choose_staff and admin_provides_service are both disabled, but using admin: {admin_username}")
            else:
                # Admin bu hizmeti veremiyorsa, personel atama yapma
                logging.info(f"ℹ️ customer_can_choose_staff and admin_provides_service are both disabled, and admin cannot provide this service")
                assigned_staff_id = None
        else:
            # customer_can_choose_staff açıksa veya admin_provides_service açıksa, otomatik atama yap
            # Bu hizmeti verebilen personelleri bul
            qualified_staff_query = {
                "organization_id": organization_id,
                "permitted_service_ids": {"$all": resolved_service_ids}
            }
            
            # Admin hizmet vermiyorsa, admin'i listeden çıkar
            if not admin_provides_service:
                qualified_staff_query["role"] = {"$ne": "admin"}
            
            qualified_staff = await db.users.find(
                qualified_staff_query,
                {"_id": 0, "username": 1, "role": 1}
            ).to_list(1000)
            
            # Eğer başka personel yoksa, admin'i personel listesine ekle (admin_provides_service açıksa)
            # (Admin hizmet vermiyor ayarı açık olsa bile, başka personel yoksa admin'i kullanabiliriz)
            if not qualified_staff:
                # Admin'in bu hizmeti verebilip veremediğini kontrol et
                admin_user = await db.users.find_one(
                    {"organization_id": organization_id, "role": "admin"},
                    {"_id": 0, "username": 1, "role": 1, "permitted_service_ids": 1}
                )
                if admin_user and all(sid in (admin_user.get('permitted_service_ids') or []) for sid in resolved_service_ids):
                    qualified_staff = [{"username": admin_user['username'], "role": "admin"}]
                    logging.info(f"⚠️ Public: No staff found, but admin can provide service. Using admin: {admin_user['username']}")
            
            if not qualified_staff:
                # Kota artırıldı ama personel bulunamadı, geri al
                plan_doc = await db.organization_plans.find_one({"organization_id": organization_id})
                if plan_doc:
                    await db.organization_plans.update_one(
                        {"organization_id": organization_id},
                        {"$inc": {"quota_usage": -1}}
                    )
                raise HTTPException(
                    status_code=400,
                    detail="Bu hizmet için uygun personel bulunamadı"
                )
            
            # Boş personel bul (duration'a göre çakışma kontrolü ile)
            for staff in qualified_staff:
                # Bu personelin o tarihteki tüm randevularını çek
                # staff_member_id=None (atanmamış) randevuları da dahil et — tek kişilik org'larda çakışma kaçmasın
                existing_appointments = await db.appointments.find(
                    {
                        "organization_id": organization_id,
                        "$or": [
                            {"staff_member_id": staff['username']},
                            {"staff_member_id": None}
                        ],
                        "appointment_date": appointment.appointment_date,
                        "status": {"$nin": ["İptal", "İptal Edildi", "Ödeme Bekleniyor"]}
                    },
                    {"_id": 0, "appointment_time": 1, "service_id": 1, "service_duration": 1}
                ).to_list(100)
                
                # Çakışma kontrolü
                has_conflict = False
                for existing_appt in existing_appointments:
                    existing_start_time = existing_appt['appointment_time']
                    existing_service_id = existing_appt.get('service_id')
                    
                    # Mevcut randevunun hizmet süresini bul — çoklu hizmette saklı TOPLAM süre
                    stored_dur = existing_appt.get('service_duration')
                    if stored_dur:
                        existing_duration = int(stored_dur)
                    elif existing_service_id:
                        existing_service = await db.services.find_one({"id": existing_service_id}, {"_id": 0, "duration": 1})
                        existing_duration = existing_service.get('duration', 30) if existing_service else 30
                    else:
                        existing_duration = 30
                    
                    # Mevcut randevunun bitiş saatini hesapla
                    existing_start_hour, existing_start_minute = map(int, existing_start_time.split(':'))
                    existing_end_minute = existing_start_minute + existing_duration
                    existing_end_hour = existing_start_hour + (existing_end_minute // 60)
                    existing_end_minute = existing_end_minute % 60
                    existing_end_time = f"{str(existing_end_hour).zfill(2)}:{str(existing_end_minute).zfill(2)}"
                    
                    # Çakışma kontrolü
                    if (appointment.appointment_time < existing_end_time and new_end_time > existing_start_time):
                        has_conflict = True
                        logging.debug(f"   ⚠️ Public: Staff {staff['username']} has conflict: {appointment.appointment_time}-{new_end_time} overlaps with {existing_start_time}-{existing_end_time}")
                        break
                
                # Mola çakışması kontrolü
                has_break_conflict = await check_break_conflict(
                    db, staff['username'], appointment.appointment_date,
                    appointment.appointment_time, new_end_time, organization_id
                )
                
                if not has_conflict and not has_break_conflict:
                    # Bu personel boş ve mola yok!
                    assigned_staff_id = staff['username']
                    logging.info(f"✅ Public booking auto-assigned to {staff['username']} for {appointment.appointment_time}")
                    break
        
        # Eğer ayarlar kapalıysa (customer_can_choose_staff ve admin_provides_service kapalı), 
        # staff_member_id None olarak kaydedilebilir (atama yapılmaz)
        if not assigned_staff_id:
            if not customer_can_choose_staff and not admin_provides_service:
                # Ayarlar kapalıysa, randevuyu staff_member_id None olarak kaydet (atama yapma)
                logging.info(f"ℹ️ Public booking: Settings disabled, saving appointment without staff assignment")
                assigned_staff_id = None  # None olarak kalacak, randevu oluşturulacak
            else:
                # Ayarlar açıksa ama personel bulunamadı, hata ver
                plan_doc = await db.organization_plans.find_one({"organization_id": organization_id})
                if plan_doc:
                    await db.organization_plans.update_one(
                        {"organization_id": organization_id},
                        {"$inc": {"quota_usage": -1}}
                    )
                raise HTTPException(
                    status_code=400,
                    detail="Bu saat dilimi doludur. Lütfen başka bir saat seçin."
                )
    
    # Randevuyu oluştur
    appointment_data = appointment.model_dump()
    appointment_data['service_id'] = resolved_service_ids[0]  # ilk hizmet (geriye dönük uyum)
    appointment_data['services'] = snapshot_lines  # çoklu hizmet snapshot kalemleri
    appointment_data['service_name'] = service['name']
    appointment_data['service_price'] = service['price']
    appointment_data['service_duration'] = service.get('duration', 30)  # Toplam hizmet süresi

    # Solo işletme (yalnızca admin hizmet veriyor) → public randevu ATANMAMIŞ
    # (staff_member_id=None) kaydedilir. Admin adının randevu kartında "personel"
    # olarak görünmesini engeller; merchant paneli ve verify-code yolu ile tutarlı.
    # Yalnızca: müşteri açıkça personel SEÇMEDİYSE + atanan admin ise + bu hizmeti
    # veren admin-dışı GERÇEK personel YOKSA uygulanır (çok personelli org'da admin
    # meşru sağlayıcıysa dokunmaz). Çakışma kontrolü zaten admin+None ile yapıldı.
    if assigned_staff_id and not appointment.staff_member_id:
        _admin_u = await db.users.find_one(
            {"organization_id": organization_id, "role": "admin"},
            {"_id": 0, "username": 1},
        )
        if _admin_u and assigned_staff_id == _admin_u.get("username"):
            _real_staff = await db.users.find_one(
                {
                    "organization_id": organization_id,
                    "role": {"$ne": "admin"},
                    "permitted_service_ids": {"$all": resolved_service_ids},
                },
                {"_id": 0, "username": 1},
            )
            if not _real_staff:
                assigned_staff_id = None
                logging.info("ℹ️ Public booking: solo işletme (yalnız admin) → randevu atanmamış (None) kaydedildi")

    appointment_data['staff_member_id'] = assigned_staff_id
    appointment_data['source'] = 'public_booking'  # Public booking'den geldiğini işaretle
    
    # Seans paketi desteği: Hizmetin session_count > 1 ise otomatik olarak ilk seans bilgilerini ekle
    service_session_count = service.get('session_count')
    if service_session_count and service_session_count > 1:
        appointment_data['session_group_id'] = str(uuid.uuid4())
        appointment_data['session_number'] = 1
        appointment_data['session_total'] = service_session_count
        appointment_data['payment_status'] = 'paid'
    
    # Online/deposit ödeme varsa: randevuyu "Ödeme Bekleniyor" olarak oluştur
    payment_rule = service.get('payment_rule', '')
    if payment_rule in ('online', 'deposit'):
        appointment_data['payment_status'] = 'pending_payment'
        appointment_data['status'] = 'Ödeme Bekleniyor'

    # Randevu durumunu kontrol et (bitiş saatine göre) — "Ödeme Bekleniyor" override edilmez
    if appointment_data.get('status') != 'Ödeme Bekleniyor':
        try:
            turkey_tz = ZoneInfo("Europe/Istanbul")
            now = datetime.now(turkey_tz)
            dt_str = f"{appointment.appointment_date} {appointment.appointment_time}"
            naive_dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
            appointment_dt = naive_dt.replace(tzinfo=turkey_tz)
            service_duration_minutes = service.get('duration', 30)
            completion_threshold = appointment_dt + timedelta(minutes=service_duration_minutes)
            if now >= completion_threshold:
                appointment_data['status'] = 'Tamamlandı'
                appointment_data['completed_at'] = datetime.now(timezone.utc).isoformat()
            else:
                appointment_data['status'] = 'Bekliyor'
        except (ValueError, TypeError) as e:
            logging.warning(f"Public randevu durumu ayarlanırken tarih hatası: {e}")
            appointment_data['status'] = 'Bekliyor'
    
    appointment_obj = Appointment(**appointment_data, organization_id=organization_id)
    doc = appointment_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.appointments.insert_one(doc)
    
    # 🔒 Redis lock'u sil (başarılı randevu — diğer müşteriler beklemek zorunda kalmasın)
    if redis_client and lock_key:
        try:
            await redis_client.delete(lock_key)
        except Exception:
            pass
    
    # 🧹 Cache temizliği (Dashboard + Müşteri listesi güncellensin)
    try:
        mock_user = type('obj', (object,), {'organization_id': organization_id})
        await invalidate_cache(request, "dashboard_stats", mock_user)
        await invalidate_cache(request, "customers_list", mock_user)
        await invalidate_availability_cache(request, organization_id)
        logging.info(f"🧹 Public randevu sonrası admin cache temizlendi: {organization_id}")
    except Exception as e:
        logging.error(f"❌ Public süpürge hatası: {e}")
    # Müşteriyi customers collection'ına ekle (eğer yoksa)
    # Format-dedup: aynı kişi farklı telefon formatıyla girilirse yeni müşteri açılmasın.
    resolved_customer_phone = appointment.phone
    try:
        # Aynı telefonun tüm format varyantlarıyla eşleşen müşterileri bul
        _phone_variants = _phone_match_variants(appointment.phone)
        customers_with_phone = await db.customers.find(
            {
                "organization_id": organization_id,
                "phone": {"$in": _phone_variants}
            },
            {"_id": 0, "name": 1, "phone": 1}
        ).to_list(100)
        
        # İsim-soyisim kontrolü (büyük-küçük harf duyarsız)
        customer_name_normalized = appointment.customer_name.strip().lower()
        existing_customer = None
        for customer in customers_with_phone:
            if customer.get("name", "").strip().lower() == customer_name_normalized:
                existing_customer = customer
                break
        if existing_customer and existing_customer.get("phone"):
            resolved_customer_phone = existing_customer.get("phone")
        
        if not existing_customer:
            # Müşteri yoksa ekle
            customer_doc = {
                "id": str(uuid.uuid4()),
                "organization_id": organization_id,
                "name": appointment.customer_name.strip(),
                "phone": appointment.phone,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "notes": ""
            }
            await db.customers.insert_one(customer_doc)
            logging.info(f"Customer auto-added from public booking: {appointment.customer_name} ({appointment.phone}) for org {organization_id}")
            
            # Cache invalidation SONRASI müşteri eklendi — tekrar temizle (race condition fix)
            try:
                mock_user = type('obj', (object,), {'organization_id': organization_id})
                await invalidate_cache(request, "customers_list", mock_user)
            except Exception:
                pass
            
            # WebSocket event gönder (müşteriler listesini güncellemek için)
            try:
                await emit_to_organization(
                    organization_id,
                    'customer_added',
                    {'customer': customer_doc}
                )
            except Exception as emit_error:
                logging.warning(f"Failed to emit customer_added event: {emit_error}")
    except Exception as e:
        logging.warning(f"Error adding customer to collection: {e}")

    # Denormalize sayaç: public randevu → +1 total (+1 completed nadiren; genellikle Bekliyor).
    # Ödeme bekleyen (pending_payment) randevu ise atla — sayaç ödeme onaylanınca
    # `_send_post_payment_notifications` içinde artırılıyor.
    try:
        _is_pending_payment_local = (appointment_data.get('payment_status') == 'pending_payment')
        if not _is_pending_payment_local:
            _apt_status = appointment_data.get('status', 'Bekliyor')
            await _apply_customer_delta(
                db,
                organization_id=organization_id,
                phone=resolved_customer_phone,
                name=appointment.customer_name,
                delta_total=1,
                delta_completed=1 if _apt_status == "Tamamlandı" else 0,
                appointment_date=appointment.appointment_date,
                appointment_time=appointment.appointment_time,
            )
    except Exception as _delta_err:
        logging.warning(f"customer delta (public/appointments) skipped: {_delta_err}")

    # WhatsApp ve Push Notification'ları background'da gönder (kullanıcıyı bekletmemek için)
    settings_data = await db.settings.find_one({"organization_id": organization_id})
    
    async def send_notifications_background():
        """Tüm bildirimleri background'da gönder"""
        try:
            if settings_data:
                company_name = settings_data.get("company_name", "İşletmeniz")
                support_phone = settings_data.get("support_phone", "Destek Hattı")
                business = settings_data or {}
                settings = business.get('settings') if isinstance(business.get('settings'), dict) else business
                location = (settings or {}).get('location') or {}
                coordinates = (location or {}).get('coordinates', {})
                lat = coordinates.get('lat')
                lng = coordinates.get('lng')
                
                # WhatsApp onay mesajı gönder (Content API)
                try:
                    _wa_msg_id = await asyncio.to_thread(
                        send_whatsapp_template,
                        appointment.phone,
                        "CONFIRMATION",
                        appointment.customer_name,
                        company_name,
                        appointment.appointment_date,
                        appointment.appointment_time,
                        service['name'],
                        support_phone,
                        business_lat=lat,
                        business_lng=lng,
                        business_address=location.get('address'),
                    )
                    logging.info(f"✓ WhatsApp confirmation sent to {appointment.phone}")
                    try:
                        import time as _time
                        _phone = str(appointment.phone).lstrip('+')
                        await db.whatsapp_message_logs.update_one(
                            {"message_id": _wa_msg_id},
                            {"$set": {
                                "message_id": _wa_msg_id,
                                "recipient": _phone,
                                "status": "sent",
                                "timestamp": int(_time.time()),
                                "recorded_at": datetime.utcnow().isoformat(),
                                "source": "public_booking"
                            }},
                            upsert=True
                        )
                    except Exception:
                        pass
                except Exception as wa_error:
                    logging.error(f"WhatsApp error: {wa_error}")
                    try:
                        import time as _time
                        _phone = str(appointment.phone).lstrip('+')
                        await db.whatsapp_message_logs.update_one(
                            {"message_id": f"fail_{_phone}_{int(_time.time())}"},
                            {"$set": {
                                "recipient": _phone,
                                "status": "failed",
                                "timestamp": int(_time.time()),
                                "recorded_at": datetime.utcnow().isoformat(),
                                "errors": [{"code": "SEND_ERROR", "title": str(wa_error)[:300]}],
                                "source": "public_booking"
                            }},
                            upsert=True
                        )
                    except Exception:
                        pass
            
            # Push notification gönder (admin ve ilgili personele)
            try:
                service_info = service.get('name', 'Randevu')
                _apt_id_push = appointment_for_emit.get('id')
                notification_data = {
                    "type": "new_appointment",
                    "appointment_id": _apt_id_push,
                    "source": "public_booking",
                    # Web push (service worker) derin bağlantısı — tıklayınca randevu detayı açılır
                    "url": f"/?randevu={_apt_id_push}" if _apt_id_push else "/",
                }
                # Bildirim metni: Müşteri / Hizmet / Tarih (7 Ağustos Cuma 09:45) alt alta
                _date_tr = format_date_tr(appointment.appointment_date)
                notification_body = (
                    f"Müşteri: {appointment.customer_name}\n"
                    f"Hizmet: {service_info}\n"
                    f"Tarih: {_date_tr} {appointment.appointment_time}"
                )
                
                # Admin'lere bildirim gönder
                admins = await db.users.find({"organization_id": organization_id, "role": "admin"}).to_list(100)
                for admin in admins:
                    await send_push_notification(
                        db=db,
                        organization_id=organization_id,
                        title="🔔 Yeni Online Randevu!",
                        body=notification_body,
                        data=notification_data,
                        user_id=admin['username']
                    )
                
                # Atanan personele bildirim gönder (admin değilse)
                if assigned_staff_id:
                    staff = await db.users.find_one({"username": assigned_staff_id, "organization_id": organization_id})
                    if staff and staff.get('role') != 'admin':
                        await send_push_notification(
                            db=db,
                            organization_id=organization_id,
                            title="🔔 Yeni Randevu Atandı!",
                            body=notification_body,
                            data=notification_data,
                            user_id=assigned_staff_id
                        )
                logging.info(f"✓ Push notifications sent for appointment {appointment_for_emit.get('id')}")
            except Exception as push_error:
                logging.error(f"Push notification error: {push_error}")
        except Exception as e:
            logging.error(f"Background notification error: {e}")
    
    is_pending_payment = appointment_data.get('payment_status') == 'pending_payment'

    if not is_pending_payment:
        appointment_for_emit = appointment_obj.model_dump()
        appointment_for_emit['created_at'] = appointment_for_emit['created_at'].isoformat()
        logger.info(f"About to emit appointment_created for org: {organization_id} (public endpoint)")
        try:
            await emit_to_organization(
                organization_id,
                'appointment_created',
                {'appointment': appointment_for_emit}
            )
            logger.info(f"Successfully emitted appointment_created for org: {organization_id} (public endpoint)")
        except Exception as emit_error:
            logger.error(f"Failed to emit appointment_created (public endpoint): {emit_error}", exc_info=True)
        asyncio.create_task(send_notifications_background())
    else:
        logging.info(f"⏳ Ödeme bekleniyor — bildirimler ödeme sonrası gönderilecek: {doc.get('id', '')}")
    
    return {"message": "Randevu başarıyla oluşturuldu", "appointment": appointment_obj}


@api_router.get("/public/verify-status")
async def check_verification_status(code: str, request: Request):
    """WhatsApp doğrulama durumunu sorgula (frontend 3sn polling)"""
    redis_client = getattr(request.app.state, 'redis_client', None)
    if not redis_client:
        raise HTTPException(status_code=503, detail="Servis geçici olarak kullanılamıyor.")

    # Webhook tarafından tamamlandı mı? (5 dk TTL flag)
    is_verified = await redis_client.get(f"plann:verified:{code}")
    if is_verified:
        return {"status": "verified", "message": "Doğrulama tamamlandı, randevunuz oluşturuldu."}

    # Pending kayıt hala var mı?
    exists = await redis_client.exists(f"plann:pending:{code}")
    if exists:
        return {"status": "pending", "message": "Doğrulama bekleniyor."}

    # Kod yok ve verified da değil → süresi dolmuş
    return {"status": "expired", "message": "Doğrulama kodunun süresi doldu. Lütfen tekrar deneyin."}


@api_router.post("/public/verify-code")
async def verify_code_and_create_appointment(request: Request):
    """Müşteri doğrulama kodunu girer → randevu oluşturulur"""
    redis_client = getattr(request.app.state, 'redis_client', None)
    if not redis_client:
        raise HTTPException(status_code=503, detail="Servis geçici olarak kullanılamıyor.")

    body = await request.json()
    code = (body.get("code") or "").strip().upper()
    entered_phone = body.get("phone", "")

    if not code:
        raise HTTPException(status_code=400, detail="Doğrulama kodu gerekli.")

    # Redis'ten pending kodu al
    pending_key = f"plann:pending:{code}"
    raw = await redis_client.hgetall(pending_key)
    if not raw:
        raise HTTPException(status_code=400, detail="Doğrulama kodu bulunamadı veya süresi doldu.")

    data = {
        (k.decode() if isinstance(k, bytes) else k): (v.decode() if isinstance(v, bytes) else v)
        for k, v in raw.items()
    }
    stored_phone = data.get("phone", "")
    org_id = data.get("org_id", "")

    # Telefon numarası eşleşmesi (son 9 hane)
    def _norm(p):
        return re.sub(r"\D", "", p).lstrip("0")

    if entered_phone and _norm(entered_phone)[-9:] != _norm(stored_phone)[-9:]:
        raise HTTPException(status_code=400, detail="Telefon numarası eşleşmiyor.")

    # Randevu oluştur
    db = await get_db_from_request(request)
    try:
        apt_dict = json.loads(data.get("appointment_data", "{}"))
    except Exception:
        raise HTTPException(status_code=500, detail="Randevu verisi işlenirken hata oluştu.")

    quota_ok, quota_error = await check_quota_and_increment(db, org_id)
    if not quota_ok:
        raise HTTPException(status_code=403, detail=quota_error)

    # Çoklu hizmet çözümleme (snapshot + toplam süre/fiyat)
    try:
        _apt_in = AppointmentCreate(**apt_dict)
        _settings_flag = await db.settings.find_one({"organization_id": org_id}, {"_id": 0, "multi_service_enabled": 1})
        _multi_enabled = bool((_settings_flag or {}).get("multi_service_enabled", False))
        snapshot_lines, total_duration, total_price, service, resolved_service_ids = await resolve_services(
            db, org_id, _apt_in, multi_enabled=_multi_enabled
        )
    except HTTPException:
        await db.organization_plans.update_one({"organization_id": org_id}, {"$inc": {"quota_usage": -1}})
        raise

    payment_rule = service.get("payment_rule", "on_site")
    apt_status = "Bekliyor"
    apt_payment_status = None
    if payment_rule in ("online", "deposit"):
        apt_status = "Ödeme Bekleniyor"
        apt_payment_status = "pending_payment"

    apt_obj = Appointment(**{
        **apt_dict,
        "organization_id": org_id,
        "service_id": resolved_service_ids[0],
        "services": snapshot_lines,
        "service_name": service.get("name", ""),
        "service_price": service.get("price", 0),
        "service_duration": service.get("duration", 30),
        "source": "public_booking_code_verified",
        "status": apt_status,
        "payment_status": apt_payment_status,
    })
    doc = apt_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()

    # ── Slot çakışma kontrolü (TOCTOU koruması) ──────────────────────────
    # /public/appointments müsaitliği doğrulama kodundan ÖNCE bakar; kod
    # girilene kadar slot dolabilir (aynı müşteri 2. kez ya da başka biri).
    # Eskiden bu yol kontrolsüz insert ediyordu → personelsiz org'da aynı
    # saate ÇİFT kayıt oluşuyordu. _check_slot_conflict tek kaynak semantiği:
    # staff_member_id=None → org-geneli overlap; atanmış personel → o personel
    # + boş randevular. Ödeme bekleyen (online/deposit) randevular slot rezerve
    # etmez → kontrolü atla (rezervasyon ödeme akışında yapılır).
    if apt_payment_status != "pending_payment":
        _conflict_dur = int(total_duration or service.get("duration", 30) or 30)
        if await _check_slot_conflict(
            db, org_id, apt_obj.staff_member_id,
            apt_obj.appointment_date, apt_obj.appointment_time, _conflict_dur,
        ):
            await db.organization_plans.update_one(
                {"organization_id": org_id}, {"$inc": {"quota_usage": -1}}
            )
            try:
                await redis_client.delete(pending_key)
                await redis_client.delete(f"plann:pending:phone:{stored_phone}:{org_id}")
            except Exception:
                pass
            raise HTTPException(
                status_code=409,
                detail="Seçtiğiniz saat az önce doldu. Lütfen sayfayı yenileyip başka bir saat seçin.",
            )

    await db.appointments.insert_one(doc)

    # Müşteriyi kaydet
    await db.customers.update_one(
        {"organization_id": org_id, "phone": stored_phone},
        {
            "$set": {"name": apt_dict.get("customer_name", ""), "phone": stored_phone, "organization_id": org_id, "updated_at": datetime.now(timezone.utc).isoformat()},
            "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": datetime.now(timezone.utc).isoformat(), "notes": ""},
        },
        upsert=True,
    )

    # Denormalize sayaç — pending_payment değilse +1 total (+1 completed nadiren).
    try:
        if apt_payment_status != "pending_payment":
            await _apply_customer_delta(
                db,
                organization_id=org_id,
                phone=stored_phone,
                name=apt_dict.get("customer_name"),
                delta_total=1,
                delta_completed=1 if apt_status == "Tamamlandı" else 0,
                appointment_date=apt_dict.get("appointment_date"),
                appointment_time=apt_dict.get("appointment_time"),
            )
    except Exception as _delta_err:
        logging.warning(f"customer delta (verify-code) skipped: {_delta_err}")

    # Redis temizle
    phone_map_key = f"plann:pending:phone:{stored_phone}:{org_id}"
    await redis_client.delete(pending_key)
    await redis_client.delete(phone_map_key)
    await redis_client.setex(f"plann:verified:{code}", 300, "1")

    # WebSocket event
    try:
        apt_clean = {k: v for k, v in doc.items() if k != "_id"}
        await emit_to_organization(org_id, "appointment_created", {"appointment": apt_clean})
    except Exception:
        pass

    # Cache temizle
    try:
        mock_user = type('obj', (object,), {'organization_id': org_id})
        await invalidate_cache(request, "dashboard_stats", mock_user)
        await invalidate_cache(request, "customers_list", mock_user)
    except Exception:
        pass

    # ── Push notifications → admin + personel ────────────────────────────
    # ÖNEMLİ: Telefon doğrulaması AÇIK işletmelerde public randevular bu yoldan
    # (verify-code) oluşur. Eskiden burada push GÖNDERİLMİYORDU → işletme sahibi
    # public randevu bildirimi alamıyordu. Diğer public yollarla aynı metin/data.
    try:
        _date_tr_vc = format_date_tr(apt_dict.get("appointment_date", ""))
        vc_body = (
            f"Müşteri: {apt_dict.get('customer_name', '')}\n"
            f"Hizmet: {service.get('name', '')}\n"
            f"Tarih: {_date_tr_vc} {apt_dict.get('appointment_time', '')}"
        )
        _apt_id_vc = doc.get("id")
        vc_data = {
            "type": "new_appointment",
            "appointment_id": _apt_id_vc,
            "source": "public_booking_code_verified",
            "url": f"/?randevu={_apt_id_vc}" if _apt_id_vc else "/",
        }
        _admins_vc = await db.users.find({"organization_id": org_id, "role": "admin"}).to_list(100)
        for _admin_vc in _admins_vc:
            await send_push_notification(
                db=db, organization_id=org_id,
                title="🔔 Yeni Online Randevu!",
                body=vc_body, data=vc_data, user_id=_admin_vc["username"],
            )
        _assigned_vc = apt_dict.get("staff_member_id")
        if _assigned_vc:
            _staff_vc = await db.users.find_one({"username": _assigned_vc, "organization_id": org_id})
            if _staff_vc and _staff_vc.get("role") != "admin":
                await send_push_notification(
                    db=db, organization_id=org_id,
                    title="🔔 Yeni Randevu Atandı!",
                    body=vc_body, data=vc_data, user_id=_assigned_vc,
                )
        logging.info(f"✓ Push (verify-code) sent for appointment {_apt_id_vc}")
    except Exception as _push_vc_err:
        logging.error(f"⚠️ Push (verify-code) failed: {_push_vc_err}", exc_info=True)

    logging.info(f"✅ Doğrulama kodu ile randevu oluşturuldu: {code} → {stored_phone}")

    # WhatsApp randevu onay mesajı gönder (background)
    async def _send_wa_confirmation():
        try:
            settings_data = await db.settings.find_one({"organization_id": org_id})
            if not settings_data:
                return
            company_name = settings_data.get("company_name", "İşletmeniz")
            support_phone = settings_data.get("support_phone", "Destek Hattı")
            location = (settings_data.get('location') or {})
            coordinates = location.get('coordinates', {})
            lat = coordinates.get('lat')
            lng = coordinates.get('lng')

            _wa_msg_id = await asyncio.to_thread(
                send_whatsapp_template,
                stored_phone,
                "CONFIRMATION",
                apt_dict.get("customer_name", ""),
                company_name,
                apt_dict.get("appointment_date", ""),
                apt_dict.get("appointment_time", ""),
                service.get('name', ''),
                support_phone,
                business_lat=lat,
                business_lng=lng,
                business_address=location.get('address'),
            )
            logging.info(f"✓ WhatsApp confirmation sent (verified): {stored_phone}")
            try:
                import time as _time
                _phone = str(stored_phone).lstrip('+')
                await db.whatsapp_message_logs.update_one(
                    {"message_id": _wa_msg_id},
                    {"$set": {
                        "message_id": _wa_msg_id,
                        "recipient": _phone,
                        "status": "sent",
                        "timestamp": int(_time.time()),
                        "recorded_at": datetime.utcnow().isoformat(),
                        "source": "public_booking_verified"
                    }},
                    upsert=True
                )
            except Exception:
                pass
        except Exception as wa_err:
            logging.error(f"WhatsApp confirmation error (verified): {wa_err}")

    if payment_rule not in ("online", "deposit"):
        asyncio.create_task(_send_wa_confirmation())

    return {
        "status": "verified",
        "message": "Doğrulama tamamlandı, randevunuz oluşturuldu.",
        "appointment_id": doc.get("id"),
        "service_payment_rule": payment_rule,
    }


# === SUPER ADMIN ENDPOINT'LERİ ===
@api_router.get("/superadmin/stats")
async def get_superadmin_stats(request: Request, current_user: UserInDB = Depends(get_superadmin_user), db = Depends(get_db)):
    """Platform özeti - Sadece superadmin"""
    try:
        now = datetime.now(timezone.utc)
        first_day_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # Bu ayın son gününü hesapla
        if now.month == 12:
            last_day_of_month = now.replace(year=now.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last_day_of_month = now.replace(month=now.month + 1, day=1) - timedelta(days=1)
        last_day_of_month = last_day_of_month.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        first_day_str = first_day_of_month.strftime("%Y-%m-%d")
        last_day_str = last_day_of_month.strftime("%Y-%m-%d")
        
        # 1. Toplam İşletme Sayısı (Settings koleksiyonundan unique organization_id sayısı)
        total_organizations = await db.settings.count_documents({})
        
        # 2. Bu Ayki Toplam Abonelik Geliri (organization_plans'den aktif planların toplamı)
        # Trial hariç, aktif planların aylık fiyatlarını topla
        active_plans = await db.organization_plans.find({}).to_list(10000)
        total_monthly_revenue = 0.0
        
        for plan in active_plans:
            plan_id = plan.get('plan_id', 'tier_trial')
            # Trial paketleri gelir sayılmaz
            if plan_id != 'tier_trial':
                plan_info = next((p for p in PLANS if p['id'] == plan_id), None)
                if plan_info:
                    # Trial bitiş tarihi kontrolü - eğer trial bitmişse ve plan aktifse
                    trial_end = plan.get('trial_end_date')
                    if trial_end:
                        if isinstance(trial_end, str):
                            trial_end = datetime.fromisoformat(trial_end.replace('Z', '+00:00'))
                        # Trial bitmişse plan fiyatını ekle
                        if trial_end < now:
                            total_monthly_revenue += plan_info.get('price_monthly', 0)
                    else:
                        # Trial yoksa direkt fiyatı ekle
                        total_monthly_revenue += plan_info.get('price_monthly', 0)
        
        # 3. Bu Ayki Toplam Randevu Sayısı (Tüm organizasyonların appointments'ları)
        total_appointments_this_month = await db.appointments.count_documents({
            "appointment_date": {"$gte": first_day_str, "$lte": last_day_str}
        })
        
        # 4. Toplam Aktif Kullanıcı (Customers + Personnel/Users)
        total_customers = await db.customers.count_documents({})
        total_personnel = await db.users.count_documents({"role": "staff"})
        total_active_users = total_customers + total_personnel
        
        return {
            "toplam_isletme": total_organizations,
            "toplam_gelir_bu_ay": round(total_monthly_revenue, 2),
            "toplam_randevu_bu_ay": total_appointments_this_month,
            "toplam_aktif_kullanici": total_active_users
        }
    except Exception as e:
        logging.error(f"Error in get_superadmin_stats: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"İstatistikler alınırken hata oluştu: {str(e)}")

@api_router.get("/superadmin/contact-requests")
async def get_superadmin_contact_requests(request: Request, current_user: UserInDB = Depends(get_superadmin_user), db = Depends(get_db)):
    """SuperAdmin için iletişim taleplerini getir"""
    try:
        contacts = await db.contact_requests.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
        return {"contacts": contacts}
    except Exception as e:
        logging.error(f"Error in get_superadmin_contact_requests: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Veriler yüklenirken hata oluştu")

@api_router.put("/superadmin/contact-requests/{contact_id}/status")
async def update_contact_status(
    request: Request,
    contact_id: str,
    status_update: ContactStatusUpdate,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """Contact request durumunu güncelle - Sadece superadmin"""
    try:
        # Contact request'i bul
        contact = await db.contact_requests.find_one({"id": contact_id}, {"_id": 0})
        if not contact:
            raise HTTPException(status_code=404, detail="İletişim talebi bulunamadı")
        
        # Durumu güncelle
        await db.contact_requests.update_one(
            {"id": contact_id},
            {"$set": {"status": status_update.status}}
        )
        
        logging.info(f"✅ [SUPERADMIN] Contact request {contact_id} durumu güncellendi: {status_update.status}")
        
        # Güncellenmiş contact request'i döndür
        updated_contact = await db.contact_requests.find_one({"id": contact_id}, {"_id": 0})
        return {"success": True, "contact": updated_contact}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"❌ [SUPERADMIN] Contact request durumu güncellenirken hata: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Durum güncellenirken hata oluştu")

@api_router.delete("/superadmin/contact-requests/{contact_id}")
async def delete_contact_request(
    request: Request,
    contact_id: str,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """Contact request'i sil - Sadece superadmin"""
    try:
        # Contact request'i bul
        contact = await db.contact_requests.find_one({"id": contact_id}, {"_id": 0})
        if not contact:
            raise HTTPException(status_code=404, detail="İletişim talebi bulunamadı")
        
        # Contact request'i sil
        await db.contact_requests.delete_one({"id": contact_id})
        
        logging.info(f"✅ [SUPERADMIN] Contact request silindi: {contact_id}")
        
        return {"success": True, "message": "İletişim talebi silindi"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"❌ [SUPERADMIN] Contact request silinirken hata: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="İletişim talebi silinirken hata oluştu")

@api_router.delete("/superadmin/contact-requests/bulk/delete-resolved")
async def delete_resolved_contacts(
    request: Request,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """Çözülen (resolved) contact request'leri toplu sil - Sadece superadmin"""
    try:
        # Resolved status'lu contact request'leri bul ve sil
        result = await db.contact_requests.delete_many({"status": "resolved"})
        
        deleted_count = result.deleted_count
        logging.info(f"✅ [SUPERADMIN] {deleted_count} adet çözülen iletişim talebi silindi")
        
        return {"success": True, "deleted_count": deleted_count, "message": f"{deleted_count} adet çözülen iletişim talebi silindi"}
    except Exception as e:
        logging.error(f"❌ [SUPERADMIN] Çözülen iletişim talepleri silinirken hata: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Çözülen iletişim talepleri silinirken hata oluştu")

# Ücretli planlarda Stripe yenilemesi tam kayıt saatinde tetiklenir ve invoice.paid
# webhook'u next_billing_date'i biraz sonra günceller. Bu boşlukta (ve Stripe dunning
# sürecinde) işletmeyi "Süresi Doldu" göstermemek için tolerans tanınır.
SUBSCRIPTION_GRACE_DAYS = 3


def _parse_plan_dt(val):
    """Plan tarihini timezone-aware datetime'a çevirir (str/datetime/None destekli)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val if val.tzinfo else val.replace(tzinfo=timezone.utc)
    if isinstance(val, str):
        try:
            return datetime.fromisoformat(val.replace("Z", "+00:00"))
        except ValueError:
            return None
    return None


def _ceil_days(delta: timedelta) -> int:
    """timedelta'yı gün olarak YUKARI yuvarlar.

    Python'da timedelta.days tabana yuvarladığı için bitiş anı (örn. 22:50)
    1 saniye geçtiğinde bile -1 döner. ceil ile bitişe saatler kala 0 yerine
    kalan tam gün gösterilir ve bitiş anında erkenden negatife düşmez.
    """
    return math.ceil(delta.total_seconds() / 86400)


def _compute_subscription_status(plan_doc: dict, now: datetime):
    """Ortak abonelik durumu hesabı → (durumu, durum_key, days_left).

    Ücretli planlarda bitiş tarihi geçse bile SUBSCRIPTION_GRACE_DAYS kadar
    tolerans tanınır; bu sürede "Yenileniyor" gösterilir (Stripe otomatik
    yeniler, webhook gecikebilir). Tolerans aşılırsa "Süresi Doldu".
    """
    plan_id = plan_doc.get("plan_id", "tier_trial")

    if plan_id == "tier_trial":
        trial_end = _parse_plan_dt(plan_doc.get("trial_end_date"))
        if not trial_end:
            return "Trial", "trial", None
        days_left = _ceil_days(trial_end - now)
        if now > trial_end:
            return "Deneme Bitti", "expired", days_left
        return f"{days_left} Gün Kaldı", "trial", days_left

    # Ücretli plan — bitiş tarihini sırayla dene
    deadline = None
    for key in ("next_billing_date", "next_payment_date", "quota_reset_date"):
        deadline = _parse_plan_dt(plan_doc.get(key))
        if deadline:
            break
    if not deadline:
        return "Aktif", "active", None

    days_left = _ceil_days(deadline - now)
    if now <= deadline:
        return "Aktif", "active", days_left

    # Bitiş geçti — grace period içindeyse henüz expired sayma.
    # Yenileme penceresinde days_left'i 0'da tut (panelde "-1 gün" karışıklığını önler).
    if (now - deadline) <= timedelta(days=SUBSCRIPTION_GRACE_DAYS):
        return "Yenileniyor", "active", max(days_left, 0)
    return "Süresi Doldu", "expired", days_left


@api_router.get("/superadmin/organizations")
async def get_superadmin_organizations(request: Request, current_user: UserInDB = Depends(get_superadmin_user), db = Depends(get_db)):
    """Detaylı işletme listesi - Sadece superadmin"""
    try:
        now = datetime.now(timezone.utc)
        first_day_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        if now.month == 12:
            last_day_of_month = now.replace(year=now.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last_day_of_month = now.replace(month=now.month + 1, day=1) - timedelta(days=1)
        last_day_of_month = last_day_of_month.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        first_day_str = first_day_of_month.strftime("%Y-%m-%d")
        last_day_str = last_day_of_month.strftime("%Y-%m-%d")
        
        all_settings = await db.settings.find({}).to_list(10000)
        org_ids = [s["organization_id"] for s in all_settings if s.get("organization_id")]

        # Batch: admin users
        admin_users_list = await db.users.find(
            {"organization_id": {"$in": org_ids}, "role": "admin"},
            {"_id": 0, "organization_id": 1, "full_name": 1, "username": 1}
        ).to_list(len(org_ids))
        admin_map = {a["organization_id"]: a for a in admin_users_list}

        # Batch: plans
        plans_list = await db.organization_plans.find(
            {"organization_id": {"$in": org_ids}}
        ).to_list(len(org_ids))
        plan_map = {p["organization_id"]: p for p in plans_list}

        # Batch: monthly appointment counts
        appt_pipeline = [
            {"$match": {"organization_id": {"$in": org_ids}, "created_at": {"$gte": first_day_str, "$lte": last_day_str}}},
            {"$group": {"_id": "$organization_id", "count": {"$sum": 1}}}
        ]
        appt_counts = await db.appointments.aggregate(appt_pipeline).to_list(len(org_ids))
        appt_map = {a["_id"]: a["count"] for a in appt_counts}

        # Batch: all-time appointment counts
        appt_total_pipeline = [
            {"$match": {"organization_id": {"$in": org_ids}}},
            {"$group": {"_id": "$organization_id", "count": {"$sum": 1}}}
        ]
        appt_total_counts = await db.appointments.aggregate(appt_total_pipeline).to_list(len(org_ids))
        appt_total_map = {a["_id"]: a["count"] for a in appt_total_counts}

        # Batch: Stripe ödemeleri (payment_logs — checkout/webhook ile dolan kayıtlar)
        pay_pipeline = [
            {"$match": {"organization_id": {"$in": org_ids}, "status": {"$in": ["active", "completed"]}}},
            {"$group": {"_id": "$organization_id", "total": {"$sum": "$amount"}}},
        ]
        pay_totals = await db.payment_logs.aggregate(pay_pipeline).to_list(len(org_ids))
        pay_map = {p["_id"]: round(p.get("total") or 0, 2) for p in pay_totals}

        # Batch: customer counts
        cust_pipeline = [
            {"$match": {"organization_id": {"$in": org_ids}}},
            {"$group": {"_id": "$organization_id", "count": {"$sum": 1}}}
        ]
        cust_counts = await db.customers.aggregate(cust_pipeline).to_list(len(org_ids))
        cust_map = {c["_id"]: c["count"] for c in cust_counts}

        # Batch: staff counts
        staff_pipeline = [
            {"$match": {"organization_id": {"$in": org_ids}, "role": "staff"}},
            {"$group": {"_id": "$organization_id", "count": {"$sum": 1}}}
        ]
        staff_counts = await db.users.aggregate(staff_pipeline).to_list(len(org_ids))
        staff_map = {s["_id"]: s["count"] for s in staff_counts}

        organizations_list = []
        
        for setting in all_settings:
            org_id = setting.get('organization_id')
            if not org_id:
                continue
            
            isletme_adi = setting.get('company_name', 'İsimsiz İşletme')
            telefon_numarasi = setting.get('support_phone', 'Telefon Yok')
            is_internal = setting.get('is_internal', False)
            
            admin_user = admin_map.get(org_id)
            admin_full_name = admin_user.get('full_name', '-') if admin_user else '-'
            admin_email = admin_user.get('username', '-') if admin_user else '-'
            
            plan_doc = plan_map.get(org_id)
            billing_cycle = "monthly"
            days_left = None
            toplam_odeme = pay_map.get(org_id, 0)
            plan_id = "tier_trial"
            abonelik_durumu_key = "none"

            if not plan_doc:
                abonelik_paketi = "Trial"
                plan_id = "tier_trial"
                abonelik_durumu = "Kayıt Yok"
            else:
                plan_id = plan_doc.get("plan_id", "tier_trial")
                plan_info = next((p for p in PLANS if p["id"] == plan_id), None)
                abonelik_paketi = plan_info.get("name", "Trial") if plan_info else "Trial"
                billing_cycle = plan_doc.get("billing_cycle", "monthly")

                abonelik_durumu, abonelik_durumu_key, days_left = _compute_subscription_status(plan_doc, now)

            organizations_list.append({
                "organization_id": org_id,
                "isletme_adi": isletme_adi,
                "telefon_numarasi": telefon_numarasi,
                "admin_full_name": admin_full_name,
                "admin_email": admin_email,
                "plan_id": plan_id,
                "abonelik_paketi": abonelik_paketi,
                "billing_cycle": billing_cycle,
                "abonelik_durumu": abonelik_durumu,
                "abonelik_durumu_key": abonelik_durumu_key,
                "days_left": days_left,
                "toplam_odeme": toplam_odeme,
                "bu_ayki_randevu_sayisi": appt_map.get(org_id, 0),
                "toplam_randevu_sayisi": appt_total_map.get(org_id, 0),
                "toplam_musteri_sayisi": cust_map.get(org_id, 0),
                "toplam_personel_sayisi": staff_map.get(org_id, 0),
                "is_internal": is_internal,
            })
        
        return {"organizations": organizations_list}
    except Exception as e:
        logging.error(f"Error in get_superadmin_organizations: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"İşletme listesi alınırken hata oluştu: {str(e)}")


class SuperAdminBroadcastEmail(BaseModel):
    subject: str
    message: str  # düz metin veya HTML — newline'lar <br/> olarak çevrilir
    organization_ids: Optional[List[str]] = None  # None/boş = tüm işletmeler
    send_to_all: bool = False


def _wrap_broadcast_email(message: str, lang: str = "tr") -> str:
    """SuperAdmin duyuru mailini ortak PLANN marka şablonuna (kayıt e-postası
    ile aynı) sarar."""
    import html as _html
    # Basit HTML algılama: tag içeriyorsa olduğu gibi kullan, yoksa escape + <br/>
    if re.search(r"<[a-zA-Z][^>]*>", message):
        body_html = message
    else:
        body_html = _html.escape(message).replace("\n", "<br/>")
    return _wrap_brand_email(body_html, lang=lang)


@api_router.post("/superadmin/broadcast-email")
async def superadmin_broadcast_email(
    payload: SuperAdminBroadcastEmail,
    request: Request,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """Tüm işletmelere veya seçili işletmelere e-posta gönder - Sadece superadmin"""
    if not payload.subject.strip() or not payload.message.strip():
        raise HTTPException(status_code=400, detail="Konu ve mesaj zorunludur.")

    # Hedef organizasyonları belirle
    query = {"role": "admin"}
    if not payload.send_to_all:
        org_ids = [o for o in (payload.organization_ids or []) if o]
        if not org_ids:
            raise HTTPException(status_code=400, detail="En az bir işletme seçin veya 'tümüne gönder' kullanın.")
        query["organization_id"] = {"$in": org_ids}

    admins = await db.users.find(query, {"_id": 0, "username": 1, "full_name": 1, "email": 1, "organization_id": 1}).to_list(10000)
    if not admins:
        raise HTTPException(status_code=404, detail="Gönderilecek işletme sahibi bulunamadı.")

    email_re = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
    html_content = _wrap_broadcast_email(payload.message)

    sent, failed, skipped = [], [], []
    seen_emails = set()
    for admin in admins:
        to_email = (admin.get("email") or admin.get("username") or "").strip()
        if not email_re.match(to_email):
            skipped.append({"organization_id": admin.get("organization_id"), "reason": "geçersiz e-posta", "value": to_email})
            continue
        if to_email.lower() in seen_emails:
            continue
        seen_emails.add(to_email.lower())
        ok = await send_email(
            to_email=to_email,
            subject=payload.subject,
            html_content=html_content,
            to_name=admin.get("full_name") or to_email,
        )
        (sent if ok else failed).append(to_email)

    logging.info(
        f"📧 [SUPERADMIN BROADCAST] '{payload.subject}' — gönderilen={len(sent)} "
        f"başarısız={len(failed)} atlanan={len(skipped)} (by {current_user.username})"
    )
    return {
        "success": True,
        "sent_count": len(sent),
        "failed_count": len(failed),
        "skipped_count": len(skipped),
        "sent": sent,
        "failed": failed,
        "skipped": skipped,
    }


class SuperAdminBroadcastPush(BaseModel):
    title: str
    body: str
    organization_ids: Optional[List[str]] = None  # None/boş = tüm işletmeler
    send_to_all: bool = False
    action: Optional[str] = None  # örn. "OPEN_STORE" — istemcide mağaza yönlendirmesi


@api_router.post("/superadmin/broadcast-push")
async def superadmin_broadcast_push(
    payload: SuperAdminBroadcastPush,
    request: Request,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """Tüm işletmelere veya seçili işletmelere push bildirim gönder - Sadece superadmin.

    action="OPEN_STORE" verilirse, bildirime dokunan kullanıcı işletim
    sistemine göre native mağaza uygulamasına yönlendirilir (istemci tarafı).
    """
    if not payload.title.strip() or not payload.body.strip():
        raise HTTPException(status_code=400, detail="Başlık ve mesaj zorunludur.")

    # Hedef abonelik sorgusu
    sub_query = {}
    if not payload.send_to_all:
        org_ids = [o for o in (payload.organization_ids or []) if o]
        if not org_ids:
            raise HTTPException(status_code=400, detail="En az bir işletme seçin veya 'tümüne gönder' kullanın.")
        sub_query["organization_id"] = {"$in": org_ids}

    subscriptions = await db.push_subscriptions.find(sub_query).to_list(None)
    if not subscriptions:
        logging.warning(
            f"🔔 [SUPERADMIN BROADCAST PUSH] Abonelik bulunamadı — "
            f"hedef={'tümü' if payload.send_to_all else payload.organization_ids} (by {current_user.username})"
        )
        raise HTTPException(
            status_code=404,
            detail="Seçilen işletmelerde push'a kayıtlı cihaz bulunamadı. Bildirim almak için kullanıcıların uygulamadan/web'den bildirim izni vermiş olması gerekir.",
        )

    # data payload — action verilirse mağaza yönlendirme bilgisi de ekle
    data = {}
    if payload.action:
        data["action"] = payload.action
        if payload.action == "OPEN_STORE":
            data["store_url_ios"] = "itms-apps://apps.apple.com/app/id6759719891"
            data["store_url_android"] = "market://details?id=co.plannapp.app"

    stats = await _deliver_push_to_subscriptions(db, subscriptions, payload.title.strip(), payload.body.strip(), data or None)

    logging.info(
        f"🔔 [SUPERADMIN BROADCAST PUSH] '{payload.title}' — hedef={'tümü' if payload.send_to_all else payload.organization_ids} "
        f"abonelik={len(subscriptions)} native={stats['native_sent']} web={stats['web_sent']} "
        f"hata={stats['failed']} (by {current_user.username})"
    )
    return {
        "success": True,
        "subscriptions": len(subscriptions),
        "native_sent": stats["native_sent"],
        "web_sent": stats["web_sent"],
        "failed": stats["failed"],
        "removed": stats["removed"],
    }


@api_router.get("/superadmin/push/subscriber-counts")
async def get_push_subscriber_counts(
    request: Request,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """İşletme bazında push'a kayıtlı cihaz sayıları - Sadece superadmin.

    SAPushBroadcast ekranında hangi işletmelerin bildirim alabileceğini
    göstermek için kullanılır.
    """
    pipeline = [{"$group": {"_id": {"org": "$organization_id", "platform": "$platform"}, "count": {"$sum": 1}}}]
    rows = await db.push_subscriptions.aggregate(pipeline).to_list(None)

    def _bucket(platform):
        # ios/android dışındaki her şey (web, None, eski kayıtlar) web kabul edilir
        p = (platform or "").lower()
        return p if p in ("ios", "android") else "web"

    counts = {}  # { org_id: {"ios": n, "android": n, "web": n, "total": n} }
    totals = {"ios": 0, "android": 0, "web": 0, "total": 0}
    unassigned = {"ios": 0, "android": 0, "web": 0, "total": 0}

    for r in rows:
        key = r.get("_id") or {}
        oid = key.get("org")
        bucket = _bucket(key.get("platform"))
        n = r.get("count", 0)

        totals[bucket] += n
        totals["total"] += n

        if oid:
            entry = counts.setdefault(oid, {"ios": 0, "android": 0, "web": 0, "total": 0})
            entry[bucket] += n
            entry["total"] += n
        else:
            unassigned[bucket] += n
            unassigned["total"] += n

    # İşletme isimlerini çöz: settings.company_name -> admin kullanıcı -> temsili abone e-postası
    count_org_ids = list(counts.keys())
    name_map = {}
    if count_org_ids:
        for s in await db.settings.find(
            {"organization_id": {"$in": count_org_ids}}, {"organization_id": 1, "company_name": 1}
        ).to_list(None):
            if s.get("company_name"):
                name_map[s["organization_id"]] = s["company_name"]

        missing = [o for o in count_org_ids if o not in name_map]
        if missing:
            for a in await db.users.find(
                {"organization_id": {"$in": missing}, "role": "admin"},
                {"organization_id": 1, "full_name": 1, "username": 1},
            ).to_list(None):
                name_map.setdefault(a["organization_id"], a.get("full_name") or a.get("username"))

        missing2 = [o for o in missing if o not in name_map]
        if missing2:
            for sub in await db.push_subscriptions.find(
                {"organization_id": {"$in": missing2}}, {"organization_id": 1, "user_id": 1}
            ).to_list(None):
                soid = sub.get("organization_id")
                if soid and soid not in name_map and sub.get("user_id"):
                    name_map[soid] = f"{sub['user_id']} (atanmamış)"

    names = {o: (name_map.get(o) or "Bilinmeyen işletme") for o in count_org_ids}

    return {"counts": counts, "totals": totals, "unassigned": unassigned, "names": names}


@api_router.get("/marketing/organizations")
async def get_marketing_organizations(request: Request, current_user: UserInDB = Depends(get_marketing_user), db = Depends(get_db)):
    """İşletme listesi - Marketing kullanıcıları için (read-only)"""
    try:
        now = datetime.now(timezone.utc)
        first_day_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        if now.month == 12:
            last_day_of_month = now.replace(year=now.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last_day_of_month = now.replace(month=now.month + 1, day=1) - timedelta(days=1)
        last_day_of_month = last_day_of_month.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        first_day_str = first_day_of_month.strftime("%Y-%m-%d")
        last_day_str = last_day_of_month.strftime("%Y-%m-%d")
        
        all_settings = await db.settings.find({}).to_list(10000)
        organizations_list = []
        
        for setting in all_settings:
            org_id = setting.get('organization_id')
            if not org_id:
                continue
            
            isletme_adi = setting.get('company_name', 'İsimsiz İşletme')
            telefon_numarasi = setting.get('support_phone', 'Telefon Yok')
            
            # Admin kullanıcıyı bul (işletme sahibi)
            admin_user = await db.users.find_one({"organization_id": org_id, "role": "admin"})
            admin_full_name = admin_user.get('full_name', '-') if admin_user else '-'
            admin_email = admin_user.get('username', '-') if admin_user else '-'
            
            # Abonelik bilgileri
            plan_doc = await db.organization_plans.find_one({"organization_id": org_id})
            billing_cycle = 'monthly'
            days_left = None
            toplam_odeme = 0
            
            if not plan_doc:
                abonelik_paketi = "Trial"
                plan_id = 'tier_trial'
                abonelik_durumu = "Kayıt Yok"
            else:
                plan_id = plan_doc.get('plan_id', 'tier_trial')
                plan_info = next((p for p in PLANS if p['id'] == plan_id), None)
                abonelik_paketi = plan_info.get('name', 'Trial') if plan_info else 'Trial'
                billing_cycle = plan_doc.get('billing_cycle', 'monthly')

                abonelik_durumu, _durum_key, days_left = _compute_subscription_status(plan_doc, now)
            
            # Toplam ödeme miktarı
            payment_logs = await db.payment_logs.find({
                "organization_id": org_id,
                "status": {"$in": ["active", "completed"]}
            }).to_list(1000)
            toplam_odeme = sum(log.get('amount', 0) for log in payment_logs)
            
            # Bu ayki randevu sayısı
            bu_ayki_randevu_sayisi = await db.appointments.count_documents({
                "organization_id": org_id,
                "appointment_date": {"$gte": first_day_str, "$lte": last_day_str}
            })
            
            # Toplam müşteri sayısı
            toplam_musteri_sayisi = await db.customers.count_documents({"organization_id": org_id})
            
            # Toplam personel sayısı
            toplam_personel_sayisi = await db.users.count_documents({"organization_id": org_id, "role": "staff"})
            
            organizations_list.append({
                "organization_id": org_id,
                "isletme_adi": isletme_adi,
                "telefon_numarasi": telefon_numarasi,
                "admin_full_name": admin_full_name,
                "admin_email": admin_email,
                "plan_id": plan_id,
                "abonelik_paketi": abonelik_paketi,
                "billing_cycle": billing_cycle,
                "abonelik_durumu": abonelik_durumu,
                "days_left": days_left,
                "toplam_odeme": toplam_odeme,
                "bu_ayki_randevu_sayisi": bu_ayki_randevu_sayisi,
                "toplam_musteri_sayisi": toplam_musteri_sayisi,
                "toplam_personel_sayisi": toplam_personel_sayisi
            })
        
        return {"organizations": organizations_list}
    except Exception as e:
        logging.error(f"Error in get_marketing_organizations: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"İşletme listesi alınırken hata oluştu: {str(e)}")

@api_router.post("/superadmin/organizations/{org_id}/assign-plan")
async def assign_plan_to_organization(
    org_id: str,
    request: Request,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """İşletmeye paket ata - Sadece superadmin"""
    try:
        body = await request.json()
        plan_id = body.get('plan_id')
        billing_cycle = body.get('billing_cycle', 'monthly')
        
        if not plan_id:
            raise HTTPException(status_code=400, detail="Plan ID gerekli")
        
        plan_info = next((p for p in PLANS if p['id'] == plan_id), None)
        if not plan_info:
            raise HTTPException(status_code=400, detail="Geçersiz plan ID")
        
        now = datetime.now(timezone.utc)
        
        if plan_id == 'tier_trial':
            # Trial paketi ata
            trial_end = now + timedelta(days=7)
            update_data = {
                "organization_id": org_id,
                "plan_id": plan_id,
                "quota_usage": 0,
                "billing_cycle": "monthly",
                "trial_start_date": now.isoformat(),
                "trial_end_date": trial_end.isoformat(),
                "quota_reset_date": trial_end.isoformat(),
                "updated_at": now.isoformat()
            }
        else:
            # Ücretli paket ata
            if billing_cycle == 'yearly':
                quota_reset = now + timedelta(days=365)
            else:
                quota_reset = now + timedelta(days=30)
            
            update_data = {
                "organization_id": org_id,
                "plan_id": plan_id,
                "quota_usage": 0,
                "billing_cycle": billing_cycle,
                "trial_start_date": None,
                "trial_end_date": None,
                "quota_reset_date": quota_reset.isoformat(),
                "updated_at": now.isoformat()
            }
        
        await db.organization_plans.update_one(
            {"organization_id": org_id},
            {"$set": update_data},
            upsert=True
        )
        
        logger.info(f"SuperAdmin paket atadı: org={org_id}, plan={plan_id}, cycle={billing_cycle}")
        return {"message": f"Paket başarıyla atandı: {plan_info.get('name')}", "plan_id": plan_id}
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error in assign_plan: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Paket atanırken hata: {str(e)}")

@api_router.delete("/superadmin/organizations/{org_id}")
async def delete_organization(
    org_id: str,
    request: Request,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """İşletmeyi tamamen sil - Sadece superadmin"""
    try:
        # Tüm ilgili verileri sil
        await db.users.delete_many({"organization_id": org_id})
        await db.appointments.delete_many({"organization_id": org_id})
        await db.customers.delete_many({"organization_id": org_id})
        await db.services.delete_many({"organization_id": org_id})
        await db.transactions.delete_many({"organization_id": org_id})
        await db.expenses.delete_many({"organization_id": org_id})
        await db.settings.delete_one({"organization_id": org_id})
        await db.organization_plans.delete_one({"organization_id": org_id})
        await db.audit_logs.delete_many({"organization_id": org_id})
        await db.customer_notes.delete_many({"organization_id": org_id})
        await db.payment_logs.delete_many({"organization_id": org_id})
        
        logger.info(f"SuperAdmin işletmeyi sildi: org={org_id}")
        return {"message": "İşletme ve tüm verileri silindi"}
    
    except Exception as e:
        logging.error(f"Error in delete_organization: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"İşletme silinirken hata: {str(e)}")

@api_router.patch("/superadmin/organizations/{org_id}/internal")
async def toggle_internal_org(
    org_id: str,
    request: Request,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """İşletmeyi iç/test olarak işaretle veya kaldır — finansal metriklerden hariç tutar"""
    try:
        body = await request.json()
        is_internal = bool(body.get("is_internal", False))
        await db.settings.update_one(
            {"organization_id": org_id},
            {"$set": {"is_internal": is_internal}},
        )
        label = "iç/test" if is_internal else "normal"
        logger.info(f"SuperAdmin org {org_id} → {label} olarak işaretledi")
        return {"message": f"İşletme {label} olarak güncellendi", "is_internal": is_internal}
    except Exception as e:
        logging.error(f"Error in toggle_internal_org: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"İşaretleme hatası: {str(e)}")

@api_router.delete("/superadmin/organizations/{org_id}/staff/{user_id}")
async def delete_staff_member(
    org_id: str,
    user_id: str,
    request: Request,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """Personeli sil - Sadece superadmin"""
    try:
        # Personeli kontrol et
        user = await db.users.find_one({"id": user_id, "organization_id": org_id, "role": "staff"})
        if not user:
            raise HTTPException(status_code=404, detail="Personel bulunamadı")
        
        await db.users.delete_one({"id": user_id})
        
        logger.info(f"SuperAdmin personeli sildi: user={user_id}, org={org_id}")
        return {"message": "Personel silindi"}
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error in delete_staff: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Personel silinirken hata: {str(e)}")

@api_router.get("/superadmin/organizations/{org_id}/staff")
async def get_organization_staff(
    org_id: str,
    request: Request,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """İşletmenin personel listesi - Sadece superadmin"""
    try:
        staff_list = await db.users.find(
            {"organization_id": org_id, "role": "staff"},
            {"_id": 0, "id": 1, "full_name": 1, "username": 1, "status": 1}
        ).to_list(1000)
        
        return {"staff": staff_list}
    
    except Exception as e:
        logging.error(f"Error in get_staff: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Personel listesi alınırken hata: {str(e)}")

@api_router.get("/superadmin/plans")
async def get_available_plans(
    request: Request,
    current_user: UserInDB = Depends(get_superadmin_user)
):
    """Mevcut planları getir - Sadece superadmin"""
    return {"plans": PLANS}

@api_router.get("/superadmin/whatsapp-logs")
async def get_whatsapp_message_logs(
    request: Request,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db),
    status_filter: Optional[str] = None,   # ?status_filter=failed
    organization_id: Optional[str] = None,  # işletme detayı — sadece bu org'a ait loglar
    limit: int = 200,
    skip: int = 0,
):
    """
    WhatsApp mesaj durum logları - Sadece superadmin.
    Kaydedilen alanlar: message_id, recipient, status (sent/delivered/read/failed),
    timestamp (Unix), errors (failed durumunda), recorded_at.
    """
    try:
        query = {}
        if status_filter:
            query["status"] = status_filter

        org_name_prefill = ""
        if organization_id:
            limit = min(max(limit, 1), 500)
            recipient_set = set()
            for coll_name in ("appointments", "customers"):
                coll = getattr(db, coll_name)
                async for doc in coll.find({"organization_id": organization_id}, {"phone": 1}):
                    p = doc.get("phone")
                    if not p:
                        continue
                    d = re.sub(r"\D", "", str(p))
                    if len(d) >= 9:
                        recipient_set.add(d)
                        if len(d) >= 10:
                            recipient_set.add(d[-10:])
            sdoc = await db.settings.find_one({"organization_id": organization_id}, {"support_phone": 1, "company_name": 1})
            if sdoc:
                org_name_prefill = sdoc.get("company_name", "") or ""
                sp = sdoc.get("support_phone")
                if sp:
                    d = re.sub(r"\D", "", str(sp))
                    if len(d) >= 9:
                        recipient_set.add(d)
                        if len(d) >= 10:
                            recipient_set.add(d[-10:])
            if recipient_set:
                query["recipient"] = {"$in": list(recipient_set)}
            else:
                logs = []
                total = 0
                failed_count = await db.whatsapp_message_logs.count_documents({"status": "failed"})
                delivered_count = await db.whatsapp_message_logs.count_documents({"status": "delivered"})
                read_count = await db.whatsapp_message_logs.count_documents({"status": "read"})
                sent_count = await db.whatsapp_message_logs.count_documents({"status": "sent"})
                return {
                    "logs": [],
                    "total": 0,
                    "skip": skip,
                    "limit": limit,
                    "summary": {"sent": sent_count, "delivered": delivered_count, "read": read_count, "failed": failed_count},
                }

        logs = await db.whatsapp_message_logs.find(
            query, {"_id": 0}
        ).sort("recorded_at", -1).skip(skip).limit(limit).to_list(limit)

        if organization_id:
            for log in logs:
                log["organization_id"] = organization_id
                log["org_name"] = org_name_prefill
        else:
            # Enrich: phone → organization via appointments
            all_phones = list({l.get("recipient", "") for l in logs if l.get("recipient")})
            phone_org_map = {}
            if all_phones:
                pipeline = [
                    {"$match": {"phone": {"$in": ["+" + p for p in all_phones] + all_phones + ["0" + p[-10:] for p in all_phones if len(p) >= 10]}}},
                    {"$group": {"_id": "$phone", "org_id": {"$first": "$organization_id"}}},
                ]
                phone_orgs = await db.appointments.aggregate(pipeline).to_list(1000)
                for po in phone_orgs:
                    norm = re.sub(r"\D", "", po["_id"]).lstrip("0")
                    phone_org_map[norm[-9:]] = po["org_id"]

                org_ids = list(set(phone_org_map.values()))
                org_name_map = {}
                if org_ids:
                    settings_docs = await db.settings.find(
                        {"organization_id": {"$in": org_ids}},
                        {"_id": 0, "organization_id": 1, "company_name": 1}
                    ).to_list(len(org_ids))
                    for s in settings_docs:
                        org_name_map[s["organization_id"]] = s.get("company_name", "")

            for log in logs:
                recip = log.get("recipient", "")
                norm_recip = re.sub(r"\D", "", recip).lstrip("0")
                oid = phone_org_map.get(norm_recip[-9:], "")
                log["organization_id"] = oid
                log["org_name"] = org_name_map.get(oid, "") if oid else ""

        total = await db.whatsapp_message_logs.count_documents(query)

        failed_count    = await db.whatsapp_message_logs.count_documents({"status": "failed"})
        delivered_count = await db.whatsapp_message_logs.count_documents({"status": "delivered"})
        read_count      = await db.whatsapp_message_logs.count_documents({"status": "read"})
        sent_count      = await db.whatsapp_message_logs.count_documents({"status": "sent"})

        return {
            "logs": logs,
            "total": total,
            "skip": skip,
            "limit": limit,
            "summary": {
                "sent":      sent_count,
                "delivered": delivered_count,
                "read":      read_count,
                "failed":    failed_count,
            },
        }
    except Exception as e:
        logging.error(f"Error in get_whatsapp_message_logs: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="WhatsApp logları alınırken hata oluştu")

# === SUPERADMIN FINANCIAL STATS ===
@api_router.get("/superadmin/financial-stats")
async def get_financial_stats(request: Request, current_user: UserInDB = Depends(get_superadmin_user), db = Depends(get_db)):
    """SaaS finansal özet: MRR, Toplam Ciro, Aktif İşletme, Churn"""
    try:
        now = datetime.now(timezone.utc)
        first_day = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        thirty_days_ago = now - timedelta(days=30)

        # İç/test işletmeleri belirle (finansal metriklerden hariç)
        internal_settings = await db.settings.find(
            {"is_internal": True}, {"_id": 0, "organization_id": 1}
        ).to_list(10000)
        internal_org_ids = {s["organization_id"] for s in internal_settings if s.get("organization_id")}

        # Tüm org plan kayıtları
        all_plans = await db.organization_plans.find({}, {"_id": 0}).to_list(10000)

        # Her org'un son ödeme tutarını al (MRR hesabı için gerçek tutar)
        last_payment_pipeline = [
            {"$match": {"status": {"$in": ["active", "completed"]}}},
            {"$sort": {"created_at": -1}},
            {"$group": {
                "_id": "$organization_id",
                "amount": {"$first": "$amount"},
                "billing_cycle": {"$first": "$billing_cycle"},
            }},
        ]
        last_payments = await db.payment_logs.aggregate(last_payment_pipeline).to_list(10000)
        last_pay_map = {lp["_id"]: lp for lp in last_payments}

        # Aktif işletme: trial OLMAYAN + süresi dolmamış + iç değil
        plan_prices = {p["id"]: p.get("price_monthly", 0) for p in PLANS}
        mrr = 0
        active_count = 0
        churn_count_calc = 0

        for p in all_plans:
            oid = p.get("organization_id")
            if oid in internal_org_ids:
                continue  # iç/test işletme — finansala dahil etme
            pid = p.get("plan_id", "tier_trial")
            if pid == "tier_trial":
                continue  # trial sayılmaz
            # Son ödeme / yenileme tarihi
            deadline = None
            for dk in ("next_billing_date", "next_payment_date", "quota_reset_date"):
                dv = p.get(dk)
                if dv:
                    if isinstance(dv, str):
                        try:
                            deadline = datetime.fromisoformat(dv.replace("Z", "+00:00"))
                        except ValueError:
                            pass
                    elif isinstance(dv, datetime):
                        deadline = dv if dv.tzinfo else dv.replace(tzinfo=timezone.utc)
                    if deadline:
                        break
            if deadline and deadline < now:
                # Süresi dolmuş — churn'e say (son 30 gün içinde ise)
                if deadline >= thirty_days_ago:
                    churn_count_calc += 1
                continue
            # Aktif ücretli plan
            active_count += 1
            # MRR: gerçek ödeme tutarından hesapla, yoksa liste fiyatı
            lp = last_pay_map.get(oid)
            if lp and lp.get("amount"):
                amt = lp["amount"]
                cycle = lp.get("billing_cycle") or p.get("billing_cycle", "monthly")
                mrr += round(amt / 12, 2) if cycle == "yearly" else amt
            else:
                mrr += plan_prices.get(pid, 0)

        # Toplam ciro (tüm zamanlar — payment_logs, iç işletmeler hariç)
        revenue_match = {"status": {"$in": ["active", "completed"]}}
        if internal_org_ids:
            revenue_match["organization_id"] = {"$nin": list(internal_org_ids)}
        revenue_pipeline = [
            {"$match": revenue_match},
            {"$group": {"_id": None, "total": {"$sum": "$amount"}}},
        ]
        rev_result = await db.payment_logs.aggregate(revenue_pipeline).to_list(1)
        total_revenue = round(rev_result[0]["total"], 2) if rev_result else 0

        # Churn: hesaplanan + iptal edilmiş kayıtlar (varsa)
        churn_cancelled = await db.organization_plans.count_documents({
            "status": {"$in": ["expired", "cancelled"]},
            "updated_at": {"$gte": thirty_days_ago.isoformat()}
        })
        churn_count = churn_count_calc + churn_cancelled

        # Bu ay yeni kayıtlar (organization_plans.created_at daha güvenilir)
        new_this_month = await db.organization_plans.count_documents({
            "created_at": {"$gte": first_day.isoformat()}
        })

        # WhatsApp başarı oranı (son 7 gün)
        seven_days_ago = now - timedelta(days=7)
        wa_total = await db.whatsapp_message_logs.count_documents({
            "recorded_at": {"$gte": seven_days_ago.isoformat()}
        })
        wa_success = await db.whatsapp_message_logs.count_documents({
            "status": {"$in": ["delivered", "read"]},
            "recorded_at": {"$gte": seven_days_ago.isoformat()}
        })
        wa_success_rate = round((wa_success / wa_total * 100) if wa_total > 0 else 0, 1)

        # Toplam randevu bu ay
        total_appointments_month = await db.appointments.count_documents({
            "created_at": {"$gte": first_day.isoformat()}
        })

        return {
            "mrr": mrr,
            "total_revenue": total_revenue,
            "active_businesses": active_count,
            "churn_last_30d": churn_count,
            "new_registrations_this_month": new_this_month,
            "wa_success_rate": wa_success_rate,
            "total_appointments_this_month": total_appointments_month,
        }
    except Exception as e:
        logging.error(f"Error in get_financial_stats: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Finansal istatistikler alınamadı")


# === SUPERADMIN — MARKETING KULLANICI YÖNETİMİ ===
class MarketingUserCreate(BaseModel):
    username: str
    password: str
    full_name: Optional[str] = None

@api_router.post("/superadmin/users/marketing")
async def create_marketing_user(
    request: Request,
    user_data: MarketingUserCreate,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """Marketing kullanıcısı oluştur - Sadece superadmin"""
    existing = await db.users.find_one({"username": user_data.username.lower().strip()})
    if existing:
        raise HTTPException(status_code=400, detail="Bu kullanıcı adı zaten kullanılıyor")
    hashed = pwd_context.hash(user_data.password)
    new_user = {
        "id": str(uuid.uuid4()),
        "username": user_data.username.lower().strip(),
        "full_name": user_data.full_name or user_data.username,
        "role": "marketing",
        "organization_id": "global",
        "hashed_password": hashed,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(new_user)
    new_user.pop("hashed_password", None)
    new_user.pop("_id", None)
    logging.info(f"Marketing kullanıcısı oluşturuldu: {new_user['username']}")
    return {"message": "Marketing kullanıcısı oluşturuldu", "user": new_user}

@api_router.get("/superadmin/users/marketing")
async def list_marketing_users(
    request: Request,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """Marketing kullanıcılarını listele"""
    users = await db.users.find({"role": "marketing"}, {"_id": 0, "hashed_password": 0}).to_list(500)
    return {"users": users}

@api_router.delete("/superadmin/users/marketing/{user_id}")
async def delete_marketing_user(
    request: Request,
    user_id: str,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """Marketing kullanıcısını sil"""
    result = await db.users.delete_one({"id": user_id, "role": "marketing"})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    return {"message": "Kullanıcı silindi"}


# === SUPERADMIN — LEAD YÖNETİMİ ===
LEAD_CLAIM_TIMEOUT_MINUTES = 15

async def _release_expired_claims(db):
    """Süresi dolmuş claim'leri havuza geri at ve sistem logu düş"""
    timeout_dt = datetime.now(timezone.utc) - timedelta(minutes=LEAD_CLAIM_TIMEOUT_MINUTES)
    expired = await db.leads.find({
        "status": "claimed",
        "claimed_at": {"$lt": timeout_dt.isoformat()}
    }, {"_id": 0}).to_list(1000)
    for lead in expired:
        system_log = {
            "action": "timeout_expired",
            "user": lead.get("claimed_by_name", "?"),
            "user_id": lead.get("claimed_by"),
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "note": f"15 dakika süre doldu, {lead.get('claimed_by_name', '?')} işlem yapmadan süresi doldu"
        }
        await db.leads.update_one(
            {"id": lead["id"]},
            {"$set": {"status": "pool", "claimed_by": None, "claimed_at": None, "claimed_by_name": None},
             "$push": {"system_logs": system_log}}
        )
    return len(expired)

@api_router.post("/superadmin/leads/upload/preview")
async def preview_lead_upload(
    request: Request,
    file: UploadFile = File(...),
    current_user: UserInDB = Depends(get_superadmin_user),
):
    """CSV/XLSX dosyasının sütunlarını ve ilk 5 satırını döndür"""
    try:
        content = await file.read()
        filename = (file.filename or "").lower()
        columns = []
        preview_rows = []

        if filename.endswith(".csv"):
            text = content.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            columns = list(reader.fieldnames or [])
            for i, row in enumerate(reader):
                if i >= 5:
                    break
                preview_rows.append(dict(row))
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                columns = [str(c) if c is not None else "" for c in rows[0]]
                for row in rows[1:6]:
                    preview_rows.append({columns[i]: (str(row[i]) if row[i] is not None else "") for i in range(len(columns))})
        else:
            raise HTTPException(status_code=400, detail="Desteklenmeyen dosya formatı. CSV veya XLSX yükleyin.")

        return {"columns": columns, "preview": preview_rows, "filename": file.filename}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Lead upload preview error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Dosya okunamadı: {str(e)}")

@api_router.post("/superadmin/leads/upload/confirm")
async def confirm_lead_upload(
    request: Request,
    file: UploadFile = File(...),
    batch_name: str = Form(...),
    company_col: str = Form(...),
    phone_col: str = Form(...),
    sector_col: str = Form(...),
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """Column mapping ile lead'leri yükle (multipart/form-data)"""
    try:
        content = await file.read()
        filename = (file.filename or "").lower()
        rows = []

        if filename.endswith(".csv"):
            text = content.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                rows.append(dict(row))
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if all_rows:
                cols = [str(c) if c is not None else "" for c in all_rows[0]]
                for row in all_rows[1:]:
                    rows.append({cols[i]: (str(row[i]) if row[i] is not None else "") for i in range(len(cols))})
        else:
            raise HTTPException(status_code=400, detail="Desteklenmeyen dosya formatı")

        batch_id = str(uuid.uuid4())
        now_iso = datetime.now(timezone.utc).isoformat()
        leads_to_insert = []
        skipped = 0
        for row in rows:
            company = str(row.get(company_col, "") or "").strip()
            phone = str(row.get(phone_col, "") or "").strip()
            sector = str(row.get(sector_col, "") or "").strip()
            if not company and not phone:
                skipped += 1
                continue
            leads_to_insert.append({
                "id": str(uuid.uuid4()),
                "batch_id": batch_id,
                "batch_name": batch_name,
                "company_name": company,
                "phone": phone,
                "sector": sector,
                "status": "pool",
                "claimed_by": None,
                "claimed_at": None,
                "claimed_by_name": None,
                "note": "",
                "system_logs": [],
                "uploaded_by": current_user.id,
                "uploaded_by_name": current_user.full_name or current_user.username,
                "created_at": now_iso,
                "updated_at": now_iso,
            })

        if leads_to_insert:
            await db.leads.insert_many(leads_to_insert)

        await db.lead_batches.insert_one({
            "id": batch_id,
            "batch_name": batch_name,
            "total": len(leads_to_insert),
            "skipped": skipped,
            "uploaded_by": current_user.id,
            "uploaded_by_name": current_user.full_name or current_user.username,
            "created_at": now_iso,
        })
        logging.info(f"Lead upload: {len(leads_to_insert)} lead yüklendi, batch={batch_id}")
        return {"message": f"{len(leads_to_insert)} lead yüklendi", "batch_id": batch_id, "total": len(leads_to_insert), "skipped": skipped}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Lead upload confirm error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Yükleme başarısız: {str(e)}")

@api_router.get("/superadmin/leads")
async def get_all_leads(
    request: Request,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db),
    status_filter: Optional[str] = None,
    batch_id: Optional[str] = None,
    assigned_to: Optional[str] = None,
    skip: int = 0,
    limit: int = 100
):
    """Tüm leadleri listele (superadmin)"""
    await _release_expired_claims(db)
    query = {}
    if status_filter:
        query["status"] = status_filter
    if batch_id:
        query["batch_id"] = batch_id
    if assigned_to:
        query["claimed_by"] = assigned_to
    total = await db.leads.count_documents(query)
    leads = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"leads": leads, "total": total}

@api_router.get("/superadmin/leads/stats")
async def get_lead_stats(
    request: Request,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """Lead dönüşüm istatistikleri"""
    await _release_expired_claims(db)
    total = await db.leads.count_documents({})
    pool = await db.leads.count_documents({"status": "pool"})
    claimed = await db.leads.count_documents({"status": "claimed"})
    interested = await db.leads.count_documents({"status": "interested"})
    unreachable = await db.leads.count_documents({"status": "unreachable"})
    not_interested = await db.leads.count_documents({"status": "not_interested"})
    waiting = await db.leads.count_documents({"status": "waiting"})
    registered = await db.leads.count_documents({"status": "registered"})
    called = total - pool
    conversion_rate = round((registered / called * 100) if called > 0 else 0, 1)
    return {
        "total": total, "pool": pool, "claimed": claimed,
        "interested": interested, "unreachable": unreachable,
        "not_interested": not_interested, "waiting": waiting,
        "registered": registered, "called": called,
        "conversion_rate": conversion_rate
    }

@api_router.get("/superadmin/leads/batches")
async def get_lead_batches(
    request: Request,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """Yükleme geçmişi"""
    batches = await db.lead_batches.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return {"batches": batches}

@api_router.delete("/superadmin/leads/batch/{batch_id}")
async def delete_lead_batch(
    request: Request,
    batch_id: str,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    """Yükleme grubunu ve tüm leadlerini sil"""
    result = await db.leads.delete_many({"batch_id": batch_id})
    await db.lead_batches.delete_one({"id": batch_id})
    return {"message": f"{result.deleted_count} lead silindi"}

@api_router.delete("/superadmin/leads/{lead_id}")
async def delete_single_lead(
    request: Request,
    lead_id: str,
    current_user: UserInDB = Depends(get_superadmin_user),
    db = Depends(get_db)
):
    await db.leads.delete_one({"id": lead_id})
    return {"message": "Lead silindi"}


# === MARKETING — TELE-SATIŞ CRM ===
@api_router.get("/marketing/leads/batches")
async def get_marketing_batches(
    request: Request,
    current_user: UserInDB = Depends(get_marketing_user),
    db = Depends(get_db)
):
    """Marketing kullanıcısı için mevcut batch listesi"""
    batches = await db.lead_batches.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return {"batches": batches}

@api_router.get("/marketing/leads/sectors")
async def get_marketing_sectors(
    request: Request,
    current_user: UserInDB = Depends(get_marketing_user),
    db = Depends(get_db)
):
    """Marketing kullanıcısı için benzersiz sektör listesi"""
    sectors = await db.leads.distinct("sector", {"sector": {"$nin": ["", None]}})
    return {"sectors": sorted(sectors)}

@api_router.get("/marketing/leads")
async def get_marketing_leads(
    request: Request,
    current_user: UserInDB = Depends(get_marketing_user),
    db = Depends(get_db),
    tab: str = "pool",
    batch_id: str = None,
    sector: str = None,
    limit: int = 200,
):
    """Havuz leadleri veya benim leadlerim (batch_id ve sector ile filtrelenebilir).

    `limit` parametresi varsayılan 200; UI tarafında pazarlamacı tek seferde
    en fazla 200 lead görür (önceki 5000 limit React render'ını donduruyordu).
    Daha fazla görmek isterse filtre (kampanya/sektör) uygulamalı.
    """
    limit = max(1, min(int(limit or 200), 1000))
    logging.info(f"[LEADS] user={current_user.username} id={current_user.id} tab={tab} batch_id={batch_id} sector={sector} limit={limit}")
    await _release_expired_claims(db)
    if tab == "mine":
        query = {"claimed_by": current_user.id, "status": {"$in": ["claimed", "interested", "unreachable", "not_interested", "waiting", "registered"]}}
        if batch_id:
            query["batch_id"] = batch_id
        if sector:
            query["sector"] = sector
        leads = await db.leads.find(query, {"_id": 0}).sort("updated_at", -1).to_list(limit)
    else:
        query = {"status": "pool"}
        if batch_id:
            query["batch_id"] = batch_id
        if sector:
            query["sector"] = sector
        leads = await db.leads.find(query, {"_id": 0}).sort("created_at", 1).to_list(limit)
    return {"leads": leads, "batch_id": batch_id, "limit": limit}

@api_router.post("/marketing/leads/{lead_id}/claim")
async def claim_lead(
    request: Request,
    lead_id: str,
    current_user: UserInDB = Depends(get_marketing_user),
    db = Depends(get_db)
):
    """Lead'i al (15dk kilit)"""
    await _release_expired_claims(db)
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead bulunamadı")
    if lead["status"] != "pool":
        raise HTTPException(status_code=409, detail="Bu lead başka biri tarafından alınmış")
    now = datetime.now(timezone.utc).isoformat()
    system_log = {
        "action": "claimed",
        "user": current_user.full_name or current_user.username,
        "user_id": current_user.id,
        "timestamp": now,
        "note": ""
    }
    await db.leads.update_one(
        {"id": lead_id, "status": "pool"},
        {"$set": {
            "status": "claimed",
            "claimed_by": current_user.id,
            "claimed_by_name": current_user.full_name or current_user.username,
            "claimed_at": now,
            "updated_at": now,
        }, "$push": {"system_logs": system_log}}
    )
    updated = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not updated or updated.get("claimed_by") != current_user.id:
        raise HTTPException(status_code=409, detail="Lead başka biri tarafından alındı")
    return {"message": "Lead alındı", "lead": updated}

@api_router.put("/marketing/leads/{lead_id}/status")
async def update_lead_status(
    request: Request,
    lead_id: str,
    current_user: UserInDB = Depends(get_marketing_user),
    db = Depends(get_db)
):
    """Lead durumunu güncelle"""
    body = await request.json()
    new_status = body.get("status")
    note = body.get("note", "").strip()
    valid_statuses = ["interested", "unreachable", "not_interested", "waiting", "registered"]
    if new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail="Geçersiz durum")
    if new_status in ["not_interested", "waiting"] and not note:
        raise HTTPException(status_code=400, detail="Bu durum için not zorunludur")
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead bulunamadı")
    if lead.get("claimed_by") != current_user.id:
        raise HTTPException(status_code=403, detail="Bu lead size ait değil")
    now = datetime.now(timezone.utc).isoformat()
    system_log = {
        "action": f"status_updated_{new_status}",
        "user": current_user.full_name or current_user.username,
        "user_id": current_user.id,
        "timestamp": now,
        "note": note
    }
    update_data = {
        "status": new_status,
        "note": note,
        "updated_at": now,
        # claimed_by ve claimed_by_name sakla — mine tab ve stats için gerekli
    }
    await db.leads.update_one(
        {"id": lead_id},
        {"$set": update_data, "$push": {"system_logs": system_log}}
    )
    updated = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    return {"message": "Durum güncellendi", "lead": updated}

@api_router.post("/marketing/leads/{lead_id}/release")
async def release_lead(
    request: Request,
    lead_id: str,
    current_user: UserInDB = Depends(get_marketing_user),
    db = Depends(get_db)
):
    """Lead'i havuza geri bırak"""
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead bulunamadı")
    if lead.get("claimed_by") != current_user.id:
        raise HTTPException(status_code=403, detail="Bu lead size ait değil")
    now = datetime.now(timezone.utc).isoformat()
    system_log = {
        "action": "manual_release",
        "user": current_user.full_name or current_user.username,
        "user_id": current_user.id,
        "timestamp": now,
        "note": "Manuel bırakıldı"
    }
    await db.leads.update_one(
        {"id": lead_id},
        {"$set": {"status": "pool", "claimed_by": None, "claimed_at": None, "claimed_by_name": None, "updated_at": now},
         "$push": {"system_logs": system_log}}
    )
    return {"message": "Lead havuza bırakıldı"}

@api_router.get("/marketing/my-stats")
async def get_my_marketing_stats(
    request: Request,
    current_user: UserInDB = Depends(get_marketing_user),
    db = Depends(get_db)
):
    """Pazarlamacının kendi istatistikleri"""
    logging.info(f"[STATS] user={current_user.username} id={current_user.id} role={current_user.role}")
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    total_called = await db.leads.count_documents({"claimed_by": current_user.id, "status": {"$nin": ["pool", "claimed"]}})
    interested = await db.leads.count_documents({"claimed_by": current_user.id, "status": "interested"})
    registered = await db.leads.count_documents({"claimed_by": current_user.id, "status": "registered"})
    # Bugün güncellenenler (yaklaşık)
    today_leads = await db.leads.count_documents({
        "claimed_by": current_user.id,
        "updated_at": {"$gte": today_start}
    })
    conversion = round((registered / total_called * 100) if total_called > 0 else 0, 1)
    return {
        "total_called": total_called,
        "today_called": today_leads,
        "interested": interested,
        "registered": registered,
        "conversion_rate": conversion
    }

@api_router.get("/marketing/profile")
async def get_marketing_profile(
    request: Request,
    current_user: UserInDB = Depends(get_marketing_user),
):
    """Marketing kullanıcısının profil bilgileri"""
    return {
        "id": current_user.id,
        "username": current_user.username,
        "full_name": current_user.full_name,
        "role": current_user.role,
    }


# === AI CHATBOT ENDPOINT ===
class AIChatRequest(BaseModel):
    message: str
    history: Optional[List[Dict]] = []
    language: Optional[str] = "tr"

@api_router.post("/ai/chat")
@rate_limit(LIMITS["ai_chat"])
async def ai_chat_endpoint(
    body: AIChatRequest,
    request: Request,
    current_user: dict = Depends(get_current_user),
    db = Depends(get_db)
):
    """
    AI Chatbot endpoint - Gemini 2.5 Flash ile sohbet
    
    Request body:
    {
        "message": "Kullanıcının mesajı",
        "history": [  // Opsiyonel chat history
            {"role": "user", "parts": [{"text": "..."}]},
            {"role": "model", "parts": [{"text": "..."}]}
        ]
    }
    """
    try:
        # current_user is UserInDB Pydantic model, not dict
        user_role = getattr(current_user, 'role', 'staff')
        username = current_user.username
        organization_id = current_user.organization_id
        user_full_name = getattr(current_user, 'full_name', username) or username
        
        # AI MESAJ KOTA KONTROLÜ
        plan_doc = await db.organization_plans.find_one({"organization_id": organization_id})
        if not plan_doc:
            # Plan yoksa default trial oluştur
            plan_doc = {
                "organization_id": organization_id,
                "plan_id": "tier_trial",
                "quota_usage": 0,
                "ai_usage_count": 0,
                "quota_reset_date": (datetime.now(timezone.utc) + timedelta(days=30)).isoformat(),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.organization_plans.insert_one(plan_doc)
        
        # Plan limitini al
        plan_id = plan_doc.get('plan_id', 'tier_trial')
        plan_info = next((p for p in PLANS if p['id'] == plan_id), None)
        
        if not plan_info:
            raise HTTPException(status_code=500, detail="Plan bilgisi bulunamadı")
        
        ai_message_limit = plan_info.get('ai_message_limit', 100)
        current_ai_usage = plan_doc.get('ai_usage_count', 0)
        
        # Limit kontrolü (Sınırsız değilse)
        if ai_message_limit != -1 and current_ai_usage >= ai_message_limit:
            remaining_pct = 0
            raise HTTPException(
                status_code=403,
                detail=f"Aylık AI kullanım limitiniz doldu ({ai_message_limit} mesaj). Kesintisiz hizmet için paketinizi yükseltin."
            )
        
        # Tier bazlı AI erişim seviyesi
        ai_tier_map = {
            "tier_trial": "basic",
            "tier_1_standard": "standard",
            "tier_2_profesyonel": "pro",
            "tier_3_premium": "pro",
            "tier_4_business": "full",
            "tier_5_enterprise": "full",
            "tier_6_kurumsal": "full",
        }
        ai_tier = ai_tier_map.get(plan_id, "basic")
        
        # Organizasyon bilgisini al
        settings = await db.settings.find_one({"organization_id": organization_id})
        organization_name = settings.get('company_name', 'İşletme') if settings else 'İşletme'
        
        # AI ile sohbet et
        result = await chat_with_ai(
            db=db,
            user_message=body.message,
            chat_history=body.history,
            user_role=user_role,
            username=username,
            organization_id=organization_id,
            organization_name=organization_name,
            language=body.language or "tr",
            can_view_all_appointments=getattr(current_user, 'can_view_all_appointments', False),
            ai_tier=ai_tier
        )
        
        if not result.get('success'):
            status_code = int(result.get('status', 500))
            detail_msg = result.get('message', 'AI hatası')
            raise HTTPException(status_code=status_code, detail=detail_msg)
        
        # Başarılı mesaj - Kullanımı artır
        await db.organization_plans.update_one(
            {"organization_id": organization_id},
            {"$inc": {"ai_usage_count": 1}}
        )
        
        # Güncel kullanım bilgisini al
        new_usage = current_ai_usage + 1
        
        return {
            "success": True,
            "message": result.get('message'),
            "history": result.get('history', []),
            "usage_info": {
                "current": new_usage,
                "limit": ai_message_limit
            }
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"AI chat endpoint error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"AI sohbet hatası: {str(e)}")

# --- Router prefix'i buraya taşındı ---
app.include_router(api_router, prefix="/api")


# ═══ iOS Universal Links — apple-app-site-association (AASA) ═══
# App kökünde (prefix'siz) servis edilir; nginx `/.well-known/apple-app-site-association`
# path'ini backend'e (8002) proxy'ler.
# Claim edilen path'ler:
#   • /api/app/open*  — e-posta / akıllı açıcı (mevcut)
#   • /appstore*      — Instagram bio Universal Link; uygulama yüklüyse app açılır,
#                       yüklü değilse GET /appstore HTTP 302 ile App Store'a gider
# Diğer https linkleri (şifre sıfırlama, public booking vb.) tarayıcıda kalsın.
# appID = <TeamID>.<bundle> = 9V42VGDN9U.co.plannapp.app
_AASA_PAYLOAD = {
    "applinks": {
        "apps": [],
        "details": [
            {
                "appID": "9V42VGDN9U.co.plannapp.app",
                "appIDs": ["9V42VGDN9U.co.plannapp.app"],
                "paths": [
                    "/api/app/open",
                    "/api/app/open/*",
                    "/appstore",
                    "/appstore/*",
                ],
                "components": [
                    {"/": "/api/app/open", "comment": "PLANN akıllı uygulama açıcı"},
                    {"/": "/api/app/open/*", "comment": "PLANN akıllı uygulama açıcı (alt path)"},
                    {"/": "/appstore", "comment": "Instagram / App Store Universal Link"},
                    {"/": "/appstore/*", "comment": "Instagram / App Store Universal Link (alt path)"},
                ],
            }
        ],
    }
}

# Instagram WebView https meta refresh'i de blokluyor → itms-apps:// native scheme.
# Otomatik JS denemesi + manuel buton (https değil, itms-apps).
_APPSTORE_ITMS_URL = "itms-apps://apps.apple.com/app/id6759719891"
_APPSTORE_META_REFRESH_HTML = """<!DOCTYPE html>
<html lang="tr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PLANN</title>
    <style>
        body { font-family: -apple-system, BlinkMacSystemFont, sans-serif; display: flex; flex-direction: column; justify-content: center; align-items: center; height: 100vh; margin: 0; background-color: #ffffff; }
        .btn { display: inline-block; padding: 14px 28px; background-color: #000000; color: #ffffff; text-decoration: none; border-radius: 8px; font-weight: 600; font-size: 16px; margin-top: 20px;}
        p { color: #333333; font-size: 18px; font-weight: 500;}
    </style>
</head>
<body>
    <p>App Store'a bağlanıyor...</p>
    <a href="itms-apps://apps.apple.com/app/id6759719891" class="btn">App Store'da Aç</a>
    
    <script>
        // Instagram'ın JS'i engellememe ihtimaline karşı otomatik fırlatma
        setTimeout(function() {
            window.location.href = "itms-apps://apps.apple.com/app/id6759719891";
        }, 300);
    </script>
</body>
</html>"""


@app.get("/.well-known/apple-app-site-association", include_in_schema=False)
@app.get("/apple-app-site-association", include_in_schema=False)
async def apple_app_site_association():
    """iOS Universal Links doğrulama dosyası. application/json + no redirect + 200."""
    return JSONResponse(
        content=_AASA_PAYLOAD,
        media_type="application/json",
        headers={"Cache-Control": "public, max-age=3600"},
    )


@app.get("/appstore", include_in_schema=False)
@app.get("/appstore/", include_in_schema=False)
async def appstore_redirect():
    """Instagram bio fallback: HTTP 200 + itms-apps:// (https meta refresh WebView'de bloklanıyor)."""
    return HTMLResponse(
        content=_APPSTORE_META_REFRESH_HTML,
        status_code=200,
        headers={"Cache-Control": "no-store, no-cache, must-revalidate", "Pragma": "no-cache"},
    )


_PLAYSTORE_REDIRECT_URL = (
    "https://play.google.com/store/apps/details?id=co.plannapp.app&pcampaignid=web_share"
)


@app.get("/playstore", include_in_schema=False)
@app.get("/playstore/", include_in_schema=False)
async def playstore_redirect():
    """Google Play Store yönlendirmesi: HTTP 302 (HTML yok)."""
    return RedirectResponse(url=_PLAYSTORE_REDIRECT_URL, status_code=302)

# --- Financial Engine Router ---
try:
    from financial.financial_endpoints import router as financial_router
    app.include_router(financial_router, tags=["Financial Engine"])
    logging.info("✅ Financial Engine router registered")
except Exception as fin_import_err:
    logging.warning(f"⚠️ Financial Engine router import failed: {fin_import_err}")

# --- PLANN Asistan Router (zeka motoru: geçmiş + superadmin önizleme/test/seed) ---
try:
    from assistant.endpoints import router as assistant_router
    app.include_router(assistant_router)
    logging.info("✅ PLANN Asistan router registered")
except Exception as asst_import_err:
    logging.warning(f"⚠️ PLANN Asistan router import failed: {asst_import_err}")

# --- Pazarlama Otopilotu Router (WhatsApp geri-kazanım ayarları) ---
try:
    from marketing_autopilot_endpoints import router as autopilot_router
    app.include_router(autopilot_router)
    logging.info("✅ Pazarlama Otopilotu router registered")
except Exception as autopilot_import_err:
    logging.warning(f"⚠️ Pazarlama Otopilotu router import failed: {autopilot_import_err}")

# --- Meta Conversions API Router (Pixel + CAPI hybrid tracking) ---
try:
    from meta_capi_endpoints import router as meta_capi_router
    app.include_router(meta_capi_router)
    logging.info("✅ Meta CAPI router registered")
except Exception as meta_import_err:
    logging.warning(f"⚠️ Meta CAPI router import failed: {meta_import_err}")

# === WhatsApp Webhook (Meta Cloud API) ===

META_WEBHOOK_VERIFY_TOKEN = os.getenv("META_WEBHOOK_VERIFY_TOKEN", "plann_meta_webhook_2025")
_META_WH_ACCESS_TOKEN = os.getenv("META_ACCESS_TOKEN")
_META_WH_PHONE_ID = os.getenv("META_PHONE_NUMBER_ID")
_META_WH_API_VERSION = os.getenv("META_API_VERSION", "v21.0")


def _send_whatsapp_text_reply(to_number: str, text: str) -> None:
    """Gelen mesaja düz metin (text) olarak otomatik yanıt gönderir."""
    if not _META_WH_ACCESS_TOKEN or not _META_WH_PHONE_ID:
        logger.warning("Meta WA credentials eksik, auto-reply gönderilemedi.")
        return
    url = f"https://graph.facebook.com/{_META_WH_API_VERSION}/{_META_WH_PHONE_ID}/messages"
    payload = {
        "messaging_product": "whatsapp",
        "to": to_number,
        "type": "text",
        "text": {"body": text},
    }
    try:
        import requests as _req
        resp = _req.post(
            url,
            headers={"Authorization": f"Bearer {_META_WH_ACCESS_TOKEN}", "Content-Type": "application/json"},
            json=payload,
            timeout=10,
        )
        if not resp.ok:
            logger.error(f"Auto-reply hata: {resp.status_code} {resp.text}")
        else:
            logger.info(f"Auto-reply gönderildi → {to_number}")
    except Exception as e:
        logger.error(f"Auto-reply exception: {e}")


@app.get("/webhook")
async def whatsapp_webhook_verify(request: Request):
    """
    Meta webhook doğrulama: GET isteğiyle hub.challenge döner.
    """
    params = request.query_params
    mode      = params.get("hub.mode")
    token     = params.get("hub.verify_token")
    challenge = params.get("hub.challenge")

    if mode == "subscribe" and token == META_WEBHOOK_VERIFY_TOKEN:
        logger.info("✅ Meta webhook doğrulandı.")
        return Response(content=challenge, media_type="text/plain")

    logger.warning(f"❌ Webhook doğrulama başarısız. token={token}")
    raise HTTPException(status_code=403, detail="Webhook doğrulama başarısız.")


@app.get("/api/webhook")
async def whatsapp_webhook_verify_api_app(request: Request):
    return await whatsapp_webhook_verify(request)


@api_router.get("/webhook")
async def whatsapp_webhook_verify_api(request: Request):
    return await whatsapp_webhook_verify(request)


async def _process_whatsapp_verification(db, redis_client, from_wa_number: str, code: str) -> dict:
    """
    Webhook'tan gelen RND kodunu doğrula ve randevuyu oluştur.
    from_wa_number: Meta API'den gelen gönderenin numarası (90XXXXXXXXXX formatı)
    """
    if not redis_client:
        return {"reply": "Sistem geçici olarak kullanılamıyor."}

    pending_key = f"plann:pending:{code}"
    raw = await redis_client.hgetall(pending_key)

    if not raw:
        lang = detect_language_from_phone(from_wa_number)
        if lang == "TR":
            return {"reply": f"❌ '{code}' kodu bulunamadı veya süresi doldu. Lütfen yeniden randevu talebinde bulunun."}
        return {"reply": f"❌ Code '{code}' not found or expired. Please submit a new appointment request."}

    data = {
        (k.decode() if isinstance(k, bytes) else k): (v.decode() if isinstance(v, bytes) else v)
        for k, v in raw.items()
    }
    stored_phone = data.get("phone", "")
    org_id = data.get("org_id", "")

    def _norm(p: str) -> str:
        return re.sub(r"\D", "", p).lstrip("0")

    if _norm(from_wa_number)[-9:] != _norm(stored_phone)[-9:]:
        lang = detect_language_from_phone(from_wa_number)
        if lang == "TR":
            return {"reply": "❌ Bu numara randevu formundaki telefon numarasıyla eşleşmiyor. Lütfen randevuyu oluşturduğunuz numara ile mesaj gönderin."}
        return {"reply": "❌ This number does not match the booking form's phone number."}

    try:
        apt_dict = json.loads(data.get("appointment_data", "{}"))
    except Exception:
        return {"reply": "❌ Randevu verisi işlenirken hata oluştu."}

    try:
        quota_ok, quota_error = await check_quota_and_increment(db, org_id)
        if not quota_ok:
            return {"reply": f"❌ {quota_error}"}

        # Çoklu hizmet çözümleme — /public/verify-code ile AYNI mantık.
        # Eskiden burada yalnızca apt_dict.service_id (ilk hizmet) okunuyordu ve
        # çoklu hizmet seçili randevularda diğer hizmetler + toplam süre/fiyat
        # KAYBOLUYORDU. Artık resolve_services ile snapshot + birleşik toplam üretilir.
        # (Tekil hizmette de sorunsuz çalışır.) Webhook 200+reply döndürmeli, bu
        # yüzden resolve_services'in HTTPException'ını yakalayıp kotayı geri alıp
        # dostça bir mesaj dönüyoruz (raise ETMİYORUZ).
        try:
            _apt_in = AppointmentCreate(**apt_dict)
            _settings_flag = await db.settings.find_one(
                {"organization_id": org_id}, {"_id": 0, "multi_service_enabled": 1}
            )
            _multi_enabled = bool((_settings_flag or {}).get("multi_service_enabled", False))
            snapshot_lines, total_duration, total_price, service, resolved_service_ids = await resolve_services(
                db, org_id, _apt_in, multi_enabled=_multi_enabled
            )
        except HTTPException as _svc_err:
            await db.organization_plans.update_one(
                {"organization_id": org_id}, {"$inc": {"quota_usage": -1}}
            )
            _detail = getattr(_svc_err, "detail", "") or ""
            if detect_language_from_phone(from_wa_number) == "TR":
                return {"reply": f"❌ {_detail or 'Hizmet bulunamadı. Lütfen tekrar randevu alın.'}"}
            return {"reply": f"❌ {_detail or 'Service not found. Please book again.'}"}

        apt_obj = Appointment(**{
            **apt_dict,
            "organization_id": org_id,
            "service_id": resolved_service_ids[0],
            "services": snapshot_lines,
            "service_name": service.get("name", ""),
            "service_price": service.get("price", 0),
            "service_duration": service.get("duration", 30),
            "source": "public_booking_wa_verified",
            "status": "Bekliyor",
        })
        doc = apt_obj.model_dump()
        doc["created_at"] = doc["created_at"].isoformat()

        # ── Slot çakışma kontrolü (TOCTOU koruması) ──────────────────────
        # /public/verify-code ile aynı: kod (WhatsApp cevabı) gelene kadar slot
        # dolabilir. Kontrolsüz insert personelsiz org'da çift kayda yol açıyordu.
        # Webhook 200+reply döndürmeli → raise etmiyoruz, kotayı geri alıp
        # pending'i temizleyip dostça mesaj dönüyoruz.
        _conflict_dur = int(total_duration or service.get("duration", 30) or 30)
        if await _check_slot_conflict(
            db, org_id, apt_obj.staff_member_id,
            apt_obj.appointment_date, apt_obj.appointment_time, _conflict_dur,
        ):
            await db.organization_plans.update_one(
                {"organization_id": org_id}, {"$inc": {"quota_usage": -1}}
            )
            try:
                await redis_client.delete(pending_key)
                await redis_client.delete(f"plann:pending:phone:{stored_phone}:{org_id}")
            except Exception:
                pass
            if detect_language_from_phone(from_wa_number) == "TR":
                return {"reply": "❌ Seçtiğiniz saat az önce doldu. Lütfen yeniden randevu oluşturun."}
            return {"reply": "❌ Your selected time slot was just taken. Please book again."}

        await db.appointments.insert_one(doc)

        await db.customers.update_one(
            {"organization_id": org_id, "phone": apt_dict.get("phone", "")},
            {
                "$set": {
                    "name": apt_dict.get("customer_name", ""),
                    "phone": apt_dict.get("phone", ""),
                    "organization_id": org_id,
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                },
                "$setOnInsert": {
                    "id": str(uuid.uuid4()),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "notes": "",
                },
            },
            upsert=True,
        )

        phone_map_key = f"plann:pending:phone:{stored_phone}:{org_id}"
        await redis_client.delete(pending_key)
        await redis_client.delete(phone_map_key)
        await redis_client.setex(f"plann:verified:{code}", 300, "1")

        try:
            apt_clean = {k: v for k, v in doc.items() if k != "_id"}
            await emit_to_organization(org_id, "appointment_created", {"appointment": apt_clean})
        except Exception:
            pass

        # ── Push notifications → admin + personel ───────────────────
        try:
            apt_date_raw = apt_dict.get("appointment_date", "")
            _date_tr_push = format_date_tr(apt_date_raw)
            # Bildirim metni: Müşteri / Hizmet / Tarih (7 Ağustos Cuma 09:45) alt alta
            notif_body = (
                f"Müşteri: {apt_dict.get('customer_name', '')}\n"
                f"Hizmet: {service.get('name', '')}\n"
                f"Tarih: {_date_tr_push} {apt_dict.get('appointment_time', '')}"
            )
            _apt_id_wa = doc.get("id")
            notif_data = {
                "type": "new_appointment",
                "appointment_id": _apt_id_wa,
                "source": "public_booking_wa_verified",
                "url": f"/?randevu={_apt_id_wa}" if _apt_id_wa else "/",
            }
            admins = await db.users.find({"organization_id": org_id, "role": "admin"}).to_list(100)
            for admin in admins:
                await send_push_notification(
                    db=db,
                    organization_id=org_id,
                    title="🔔 Yeni Online Randevu!",
                    body=notif_body,
                    data=notif_data,
                    user_id=admin["username"],
                )
            assigned = apt_dict.get("staff_member_id")
            if assigned:
                staff_doc = await db.users.find_one({"username": assigned, "organization_id": org_id})
                if staff_doc and staff_doc.get("role") != "admin":
                    await send_push_notification(
                        db=db,
                        organization_id=org_id,
                        title="🔔 Yeni Randevu Atandı!",
                        body=notif_body,
                        data=notif_data,
                        user_id=assigned,
                    )
            logging.info(f"✓ Push notifications sent for WA-verified appointment {doc.get('id')}")
        except Exception as _push_err:
            logging.error(f"WA verified push notification hatası: {_push_err}")

        settings_data = await db.settings.find_one({"organization_id": org_id})
        company_name  = (settings_data or {}).get("company_name", "İşletmeniz")
        support_phone = (settings_data or {}).get("support_phone", "")
        location      = (settings_data or {}).get("location") or {}
        coords        = location.get("coordinates") or {}
        lat           = coords.get("lat")
        lng           = coords.get("lng")
        address       = location.get("address", "")

        try:
            await asyncio.to_thread(
                send_whatsapp_template,
                apt_dict.get("phone", ""),
                "CONFIRMATION",
                apt_dict.get("customer_name", ""),
                company_name,
                apt_dict.get("appointment_date", ""),
                apt_dict.get("appointment_time", ""),
                service.get("name", ""),
                support_phone,
                business_lat=lat,
                business_lng=lng,
                business_address=address,
            )
        except Exception as _wa_err:
            logging.warning(f"WA onay template gönderilemedi: {_wa_err}")

        return {"reply": ""}  # Template gönderildi, ek düz metin gerekmez

    except Exception as e:
        logging.error(f"WA doğrulama randevu hatası: {e}", exc_info=True)
        return {"reply": "❌ Randevu oluşturulurken bir hata oluştu. Lütfen tekrar deneyin."}


@app.post("/webhook")
async def whatsapp_webhook(request: Request):
    """
    Meta'dan gelen mesajları ve durum güncellemelerini dinler.
    - messages  → dil tespiti yaparak otomatik yanıt gönderir
    - statuses  → sent/delivered/read/failed durumlarını DB'ye kaydeder
    """
    try:
        body = await request.json()
    except Exception:
        return Response(status_code=200)

    logger.info(f"📩 Meta Webhook gövdesi: {body}")

    db = await get_db_from_request(request)

    try:
        for entry in body.get("entry", []):
            for change in entry.get("changes", []):
                value = change.get("value", {})

                # ── 1. Gelen müşteri mesajları ──────────────────────────────
                for message in value.get("messages", []):
                    from_number = message.get("from")
                    msg_type    = message.get("type", "")
                    if not from_number or msg_type not in (
                        "text", "audio", "image", "video", "document",
                        "sticker", "location", "contacts", "interactive", "button",
                    ):
                        continue
                    logger.info(f"📩 Gelen WA mesajı: from={from_number} type={msg_type}")

                    # ── Doğrulama kodu kontrolü (6 haneli sayı) ──────────────────────────
                    if msg_type == "text":
                        text_body = message.get("text", {}).get("body", "").strip()
                        rnd_match = re.search(r"\b\d{6}\b", text_body)
                        if rnd_match:
                            redis_client_wh = getattr(request.app.state, "redis_client", None)
                            result = await _process_whatsapp_verification(
                                db, redis_client_wh, from_number, rnd_match.group(0)
                            )
                            if result.get("reply"):
                                _send_whatsapp_text_reply(from_number, result["reply"])
                            continue

                    # ── Varsayılan otomatik yanıt ────────────────────────────
                    lang = detect_language_from_phone(from_number)
                    if lang == "TR":
                        reply = (
                            "Bu numara sadece bilgilendirme amaçlıdır, "
                            "lütfen doğrudan işletmeyle iletişime geçin."
                        )
                    else:
                        reply = (
                            "This number is for information purposes only, "
                            "please contact the business directly."
                        )
                    _send_whatsapp_text_reply(from_number, reply)

                # ── 2. Mesaj durumu güncellemeleri → DB'ye kaydet ────────────
                for status_obj in value.get("statuses", []):
                    message_id   = status_obj.get("id")
                    recipient    = status_obj.get("recipient_id")
                    wa_status    = status_obj.get("status")        # sent/delivered/read/failed
                    timestamp_ms = status_obj.get("timestamp")
                    errors       = status_obj.get("errors", [])

                    if wa_status == "failed" and errors:
                        logger.error(
                            f"❌ WA mesaj gönderilemedi | id={message_id} "
                            f"to={recipient} errors={errors}"
                        )
                    else:
                        logger.info(
                            f"📬 WA durum: {wa_status} | id={message_id} to={recipient}"
                        )

                    if db is not None:
                        try:
                            log_doc = {
                                "message_id": message_id,
                                "recipient":  recipient,
                                "status":     wa_status,
                                "timestamp":  int(timestamp_ms) if timestamp_ms else None,
                                "errors":     errors,
                                "recorded_at": datetime.now(timezone.utc).isoformat(),
                            }
                            await db.whatsapp_message_logs.update_one(
                                {"message_id": message_id, "status": wa_status},
                                {"$set": log_doc},
                                upsert=True,
                            )
                        except Exception as db_err:
                            logger.error(f"WA log DB kayıt hatası: {db_err}")

    except Exception as e:
        logger.error(f"Webhook işleme hatası: {e}")

    return Response(status_code=200)


@app.post("/api/webhook")
async def whatsapp_webhook_api_app(request: Request):
    return await whatsapp_webhook(request)


@api_router.post("/webhook")
async def whatsapp_webhook_api(request: Request):
    return await whatsapp_webhook(request)

# Static files serving for logos (must be after router)
static_files_dir = str(ROOT_DIR / "static")
app.mount("/api/static", StaticFiles(directory=static_files_dir), name="static")

# --- CORS Ayarı ---
cors_origins_str = os.environ.get('CORS_ORIGINS', '*')
if cors_origins_str == '*':
    cors_origins = ['*']
else:
    # Virgülle ayrılmış origin'leri parse et ve boşlukları temizle
    cors_origins = [origin.strip() for origin in cors_origins_str.split(',') if origin.strip()]
logging.info(f"CORS origins configured: {cors_origins}")
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=cors_origins,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"], 
    allow_headers=["*"],
)

# Export socket_app as the main application for ASGI servers
# This allows both FastAPI and Socket.IO to work together
application = socket_app

async def run_stripe_audit():
    """
    COMPREHENSIVE STRIPE AUDIT
    Verifies that every plan, cycle, and currency combination exists in Stripe
    AND has the correct appointment_limit in its metadata.
    """
    if not STRIPE_SECRET_KEY:
        logger.error("❌ STRIPE_SECRET_KEY tanımlı değil! Audit çalıştırılamıyor.")
        return
    
    # Expected Source of Truth
    EXPECTED_CONFIG = {
        "standart": 100,  # Not: "standard" değil, "standart" (Türkçe)
        "professional": 300,
        "premium": 600,
        "business": 900,
        "enterprise": 1200,
        "corporate": -1
    }
    CYCLES = ["monthly", "yearly"]
    CURRENCIES = ["gbp", "try"]
    
    logger.info("=" * 80)
    logger.info("🔍 STRIPE COMPREHENSIVE AUDIT - Starting...")
    logger.info("=" * 80)
    
    results = {
        "pass": [],
        "fail": [],
        "warn": []
    }
    
    total_checks = 0
    
    # Iterate through every Plan x Cycle x Currency combination
    for plan_name, expected_limit in EXPECTED_CONFIG.items():
        for cycle in CYCLES:
            for currency in CURRENCIES:
                total_checks += 1
                lookup_key = f"{plan_name}_{cycle}_{currency}"
                
                try:
                    # Call Stripe API with product expansion for fallback
                    prices = stripe.Price.list(
                        lookup_keys=[lookup_key],
                        active=True,
                        limit=1,
                        expand=['data.product']
                    )
                    
                    # Check 1: Existence
                    if not prices.data or len(prices.data) == 0:
                        error_msg = f"Price not found in Stripe"
                        results["fail"].append({
                            "lookup_key": lookup_key,
                            "error": error_msg,
                            "expected_limit": expected_limit
                        })
                        logger.error(f"[❌ FAIL] {lookup_key:35} | Expected Limit: {expected_limit:4} | Error: {error_msg}")
                        continue
                    
                    price = prices.data[0]
                    
                    # Check 2: Metadata Presence (FALLBACK STRATEGY)
                    # Step 1: Check Price metadata
                    price_metadata = price.metadata or {}
                    appointment_limit = price_metadata.get('appointment_limit')
                    metadata_source = "Price"
                    
                    # Step 2: If not found in price, check product metadata
                    if not appointment_limit:
                        product_obj = price.product
                        if product_obj:
                            # product_obj can be a string (ID) or dict (expanded)
                            if isinstance(product_obj, dict):
                                product_metadata = product_obj.get('metadata', {})
                                appointment_limit = product_metadata.get('appointment_limit')
                                if appointment_limit:
                                    metadata_source = "Product"
                                    logger.info(f"✅ Found appointment_limit in Product metadata (fallback): {appointment_limit}")
                            elif isinstance(product_obj, str):
                                # Product not expanded, retrieve it
                                try:
                                    product_retrieved = stripe.Product.retrieve(product_obj)
                                    product_metadata = _stripe_metadata_dict(product_retrieved)
                                    appointment_limit = product_metadata.get('appointment_limit')
                                    if appointment_limit:
                                        metadata_source = "Product"
                                        logger.info(f"✅ Found appointment_limit in Product metadata (fallback, retrieved): {appointment_limit}")
                                except Exception as e:
                                    logger.warning(f"⚠️ Product retrieve hatası: {e}")
                    
                    if not appointment_limit:
                        error_msg = f"Metadata 'appointment_limit' missing in both Price and Product"
                        results["warn"].append({
                            "lookup_key": lookup_key,
                            "error": error_msg,
                            "expected_limit": expected_limit,
                            "price_id": price.id,
                            "currency": price.currency.upper()
                        })
                        logger.warning(f"[⚠️ WARN] {lookup_key:35} | Expected Limit: {expected_limit:4} | Error: {error_msg} | Price ID: {price.id}")
                        continue
                    
                    # Check 3: Metadata Accuracy
                    try:
                        actual_limit = int(appointment_limit)
                    except (ValueError, TypeError):
                        error_msg = f"Metadata 'appointment_limit' is not a valid integer: '{appointment_limit}'"
                        results["fail"].append({
                            "lookup_key": lookup_key,
                            "error": error_msg,
                            "expected_limit": expected_limit,
                            "price_id": price.id
                        })
                        logger.error(f"[❌ FAIL] {lookup_key:35} | Expected Limit: {expected_limit:4} | Error: {error_msg} | Price ID: {price.id}")
                        continue
                    
                    # Yearly plans must also have the MONTHLY limit (not multiplied by 12)
                    if actual_limit == expected_limit:
                        results["pass"].append({
                            "lookup_key": lookup_key,
                            "expected_limit": expected_limit,
                            "actual_limit": actual_limit,
                            "price_id": price.id,
                            "currency": price.currency.upper(),
                            "amount": price.unit_amount / 100,
                            "metadata_source": metadata_source
                        })
                        logger.info(f"[✅ PASS] {lookup_key:35} | Limit: {actual_limit:4} (Expected: {expected_limit:4}) | Source: {metadata_source:7} | {price.currency.upper():4} | {price.unit_amount/100:8.2f} | Price ID: {price.id}")
                    else:
                        error_msg = f"Limit mismatch: Expected {expected_limit}, Found {actual_limit}"
                        results["fail"].append({
                            "lookup_key": lookup_key,
                            "error": error_msg,
                            "expected_limit": expected_limit,
                            "actual_limit": actual_limit,
                            "price_id": price.id
                        })
                        logger.error(f"[❌ FAIL] {lookup_key:35} | Expected Limit: {expected_limit:4} | Actual Limit: {actual_limit:4} | Error: {error_msg} | Price ID: {price.id}")
                
                except stripe.error.StripeError as e:
                    error_msg = f"Stripe API error: {str(e)}"
                    results["fail"].append({
                        "lookup_key": lookup_key,
                        "error": error_msg,
                        "expected_limit": expected_limit
                    })
                    logger.error(f"[❌ FAIL] {lookup_key:35} | Expected Limit: {expected_limit:4} | Error: {error_msg}")
                except Exception as e:
                    error_msg = f"Unexpected error: {str(e)}"
                    results["fail"].append({
                        "lookup_key": lookup_key,
                        "error": error_msg,
                        "expected_limit": expected_limit
                    })
                    logger.error(f"[❌ FAIL] {lookup_key:35} | Expected Limit: {expected_limit:4} | Error: {error_msg}")
    
    # Summary
    logger.info("=" * 80)
    logger.info("📊 AUDIT SUMMARY")
    logger.info("=" * 80)
    logger.info(f"Total Checks: {total_checks}")
    logger.info(f"✅ PASS: {len(results['pass'])}")
    logger.info(f"⚠️  WARN: {len(results['warn'])}")
    logger.info(f"❌ FAIL: {len(results['fail'])}")
    logger.info("=" * 80)
    
    if results["fail"]:
        logger.error("\n❌ FAILED ITEMS:")
        for item in results["fail"]:
            logger.error(f"   - {item['lookup_key']}: {item['error']}")
    
    if results["warn"]:
        logger.warning("\n⚠️  WARNINGS:")
        for item in results["warn"]:
            logger.warning(f"   - {item['lookup_key']}: {item['error']}")
    
    if len(results["pass"]) == total_checks:
        logger.info("\n✅ ALL CHECKS PASSED! Stripe configuration is perfect!")
    else:
        logger.warning(f"\n⚠️  {len(results['fail']) + len(results['warn'])} issues found. Please fix them in Stripe Dashboard.")
    
    logger.info("=" * 80)
    
    return results
